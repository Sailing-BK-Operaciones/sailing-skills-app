"""
Risk Monitoring Client — versión web (Streamlit).
Recibe todos los archivos como UploadedFile; devuelve BytesIO con el Excel.
"""

import csv
import re
from io import BytesIO
from datetime import datetime

import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


def generar_reporte(
    csv_grouping_file,
    csv_accounts_file,
    pc_file,
    especies_file,
    sagaclte_file,
) -> BytesIO:
    # ── Tabla de cuentas ─────────────────────────────────────────────────────
    accounts_map = {}
    content = csv_accounts_file.read().decode("utf-8-sig")
    for line in content.splitlines():
        line = line.strip().strip('"')
        parts = line.split(";")
        if len(parts) != 3 or parts[0] == "AGENTE":
            continue
        agente, ctte, id_clearing = parts[0].strip(), parts[1].strip(), parts[2].strip()
        if id_clearing:
            accounts_map[id_clearing] = {"comitente": ctte, "name": ""}

    # ── Datos de riesgo ───────────────────────────────────────────────────────
    risk_data = []
    decoded = csv_grouping_file.read().decode("utf-8-sig").splitlines()
    reader = csv.DictReader(decoded)
    for row in reader:
        acct_id = str(row["ACCOUNT ID"]).strip()
        por_risk = float(row["Portfolio Risk"] or 0)
        por_val  = float(row["Portfolio Value"] or 0)
        var_mg   = float(row["Variation Margin"] or 0)
        currency = str(row["Currency"]).strip()
        deficit  = por_risk - por_val - var_mg
        info = accounts_map.get(acct_id, {})
        risk_data.append({
            "account_id":       acct_id,
            "comitente":        info.get("comitente", ""),
            "portfolio_risk":   por_risk,
            "portfolio_value":  por_val,
            "variation_margin": var_mg,
            "deficit":          deficit,
            "currency":         currency,
        })

    risk_data.sort(key=lambda x: (-x["deficit"], x["account_id"]))

    grand_total    = sum(r["deficit"]          for r in risk_data)
    total_risk     = sum(r["portfolio_risk"]   for r in risk_data)
    total_garantias_preview = 0.0   # se calcula tras cargar SAGACLTE

    # ── Fecha desde nombre de archivo ─────────────────────────────────────────
    ts_match = re.search(r"(\d{8})-(\d{6})", csv_grouping_file.name)
    if ts_match:
        fecha_dt = datetime.strptime(ts_match.group(1), "%Y%m%d")
        hora_fmt = "{0}:{1}:{2}".format(*[ts_match.group(2)[i:i+2] for i in (0, 2, 4)])
    else:
        fecha_dt = datetime.today()
        hora_fmt = datetime.now().strftime("%H:%M:%S")

    fecha_display = fecha_dt.strftime("%d/%m/%Y")
    fecha_output  = fecha_dt.strftime("%d-%m-%Y")

    # ── Precios de cierre (PC*.XLS) ───────────────────────────────────────────
    precios = {}
    wb_pc = xlrd.open_workbook(file_contents=pc_file.read())
    ws_pc = (wb_pc.sheet_by_name("Precios_de_Cierre")
             if "Precios_de_Cierre" in wb_pc.sheet_names()
             else wb_pc.sheet_by_index(0))
    for r in range(1, ws_pc.nrows):
        especie_str = str(ws_pc.cell_value(r, 0)).strip()
        code5 = especie_str[:5].strip().lstrip("'").zfill(5)
        price = float(ws_pc.cell_value(r, 1) or 0)
        if code5:
            precios[code5] = price

    # ── Tipos de precio y aforos BYMA (ESPECIES.XLS) ─────────────────────────
    # Col 26 (Aforo) = haircut BYMA API; aforo_byma = (100 - haircut) / 100.0
    tipo_precio_map = {}
    byma_aforos     = {}
    wb_esp = xlrd.open_workbook(file_contents=especies_file.read())
    ws_esp = wb_esp.sheet_by_name("Datos_Fijos_Especies")
    for r in range(1, ws_esp.nrows):
        row   = ws_esp.row_values(r)
        code  = str(row[0]).strip().lstrip("'").zfill(5)
        if not code:
            continue
        tipo  = str(row[4]).strip()
        tipo_precio_map[code] = tipo
        try:
            haircut = int(float(row[26])) if len(row) > 26 and row[26] else 0
        except (ValueError, TypeError):
            haircut = 0
        if haircut > 0:
            byma_aforos[code] = (100 - haircut) / 100.0

    # ── Garantías por comitente (SAGACLTE.XLS) ────────────────────────────────
    garantias_by_ctte = {}
    wb_saga = xlrd.open_workbook(file_contents=sagaclte_file.read())
    ws_saga = wb_saga.sheet_by_name("Saldos_de_Garantias")

    for r in range(1, ws_saga.nrows):
        ctte_raw = ws_saga.cell_value(r, 0)
        if ctte_raw == "" or ctte_raw is None:
            continue
        ctte = (
            str(int(float(str(ctte_raw).lstrip("'"))))
            if str(ctte_raw).lstrip("'").replace(".", "", 1).isdigit()
            else str(ctte_raw).strip().lstrip("'")
        )
        code_raw = str(ws_saga.cell_value(r, 2)).strip().lstrip("'")
        code5    = code_raw.zfill(5) if code_raw else ""
        cantidad = float(ws_saga.cell_value(r, 6) or 0)

        if not ctte or not code5 or cantidad <= 0:
            continue

        precio = precios.get(code5, 0.0)
        tipo   = tipo_precio_map.get(code5, "Normal")
        aforo  = byma_aforos.get(code5, 0.0)

        if precio <= 0 or aforo <= 0:
            continue

        monto = cantidad * precio / 100.0 if tipo.lower().startswith("porc") else cantidad * precio
        garantias_by_ctte[ctte] = garantias_by_ctte.get(ctte, 0.0) + monto * aforo

    for row in risk_data:
        row["garantias"] = garantias_by_ctte.get(row["comitente"], 0.0)

    total_garantias = sum(r["garantias"] for r in risk_data)

    # ── Estilos ───────────────────────────────────────────────────────────────
    C_DARK_BLUE  = "1F4E79"
    C_MED_BLUE   = "2E75B6"
    C_LIGHT_BLUE = "D6E4F7"
    C_ALT_ROW    = "EBF3FB"
    C_WHITE      = "FFFFFF"
    C_SUBTOTAL   = "D0E4F5"
    C_RED_FILL   = "FCE4D6"
    C_GREEN_FILL = "E2EFDA"
    C_ORANGE     = "C55A11"
    C_GREEN_TXT  = "375623"
    C_RED_TXT    = "9C0006"
    C_HEADER_FNT = "FFFFFF"
    C_GARA_HDR   = "375623"
    NUM_FMT      = "#,##0.00"
    PCT_FMT      = "0.00%"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border_thin(color="BFBFBF"):
        s = Side(style="thin", color=color)
        return Border(bottom=s)

    def border_all(color="BFBFBF"):
        s = Side(style="thin", color=color)
        return Border(top=s, bottom=s, left=s, right=s)

    # ── Libro Excel ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Monitoring"

    HEADER_ROW   = 6
    SUBTOTAL_ROW = 7
    DATA_START   = 8
    n_data       = len(risk_data)
    DATA_END     = DATA_START + n_data - 1

    COLS = [
        "Comitente",
        "Account ID (BYMA)",
        "Portfolio Risk",
        "Portfolio Value",
        "Variation Margin",
        "Total Margin Deficit",
        "Moneda",
        "",
        "Estado",
        "Garantias integradas a aforo BYMA",
        "Margen de cobertura",
        "% Margen de cobertura",
    ]
    COL_WIDTHS = [13, 20, 22, 22, 22, 24, 10, 3, 12, 28, 24, 20]
    N_COLS   = len(COLS)
    last_col = get_column_letter(N_COLS)

    # Fila 1: título
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = "Risk Monitoring Client — BYMA Clearing"
    c.font      = Font(bold=True, color=C_HEADER_FNT, size=16)
    c.fill      = fill(C_MED_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Fila 2: fecha/hora
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value     = f"Fecha: {fecha_display}   |   Hora descarga: {hora_fmt}"
    c.font      = Font(color=C_DARK_BLUE, size=10)
    c.fill      = fill(C_LIGHT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6

    # Fila 4: resumen total
    ws.row_dimensions[4].height = 26
    ws.merge_cells("A4:C4")
    c = ws["A4"]
    c.value     = "TOTAL MARGIN DEFICIT (ARS)"
    c.font      = Font(bold=True, color=C_HEADER_FNT, size=12)
    c.fill      = fill(C_ORANGE if grand_total > 0 else C_GREEN_TXT)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("D4:F4")
    c = ws["D4"]
    c.value         = f"=F{SUBTOTAL_ROW}"
    c.number_format = NUM_FMT
    c.font          = Font(bold=True, size=13, color=C_RED_TXT if grand_total > 0 else C_GREEN_TXT)
    c.alignment     = Alignment(horizontal="right", vertical="center")
    c.fill          = fill(C_RED_FILL if grand_total > 0 else C_GREEN_FILL)

    for col in [7, 8, 9]:
        ws.cell(row=4, column=col).fill = fill(C_WHITE if col == 8 else (C_RED_FILL if grand_total > 0 else C_GREEN_FILL))

    c = ws["J4"]
    c.value         = f"=J{SUBTOTAL_ROW}"
    c.number_format = NUM_FMT
    c.font          = Font(bold=True, size=12, color=C_HEADER_FNT)
    c.fill          = fill(C_GARA_HDR)
    c.alignment     = Alignment(horizontal="right", vertical="center")

    c = ws["K4"]
    c.value         = f"=K{SUBTOTAL_ROW}"
    c.number_format = NUM_FMT
    c.font          = Font(bold=True, size=12,
                           color=C_RED_TXT if (grand_total - total_garantias) > 0 else C_GREEN_TXT)
    c.fill          = fill(C_RED_FILL if (grand_total - total_garantias) > 0 else C_GREEN_FILL)
    c.alignment     = Alignment(horizontal="right", vertical="center")

    c = ws["L4"]
    c.value         = f"=IF(C{SUBTOTAL_ROW}=0,\"\",J{SUBTOTAL_ROW}/C{SUBTOTAL_ROW})"
    c.number_format = PCT_FMT
    c.font          = Font(bold=True, size=12)
    c.fill          = fill(C_LIGHT_BLUE)
    c.alignment     = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[5].height = 6

    # Fila 6: encabezados
    for ci, col_name in enumerate(COLS, 1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=col_name)
        if ci == 8:
            c.fill = fill(C_WHITE)
            continue
        hdr_color = C_GARA_HDR if ci >= 9 else C_DARK_BLUE
        c.font      = Font(bold=True, color=C_HEADER_FNT, size=10)
        c.fill      = fill(hdr_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border_all(hdr_color)
    ws.row_dimensions[HEADER_ROW].height = 36

    # Fila 7: SUBTOTAL
    ws.cell(row=SUBTOTAL_ROW, column=1, value="SUBTOTAL").font = Font(bold=True, size=10)
    ws.cell(row=SUBTOTAL_ROW, column=1).fill      = fill(C_SUBTOTAL)
    ws.cell(row=SUBTOTAL_ROW, column=1).alignment = Alignment(horizontal="center")
    for col in [2, 7, 9]:
        ws.cell(row=SUBTOTAL_ROW, column=col).fill = fill(C_SUBTOTAL)
    ws.cell(row=SUBTOTAL_ROW, column=8).fill = fill(C_WHITE)

    for num_col in [3, 4, 5, 6]:
        col_ltr = get_column_letter(num_col)
        c = ws.cell(row=SUBTOTAL_ROW, column=num_col,
                    value=f"=SUBTOTAL(9,{col_ltr}{DATA_START}:{col_ltr}{DATA_END})")
        c.number_format = NUM_FMT
        c.font          = Font(bold=True, size=10)
        c.fill          = fill(C_SUBTOTAL)
        c.alignment     = Alignment(horizontal="right")

    c = ws.cell(row=SUBTOTAL_ROW, column=10,
                value=f"=SUBTOTAL(9,J{DATA_START}:J{DATA_END})")
    c.number_format = NUM_FMT; c.font = Font(bold=True, size=10)
    c.fill = fill(C_SUBTOTAL); c.alignment = Alignment(horizontal="right")

    c = ws.cell(row=SUBTOTAL_ROW, column=11,
                value=f"=SUBTOTAL(9,K{DATA_START}:K{DATA_END})")
    c.number_format = NUM_FMT; c.font = Font(bold=True, size=10)
    c.fill = fill(C_SUBTOTAL); c.alignment = Alignment(horizontal="right")

    c = ws.cell(row=SUBTOTAL_ROW, column=12,
                value=f"=IF(C{SUBTOTAL_ROW}=0,\"\",J{SUBTOTAL_ROW}/C{SUBTOTAL_ROW})")
    c.number_format = PCT_FMT; c.font = Font(bold=True, size=10)
    c.fill = fill(C_SUBTOTAL); c.alignment = Alignment(horizontal="right")

    ws.row_dimensions[SUBTOTAL_ROW].height = 20

    # Filas de datos
    for i, row in enumerate(risk_data):
        r = DATA_START + i
        row_fill = fill(C_ALT_ROW if i % 2 == 0 else C_WHITE)

        def cell(col, value, fmt=None, bold=False, color=None, align="right", bord=True):
            c = ws.cell(row=r, column=col, value=value)
            if fmt:
                c.number_format = fmt
            c.font      = Font(bold=bold, size=10, color=color or "000000")
            c.alignment = Alignment(horizontal=align)
            c.fill      = row_fill
            if bord:
                c.border = border_thin()
            return c

        cell(1, row["comitente"],        align="center")
        cell(2, row["account_id"],       align="center")
        cell(3, row["portfolio_risk"],   fmt=NUM_FMT)
        cell(4, row["portfolio_value"],  fmt=NUM_FMT)
        cell(5, row["variation_margin"], fmt=NUM_FMT)

        deficit = row["deficit"]
        if deficit > 0:
            def_fill = fill(C_RED_FILL)
            def_font = Font(bold=True, size=10, color=C_RED_TXT)
        elif deficit < 0:
            def_fill = fill(C_GREEN_FILL)
            def_font = Font(bold=True, size=10, color=C_GREEN_TXT)
        else:
            def_fill = row_fill
            def_font = Font(size=10)
        c6 = ws.cell(row=r, column=6, value=deficit)
        c6.number_format = NUM_FMT; c6.alignment = Alignment(horizontal="right")
        c6.fill = def_fill; c6.font = def_font; c6.border = border_thin()

        cell(7, row["currency"], align="center")
        ws.cell(row=r, column=8).fill = fill(C_WHITE)

        c9 = ws.cell(row=r, column=9, value=f'=IF(J{r}>F{r},"CUBRE","NO CUBRE")')
        c9.alignment = Alignment(horizontal="center")
        c9.font = Font(bold=True, size=10); c9.border = border_thin()

        c10 = ws.cell(row=r, column=10, value=row["garantias"])
        c10.number_format = NUM_FMT; c10.alignment = Alignment(horizontal="right")
        c10.fill = row_fill; c10.font = Font(size=10); c10.border = border_thin()

        c11 = ws.cell(row=r, column=11, value=f"=F{r}-J{r}")
        c11.number_format = NUM_FMT; c11.alignment = Alignment(horizontal="right")
        c11.font = Font(bold=True, size=10); c11.border = border_thin()

        c12 = ws.cell(row=r, column=12, value=f'=IF(C{r}=0,"",J{r}/C{r})')
        c12.number_format = PCT_FMT; c12.alignment = Alignment(horizontal="right")
        c12.fill = row_fill; c12.font = Font(size=10); c12.border = border_thin()

    # Formato condicional columna I
    red_cf   = PatternFill(start_color=C_RED_FILL,   end_color=C_RED_FILL,   fill_type="solid")
    green_cf = PatternFill(start_color=C_GREEN_FILL, end_color=C_GREEN_FILL, fill_type="solid")
    i_range  = f"I{DATA_START}:I{DATA_END}"
    ws.conditional_formatting.add(i_range, CellIsRule(
        operator="equal", formula=['"CUBRE"'],
        font=Font(bold=True, color=C_GREEN_TXT), fill=green_cf))
    ws.conditional_formatting.add(i_range, CellIsRule(
        operator="equal", formula=['"NO CUBRE"'],
        font=Font(bold=True, color=C_RED_TXT), fill=red_cf))

    # Formato condicional columna K
    k_range = f"K{DATA_START}:K{DATA_END}"
    ws.conditional_formatting.add(k_range, CellIsRule(
        operator="greaterThan", formula=["0"],
        font=Font(bold=True, color=C_RED_TXT), fill=red_cf))
    ws.conditional_formatting.add(k_range, CellIsRule(
        operator="lessThanOrEqual", formula=["0"],
        font=Font(bold=True, color=C_GREEN_TXT), fill=green_cf))

    ws.auto_filter.ref = f"A{HEADER_ROW}:L{DATA_END}"
    ws.freeze_panes   = f"A{DATA_START}"
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, fecha_output, n_data, grand_total, total_garantias
