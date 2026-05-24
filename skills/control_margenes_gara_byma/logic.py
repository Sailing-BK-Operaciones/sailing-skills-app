"""
Control Márgenes Gara BYMA
Valoriza posiciones en garantía por comitente, detecta VTAs y disponibles,
compara contra requerimiento y determina CUBIERTO / DESCUBIERTO.
Adaptado de run_control_posiciones.py para funcionar con archivos en memoria.
"""

import io
import re
import csv
from collections import OrderedDict, defaultdict
from itertools import groupby
from datetime import date, datetime

import xlrd
import openpyxl
import pdfplumber
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Estilos ─────────────────────────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
HEADER_FONT    = Font(color="FFFFFF", bold=True)
GREEN_FILL     = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT     = Font(color="276221", bold=True)
YELLOW_FILL    = PatternFill("solid", fgColor="FFEB9C")
YELLOW_FONT    = Font(color="9C6500", bold=True)
GRAY_FILL      = PatternFill("solid", fgColor="D9D9D9")
GRAY_FONT      = Font(color="595959", italic=True)
RED_FILL       = PatternFill("solid", fgColor="FFC7CE")
RED_FONT       = Font(color="9C0006", bold=True)
ORANGE_FILL    = PatternFill("solid", fgColor="FCE4D6")
ALT_FILL       = PatternFill("solid", fgColor="F2F2F2")
BOLD_FONT      = Font(bold=True)
VTA_FILL       = PatternFill("solid", fgColor="FFEB9C")
VTA_FONT       = Font(color="9C6500", bold=True)
CI_COMPRA_FILL = PatternFill("solid", fgColor="E8DAEF")
CI_COMPRA_FONT = Font(color="6C3483", bold=True)
SUB_FILL       = PatternFill("solid", fgColor="D6E4F0")
SUB_FONT       = Font(bold=True, color="1F3864")

FMT_INT   = '#,##0'
FMT_PRICE = '#,##0.00'
FMT_PCT   = '0.00"%"'


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header(ws, row_idx, col_idx, value):
    cell = ws.cell(row=row_idx, column=col_idx, value=value)
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    return cell


def _parse_fecha_cell(raw_val, datemode):
    """Convierte valor de celda (serial Excel o string) a date, o None."""
    if not raw_val and raw_val != 0:
        return None
    try:
        fval = float(raw_val)
        if fval > 1:
            tup = xlrd.xldate_as_tuple(fval, datemode)
            return date(tup[0], tup[1], tup[2])
    except Exception:
        pass
    s = str(raw_val).strip()
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):   # %y primero: "01/04/26" → 2026
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def generar_control(
    saldos_file,      # SALDOS DEUDORES.xlsx  — obligatorio
    contbole_file,    # CONTBOLE.XLS           — obligatorio
    sagaclte_file,    # SAGACLTE.XLS           — shared
    sateclte_file,    # SATECLTE.XLS           — shared
    especies_file,    # ESPECIES.XLS           — shared
    tabcompb_file,    # TABCOMPB.XLS           — shared
    pc_file,          # PC*.XLS                — shared
    pdf_aforos_file,  # PDF aforos BYMA        — shared
    accounts_file,    # table-accounts_*.csv   — shared
):
    """
    Retorna
    -------
    xlsx_bytes  : bytes
    resumen     : dict  (conteos y estados para mostrar en la UI)
    advertencias: list[str]
    """
    advertencias = []
    FECHA_PROCESO = date.today()
    fecha_str     = FECHA_PROCESO.strftime("%d-%m-%Y")

    # ── 1. ESPECIES.XLS ───────────────────────────────────────────────────────
    especies_file.seek(0)
    wb_esp = xlrd.open_workbook(file_contents=especies_file.read())
    hoja_esp = ("Datos_Fijos_Especies"
                if "Datos_Fijos_Especies" in wb_esp.sheet_names()
                else wb_esp.sheet_names()[0])
    sh_esp = wb_esp.sheet_by_name(hoja_esp)
    especies = {}
    for r in range(1, sh_esp.nrows):
        row  = sh_esp.row_values(r)
        cod  = str(row[0]).strip().zfill(5)
        nom  = str(row[1]).strip()
        tprc = str(row[4]).strip()
        nemo = str(row[9]).strip()
        try:
            lam_min = float(str(row[6]).strip()) if row[6] else 0.0
        except (ValueError, TypeError):
            lam_min = 0.0
        especies[cod] = {"ticker": nemo or cod, "nombre": nom,
                         "tipo_precio": tprc, "lam_min": lam_min}

    # ── 2. TABCOMPB.XLS ───────────────────────────────────────────────────────
    op_tipo = {}  # abrev -> 'COMPRA' | 'VENTA'
    try:
        tabcompb_file.seek(0)
        wb_tab = xlrd.open_workbook(file_contents=tabcompb_file.read())
        hoja_tab = ("Tabla_Comprobantes"
                    if "Tabla_Comprobantes" in wb_tab.sheet_names()
                    else wb_tab.sheet_names()[0])
        sh_tab = wb_tab.sheet_by_name(hoja_tab)
        for r in range(1, sh_tab.nrows):
            row_t  = sh_tab.row_values(r)
            nombre = str(row_t[1]).upper()
            abrev  = str(row_t[2]).strip()
            if "COMPRA" in nombre:
                op_tipo[abrev] = "COMPRA"
            elif "VENTA" in nombre:
                op_tipo[abrev] = "VENTA"
    except Exception as e:
        advertencias.append(f"No se pudo leer TABCOMPB.XLS: {e}")

    # ── 3. PDF aforos BYMA ────────────────────────────────────────────────────
    byma_dict = {}
    pdf_aforos_file.seek(0)
    all_text = []
    with pdfplumber.open(pdf_aforos_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                all_text.append(text)
    full_pdf = "\n".join(all_text)
    pat_pdf = re.compile(r"([A-Za-z0-9][A-Za-z0-9._-]*)\s+(\d{2,6})\s+(\d+)%\s+\d+%")
    for m in pat_pdf.finditer(full_pdf):
        codigo = m.group(2).zfill(5)
        aforo  = int(m.group(3)) / 100.0
        byma_dict[codigo] = (m.group(1), aforo)
    if not byma_dict:
        advertencias.append("No se extrajeron especies del PDF de aforos. Verificar formato del PDF.")

    # ── 4. PC*.XLS ────────────────────────────────────────────────────────────
    pc_file.seek(0)
    wb_pc = xlrd.open_workbook(file_contents=pc_file.read())
    hoja_pc = ("Precios_de_Cierre"
               if "Precios_de_Cierre" in wb_pc.sheet_names()
               else wb_pc.sheet_names()[0])
    sh_pc = wb_pc.sheet_by_name(hoja_pc)
    precios = {}
    for r in range(1, sh_pc.nrows):
        row = sh_pc.row_values(r)
        raw = str(row[0]).strip()
        m   = re.match(r"(\d{5})", raw)
        if m:
            cod    = m.group(1).zfill(5)
            precio = float(row[1]) if row[1] else 0.0
            precios[cod] = precio

    # ── 5. SAGACLTE.XLS ───────────────────────────────────────────────────────
    sagaclte_file.seek(0)
    wb_saga = xlrd.open_workbook(file_contents=sagaclte_file.read())
    hoja_saga = ("Saldos_de_Garantias"
                 if "Saldos_de_Garantias" in wb_saga.sheet_names()
                 else wb_saga.sheet_names()[0])
    sh_saga = wb_saga.sheet_by_name(hoja_saga)
    saga_by_ctte = {}
    for r in range(1, sh_saga.nrows):
        row = sh_saga.row_values(r)
        if not row[0]:
            continue
        agente = str(row[5]).strip()
        if agente:
            continue   # solo posiciones directas BYMA (agente vacío)
        ctte        = str(int(float(row[0])))
        cod         = str(row[2]).strip().zfill(5)
        qty         = float(row[6]) if row[6] else 0.0
        nombre_ctte = str(row[1]).strip()
        if qty <= 0:
            continue
        raw_liq   = row[7] if len(row) > 7 else ""
        fecha_liq = _parse_fecha_cell(raw_liq, wb_saga.datemode)
        if ctte not in saga_by_ctte:
            saga_by_ctte[ctte] = {"nombre": nombre_ctte, "posiciones": []}
        saga_by_ctte[ctte]["posiciones"].append(
            {"codigo": cod, "qty": qty, "fecha_liq": fecha_liq}
        )

    # ── 6. SATECLTE.XLS ───────────────────────────────────────────────────────
    sateclte_ten = {}   # ctte -> {"vtas": {}, "disponibles": {}, "cpr": {}}

    def _sat_is_cpra(d):
        u = d.upper()
        return "CPRA" in u or "CPU$" in u or "CCFP" in u

    def _sat_is_vtas(d):
        u = d.upper()
        return "VTAS" in u or "VTU$" in u or "VTTR" in u

    def _sat_add(dest, ctte, field, cod, qty):
        dest.setdefault(ctte, {"vtas": {}, "disponibles": {}, "cpr": {}})
        dest[ctte][field][cod] = dest[ctte][field].get(cod, 0.0) + qty

    sateclte_file.seek(0)
    wb_sat = xlrd.open_workbook(file_contents=sateclte_file.read())
    hoja_sat = ("Saldos_de_Tenencia"
                if "Saldos_de_Tenencia" in wb_sat.sheet_names()
                else wb_sat.sheet_names()[0])
    sh_sat = wb_sat.sheet_by_name(hoja_sat)

    _sat_rows = []
    _cur_ctte_sat = ""
    for r in range(1, sh_sat.nrows):
        row = sh_sat.row_values(r)
        if row[0]:
            _cur_ctte_sat = str(int(float(str(row[0]).strip())))
        _sat_rows.append((_cur_ctte_sat, str(row[2]).strip(), row[3], row[4], row[5]))

    _last_code_sat = {}
    for i, (ctte, desc, cant_t, cant_cv, total) in enumerate(_sat_rows):
        m_cod = re.match(r'^(\d{5})\s', desc)
        if m_cod:
            cod = m_cod.group(1)
            _last_code_sat[ctte] = cod
            vn = float(total) if total and float(total) > 0 else (float(cant_cv) if cant_cv else 0.0)
            if vn > 0:
                _sat_add(sateclte_ten, ctte, "disponibles", cod, vn)
            continue
        if "GTIA" in desc.upper() or desc.upper() == "TOTAL":
            continue
        qty_cv = float(cant_cv) if cant_cv else 0.0
        if qty_cv == 0.0:
            continue
        if _sat_is_cpra(desc) or _sat_is_vtas(desc):
            cod = _last_code_sat.get(ctte)
            if cod is None:
                for j in range(i + 1, len(_sat_rows)):
                    if _sat_rows[j][0] != ctte:
                        break
                    m_la = re.match(r'^(\d{5})\s', _sat_rows[j][1])
                    if m_la:
                        cod = m_la.group(1)
                        _last_code_sat[ctte] = cod
                        break
            if cod is None:
                continue
            if _sat_is_cpra(desc) and qty_cv > 0:
                _sat_add(sateclte_ten, ctte, "cpr", cod, qty_cv)
            elif _sat_is_vtas(desc):
                _sat_add(sateclte_ten, ctte, "vtas", cod, abs(qty_cv))

    # ── 7. table-accounts CSV → account_id por comitente ─────────────────────
    account_id_by_ctte = {}
    try:
        accounts_file.seek(0)
        text = accounts_file.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text), quotechar='"', delimiter=',')
        next(reader, None)   # skip header
        for row in reader:
            if len(row) < 4:
                continue
            netting_type = row[0].strip()
            account_id   = row[1].strip()
            account_name = row[3].strip()
            if netting_type == "NORMAL" and account_name and account_id:
                account_id_by_ctte[account_name] = account_id
    except Exception as e:
        advertencias.append(f"No se pudo leer table-accounts: {e}")

    # ── 8. SALDOS DEUDORES.xlsx ───────────────────────────────────────────────
    saldos_file.seek(0)
    wb_saldo  = openpyxl.load_workbook(saldos_file)
    ws_saldo  = wb_saldo.active
    comitentes = []
    for row in ws_saldo.iter_rows(min_row=3, values_only=True):
        if row[0] and row[1] and not isinstance(row[0], str):
            raw_vto = row[2] if len(row) > 2 else None
            if hasattr(raw_vto, "date"):
                fecha_vto = raw_vto.date()
                label = fecha_vto.strftime("%d/%m/%y")
            elif raw_vto is not None:
                s = str(raw_vto).strip()
                fecha_vto = None
                for _fmt in ("%d/%m/%Y", "%d/%m/%y"):
                    try:
                        fecha_vto = datetime.strptime(s, _fmt).date()
                        label = fecha_vto.strftime("%d/%m/%y")
                        break
                    except Exception:
                        pass
                if fecha_vto is None:
                    label = s
            else:
                fecha_vto = None
                label = ""
            comitentes.append({
                "ctte": str(int(row[0])),
                "requerido": float(row[1]),
                "label": label,
                "fecha_vto": fecha_vto,
            })

    if not comitentes:
        raise ValueError("SALDOS DEUDORES.xlsx no tiene filas de datos (fila 3 en adelante).")

    # Agrupar por ctte
    _ctte_groups = OrderedDict()
    for _c in comitentes:
        _ctte_groups.setdefault(_c["ctte"], []).append({
            "requerido": _c["requerido"],
            "label":     _c["label"],
            "fecha_vto": _c["fecha_vto"],
        })
    comitentes_grouped = [{"ctte": k, "reqs": v} for k, v in _ctte_groups.items()]
    cttes_scope = {c["ctte"] for c in comitentes}

    # ── 9. CONTBOLE.XLS — operaciones CI del día ──────────────────────────────
    ci_ops        = {}   # {ctte: {codigo_cvsa: net_qty}}
    fecha_hoy_str = FECHA_PROCESO.strftime("%d/%m/%y")
    ci_rows_today = 0
    ci_rows_other = 0

    contbole_file.seek(0)
    wb_cont = xlrd.open_workbook(file_contents=contbole_file.read())
    hoja_cont = ("Control_de_Boletos"
                 if "Control_de_Boletos" in wb_cont.sheet_names()
                 else wb_cont.sheet_names()[0])
    sh_cont = wb_cont.sheet_by_name(hoja_cont)

    for r in range(1, sh_cont.nrows):
        row_c   = sh_cont.row_values(r)
        fec_ope = str(row_c[2]).strip()
        fec_liq = str(row_c[3]).strip()
        if fec_ope != fec_liq:
            continue
        if fec_ope == fecha_hoy_str:
            ci_rows_today += 1
        else:
            ci_rows_other += 1
            continue
        try:
            ctte = str(int(float(row_c[4])))
        except Exception:
            continue
        if ctte not in cttes_scope:
            continue
        op_abrev = str(row_c[1]).strip()
        tipo = op_tipo.get(op_abrev)
        if tipo not in ("COMPRA", "VENTA"):
            continue
        raw_cod = str(row_c[31]).strip()
        if raw_cod and raw_cod != "None":
            cod_cvsa = raw_cod.zfill(5)
        else:
            ticker_fb = str(row_c[6]).strip()
            cod_cvsa  = next(
                (c for c, e in especies.items() if e["ticker"] == ticker_fb), None
            )
            if not cod_cvsa:
                advertencias.append(
                    f"CI: especie '{row_c[6]}' no encontrada en maestro (ctte {ctte})"
                )
                continue
        try:
            vn = float(row_c[8])
        except Exception:
            continue
        if vn <= 0:
            continue
        ci_ops.setdefault(ctte, {}).setdefault(cod_cvsa, 0.0)
        ci_ops[ctte][cod_cvsa] += vn if tipo == "COMPRA" else -vn

    # Limpiar netos ≈ 0
    for _ct in list(ci_ops.keys()):
        ci_ops[_ct] = {c: q for c, q in ci_ops[_ct].items() if abs(q) > 0.001}
        if not ci_ops[_ct]:
            del ci_ops[_ct]

    if ci_rows_today == 0 and ci_rows_other > 0:
        advertencias.append(
            f"CONTBOLE.XLS tiene operaciones CI pero NO de hoy ({fecha_hoy_str}). "
            "Re-descargarlo actualizado desde Gallo."
        )
    elif ci_rows_today == 0:
        advertencias.append(
            f"No se encontraron operaciones CI en CONTBOLE.XLS para hoy ({fecha_hoy_str}). "
            "Verificar si hubo operaciones CI en la rueda."
        )

    # ── 10. Helper: valorizar ─────────────────────────────────────────────────
    def valorizar(vn, codigo):
        precio    = precios.get(codigo, 0.0)
        byma_info = byma_dict.get(codigo)
        aforo     = byma_info[1] if byma_info else 0.0
        if precio == 0 or aforo == 0:
            return 0.0
        esp  = especies.get(codigo, {})
        tipo = esp.get("tipo_precio", "Normal")
        if tipo.lower().startswith("porc"):
            return vn * (precio / 100.0) * aforo
        return vn * precio * aforo

    # ── 11. Construcción del workbook ─────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    summary_data = []

    for ctte_info in comitentes_grouped:
        ctte      = ctte_info["ctte"]
        reqs      = ctte_info["reqs"]
        requerido = sum(r["requerido"] for r in reqs)
        multi_req = len(reqs) > 1

        ctte_data   = saga_by_ctte.get(ctte)
        nombre_ctte = ctte_data["nombre"] if ctte_data else ""
        posiciones  = ctte_data["posiciones"] if ctte_data else []

        if not posiciones:
            advertencias.append(f"Comitente {ctte}: sin posiciones en SAGACLTE (BYMA).")

        # Agregar SAGACLTE por (código, fecha_liq)
        agg = {}
        for p in posiciones:
            key = (p["codigo"], p.get("fecha_liq"))
            agg[key] = agg.get(key, 0.0) + p["qty"]

        # Tenencia desde SATECLTE
        tenencia_data = sateclte_ten.get(ctte)
        vtas_dict = tenencia_data["vtas"]        if tenencia_data else {}
        disp_dict = tenencia_data["disponibles"] if tenencia_data else {}
        cpr_dict  = tenencia_data["cpr"]         if tenencia_data else {}
        tiene_vtas = bool(vtas_dict)

        # CI ops
        ci_ctte       = ci_ops.get(ctte, {})
        ci_ventas_ci  = {cod: abs(qty) for cod, qty in ci_ctte.items() if qty < 0}
        ci_compras_ci = {cod: qty       for cod, qty in ci_ctte.items() if qty > 0}

        # Tabla de species
        species_table = []
        for key in sorted(agg.keys(), key=lambda k: (k[1] is None, k[1] or date.min, k[0])):
            cod, fecha_liq = key
            qty     = agg[key]
            esp     = especies.get(cod, {})
            ticker  = esp.get("ticker", cod)
            nombre_esp  = esp.get("nombre", "")
            tipo_precio = esp.get("tipo_precio", "Normal")
            byma_info   = byma_dict.get(cod)
            aforo_byma  = byma_info[1] if byma_info else 0.0
            precio      = precios.get(cod, 0.0)

            vta_qty         = min(vtas_dict.get(cod, 0.0), qty)
            qty_after_tfile = qty - vta_qty
            vta_ci_qty      = min(ci_ventas_ci.get(cod, 0.0), qty_after_tfile)
            qty_neto        = qty_after_tfile - vta_ci_qty

            monto_bruto = monto_vta_tfile = monto_vta_ci = monto_neto = 0.0
            estado_esp = "OK"

            if precio == 0:
                estado_esp = "sin precio"
                advertencias.append(f"Ctte {ctte}: {ticker} ({cod}) sin precio en PC.")
            elif aforo_byma == 0:
                estado_esp = "sin aforo BYMA"
            else:
                factor          = (precio / 100.0) * aforo_byma if tipo_precio.lower().startswith("porc") else precio * aforo_byma
                monto_bruto     = qty          * factor
                monto_vta_tfile = vta_qty      * factor
                monto_vta_ci    = vta_ci_qty   * factor
                monto_neto      = qty_neto     * factor

            species_table.append({
                "codigo": cod, "ticker": ticker, "nombre": nombre_esp,
                "aforo_byma": aforo_byma, "fecha_liq": fecha_liq,
                "qty_garantia": qty,
                "vta_qty": vta_qty, "vta_ci_qty": vta_ci_qty, "qty_neto": qty_neto,
                "precio": precio, "tipo_precio": tipo_precio,
                "monto_bruto": monto_bruto, "monto_vta_tfile": monto_vta_tfile,
                "monto_vta_ci": monto_vta_ci, "monto_neto": monto_neto,
                "estado": estado_esp,
                "es_vta": vta_qty > 0 or vta_ci_qty > 0,
            })

        total_bruto      = sum(s["monto_bruto"]      for s in species_table)
        total_vtas_tfile = sum(s["monto_vta_tfile"]  for s in species_table)
        total_vtas_ci    = sum(s["monto_vta_ci"]     for s in species_table)
        total_vtas       = total_vtas_tfile + total_vtas_ci
        total_neto       = total_bruto - total_vtas
        diferencia       = total_neto - requerido

        # Totales por fecha_liq para panel multi-req
        _date_bruto = {}
        for s in species_table:
            fd = s["fecha_liq"]
            _date_bruto[fd] = _date_bruto.get(fd, 0.0) + s["monto_neto"]
        for req_item in reqs:
            fd = req_item.get("fecha_vto")
            req_item["garantia_fecha"]  = _date_bruto.get(fd, 0.0)
            dif_f = req_item["garantia_fecha"] - req_item["requerido"]
            req_item["diferencia_fecha"] = dif_f
            req_item["cubierto_fecha"]   = dif_f >= 0

        cubierto    = diferencia >= 0
        pct_dif     = (abs(diferencia) / requerido * 100) if requerido > 0 else 0.0
        tiene_vtas_ci  = any(s["vta_ci_qty"] > 0 for s in species_table)
        tiene_any_vta  = tiene_vtas or tiene_vtas_ci

        # Tabla de reemplazo
        def make_reemplazo_row(cod, qty_r, origen):
            esp_r    = especies.get(cod, {})
            byma_r   = byma_dict.get(cod)
            aforo_r  = byma_r[1] if byma_r else 0.0
            precio_r = precios.get(cod, 0.0)
            return {
                "origen": origen, "codigo": cod,
                "ticker":     esp_r.get("ticker", cod),
                "nombre":     esp_r.get("nombre", ""),
                "tipo_precio": esp_r.get("tipo_precio", "Normal"),
                "lam_min":    esp_r.get("lam_min", 0.0),
                "aforo_byma": aforo_r,
                "qty":        qty_r,
                "precio":     precio_r,
                "monto":      valorizar(qty_r, cod),
                "acepta_byma": bool(byma_r),
            }

        def _reemplazo_sort(x):
            return (x["lam_min"] == 1.0, x["monto"])

        disp_rows      = sorted([make_reemplazo_row(c, q, "DISPONIBLE")      for c, q in disp_dict.items()],      key=_reemplazo_sort, reverse=True)
        cpr_rows       = sorted([make_reemplazo_row(c, q, "CPR A LIQUIDAR")  for c, q in cpr_dict.items()],       key=_reemplazo_sort, reverse=True)
        ci_compra_rows = sorted([make_reemplazo_row(c, q, "COMPRA CI")       for c, q in ci_compras_ci.items()],  key=_reemplazo_sort, reverse=True)
        reemplazo_table = disp_rows + cpr_rows + ci_compra_rows
        total_reemplazo_byma = sum(r["monto"] for r in reemplazo_table if r["acepta_byma"])

        # ── Crear hoja C-{CTTE} ───────────────────────────────────────────────
        ws = wb_out.create_sheet(title=f"C-{ctte}")

        titulo = f"Comitente {ctte}" + (f" - {nombre_ctte}" if nombre_ctte else "")
        ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=12)

        # Panel
        panel_rows = []
        if not multi_req:
            panel_rows.append(("Monto Requerido",        requerido,   "ARS", None))
            panel_rows.append(("Garantia Bruta (BYMA)",  total_bruto, "ARS", None))
            if tiene_vtas:
                panel_rows.append(("Ventas a Liquidar (impacto)",    -total_vtas_tfile, "ARS", "VTA"))
            if tiene_vtas_ci:
                panel_rows.append(("Ventas CI a Liquidar (impacto)", -total_vtas_ci,    "ARS", "VTA_CI"))
            if tiene_any_vta:
                panel_rows.append(("Garantia Neta",                   total_neto,       "ARS", None))
            panel_rows.append(("Diferencia",             diferencia,  "ARS", "DIF"))
            panel_rows.append(("% s/Requerimiento",      round(pct_dif, 2), "PCT", None))
            panel_rows.append(("Estado",                 "CUBIERTO" if cubierto else "DESCUBIERTO", "EST", None))
        else:
            if tiene_vtas:
                panel_rows.append(("Garantia Bruta (BYMA)",          total_bruto,        "ARS", None))
                panel_rows.append(("Ventas a Liquidar (impacto)",    -total_vtas_tfile,  "ARS", "VTA"))
            if tiene_vtas_ci:
                if not tiene_vtas:
                    panel_rows.append(("Garantia Bruta (BYMA)",      total_bruto,        "ARS", None))
                panel_rows.append(("Ventas CI a Liquidar (impacto)", -total_vtas_ci,    "ARS", "VTA_CI"))
            if tiene_any_vta:
                panel_rows.append(("Garantia Neta Total",             total_neto,        "ARS", None))
            for idx, req_item in enumerate(reqs, start=1):
                lbl   = req_item["label"] if req_item["label"] else f"#{idx}"
                cub_f = req_item["cubierto_fecha"]
                panel_rows.append((f"Requerido {lbl}",  req_item["requerido"],      "ARS", "REQ_ITEM"))
                panel_rows.append((f"Cobertura {lbl}",  req_item["garantia_fecha"], "ARS", "DIF_FECHA_OK" if cub_f else "DIF_FECHA_DESC"))
                panel_rows.append((f"Estado {lbl}",     "CUBIERTO" if cub_f else "DESCUBIERTO", "EST", "EST_FECHA_OK" if cub_f else "EST_FECHA_DESC"))
            panel_rows.append(("Total Requerido",         requerido,  "ARS", None))
            panel_rows.append(("Diferencia Total",        diferencia, "ARS", "DIF"))
            panel_rows.append(("% s/Total Requerido",     round(pct_dif, 2), "PCT", None))
            panel_rows.append(("Estado Global",           "CUBIERTO" if cubierto else "DESCUBIERTO", "EST", None))

        for i, (lbl, val, fmt, tag) in enumerate(panel_rows):
            r_idx  = i + 2
            cell_a = ws.cell(row=r_idx, column=1, value=lbl)
            cell_b = ws.cell(row=r_idx, column=2, value=val)
            cell_a.border = _thin_border()
            cell_b.border = _thin_border()

            if tag in ("VTA", "VTA_CI"):
                for c in (cell_a, cell_b):
                    c.fill = ORANGE_FILL
                cell_a.font = VTA_FONT
                cell_b.font = Font(bold=True, color="9C6500")
                cell_b.number_format = FMT_INT
                cell_b.alignment = Alignment(horizontal="right")
            elif tag == "REQ_ITEM":
                cell_a.fill = ALT_FILL; cell_b.fill = ALT_FILL
                cell_a.font = BOLD_FONT; cell_b.font = BOLD_FONT
                cell_b.number_format = FMT_INT
                cell_b.alignment = Alignment(horizontal="right")
            elif tag in ("DIF_FECHA_OK", "DIF_FECHA_DESC"):
                _ok   = tag == "DIF_FECHA_OK"
                _fill = PatternFill("solid", fgColor="EAF4E8") if _ok else PatternFill("solid", fgColor="FCE4D6")
                cell_a.fill = _fill; cell_b.fill = _fill
                cell_a.font = Font(bold=True, color="276221" if _ok else "9C0006")
                cell_b.font = Font(bold=True, color="276221" if _ok else "9C0006")
                cell_b.number_format = FMT_INT
                cell_b.alignment = Alignment(horizontal="right")
            elif tag in ("EST_FECHA_OK", "EST_FECHA_DESC"):
                _ok   = tag == "EST_FECHA_OK"
                _fill = GREEN_FILL if _ok else RED_FILL
                cell_a.fill = _fill; cell_b.fill = _fill
                cell_a.font = Font(italic=True, bold=True, color="276221" if _ok else "9C0006")
                cell_b.font = Font(italic=True, bold=True, color="276221" if _ok else "9C0006")
            elif tag == "EST":
                fill = GREEN_FILL if cubierto else RED_FILL
                font = GREEN_FONT if cubierto else RED_FONT
                for c in (cell_a, cell_b):
                    c.fill = fill; c.font = font
            elif tag == "DIF":
                cell_a.fill = HEADER_FILL; cell_a.font = HEADER_FONT
                cell_a.alignment = Alignment(horizontal="left")
                cell_b.font = Font(bold=True, color="276221" if cubierto else "9C0006")
                cell_b.number_format = FMT_INT
                cell_b.alignment = Alignment(horizontal="right")
            else:
                cell_a.fill = HEADER_FILL; cell_a.font = HEADER_FONT
                cell_a.alignment = Alignment(horizontal="left")
                cell_b.font = BOLD_FONT
                cell_b.alignment = Alignment(horizontal="right")
                if fmt == "ARS":
                    cell_b.number_format = FMT_INT
                elif fmt == "PCT":
                    cell_b.number_format = FMT_PCT

        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 22

        # Account ID (cols D-E, fila 2)
        account_id_val = account_id_by_ctte.get(ctte, "N/D")
        cell_d2 = ws.cell(row=2, column=4, value="Account ID")
        cell_d2.fill = HEADER_FILL; cell_d2.font = HEADER_FONT
        cell_d2.alignment = Alignment(horizontal="left"); cell_d2.border = _thin_border()
        cell_e2 = ws.cell(row=2, column=5, value=account_id_val)
        cell_e2.font = BOLD_FONT; cell_e2.alignment = Alignment(horizontal="left"); cell_e2.border = _thin_border()
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14

        current_row = 2 + len(panel_rows) + 2

        # ── Tabla POSICIONES EN GARANTIA ──────────────────────────────────────
        ws.cell(row=current_row, column=1, value="POSICIONES EN GARANTIA BYMA").font = BOLD_FONT
        current_row += 1

        if tiene_vtas and tiene_vtas_ci:
            sp_headers = ["Ticker", "Codigo CVSA", "Nombre Especie", "Aforo BYMA",
                          "VN en Garantia", "VTA a Liquidar (VN)", "VTA CI (VN)", "VN Neto",
                          "Precio Cierre", "Tipo Precio", "Monto Bruto", "Monto Neto", "Estado"]
            COL_AFORO = 4; COL_VN = 5; COL_VTA = 6; COL_VTA_CI = 7; COL_VNETO = 8
            COL_PRECIO = 9; COL_MBRUTO = 11; COL_MNETO = 12; COL_ESTADO = 13
        elif tiene_vtas:
            sp_headers = ["Ticker", "Codigo CVSA", "Nombre Especie", "Aforo BYMA",
                          "VN en Garantia", "VTA a Liquidar (VN)", "VN Neto",
                          "Precio Cierre", "Tipo Precio", "Monto Bruto", "Monto Neto", "Estado"]
            COL_AFORO = 4; COL_VN = 5; COL_VTA = 6; COL_VTA_CI = None; COL_VNETO = 7
            COL_PRECIO = 8; COL_MBRUTO = 10; COL_MNETO = 11; COL_ESTADO = 12
        elif tiene_vtas_ci:
            sp_headers = ["Ticker", "Codigo CVSA", "Nombre Especie", "Aforo BYMA",
                          "VN en Garantia", "VTA CI (VN)", "VN Neto",
                          "Precio Cierre", "Tipo Precio", "Monto Bruto", "Monto Neto", "Estado"]
            COL_AFORO = 4; COL_VN = 5; COL_VTA = None; COL_VTA_CI = 6; COL_VNETO = 7
            COL_PRECIO = 8; COL_MBRUTO = 10; COL_MNETO = 11; COL_ESTADO = 12
        else:
            sp_headers = ["Ticker", "Codigo CVSA", "Nombre Especie", "Aforo BYMA",
                          "VN en Garantia", "Precio Cierre", "Tipo Precio",
                          "Monto Valorizado", "Estado"]
            COL_AFORO = 4; COL_VN = 5; COL_VTA = None; COL_VTA_CI = None; COL_VNETO = None
            COL_PRECIO = 6; COL_MBRUTO = 8; COL_MNETO = None; COL_ESTADO = 9

        for ci_h, h in enumerate(sp_headers, start=1):
            _apply_header(ws, current_row, ci_h, h)
        n_cols_table = len(sp_headers)

        def _write_species_row(ws, row_idx, vals, fill_row, es_vta):
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.border = _thin_border()
                if es_vta:
                    cell.fill = VTA_FILL
                elif fill_row:
                    cell.fill = ALT_FILL
                if ci == COL_AFORO and v != "":
                    cell.number_format = '0"%"'; cell.alignment = Alignment(horizontal="right")
                elif ci == COL_VN and v != "":
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                elif COL_VTA and ci == COL_VTA and v != "":
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                    cell.font = Font(color="9C0006", bold=True)
                elif COL_VTA_CI and ci == COL_VTA_CI and v != "":
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                    cell.font = Font(color="6C3483", bold=True)
                elif COL_VNETO and ci == COL_VNETO and v != "":
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                elif ci == COL_PRECIO and v != "":
                    cell.number_format = FMT_PRICE; cell.alignment = Alignment(horizontal="right")
                elif ci == COL_MBRUTO and v != "":
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                elif COL_MNETO and ci == COL_MNETO and v != "":
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                if ci == COL_ESTADO and v not in ("OK", ""):
                    cell.font = Font(color="9C0006", italic=True)

        def _build_vals(s):
            if tiene_vtas and tiene_vtas_ci:
                return [s["ticker"], s["codigo"], s["nombre"],
                        int(s["aforo_byma"] * 100) if s["aforo_byma"] else "",
                        int(s["qty_garantia"]),
                        int(s["vta_qty"])    if s["vta_qty"]    else "",
                        int(s["vta_ci_qty"]) if s["vta_ci_qty"] else "",
                        int(s["qty_neto"])   if s["qty_neto"]   else "",
                        s["precio"] if s["precio"] else "", s["tipo_precio"],
                        int(round(s["monto_bruto"])) if s["monto_bruto"] else "",
                        int(round(s["monto_neto"]))  if s["monto_neto"]  else "",
                        s["estado"]]
            elif tiene_vtas:
                return [s["ticker"], s["codigo"], s["nombre"],
                        int(s["aforo_byma"] * 100) if s["aforo_byma"] else "",
                        int(s["qty_garantia"]),
                        int(s["vta_qty"])  if s["vta_qty"]  else "",
                        int(s["qty_neto"]) if s["qty_neto"] else "",
                        s["precio"] if s["precio"] else "", s["tipo_precio"],
                        int(round(s["monto_bruto"])) if s["monto_bruto"] else "",
                        int(round(s["monto_neto"]))  if s["monto_neto"]  else "",
                        s["estado"]]
            elif tiene_vtas_ci:
                return [s["ticker"], s["codigo"], s["nombre"],
                        int(s["aforo_byma"] * 100) if s["aforo_byma"] else "",
                        int(s["qty_garantia"]),
                        int(s["vta_ci_qty"]) if s["vta_ci_qty"] else "",
                        int(s["qty_neto"])   if s["qty_neto"]   else "",
                        s["precio"] if s["precio"] else "", s["tipo_precio"],
                        int(round(s["monto_bruto"])) if s["monto_bruto"] else "",
                        int(round(s["monto_neto"]))  if s["monto_neto"]  else "",
                        s["estado"]]
            else:
                return [s["ticker"], s["codigo"], s["nombre"],
                        int(s["aforo_byma"] * 100) if s["aforo_byma"] else "",
                        int(s["qty_garantia"]),
                        s["precio"] if s["precio"] else "", s["tipo_precio"],
                        int(round(s["monto_bruto"])) if s["monto_bruto"] else "",
                        s["estado"]]

        def _write_subtotal_row(ws, row_idx, lbl, sub_vn, sub_vta, sub_vta_ci,
                                sub_vneto, sub_mbruto, sub_mneto):
            if tiene_vtas and tiene_vtas_ci:
                sub_vals = [lbl, "", "", "", int(sub_vn),
                            int(sub_vta) if sub_vta else "", int(sub_vta_ci) if sub_vta_ci else "",
                            int(sub_vneto), "", "", int(round(sub_mbruto)), int(round(sub_mneto)), ""]
                _int_cols = {5, 8, 11, 12}
            elif tiene_vtas or tiene_vtas_ci:
                _vta = sub_vta if tiene_vtas else sub_vta_ci
                sub_vals = [lbl, "", "", "", int(sub_vn),
                            int(_vta) if _vta else "", int(sub_vneto),
                            "", "", int(round(sub_mbruto)), int(round(sub_mneto)), ""]
                _int_cols = {5, 7, 10, 11}
            else:
                sub_vals = [lbl, "", "", "", int(sub_vn), "", "", int(round(sub_mbruto)), ""]
                _int_cols = {5, 8}
            for ci, v in enumerate(sub_vals, start=1):
                cell = ws.cell(row=row_idx, column=ci, value=v)
                cell.fill = SUB_FILL; cell.font = SUB_FONT; cell.border = _thin_border()
                if ci in _int_cols and isinstance(v, int):
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")

        total_vn_gara = total_vn_vta = total_vn_vta_ci = total_vn_neto = 0.0
        total_m_bruto = total_m_neto = 0.0

        for fecha_grupo, rows_iter in groupby(species_table, key=lambda x: x["fecha_liq"]):
            rows_grupo = list(rows_iter)
            current_row += 1

            _fecha_txt = fecha_grupo.strftime("%d/%m/%y") if fecha_grupo else "Sin fecha"
            _req_match = next((r for r in reqs if r.get("fecha_vto") == fecha_grupo), None)
            _req_txt   = f"  |  Req: ARS {int(_req_match['requerido']):,}" if _req_match else ""
            _sec_label = f"Vencimiento {_fecha_txt}{_req_txt}"

            DATE_HDR_FILL = PatternFill("solid", fgColor="2E5C8A")
            for ci in range(1, n_cols_table + 1):
                cell = ws.cell(row=current_row, column=ci,
                               value=_sec_label if ci == 1 else "")
                cell.fill = DATE_HDR_FILL
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.border = _thin_border()
                if ci == 1:
                    cell.alignment = Alignment(horizontal="left")

            alt = False
            sub_vn = sub_vta = sub_vta_ci = sub_vneto = sub_mbruto = sub_mneto = 0.0
            for s in rows_grupo:
                current_row += 1
                alt = not alt
                _write_species_row(ws, current_row, _build_vals(s), alt, s["es_vta"])
                sub_vn     += s["qty_garantia"]
                sub_vta    += s["vta_qty"]
                sub_vta_ci += s["vta_ci_qty"]
                sub_vneto  += s["qty_neto"]
                sub_mbruto += s["monto_bruto"]
                sub_mneto  += s["monto_neto"]

            current_row += 1
            _write_subtotal_row(ws, current_row, f"Subtotal {_fecha_txt}",
                                sub_vn, sub_vta, sub_vta_ci, sub_vneto, sub_mbruto, sub_mneto)
            total_vn_gara   += sub_vn
            total_vn_vta    += sub_vta
            total_vn_vta_ci += sub_vta_ci
            total_vn_neto   += sub_vneto
            total_m_bruto   += sub_mbruto
            total_m_neto    += sub_mneto

        # Fila TOTALES
        current_row += 1
        if tiene_vtas and tiene_vtas_ci:
            tot_g = ["TOTALES", "", "", "", int(total_vn_gara),
                     int(total_vn_vta) if total_vn_vta else "",
                     int(total_vn_vta_ci) if total_vn_vta_ci else "",
                     int(total_vn_neto), "", "",
                     int(round(total_m_bruto)), int(round(total_m_neto)), ""]
            int_cols_g = {5, 8, 11, 12}
        elif tiene_vtas or tiene_vtas_ci:
            vn_vta = total_vn_vta if tiene_vtas else total_vn_vta_ci
            tot_g = ["TOTALES", "", "", "", int(total_vn_gara),
                     int(vn_vta) if vn_vta else "", int(total_vn_neto),
                     "", "", int(round(total_m_bruto)), int(round(total_m_neto)), ""]
            int_cols_g = {5, 7, 10, 11}
        else:
            tot_g = ["TOTALES", "", "", "", int(total_vn_gara), "", "",
                     int(round(total_m_bruto)), ""]
            int_cols_g = {5, 8}
        for ci, v in enumerate(tot_g, start=1):
            cell = ws.cell(row=current_row, column=ci, value=v)
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER_FILL; cell.border = _thin_border()
            if ci in int_cols_g and isinstance(v, int):
                cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")

        # Tabla GESTION REQUERIDA
        if tiene_any_vta and reemplazo_table:
            current_row += 2
            gap_label = f"GESTION REQUERIDA -- Reemplazar en garantia: ARS {int(round(total_vtas)):,}"
            ws.cell(row=current_row, column=1, value=gap_label).font = Font(bold=True, color="9C6500", size=11)
            current_row += 1
            rem_headers = ["Origen", "Codigo CVSA", "Ticker", "Nombre Especie",
                           "Aforo BYMA", "VN Disponible", "Precio Cierre", "Tipo Precio",
                           "Monto Garantizable", "Acepta BYMA"]
            for ci, h in enumerate(rem_headers, start=1):
                _apply_header(ws, current_row, ci, h)
            for row_data in reemplazo_table:
                current_row += 1
                acepta  = row_data["acepta_byma"]
                origen  = row_data["origen"]
                if not acepta:
                    row_fill = GRAY_FILL; row_font = GRAY_FONT
                elif origen == "COMPRA CI":
                    row_fill = CI_COMPRA_FILL; row_font = CI_COMPRA_FONT
                elif origen == "DISPONIBLE":
                    row_fill = GREEN_FILL; row_font = Font(color="276221")
                else:
                    row_fill = YELLOW_FILL; row_font = Font(color="9C6500")
                dvals = [origen, row_data["codigo"], row_data["ticker"], row_data["nombre"],
                         int(row_data["aforo_byma"] * 100) if row_data["aforo_byma"] else "",
                         int(row_data["qty"]) if row_data["qty"] else "",
                         row_data["precio"] if row_data["precio"] else "",
                         row_data["tipo_precio"],
                         int(round(row_data["monto"])) if row_data["monto"] else "",
                         "SI" if acepta else "NO ACEPTA"]
                for ci, v in enumerate(dvals, start=1):
                    cell = ws.cell(row=current_row, column=ci, value=v)
                    cell.border = _thin_border(); cell.fill = row_fill; cell.font = row_font
                    if ci == 5 and v != "":
                        cell.number_format = '0"%"'; cell.alignment = Alignment(horizontal="right")
                    elif ci == 6 and v != "":
                        cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                    elif ci == 7 and v != "":
                        cell.number_format = FMT_PRICE; cell.alignment = Alignment(horizontal="right")
                    elif ci == 9 and v != "":
                        cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                    elif ci == 10:
                        cell.alignment = Alignment(horizontal="center")
                        if not acepta:
                            cell.font = Font(color="9C0006", bold=True, italic=True)
            current_row += 1
            tot_rem_vals = ["TOTAL GARANTIZABLE", "", "", "", "", "", "", "",
                            int(round(total_reemplazo_byma)), ""]
            for ci, v in enumerate(tot_rem_vals, start=1):
                cell = ws.cell(row=current_row, column=ci, value=v)
                cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER_FILL; cell.border = _thin_border()
                if ci == 9 and isinstance(v, int):
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
            current_row += 1
            gap_cubierto = total_reemplazo_byma >= total_vtas
            conclusion_txt = (
                f"Disponible garantizable: ARS {int(total_reemplazo_byma):,}  /  "
                f"Reemplazo necesario: ARS {int(total_vtas):,}  "
                + ("  ->  ALCANZA para reemplazar las VTAs"
                   if gap_cubierto else
                   "  ->  INSUFICIENTE -- gestionar activos adicionales")
            )
            ws.cell(row=current_row, column=1, value=conclusion_txt).font = Font(
                bold=True, color="276221" if gap_cubierto else "9C0006"
            )

        # Autofit columnas
        for col in ws.columns:
            max_len    = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val     = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 34)
        ws.freeze_panes = "A" + str(2 + len(panel_rows) + 1)

        summary_data.append({
            "ctte": ctte, "nombre": nombre_ctte,
            "requerido": requerido, "bruto": total_bruto,
            "vtas": total_vtas_tfile, "vtas_ci": total_vtas_ci,
            "neto": total_neto, "diferencia": diferencia,
            "pct": round(pct_dif, 2),
            "estado": "CUBIERTO" if cubierto else "DESCUBIERTO",
            "tiene_vtas": tiene_vtas, "tiene_vtas_ci": tiene_vtas_ci,
        })

    # ── Hoja RESUMEN ──────────────────────────────────────────────────────────
    ws_res = wb_out.create_sheet(title="RESUMEN", index=0)
    res_headers = ["Nro Comitente", "Nombre", "Monto Requerido",
                   "Garantia Bruta (BYMA)", "Ventas a Liquidar", "Ventas CI a Liquidar",
                   "Garantia Neta", "Diferencia", "% Dif. s/Req.", "Estado"]
    for ci, h in enumerate(res_headers, start=1):
        _apply_header(ws_res, 1, ci, h)
    _apply_header(ws_res, 1, 12, "Account ID")

    alt = False
    for i, sd in enumerate(summary_data):
        r   = i + 2
        alt = not alt
        vals = [sd["ctte"], sd["nombre"],
                int(sd["requerido"]), int(sd["bruto"]),
                int(sd["vtas"])    if sd["vtas"]    else "",
                int(sd["vtas_ci"]) if sd["vtas_ci"] else "",
                int(sd["neto"]), int(sd["diferencia"]),
                sd["pct"], sd["estado"]]
        for ci, v in enumerate(vals, start=1):
            cell = ws_res.cell(row=r, column=ci, value=v)
            cell.border = _thin_border()
            if alt:
                cell.fill = ALT_FILL
            if ci in (3, 4, 7, 8):
                cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
            if ci == 5 and v != "":
                cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                cell.font = Font(color="9C0006", bold=True)
            if ci == 6 and v != "":
                cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                cell.font = Font(color="6C3483", bold=True)
            if ci == 9:
                cell.number_format = FMT_PCT; cell.alignment = Alignment(horizontal="right")
            if ci == 10:
                cell.fill = (GREEN_FILL if v == "CUBIERTO" else RED_FILL)
                cell.font = (Font(bold=True, color="276221") if v == "CUBIERTO"
                             else Font(bold=True, color="9C0006"))
        acc_id = account_id_by_ctte.get(sd["ctte"], "N/D")
        cell_acc = ws_res.cell(row=r, column=12, value=acc_id)
        cell_acc.border = _thin_border()
        if alt:
            cell_acc.fill = ALT_FILL

    r_tot = len(summary_data) + 2
    tot_vtas    = sum(s["vtas"]    for s in summary_data)
    tot_vtas_ci = sum(s["vtas_ci"] for s in summary_data)
    tot_vals = ["TOTALES", "",
                int(sum(s["requerido"]  for s in summary_data)),
                int(sum(s["bruto"]      for s in summary_data)),
                int(tot_vtas)    if tot_vtas    else "",
                int(tot_vtas_ci) if tot_vtas_ci else "",
                int(sum(s["neto"]       for s in summary_data)),
                int(sum(s["diferencia"] for s in summary_data)),
                "", ""]
    for ci, v in enumerate(tot_vals, start=1):
        cell = ws_res.cell(row=r_tot, column=ci, value=v)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER_FILL; cell.border = _thin_border()
        if ci in (3, 4, 7, 8) and isinstance(v, int):
            cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
        if ci in (5, 6) and isinstance(v, int):
            cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")

    for col in ws_res.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val     = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws_res.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)
    ws_res.freeze_panes = "A2"

    # ── Serializar ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb_out.save(buf)
    xlsx_bytes = buf.getvalue()

    n_cubiertos    = sum(1 for s in summary_data if s["estado"] == "CUBIERTO")
    n_descubiertos = len(summary_data) - n_cubiertos
    resumen = {
        "fecha":          fecha_str,
        "n_comitentes":   len(summary_data),
        "n_cubiertos":    n_cubiertos,
        "n_descubiertos": n_descubiertos,
        "total_requerido": sum(s["requerido"]  for s in summary_data),
        "total_neto":      sum(s["neto"]       for s in summary_data),
        "total_diferencia": sum(s["diferencia"] for s in summary_data),
    }

    return xlsx_bytes, resumen, advertencias
