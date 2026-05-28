#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Distribución Garantías BYMA — adaptación web (Streamlit).
"""

from io import BytesIO, StringIO
import re
import math
import csv
import xlrd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime


def generar_reporte(
    sagaclte_file,
    sateclte_file,
    especies_file,
    pc_file,
    saldos_file,
    contbole_file,
    aforo_sail_file,
    accounts_file,
    tabcompb_file=None,
    rmc_file=None,
    rescates_texto="",
    fecha_proceso=None,
):
    """
    Returns (BytesIO xlsx, resumen_dict).
    Raises ValueError on unrecoverable errors.
    """
    if fecha_proceso is None:
        fecha_proceso = date.today()

    FECHA_PROCESO = fecha_proceso

    # ── Styles ──────────────────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT   = Font(color="FFFFFF", bold=True)
    GREEN_FILL    = PatternFill("solid", fgColor="C6EFCE")
    GREEN_FONT    = Font(color="276221", bold=True)
    RED_FILL      = PatternFill("solid", fgColor="FFC7CE")
    RED_FONT      = Font(color="9C0006", bold=True)
    GREY_FONT     = Font(color="AAAAAA")
    BOLD_FONT     = Font(bold=True)
    ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")
    ENGARA_FILL   = PatternFill("solid", fgColor="BDD7EE")
    ENVIAR_FILL   = PatternFill("solid", fgColor="FCE4D6")
    STOCK_FILL    = PatternFill("solid", fgColor="D9D9D9")
    DEVOLUCION_FILL = PatternFill("solid", fgColor="E2EFDA")
    VTO_TITLE_FILL  = PatternFill("solid", fgColor="C55A11")
    VTO_HEADER_FILL = PatternFill("solid", fgColor="F4B942")
    VTO_SUB_FILL    = PatternFill("solid", fgColor="FCE4D6")
    VTO_TAB_COLOR   = "FF6600"
    CI_COMPRA_FILL      = PatternFill("solid", fgColor="E8DAEF")
    CI_COMPRA_FONT      = Font(color="6C3483", bold=True)
    CI_HEADER_FILL      = PatternFill("solid", fgColor="9B59B6")
    CI_VENTA_ALERT_FILL = PatternFill("solid", fgColor="FFCCCC")
    CI_VENTA_TITLE_FILL = PatternFill("solid", fgColor="C0392B")
    CI_VENTA_FONT       = Font(color="922B21", bold=True)
    FMT_INT   = '#,##0'
    FMT_PRICE = '#,##0.00'

    def thin_border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def apply_header(ws, row_idx, col_idx, value):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        return cell

    # ── STEP 1: Aforo SAIL por lista ────────────────────────────────────────────
    aforo_sail_file.seek(0)
    wb_aforo = openpyxl.load_workbook(BytesIO(aforo_sail_file.read()))
    ws_aforo = wb_aforo.active
    _AFORO_SAIL_BY_INT = {}
    TABLE_COLS = [(1, 3), (5, 7), (9, 11), (13, 15)]
    for row in ws_aforo.iter_rows(min_row=3, values_only=True):
        for col_lista, col_sail in TABLE_COLS:
            lista_val = row[col_lista - 1] if len(row) >= col_lista else None
            sail_val  = row[col_sail - 1]  if len(row) >= col_sail  else None
            if lista_val is not None and sail_val is not None:
                try:
                    _AFORO_SAIL_BY_INT[int(lista_val)] = float(sail_val) / 100.0
                except (ValueError, TypeError):
                    pass

    def get_aforo_sail(lista_str):
        try:
            return _AFORO_SAIL_BY_INT.get(int(lista_str), 0.0)
        except (ValueError, TypeError):
            return 0.0

    # ── STEP 2: Account IDs ─────────────────────────────────────────────────────
    accounts_file.seek(0)
    _ACCOUNT_ID_BY_CTTE = {}
    content = accounts_file.read().decode("utf-8-sig")
    reader = csv.reader(StringIO(content), quotechar='"', delimiter=',')
    next(reader, None)
    for row in reader:
        raw = row[0].strip() if row else ''
        parts = raw.split(';')
        if len(parts) >= 3:
            ctte       = parts[1].strip()
            account_id = parts[2].strip()
            if ctte and account_id:
                _ACCOUNT_ID_BY_CTTE[ctte] = account_id

    # ── STEP 3: ESPECIES.XLS ────────────────────────────────────────────────────
    especies_file.seek(0)
    wb_esp = xlrd.open_workbook(file_contents=especies_file.read())
    sh_esp = wb_esp.sheet_by_name("Datos_Fijos_Especies")
    especies = {}
    for r in range(1, sh_esp.nrows):
        row  = sh_esp.row_values(r)
        cod  = str(row[0]).strip().zfill(5)
        nom  = str(row[1]).strip()
        tprc = str(row[4]).strip()
        lista = str(int(float(str(row[5]).strip()))) if row[5] else "0"
        if lista in ("0", ""):
            lista = "0"
        nemo  = str(row[9]).strip()
        aforo = get_aforo_sail(lista)
        try:
            lam_min = float(str(row[6]).strip()) if row[6] else 0.0
        except (ValueError, TypeError):
            lam_min = 0.0
        # Col 26: haircut BYMA API → aforo_byma = (100 - haircut) / 100.0
        try:
            haircut = int(float(row[26])) if len(row) > 26 and row[26] else 0
        except (ValueError, TypeError):
            haircut = 0
        especies[cod] = {
            "ticker": nemo or cod,
            "nombre": nom,
            "lista":  lista,
            "tipo_precio": tprc,
            "aforo_sail": aforo,
            "lam_min": lam_min,
            "aforo_byma": (100 - haircut) / 100.0 if haircut > 0 else 0.0,
        }

    # ── STEP 4: Aforos BYMA desde ESPECIES.XLS col 26 ───────────────────────────
    # byma_dict[cod] = (ticker, aforo_float)
    byma_dict = {
        cod: (esp["ticker"], esp["aforo_byma"])
        for cod, esp in especies.items()
        if esp["aforo_byma"] > 0.0
    }

    # ── STEP 5: SAGACLTE.XLS ────────────────────────────────────────────────────
    sagaclte_file.seek(0)
    wb_saga = xlrd.open_workbook(file_contents=sagaclte_file.read())
    sh_saga = wb_saga.sheet_by_name("Saldos_de_Garantias")
    saga_rows = []
    for r in range(1, sh_saga.nrows):
        row    = sh_saga.row_values(r)
        ctte   = str(int(float(row[0]))) if row[0] else ""
        agente = str(row[5]).strip()
        if agente:
            continue
        cod  = str(row[2]).strip().zfill(5)
        qty  = float(row[6]) if row[6] else 0.0
        liq  = str(row[7]).strip()
        liq_date = None
        if liq:
            for fmt in ("%d/%m/%y", "%d/%m/%Y"):
                try:
                    liq_date = datetime.strptime(liq, fmt).date()
                    break
                except Exception:
                    pass
        estado = "RETIRABLE" if (liq_date is not None and liq_date <= FECHA_PROCESO) else "UTILIZADA"
        saga_rows.append({"ctte": ctte, "codigo": cod, "qty": qty,
                          "liq_date": liq_date, "estado": estado})

    # ── STEP 6: SATECLTE.XLS ────────────────────────────────────────────────────
    def _is_cpra(desc):
        u = desc.upper()
        return "CPRA" in u or "CPU$" in u or "CCFP" in u

    def _is_vtas(desc):
        u = desc.upper()
        return "VTAS" in u or "VTU$" in u or "VTTR" in u

    sateclte_tenencia = {}

    def _add_tenencia(d, ctte, cod, field, qty):
        d.setdefault(ctte, {}).setdefault(cod, {"disp": 0.0, "cpr": 0.0, "venta": 0.0})
        d[ctte][cod][field] += qty

    sateclte_file.seek(0)
    wb_sat = xlrd.open_workbook(file_contents=sateclte_file.read())
    sh_sat = wb_sat.sheet_by_name("Saldos_de_Tenencia")

    _sat_rows = []
    _cur_ctte = ""
    for r in range(1, sh_sat.nrows):
        row = sh_sat.row_values(r)
        if row[0]:
            _cur_ctte = str(int(float(str(row[0]).strip())))
        fecha_raw = row[6] if len(row) > 6 else ""
        _sat_rows.append((_cur_ctte, str(row[2]).strip(), row[3], row[4], row[5], fecha_raw))

    def _has_fecha(v):
        if not v and v != 0:
            return False
        try:
            return float(v) > 0
        except (ValueError, TypeError):
            return False

    _last_code    = {}
    _in_gtia_grp  = {}
    _sat_unresolved = {}

    for i, (ctte, desc, cant_t, cant_cv, total, fecha) in enumerate(_sat_rows):
        m_cod = re.match(r'^(\d{5})\s', desc)
        if m_cod:
            cod = m_cod.group(1)
            _last_code[ctte]   = cod
            _in_gtia_grp[ctte] = False
            vn = float(total) if total and float(total) > 0 else (float(cant_cv) if cant_cv else 0.0)
            if vn > 0:
                if _has_fecha(fecha) and "GTIA" not in desc.upper():
                    _add_tenencia(sateclte_tenencia, ctte, cod, "cpr", vn)
                else:
                    _add_tenencia(sateclte_tenencia, ctte, cod, "disp", vn)
            elif vn < 0:
                _add_tenencia(sateclte_tenencia, ctte, cod, "venta", abs(vn))
            continue

        if "GTIA" in desc.upper() or desc.upper() == "TOTAL":
            if desc.upper() != "GTIA.PLAZO" and "GTIA" in desc.upper():
                _in_gtia_grp[ctte] = True
            continue

        qty_cv = float(cant_cv) if cant_cv else 0.0
        if qty_cv == 0.0:
            continue

        if _is_cpra(desc) or _is_vtas(desc):
            if _in_gtia_grp.get(ctte, False):
                _sat_unresolved.setdefault(ctte, []).append({
                    "tipo": "VTAS" if _is_vtas(desc) else "CPRA",
                    "qty":  abs(qty_cv),
                    "cod":  _last_code.get(ctte),
                })
                continue
            cod = _last_code.get(ctte)
            if cod is None:
                for j in range(i + 1, len(_sat_rows)):
                    if _sat_rows[j][0] != ctte:
                        break
                    m_la = re.match(r'^(\d{5})\s', _sat_rows[j][1])
                    if m_la:
                        cod = m_la.group(1)
                        _last_code[ctte] = cod
                        break
            if cod is None:
                continue
            if _is_cpra(desc) and qty_cv > 0:
                _add_tenencia(sateclte_tenencia, ctte, cod, "cpr", qty_cv)
            elif _is_vtas(desc):
                _add_tenencia(sateclte_tenencia, ctte, cod, "venta", abs(qty_cv))

    # ── STEP 7: Rescates / Amortizaciones ───────────────────────────────────────
    _ticker_to_cod   = {e["ticker"].upper(): c for c, e in especies.items()}
    _rescate_ajustes = {}
    for line in rescates_texto.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.upper().split()
        if len(parts) < 2:
            continue
        ticker_input = parts[0]
        tipo         = parts[1]
        cod_match    = _ticker_to_cod.get(ticker_input)
        if not cod_match:
            continue
        if tipo == "RESCATE":
            _rescate_ajustes[cod_match] = 0.0
        elif tipo == "AMORTIZACION" and len(parts) >= 3:
            try:
                pct = float(parts[2].replace(',', '.').replace('%', ''))
                _rescate_ajustes[cod_match] = 1.0 - pct / 100.0
            except ValueError:
                pass

    if _rescate_ajustes:
        for r in saga_rows:
            factor = _rescate_ajustes.get(r["codigo"])
            if factor is not None:
                r["qty"] = r["qty"] * factor
        for _ctte_t, _cod_dict in sateclte_tenencia.items():
            for _cod, _vdict in _cod_dict.items():
                factor = _rescate_ajustes.get(_cod)
                if factor is not None:
                    _vdict["disp"] = _vdict.get("disp", 0.0) * factor

    # ── STEP 8: Precios de Cierre ────────────────────────────────────────────────
    pc_file.seek(0)
    wb_pc = xlrd.open_workbook(file_contents=pc_file.read())
    sh_pc = (wb_pc.sheet_by_name("Precios_de_Cierre")
             if "Precios_de_Cierre" in wb_pc.sheet_names()
             else wb_pc.sheet_by_index(0))
    precios = {}
    for r in range(1, sh_pc.nrows):
        row = sh_pc.row_values(r)
        m   = re.match(r"(\d{5})", str(row[0]).strip())
        if m:
            precios[m.group(1).zfill(5)] = float(row[1]) if row[1] else 0.0

    # ── STEP 9: Saldos Gara a cubrir ─────────────────────────────────────────────
    saldos_file.seek(0)
    wb_saldo = openpyxl.load_workbook(BytesIO(saldos_file.read()))
    sh_saldo = wb_saldo.active
    comitentes = []
    for row in sh_saldo.iter_rows(min_row=3, values_only=True):
        if not (row[0] and isinstance(row[0], (int, float)) and
                row[1] is not None and isinstance(row[1], (int, float))):
            continue
        fecha_vto = None
        if len(row) > 2 and row[2] is not None:
            raw_fv = row[2]
            if isinstance(raw_fv, datetime):
                fecha_vto = raw_fv.date()
            elif isinstance(raw_fv, date):
                fecha_vto = raw_fv
            else:
                for fmt in ("%d/%m/%Y", "%d/%m/%y"):
                    try:
                        fecha_vto = datetime.strptime(str(raw_fv).strip(), fmt).date()
                        break
                    except Exception:
                        pass
        comitentes.append({
            "ctte":      str(int(row[0])),
            "requerido": float(row[1]),
            "fecha_vto": fecha_vto,
        })

    # ── STEP 9.1: Risk Monitoring Client (opcional) ───────────────────────────────
    _rmc_deficit = {}
    _rmc_rows    = []
    if rmc_file is not None:
        rmc_file.seek(0)
        _wb_rmc = openpyxl.load_workbook(BytesIO(rmc_file.read()), data_only=True)
        _ws_rmc = _wb_rmc.active
        for _r in _ws_rmc.iter_rows(min_row=1, values_only=True):
            if _r[1] is None or _r[5] is None:
                continue
            if not isinstance(_r[5], (int, float)):
                continue
            if not str(_r[1]).strip().lstrip('-').isdigit():
                continue
            _acct = str(_r[1]).strip()
            _def  = float(_r[5])
            _cte  = str(_r[0]).strip() if _r[0] else ""
            _rmc_deficit[_acct] = _def
            _rmc_rows.append({"account_id": _acct, "comitente_rmc": _cte, "deficit": _def})

    _acct_to_ctte   = {v: k for k, v in _ACCOUNT_ID_BY_CTTE.items()}
    _cttes_en_saldos = {c["ctte"] for c in comitentes}
    _rmc_extra = []
    for _row in _rmc_rows:
        _ctte_mapped = _acct_to_ctte.get(_row["account_id"], _row["comitente_rmc"])
        if _ctte_mapped and _ctte_mapped not in _cttes_en_saldos and _row["deficit"] > 0:
            _rmc_extra.append({
                "ctte":       _ctte_mapped,
                "account_id": _row["account_id"],
                "deficit":    _row["deficit"],
            })

    _validacion_issues = []
    for _c in comitentes:
        _acct_str = _ACCOUNT_ID_BY_CTTE.get(_c["ctte"])
        _deficit  = _rmc_deficit.get(_acct_str) if _acct_str else None
        if _deficit is not None:
            _validacion = _deficit - _c["requerido"]
            if _validacion > 0:
                _validacion_issues.append({
                    "ctte":       _c["ctte"],
                    "account_id": _acct_str,
                    "importe_g":  _c["requerido"],
                    "deficit_bc": _deficit,
                    "diferencia": _validacion,
                })

    # ── STEP 9.5: VTO comitentes ─────────────────────────────────────────────────
    cttes_today = {c["ctte"] for c in comitentes}
    vto_cttes   = []
    for r in saga_rows:
        if r["estado"] == "RETIRABLE" and r["ctte"] not in cttes_today and r["ctte"] not in vto_cttes:
            vto_cttes.append(r["ctte"])

    # ── STEP 10: TABCOMPB (opcional) + CONTBOLE ──────────────────────────────────
    _op_tipo = {}
    if tabcompb_file is not None:
        tabcompb_file.seek(0)
        try:
            wb_tab = xlrd.open_workbook(file_contents=tabcompb_file.read())
            sh_tab = wb_tab.sheet_by_name("Tabla_Comprobantes")
            for r in range(1, sh_tab.nrows):
                row_t  = sh_tab.row_values(r)
                nombre = str(row_t[1]).upper()
                abrev  = str(row_t[2]).strip()
                if "COMPRA" in nombre:
                    _op_tipo[abrev] = "COMPRA"
                elif "VENTA" in nombre:
                    _op_tipo[abrev] = "VENTA"
        except Exception:
            pass

    ci_ops         = {}
    _fecha_hoy_str = FECHA_PROCESO.strftime("%d/%m/%y")
    _cttes_scope   = {c["ctte"] for c in comitentes}
    _ci_rows_today = 0
    _ci_rows_other = 0

    contbole_file.seek(0)
    wb_cont = xlrd.open_workbook(file_contents=contbole_file.read())
    sh_cont = wb_cont.sheet_by_name("Control_de_Boletos")
    for r in range(1, sh_cont.nrows):
        row_c   = sh_cont.row_values(r)
        fec_ope = str(row_c[2]).strip()
        fec_liq = str(row_c[3]).strip()
        if fec_ope != fec_liq:
            continue
        if fec_ope == _fecha_hoy_str:
            _ci_rows_today += 1
        else:
            _ci_rows_other += 1
            continue
        try:
            ctte = str(int(float(row_c[4])))
        except Exception:
            continue
        if ctte not in _cttes_scope:
            continue
        op_abrev = str(row_c[1]).strip()
        tipo = _op_tipo.get(op_abrev)
        if tipo not in ("COMPRA", "VENTA"):
            continue
        raw_cod = str(row_c[31]).strip()
        if raw_cod and raw_cod != "None":
            cod_cvsa = raw_cod.zfill(5)
        else:
            ticker_fb = str(row_c[6]).strip()
            cod_cvsa  = next((c for c, e in especies.items() if e["ticker"] == ticker_fb), None)
            if not cod_cvsa:
                continue
        try:
            vn = float(row_c[8])
        except Exception:
            continue
        if vn <= 0:
            continue
        ci_ops.setdefault(ctte, {}).setdefault(cod_cvsa, 0.0)
        ci_ops[ctte][cod_cvsa] += vn if tipo == "COMPRA" else -vn

    for _ct in list(ci_ops.keys()):
        ci_ops[_ct] = {c: q for c, q in ci_ops[_ct].items() if abs(q) > 0.001}
        if not ci_ops[_ct]:
            del ci_ops[_ct]

    contbole_warning = None
    if _ci_rows_today == 0 and _ci_rows_other > 0:
        contbole_warning = (
            f"CONTBOLE tiene operaciones CI pero NO de hoy ({_fecha_hoy_str}). "
            "Re-descargarlo actualizado desde Gallo."
        )
    elif _ci_rows_today == 0:
        contbole_warning = (
            f"No se encontraron operaciones CI en CONTBOLE para hoy ({_fecha_hoy_str}). "
            "Verificar si hubo operaciones CI en la rueda de hoy."
        )

    # ── STEP 11: Helper functions ─────────────────────────────────────────────────
    def monto_garantizable(vn, codigo, aforo):
        precio = precios.get(codigo, 0.0)
        if precio == 0 or aforo == 0:
            return 0.0
        tipo = especies.get(codigo, {}).get("tipo_precio", "Normal")
        if tipo.lower().startswith("porc"):
            return vn * (precio / 100.0) * aforo
        return vn * precio * aforo

    def _vn_para_cubrir(entry, faltante):
        precio = entry["precio"]
        aforo  = entry["aforo_sail"]
        if aforo == 0 or precio == 0:
            return None
        tipo = entry["tipo_precio"]
        vn_n = (faltante / ((precio / 100.0) * aforo)
                if tipo.lower().startswith("porc")
                else faltante / (precio * aforo))
        vn_n = math.ceil(vn_n)
        lam  = int(entry["lam_min"]) if entry.get("lam_min", 0) > 1 else 1
        if lam > 1:
            vn_n = math.ceil(vn_n / lam) * lam
        return vn_n if vn_n <= entry["vn"] else None

    def _monto_vn(entry, vn):
        precio = entry["precio"]
        aforo  = entry["aforo_sail"]
        tipo   = entry["tipo_precio"]
        if tipo.lower().startswith("porc"):
            return vn * (precio / 100.0) * aforo
        return vn * precio * aforo

    def _mejor_unico_cobertor(stage_list, faltante):
        for e in sorted(stage_list, key=lambda x: x["monto_sail_total"], reverse=True):
            vn_n = _vn_para_cubrir(e, faltante)
            if vn_n is not None:
                return e, vn_n
        return None, None

    def _simular_activos_greedy(stage_list, acumulado_ini, req):
        count = 0
        sim   = acumulado_ini
        for e in sorted(stage_list, key=lambda x: x["monto_sail_total"], reverse=True):
            if sim >= req:
                break
            if e["aforo_sail"] == 0 or e["precio"] == 0:
                continue
            count += 1
            vn_n = _vn_para_cubrir(e, req - sim)
            sim  = req if vn_n is not None else sim + _monto_vn(e, e["vn"])
        return count

    def _agregar_remanente(split_rows, entry, acumulado):
        if entry["aforo_sail"] == 0 or entry["precio"] == 0:
            split_rows.append({**entry, "vn_usado": entry["vn"], "monto_usado": 0,
                               "acumulado": acumulado, "alcanza": False,
                               "is_remanente": True, "is_skipped": True})
        else:
            split_rows.append({**entry, "vn_usado": entry["vn"],
                               "monto_usado": entry["monto_sail_total"],
                               "acumulado": acumulado, "alcanza": False,
                               "is_remanente": True, "is_skipped": False})

    def _greedy_stage(split_rows, stage_list, req, acumulado, covered_row):
        for entry in sorted(stage_list, key=lambda x: x["monto_sail_total"], reverse=True):
            vn_total = entry["vn"]
            precio   = entry["precio"]
            aforo    = entry["aforo_sail"]
            if aforo == 0 or precio == 0:
                split_rows.append({**entry, "vn_usado": vn_total, "monto_usado": 0,
                                   "acumulado": acumulado, "alcanza": False,
                                   "is_remanente": True, "is_skipped": True})
                continue
            if covered_row is not None:
                split_rows.append({**entry, "vn_usado": vn_total,
                                   "monto_usado": entry["monto_sail_total"],
                                   "acumulado": acumulado, "alcanza": False,
                                   "is_remanente": True, "is_skipped": False})
                continue
            faltante = req - acumulado
            vn_n     = _vn_para_cubrir(entry, faltante)
            if vn_n is not None:
                monto_u   = _monto_vn(entry, vn_n)
                acumulado += monto_u
                split_rows.append({**entry, "vn_usado": vn_n, "monto_usado": monto_u,
                                   "acumulado": acumulado, "alcanza": True,
                                   "is_remanente": False, "is_skipped": False})
                covered_row = len(split_rows) - 1
                vn_rem = vn_total - vn_n
                if vn_rem > 0:
                    split_rows.append({**entry,
                                       "origen": entry["origen"] + " (remanente)",
                                       "vn_usado": vn_rem, "monto_usado": _monto_vn(entry, vn_rem),
                                       "acumulado": acumulado, "alcanza": False,
                                       "is_remanente": True, "is_skipped": False})
            else:
                monto_u   = _monto_vn(entry, vn_total)
                acumulado += monto_u
                split_rows.append({**entry, "vn_usado": vn_total, "monto_usado": monto_u,
                                   "acumulado": acumulado, "alcanza": False,
                                   "is_remanente": False, "is_skipped": False})
        return acumulado, covered_row

    # ── STEP 12: Build workbook ───────────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    summary_data     = []
    gallo_rows       = []
    cpr_found_global = []
    _alerts_gara     = []

    for ctte_info in comitentes:
        ctte      = ctte_info["ctte"]
        requerido = ctte_info["requerido"]

        tenencia    = sateclte_tenencia.get(ctte, {})
        all_codigos = set(tenencia.keys())

        saga_ctte = [r for r in saga_rows if r["ctte"] == ctte]
        for r in saga_ctte:
            all_codigos.add(r["codigo"])

        ci_ctte    = ci_ops.get(ctte, {})
        ci_compras = {cod: qty for cod, qty in ci_ctte.items() if qty > 0}
        ci_ventas  = {cod: abs(qty) for cod, qty in ci_ctte.items() if qty < 0}
        for cod in ci_compras:
            all_codigos.add(cod)

        stock = {cod: {"disp": 0.0, "util": 0.0, "ret": 0.0, "cpr": 0.0, "ci_compra": 0.0}
                 for cod in all_codigos}

        for cod, vdict in tenencia.items():
            stock[cod]["disp"] += vdict.get("disp", 0.0)
            stock[cod]["cpr"]  += vdict.get("cpr",  0.0)

        for r in saga_ctte:
            cod = r["codigo"]
            if r["estado"] == "UTILIZADA":
                stock[cod]["util"] += r["qty"]
            else:
                stock[cod]["ret"]  += r["qty"]

        for cod, qty in ci_compras.items():
            stock[cod]["ci_compra"] = qty

        # VTAS/CPRAS en sección GTIA de SATECLTE
        for _ur in _sat_unresolved.get(ctte, []):
            _ur_cod = _ur.get("cod")
            if _ur_cod:
                if _ur["tipo"] == "VTAS":
                    ret_qty = stock.get(_ur_cod, {}).get("ret", 0.0)
                    if ret_qty > 0:
                        forced = min(_ur["qty"], ret_qty)
                        stock.setdefault(_ur_cod, {"disp": 0.0, "util": 0.0, "ret": 0.0, "cpr": 0.0, "ci_compra": 0.0})
                        stock[_ur_cod]["ret"] -= forced
                    disp_qty = stock.get(_ur_cod, {}).get("disp", 0.0)
                    if disp_qty > 0:
                        stock[_ur_cod]["disp"] = max(0.0, disp_qty - min(_ur["qty"], disp_qty))
            else:
                _alert_entry = {
                    "tipo":   _ur["tipo"] + " (especie no identificable — seccion GTIA en SATECLTE)",
                    "ctte":   ctte, "ticker": "???", "cod": "?????", "util": 0.0,
                }
                _alert_entry["venta" if _ur["tipo"] == "VTAS" else "cpr"] = _ur["qty"]
                _alerts_gara.append(_alert_entry)

        # Ventas CI que están en garantía
        ventas_ci_en_gara = {}
        for cod, qty_v in ci_ventas.items():
            util_qty = stock.get(cod, {}).get("util", 0.0)
            if util_qty > 0:
                ventas_ci_en_gara[cod] = min(qty_v, util_qty)

        # Ventas SATECLTE en retirables
        ventas_sat_en_retirables = {}
        for cod, vdict in tenencia.items():
            qty_v   = vdict.get("venta", 0.0)
            ret_qty = stock.get(cod, {}).get("ret", 0.0) if qty_v > 0 else 0.0
            if qty_v > 0 and ret_qty > 0:
                forced = min(qty_v, ret_qty)
                ventas_sat_en_retirables[cod] = forced
                stock[cod]["ret"] -= forced

        # Ventas CI en retirables
        ventas_ci_en_retirables = {}
        for cod, qty_v in ci_ventas.items():
            ret_qty = stock.get(cod, {}).get("ret", 0.0)
            if ret_qty > 0:
                forced = min(qty_v, ret_qty)
                ventas_ci_en_retirables[cod] = forced
                stock[cod]["ret"] -= forced

        # Ventas SATECLTE reducen disponible y generan alertas
        for cod, vdict in tenencia.items():
            qty_v = vdict.get("venta", 0.0)
            if qty_v == 0.0:
                continue
            disp_qty = stock.get(cod, {}).get("disp", 0.0)
            if disp_qty > 0:
                stock[cod]["disp"] = max(0.0, disp_qty - min(qty_v, disp_qty))
            if stock.get(cod, {}).get("util", 0.0) > 0:
                _alerts_gara.append({
                    "tipo": "VTAS", "ctte": ctte,
                    "ticker": especies.get(cod, {}).get("ticker", cod),
                    "cod": cod, "venta": qty_v, "util": stock[cod]["util"],
                })

        # Alerta CPRA sobre UTILIZADA
        for cod, vdict in tenencia.items():
            qty_cpr = vdict.get("cpr", 0.0)
            if qty_cpr > 0 and stock.get(cod, {}).get("util", 0.0) > 0:
                _alerts_gara.append({
                    "tipo": "CPRA", "ctte": ctte,
                    "ticker": especies.get(cod, {}).get("ticker", cod),
                    "cod": cod, "cpr": qty_cpr, "util": stock[cod]["util"],
                })

        # Ventas CI reducen disponible
        for cod, qty_v in ci_ventas.items():
            disp_qty = stock.get(cod, {}).get("disp", 0.0)
            if disp_qty > 0:
                stock[cod]["disp"] = max(0.0, disp_qty - min(qty_v, disp_qty))

        # Build species table
        species_table = []
        for cod in sorted(all_codigos):
            esp         = especies.get(cod, {})
            byma_info   = byma_dict.get(cod)
            aforo_byma  = byma_info[1] if byma_info else 0.0
            aforo_sail  = esp.get("aforo_sail", 0.0)
            tipo_precio = esp.get("tipo_precio", "Normal")
            precio      = precios.get(cod, 0.0)

            vn_disp      = stock[cod]["disp"]
            vn_util      = stock[cod]["util"]
            vn_ret       = stock[cod]["ret"]
            vn_cpr       = stock[cod]["cpr"]
            vn_ci_compra = stock[cod]["ci_compra"]

            msd = msb = 0.0
            estado_esp = "OK"
            if vn_disp > 0:
                if precio == 0:
                    estado_esp = "sin precio"
                elif aforo_sail == 0 and aforo_byma == 0:
                    estado_esp = "sin aforo"
                else:
                    msd = monto_garantizable(vn_disp, cod, aforo_sail)
                    msb = monto_garantizable(vn_disp, cod, aforo_byma)

            msr = mbr = msc = mbc = msi = mbi = 0.0
            if vn_ret       > 0: msr = monto_garantizable(vn_ret,       cod, aforo_sail); mbr = monto_garantizable(vn_ret,       cod, aforo_byma)
            if vn_cpr       > 0: msc = monto_garantizable(vn_cpr,       cod, aforo_sail); mbc = monto_garantizable(vn_cpr,       cod, aforo_byma)
            if vn_ci_compra > 0: msi = monto_garantizable(vn_ci_compra, cod, aforo_sail); mbi = monto_garantizable(vn_ci_compra, cod, aforo_byma)

            species_table.append({
                "codigo": cod, "ticker": esp.get("ticker", cod),
                "nemo": esp.get("nombre", ""), "lista": esp.get("lista", "0"),
                "lam_min": esp.get("lam_min", 0.0),
                "aforo_byma": aforo_byma, "aforo_sail": aforo_sail,
                "vn_disp": vn_disp, "vn_util": vn_util, "vn_ret": vn_ret,
                "vn_cpr": vn_cpr, "vn_ci_compra": vn_ci_compra,
                "vn_disponible_combined": vn_disp + vn_cpr,
                "precio": precio, "tipo_precio": tipo_precio,
                "monto_sail_disp": msd, "monto_byma_disp": msb,
                "monto_sail_ret": msr,  "monto_byma_ret": mbr,
                "monto_sail_cpr": msc,  "monto_byma_cpr": mbc,
                "monto_sail_ci":  msi,  "monto_byma_ci":  mbi,
                "monto_sail": msd + msr + msc + msi,
                "monto_byma": msb + mbr + mbc + mbi,
                "monto_sail_tabla": msd + msc,
                "monto_byma_tabla": msb + mbc,
                "estado": estado_esp,
            })

        total_garantizable_sail = sum(s["monto_sail"] for s in species_table)
        total_garantizable_byma = sum(s["monto_byma"] for s in species_table)
        diferencia_sail         = total_garantizable_sail - requerido
        superavit               = diferencia_sail >= 0

        cpr_items = [(s["ticker"], int(s["vn_cpr"])) for s in species_table if s["vn_cpr"] > 0]
        if cpr_items:
            cpr_found_global.append({"ctte": ctte, "items": cpr_items})

        # ── Split ─────────────────────────────────────────────────────────────
        def _mk_split_entry(origen, s, vn_key, monto_key):
            return {"origen": origen, "codigo": s["codigo"], "ticker": s["ticker"],
                    "vn": s[vn_key], "precio": s["precio"], "aforo_sail": s["aforo_sail"],
                    "monto_sail_total": s[monto_key], "tipo_precio": s["tipo_precio"],
                    "lam_min": s["lam_min"]}

        retirables_list  = [_mk_split_entry("Retirable",  s, "vn_ret",       "monto_sail_ret")
                            for s in species_table if s["vn_ret"]  > 0 and s["monto_sail_ret"]  > 0]
        disponibles_list = [_mk_split_entry("Disponible", s, "vn_disp",      "monto_sail_disp")
                            for s in species_table if s["vn_disp"] > 0 and s["monto_sail_disp"] > 0]
        cpr_list         = [_mk_split_entry("CPR",        s, "vn_cpr",       "monto_sail_cpr")
                            for s in species_table if s["vn_cpr"]  > 0 and s["monto_sail_cpr"]  > 0]
        ci_compra_list   = [_mk_split_entry("Compra CI",  s, "vn_ci_compra", "monto_sail_ci")
                            for s in species_table if s["vn_ci_compra"] > 0 and s["monto_sail_ci"] > 0]

        def _sort_key(e):
            return (e["lam_min"] == 1.0, e["monto_sail_total"])

        retirables_list.sort(key=_sort_key, reverse=True)

        codigos_retirables = {e["codigo"] for e in retirables_list}
        codigos_con_util   = {s["codigo"] for s in species_table if s["vn_util"] > 0}
        disp_prioritarios  = [e for e in disponibles_list if e["codigo"] in codigos_retirables]
        disp_con_util      = [e for e in disponibles_list if e["codigo"] not in codigos_retirables and e["codigo"] in codigos_con_util]
        disp_resto         = [e for e in disponibles_list if e["codigo"] not in codigos_retirables and e["codigo"] not in codigos_con_util]
        for lst in (disp_prioritarios, disp_con_util, disp_resto):
            lst.sort(key=_sort_key, reverse=True)
        disponibles_ordered = disp_prioritarios + disp_con_util + disp_resto

        codigos_previos  = codigos_retirables | {e["codigo"] for e in disponibles_list}
        cpr_prioritarios = [e for e in cpr_list if e["codigo"] in codigos_previos]
        cpr_resto        = [e for e in cpr_list if e["codigo"] not in codigos_previos]
        cpr_prioritarios.sort(key=_sort_key, reverse=True)
        cpr_resto.sort(key=_sort_key, reverse=True)
        cpr_ordered = cpr_prioritarios + cpr_resto

        ci_compra_list.sort(key=_sort_key, reverse=True)

        split_rows  = []
        acumulado   = 0.0
        covered_row = None

        # Phase 1: Retirables → Disponibles
        for stage_list in [retirables_list, disponibles_ordered]:
            if covered_row is not None:
                for entry in stage_list:
                    _agregar_remanente(split_rows, entry, acumulado)
                continue
            faltante  = requerido - acumulado
            best, bvn = _mejor_unico_cobertor(stage_list, faltante)
            n_greedy  = _simular_activos_greedy(stage_list, acumulado, requerido) if best else 0
            if best is not None and n_greedy > 1:
                monto_u    = _monto_vn(best, bvn)
                acumulado += monto_u
                split_rows.append({**best, "vn_usado": bvn, "monto_usado": monto_u,
                                   "acumulado": acumulado, "alcanza": True,
                                   "is_remanente": False, "is_skipped": False})
                covered_row = len(split_rows) - 1
                vn_rem = best["vn"] - bvn
                if vn_rem > 0:
                    split_rows.append({**best, "origen": best["origen"] + " (remanente)",
                                       "vn_usado": vn_rem, "monto_usado": _monto_vn(best, vn_rem),
                                       "acumulado": acumulado, "alcanza": False,
                                       "is_remanente": True, "is_skipped": False})
                for entry in stage_list:
                    if entry["codigo"] != best["codigo"]:
                        _agregar_remanente(split_rows, entry, acumulado)
            else:
                acumulado, covered_row = _greedy_stage(split_rows, stage_list, requerido, acumulado, covered_row)

        # Append devolucion forzada rows
        def _append_devolucion(split_rows, cod_v, forced_qty, origen_label):
            esp_v  = especies.get(cod_v, {})
            prec_v = precios.get(cod_v, 0.0)
            afs_v  = esp_v.get("aforo_sail", 0.0)
            tipo_v = esp_v.get("tipo_precio", "Normal")
            monto_v = forced_qty * (prec_v / 100.0 if tipo_v.lower().startswith("porc") else prec_v) * afs_v
            split_rows.append({
                "origen": origen_label, "codigo": cod_v,
                "ticker": esp_v.get("ticker", cod_v), "vn": forced_qty,
                "precio": prec_v, "aforo_sail": afs_v,
                "monto_sail_total": monto_v, "tipo_precio": tipo_v,
                "lam_min": esp_v.get("lam_min", 0.0),
                "vn_usado": forced_qty, "monto_usado": monto_v,
                "acumulado": acumulado, "alcanza": False,
                "is_remanente": True, "is_skipped": False,
            })

        for cod_v, qty in ventas_sat_en_retirables.items():
            _append_devolucion(split_rows, cod_v, qty, "Retirable (venta T)")
        for cod_v, qty in ventas_ci_en_retirables.items():
            _append_devolucion(split_rows, cod_v, qty, "Retirable (venta CI)")

        # Phase 2: Boost BYMA
        if covered_row is None and acumulado < requerido:
            faltante_boost = requerido - acumulado
            boost_cands    = []
            for s in species_table:
                if s["aforo_byma"] <= s["aforo_sail"]:
                    continue
                vn_avail = s["vn_disp"] + s["vn_ret"]
                if vn_avail <= 0 or s["precio"] <= 0:
                    continue
                delta = s["aforo_byma"] - s["aforo_sail"]
                extra = (vn_avail * (s["precio"] / 100.0) * delta
                         if s["tipo_precio"].lower().startswith("porc")
                         else vn_avail * s["precio"] * delta)
                if extra > 0:
                    boost_cands.append({
                        "codigo": s["codigo"], "ticker": s["ticker"],
                        "vn_avail": vn_avail, "aforo_sail_orig": s["aforo_sail"],
                        "aforo_byma": s["aforo_byma"], "precio": s["precio"],
                        "tipo_precio": s["tipo_precio"], "lam_min": s.get("lam_min", 1),
                        "extra_capacity": extra,
                    })
            boost_cands.sort(key=lambda x: x["extra_capacity"], reverse=True)
            for bc in boost_cands:
                if faltante_boost <= 0:
                    break
                aplicado       = min(bc["extra_capacity"], faltante_boost)
                faltante_boost -= aplicado
                acumulado      += aplicado
                monto_bruto    = (bc["vn_avail"] * (bc["precio"] / 100.0)
                                  if bc["tipo_precio"].lower().startswith("porc")
                                  else bc["vn_avail"] * bc["precio"])
                aforo_ef = (min(bc["aforo_sail_orig"] + aplicado / monto_bruto, bc["aforo_byma"])
                            if monto_bruto > 0 else bc["aforo_byma"])
                split_rows.append({
                    "origen": f"Boost BYMA ({int(bc['aforo_sail_orig']*100)}%->{int(round(aforo_ef*100))}%)",
                    "codigo": bc["codigo"], "ticker": bc["ticker"],
                    "precio": bc["precio"], "aforo_sail": aforo_ef,
                    "aforo_sail_orig": bc["aforo_sail_orig"],
                    "tipo_precio": bc["tipo_precio"], "lam_min": bc.get("lam_min", 1),
                    "monto_sail_total": bc["extra_capacity"],
                    "vn_usado": bc["vn_avail"], "monto_usado": aplicado,
                    "acumulado": acumulado, "alcanza": faltante_boost <= 0,
                    "is_remanente": False, "is_skipped": False,
                    "_boost_new_vn": bc["aforo_sail_orig"] == 0,
                })
                if faltante_boost <= 0:
                    covered_row = len(split_rows) - 1

        # Phase 3: CI Compra
        if covered_row is None and acumulado < requerido:
            faltante  = requerido - acumulado
            best, bvn = _mejor_unico_cobertor(ci_compra_list, faltante)
            n_greedy  = _simular_activos_greedy(ci_compra_list, acumulado, requerido) if best else 0
            if best is not None and n_greedy > 1:
                monto_u    = _monto_vn(best, bvn)
                acumulado += monto_u
                split_rows.append({**best, "vn_usado": bvn, "monto_usado": monto_u,
                                   "acumulado": acumulado, "alcanza": True,
                                   "is_remanente": False, "is_skipped": False})
                covered_row = len(split_rows) - 1
                vn_rem = best["vn"] - bvn
                if vn_rem > 0:
                    split_rows.append({**best, "origen": best["origen"] + " (remanente)",
                                       "vn_usado": vn_rem, "monto_usado": _monto_vn(best, vn_rem),
                                       "acumulado": acumulado, "alcanza": False,
                                       "is_remanente": True, "is_skipped": False})
                for entry in ci_compra_list:
                    if entry["codigo"] != best["codigo"]:
                        _agregar_remanente(split_rows, entry, acumulado)
            else:
                acumulado, covered_row = _greedy_stage(split_rows, ci_compra_list, requerido, acumulado, covered_row)

        # Phase 4: CPR A LIQUIDAR
        if covered_row is None and acumulado < requerido:
            for entry in cpr_ordered:
                precio   = entry["precio"]
                aforo    = entry["aforo_sail"]
                tipo     = entry["tipo_precio"]
                vn_total = entry["vn"]
                if aforo == 0 or precio == 0:
                    split_rows.append({**entry, "origen": "CPR A LIQUIDAR",
                                       "vn_usado": vn_total, "monto_usado": 0,
                                       "acumulado": acumulado, "alcanza": False,
                                       "is_remanente": True, "is_skipped": True})
                    continue
                if covered_row is not None:
                    split_rows.append({**entry, "origen": "CPR A LIQUIDAR (remanente)",
                                       "vn_usado": vn_total, "monto_usado": entry["monto_sail_total"],
                                       "acumulado": acumulado, "alcanza": False,
                                       "is_remanente": True, "is_skipped": False})
                    continue
                faltante = requerido - acumulado
                vn_needed = (faltante / ((precio / 100.0) * aforo)
                             if tipo.lower().startswith("porc") and precio > 0 and aforo > 0
                             else faltante / (precio * aforo) if precio > 0 and aforo > 0
                             else vn_total)
                vn_ceil = math.ceil(vn_needed)
                lam = int(entry["lam_min"]) if entry.get("lam_min", 0) > 1 else 1
                if lam > 1:
                    vn_ceil = math.ceil(vn_ceil / lam) * lam
                if vn_ceil <= vn_total:
                    monto_u    = vn_ceil * (precio / 100.0 if tipo.lower().startswith("porc") else precio) * aforo
                    acumulado += monto_u
                    split_rows.append({**entry, "origen": "CPR A LIQUIDAR",
                                       "vn_usado": vn_ceil, "monto_usado": monto_u,
                                       "acumulado": acumulado, "alcanza": True,
                                       "is_remanente": False, "is_skipped": False})
                    covered_row = len(split_rows) - 1
                    vn_rem = vn_total - vn_ceil
                    if vn_rem > 0:
                        monto_rem = vn_rem * (precio / 100.0 if tipo.lower().startswith("porc") else precio) * aforo
                        split_rows.append({**entry, "origen": "CPR A LIQUIDAR (remanente)",
                                           "vn_usado": vn_rem, "monto_usado": monto_rem,
                                           "acumulado": acumulado, "alcanza": False,
                                           "is_remanente": True, "is_skipped": False})
                else:
                    monto_u    = vn_total * (precio / 100.0 if tipo.lower().startswith("porc") else precio) * aforo
                    acumulado += monto_u
                    split_rows.append({**entry, "origen": "CPR A LIQUIDAR",
                                       "vn_usado": vn_total, "monto_usado": monto_u,
                                       "acumulado": acumulado, "alcanza": False,
                                       "is_remanente": False, "is_skipped": False})

        split_cubierto = covered_row is not None

        # Collect Gallo rows
        for srow in split_rows:
            if srow.get("is_skipped") or srow["is_remanente"]:
                continue
            origen_base = srow["origen"].replace(" (remanente)", "").replace(" (venta T)", "")
            if origen_base == "Retirable":
                accion_g = "EN GARA"
            elif origen_base in ("Disponible", "CPR", "CPR A LIQUIDAR"):
                accion_g = "ENVIAR"
            elif origen_base.startswith("Boost BYMA") and srow.get("_boost_new_vn"):
                accion_g = "ENVIAR"
            else:
                continue
            gallo_rows.append({
                "ctte": ctte, "ticker": srow["ticker"], "codigo": srow["codigo"],
                "vn": int(srow["vn_usado"]), "fecha_vto": ctte_info.get("fecha_vto"),
            })

        # ── Build per-comitente sheet ───────────────────────────────────────
        ws = wb_out.create_sheet(title=f"C-{ctte}")

        fecha_vto_str = ctte_info["fecha_vto"].strftime("%d/%m/%Y") if ctte_info.get("fecha_vto") else "-"
        panel = [
            ("Monto Requerido",           requerido),
            ("Garantizable c/Aforo SAIL", total_garantizable_sail),
            ("Garantizable c/Aforo BYMA", total_garantizable_byma),
            ("Diferencia SAIL",           diferencia_sail),
            ("Estado",                    "SUPERAVIT" if superavit else "DEFICIT"),
            ("Fecha Vto Caucion",         fecha_vto_str),
        ]
        for i, (lbl, val) in enumerate(panel):
            r_idx = i + 1
            cell_a = ws.cell(row=r_idx, column=1, value=lbl)
            cell_a.fill = HEADER_FILL; cell_a.font = HEADER_FONT
            cell_a.alignment = Alignment(horizontal="left")
            cell_a.border = thin_border()
            cell_b = ws.cell(row=r_idx, column=2, value=val)
            cell_b.font = Font(bold=True)
            cell_b.alignment = Alignment(horizontal="right")
            cell_b.border = thin_border()
            if i < 4:
                cell_b.number_format = FMT_INT
            if i == 3:
                cell_b.font = Font(bold=True, color="276221" if superavit else "9C0006")
            if i == 4:
                fill_ = GREEN_FILL if superavit else RED_FILL
                font_ = GREEN_FONT if superavit else RED_FONT
                cell_a.fill = fill_; cell_a.font = font_
                cell_b.fill = fill_; cell_b.font = font_
            if i == 5:
                cell_b.alignment = Alignment(horizontal="left")
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 22

        account_id_val = _ACCOUNT_ID_BY_CTTE.get(ctte, "N/D")
        cell_d1 = ws.cell(row=1, column=4, value="Account ID")
        cell_d1.fill = HEADER_FILL; cell_d1.font = HEADER_FONT
        cell_d1.alignment = Alignment(horizontal="left"); cell_d1.border = thin_border()
        cell_e1 = ws.cell(row=1, column=5, value=account_id_val)
        cell_e1.font = Font(bold=True); cell_e1.alignment = Alignment(horizontal="left")
        cell_e1.border = thin_border()
        ws.column_dimensions["D"].width = 14; ws.column_dimensions["E"].width = 14

        current_row = 8

        # Stock table
        ws.cell(row=current_row, column=1, value="STOCK POR ESTADO").font = BOLD_FONT
        current_row += 1
        for ci, h in enumerate(["Ticker", "Codigo CVSA", "CPR a Liquidar", "DISPONIBLES", "UTILIZADAS", "RETIRABLES"], 1):
            apply_header(ws, current_row, ci, h)
        tot_cpr = tot_disp = tot_util = tot_ret = 0
        alt = False
        for s in species_table:
            current_row += 1
            alt = not alt
            vals = [s["ticker"], s["codigo"],
                    int(s["vn_cpr"])  if s["vn_cpr"]  else "",
                    int(s["vn_disp"]) if s["vn_disp"] else "",
                    int(s["vn_util"]) if s["vn_util"] else "",
                    int(s["vn_ret"])  if s["vn_ret"]  else ""]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=current_row, column=ci, value=v)
                cell.border = thin_border()
                if ci >= 3 and v != "":
                    cell.number_format = FMT_INT
                    cell.alignment = Alignment(horizontal="right")
                if alt:
                    cell.fill = ALT_FILL
            tot_cpr += s["vn_cpr"]; tot_disp += s["vn_disp"]
            tot_util += s["vn_util"]; tot_ret += s["vn_ret"]
        current_row += 1
        for ci, v in enumerate(["TOTALES", "", int(tot_cpr), int(tot_disp), int(tot_util), int(tot_ret)], 1):
            cell = ws.cell(row=current_row, column=ci, value=v)
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER_FILL; cell.border = thin_border()
            if ci >= 3:
                cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")

        # CI compra subsection
        ci_compra_stock = [(s["ticker"], s["codigo"], s["vn_ci_compra"])
                           for s in species_table if s["vn_ci_compra"] > 0]
        if ci_compra_stock:
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            tc = ws.cell(row=current_row, column=1,
                         value="COMPRAS CI — Contado Inmediato (prioridad minima en split)")
            tc.fill = CI_HEADER_FILL; tc.font = Font(color="FFFFFF", bold=True)
            tc.alignment = Alignment(horizontal="left"); tc.border = thin_border()
            current_row += 1
            for tick_ci, cod_ci, vn_ci in ci_compra_stock:
                for col_ci in range(1, 7):
                    ws.cell(row=current_row, column=col_ci).fill = CI_COMPRA_FILL
                    ws.cell(row=current_row, column=col_ci).border = thin_border()
                ws.cell(row=current_row, column=1, value=tick_ci).font = CI_COMPRA_FONT
                ws.cell(row=current_row, column=2, value=cod_ci).font = CI_COMPRA_FONT
                c3 = ws.cell(row=current_row, column=3, value=int(vn_ci) if vn_ci else "")
                c3.number_format = FMT_INT; c3.fill = CI_COMPRA_FILL; c3.border = thin_border()
                c3.font = CI_COMPRA_FONT; c3.alignment = Alignment(horizontal="right")
                current_row += 1
        current_row += 1

        # Ventas CI en garantía alert
        if ventas_ci_en_gara:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            atc = ws.cell(row=current_row, column=1,
                value="*** ATENCION: VENTAS CI EN GARANTIA — estas especies deben BAJAR DE GARANTIA para ser entregadas al mercado ***")
            atc.fill = CI_VENTA_TITLE_FILL; atc.font = Font(color="FFFFFF", bold=True)
            atc.alignment = Alignment(horizontal="center"); atc.border = thin_border()
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            for ci, h in enumerate(["Ticker", "Codigo CVSA", "VN Vendido CI", "VN en Gara", "Accion"], 1):
                cell = ws.cell(row=current_row, column=ci, value=h)
                cell.fill = CI_VENTA_TITLE_FILL; cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center"); cell.border = thin_border()
            current_row += 1
            for cod_v, qty_en_gara in ventas_ci_en_gara.items():
                esp_v = especies.get(cod_v, {})
                qty_vci = ci_ventas.get(cod_v, qty_en_gara)
                for ci, v in enumerate([esp_v.get("ticker", cod_v), cod_v,
                                        int(qty_vci), int(qty_en_gara), "BAJAR DE GARA"], 1):
                    cell = ws.cell(row=current_row, column=ci, value=v)
                    cell.fill = CI_VENTA_ALERT_FILL; cell.font = CI_VENTA_FONT; cell.border = thin_border()
                    if ci in (3, 4):
                        cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                    if ci == 5:
                        cell.alignment = Alignment(horizontal="center")
                current_row += 1
            current_row += 1

        # Species table
        ws.cell(row=current_row, column=1, value="TABLA DE ESPECIES").font = BOLD_FONT
        current_row += 1
        for ci, h in enumerate(["Codigo CVSA", "Ticker", "Nombre", "Lista",
                                 "Aforo BYMA", "Aforo SAIL", "VN Disponible",
                                 "Precio Cierre", "Tipo Precio",
                                 "Monto Garantizable SAIL", "Monto Garantizable BYMA", "Estado"], 1):
            apply_header(ws, current_row, ci, h)
        alt = False
        for s in species_table:
            current_row += 1
            alt = not alt
            vals = [
                s["codigo"], s["ticker"], s["nemo"], s["lista"],
                int(s["aforo_byma"] * 100) if s["aforo_byma"] else "",
                int(s["aforo_sail"] * 100) if s["aforo_sail"] else "",
                int(s["vn_disponible_combined"]) if s["vn_disponible_combined"] else "",
                s["precio"] if s["precio"] else "", s["tipo_precio"],
                int(s["monto_sail_tabla"]) if s["monto_sail_tabla"] else "",
                int(s["monto_byma_tabla"]) if s["monto_byma_tabla"] else "",
                s["estado"],
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=current_row, column=ci, value=v)
                cell.border = thin_border()
                if alt:
                    cell.fill = ALT_FILL
                if ci in (5, 6) and v != "":
                    cell.number_format = '0"%"'
                elif ci == 8 and v != "":
                    cell.number_format = FMT_PRICE
                elif ci in (7, 10, 11) and v != "":
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
        current_row += 1
        for ci, v in enumerate(["TOTALES", "", "", "", "", "", "", "", "",
                                  int(total_garantizable_sail), int(total_garantizable_byma), ""], 1):
            cell = ws.cell(row=current_row, column=ci, value=v)
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER_FILL; cell.border = thin_border()
            if ci in (10, 11) and v != "":
                cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")

        current_row += 2

        # Split table
        if not split_cubierto:
            ws.cell(row=current_row, column=1,
                value="COBERTURA: DEFICIT TOTAL -- Retirables + Disponibles + Boost BYMA + CPR no alcanzan el requerimiento"
            ).font = Font(bold=True, color="9C0006")
        else:
            ws.cell(row=current_row, column=1,
                value="TABLA DE SPLIT -- COBERTURA DEL REQUERIMIENTO").font = BOLD_FONT
            current_row += 1
            for ci, h in enumerate(["Origen", "Ticker", "Codigo CVSA", "VN",
                                     "Precio Cierre", "Aforo Aplic.",
                                     "Monto Garantizable", "Acumulado", "Alcanza?", "Accion"], 1):
                apply_header(ws, current_row, ci, h)
            alt = False
            for srow in split_rows:
                if srow.get("is_skipped"):
                    continue
                current_row += 1
                is_rem   = srow["is_remanente"]
                alcanza  = srow["alcanza"]
                origen_raw  = srow["origen"]
                origen_base = origen_raw.replace(" (remanente)", "").replace(" (venta T)", "").replace(" (venta CI)", "")
                es_ret   = origen_base == "Retirable"
                es_disp  = origen_base == "Disponible"
                es_cpr   = origen_base == "CPR"
                es_ci    = origen_base == "Compra CI"
                es_boost = origen_base.startswith("Boost BYMA")
                es_cpl   = origen_base == "CPR A LIQUIDAR"
                if is_rem:
                    accion      = "DEVOLUCION" if es_ret else "STOCK"
                    accion_fill = DEVOLUCION_FILL if es_ret else (CI_COMPRA_FILL if es_ci else STOCK_FILL)
                else:
                    if es_ret:
                        accion = "EN GARA";        accion_fill = ENGARA_FILL
                    elif es_disp or es_cpr:
                        accion = "ENVIAR";         accion_fill = ENVIAR_FILL
                    elif es_boost:
                        accion = "AFORO BYMA";     accion_fill = PatternFill("solid", fgColor="D9C3E8")
                    elif es_cpl:
                        accion = "CPR A LIQUIDAR"; accion_fill = PatternFill("solid", fgColor="FFEB9C")
                    else:
                        accion = "ENVIAR CI";      accion_fill = CI_COMPRA_FILL

                if not is_rem:
                    if accion == "ENVIAR" and es_disp:
                        row_fill = PatternFill("solid", fgColor="C6EFCE")
                    elif accion in ("ENVIAR", "CPR A LIQUIDAR") and (es_cpr or es_cpl):
                        row_fill = PatternFill("solid", fgColor="FFEB9C")
                    elif accion == "AFORO BYMA":
                        row_fill = PatternFill("solid", fgColor="D9C3E8")
                    elif accion == "ENVIAR CI":
                        row_fill = CI_COMPRA_FILL
                    else:
                        row_fill = None
                else:
                    row_fill = None

                vals = [
                    origen_raw, srow["ticker"], srow["codigo"],
                    int(srow["vn_usado"]),
                    srow["precio"] if srow["precio"] else "",
                    int(srow["aforo_sail"] * 100) if srow["aforo_sail"] else "",
                    int(round(srow["monto_usado"])) if srow["monto_usado"] else "",
                    int(round(srow["acumulado"])) if srow["acumulado"] else "",
                    "CUBIERTO" if alcanza else "",
                    accion,
                ]
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=current_row, column=ci, value=v)
                    cell.border = thin_border()
                    if is_rem:
                        cell.font = GREY_FONT; cell.fill = ALT_FILL
                    else:
                        if ci != 9:
                            if row_fill is not None:
                                cell.fill = row_fill
                            elif alt:
                                cell.fill = ALT_FILL
                    if ci == 4 and v != "":
                        cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                    elif ci == 5 and v != "":
                        cell.number_format = FMT_PRICE; cell.alignment = Alignment(horizontal="right")
                    elif ci == 6 and v != "":
                        cell.number_format = '0"%"'
                    elif ci in (7, 8) and v != "":
                        cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                    if alcanza and ci == 9:
                        cell.font = Font(bold=True, color="276221")
                if not is_rem:
                    alt = not alt
                cell_acc = ws.cell(row=current_row, column=10)
                cell_acc.value = accion; cell_acc.fill = accion_fill
                cell_acc.border = thin_border()
                cell_acc.alignment = Alignment(horizontal="center", vertical="center")
                cell_acc.font = GREY_FONT if is_rem else Font(bold=True)
                if alcanza:
                    ws.cell(row=current_row, column=9).font = Font(bold=True, color="276221")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value) if cell.value is not None else ""))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 45)
        ws.freeze_panes = "A7"

        summary_data.append({
            "ctte": ctte, "requerido": requerido,
            "garantizable_sail": total_garantizable_sail,
            "garantizable_byma": total_garantizable_byma,
            "diferencia_sail": diferencia_sail,
            "estado": "SUPERAVIT" if superavit else "DEFICIT",
        })

    # ── Validate alerts ─────────────────────────────────────────────────────────
    if _alerts_gara:
        msgs = []
        for _a in _alerts_gara:
            vn_op = _a.get("venta") or _a.get("cpr", 0)
            msgs.append(f"Ctte {_a['ctte']} | {_a['tipo']} | {_a['ticker']} ({_a['cod']}) VN={int(vn_op):,} Util={int(_a['util']):,}")
        raise ValueError(
            "VTAS/CPRAS sobre títulos EN GARANTIA — corregir SATECLTE.XLS y volver a ejecutar:\n"
            + "\n".join(msgs)
        )

    # ── VTO sheets ──────────────────────────────────────────────────────────────
    for vto_ctte in vto_cttes:
        saga_vto = [r for r in saga_rows if r["ctte"] == vto_ctte and r["estado"] == "RETIRABLE"]
        ws_vto = wb_out.create_sheet(title=f"VTO-{vto_ctte}")
        ws_vto.sheet_properties.tabColor = VTO_TAB_COLOR

        ws_vto.merge_cells("A1:D1")
        tc = ws_vto["A1"]
        tc.value = f"*** VTO DE CAUCION  ---  Comitente {vto_ctte}  ---  Garantias a recuperar ***"
        tc.fill = VTO_TITLE_FILL; tc.font = Font(color="FFFFFF", bold=True, size=12)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws_vto.row_dimensions[1].height = 22

        ws_vto.merge_cells("A2:D2")
        sc = ws_vto["A2"]
        sc.value = "El comitente cancelo la caucion. Las especies en garantia quedan DISPONIBLES para retiro."
        sc.fill = VTO_SUB_FILL; sc.font = Font(italic=True, color="843C0C", size=10)
        sc.alignment = Alignment(horizontal="center", vertical="center")
        ws_vto.row_dimensions[2].height = 16

        for ci, h in enumerate(["Ticker", "Codigo CVSA", "VN en Gara", "Accion"], 1):
            cell = ws_vto.cell(row=4, column=ci, value=h)
            cell.fill = VTO_HEADER_FILL; cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()

        current_vto_row = 5
        alt = False
        for r in sorted(saga_vto, key=lambda x: x["codigo"]):
            alt = not alt
            esp = especies.get(r["codigo"], {})
            for ci, v in enumerate([esp.get("ticker", r["codigo"]), r["codigo"], int(r["qty"]), "DEVOLUCION"], 1):
                cell = ws_vto.cell(row=current_vto_row, column=ci, value=v)
                cell.border = thin_border()
                if alt:
                    cell.fill = ALT_FILL
                if ci == 3:
                    cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
                if ci == 4:
                    cell.fill = GREEN_FILL; cell.font = Font(bold=True, color="276221")
                    cell.alignment = Alignment(horizontal="center")
            current_vto_row += 1

        total_vn = sum(int(r["qty"]) for r in saga_vto)
        for ci, v in enumerate(["TOTALES", "", total_vn, ""], 1):
            cell = ws_vto.cell(row=current_vto_row, column=ci, value=v)
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = VTO_HEADER_FILL
            cell.border = thin_border()
            if ci == 3:
                cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")

        for col_letter, width in [("A", 14), ("B", 14), ("C", 16), ("D", 16)]:
            ws_vto.column_dimensions[col_letter].width = width

    # ── RESUMEN sheet ────────────────────────────────────────────────────────────
    ws_res = wb_out.create_sheet(title="RESUMEN", index=0)
    for ci, h in enumerate(["Nro Comitente", "Monto Requerido",
                              "Garantizable SAIL", "Garantizable BYMA",
                              "Diferencia SAIL", "Estado"], 1):
        apply_header(ws_res, 1, ci, h)
    apply_header(ws_res, 1, 8, "Account ID")
    apply_header(ws_res, 1, 9, "BC - Total Margin Deficit")
    apply_header(ws_res, 1, 11, "Validacion BC vs Gallo")

    _rmc_by_ctte = {}
    for _c in comitentes:
        _acct = _ACCOUNT_ID_BY_CTTE.get(_c["ctte"])
        if _acct and _acct in _rmc_deficit:
            _rmc_by_ctte[_c["ctte"]] = _rmc_deficit[_acct]

    alt = False
    for i, sd in enumerate(summary_data):
        r_idx = i + 2
        alt   = not alt
        for ci, v in enumerate([sd["ctte"], int(sd["requerido"]),
                                  int(sd["garantizable_sail"]), int(sd["garantizable_byma"]),
                                  int(sd["diferencia_sail"]), sd["estado"]], 1):
            cell = ws_res.cell(row=r_idx, column=ci, value=v)
            cell.border = thin_border()
            if alt:
                cell.fill = ALT_FILL
            if ci in (2, 3, 4, 5):
                cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")
            if ci == 6:
                cell.fill = (GREEN_FILL if v == "SUPERAVIT" else RED_FILL)
                cell.font = (Font(bold=True, color="276221") if v == "SUPERAVIT" else Font(bold=True, color="9C0006"))
        acc_id = _ACCOUNT_ID_BY_CTTE.get(sd["ctte"], "N/D")
        c8 = ws_res.cell(row=r_idx, column=8, value=acc_id)
        c8.border = thin_border()
        if alt:
            c8.fill = ALT_FILL
        _def_rmc = _rmc_by_ctte.get(sd["ctte"])
        c9 = ws_res.cell(row=r_idx, column=9,
                          value=_def_rmc if _def_rmc is not None else "SIN DATO RMC")
        c9.border = thin_border(); c9.alignment = Alignment(horizontal="right")
        if _def_rmc is not None:
            c9.number_format = FMT_PRICE
        if alt:
            c9.fill = ALT_FILL
        if _def_rmc is not None:
            val_k = _def_rmc - sd["requerido"]
            ck = ws_res.cell(row=r_idx, column=11, value=val_k)
            ck.number_format = FMT_PRICE; ck.border = thin_border()
            ck.alignment = Alignment(horizontal="right")
            if val_k > 0:
                ck.fill = PatternFill("solid", fgColor="FCE4D6"); ck.font = Font(bold=True, color="9C0006")
            else:
                ck.fill = PatternFill("solid", fgColor="E2EFDA"); ck.font = Font(color="375623")
        else:
            ck = ws_res.cell(row=r_idx, column=11, value="")
            ck.border = thin_border()
            if alt:
                ck.fill = ALT_FILL

    r_tot = len(summary_data) + 2
    for ci, v in enumerate(["TOTALES",
                              int(sum(s["requerido"] for s in summary_data)),
                              int(sum(s["garantizable_sail"] for s in summary_data)),
                              int(sum(s["garantizable_byma"] for s in summary_data)),
                              int(sum(s["diferencia_sail"] for s in summary_data)), ""], 1):
        cell = ws_res.cell(row=r_tot, column=ci, value=v)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER_FILL; cell.border = thin_border()
        if ci in (2, 3, 4, 5) and v != "":
            cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")

    # RMC extra section
    if _rmc_extra:
        _EXTRA_TITLE_FILL = PatternFill("solid", fgColor="2E4057")
        _EXTRA_HDR_FILL   = PatternFill("solid", fgColor="48688A")
        r_extra = len(summary_data) + 5
        ws_res.merge_cells(f"A{r_extra}:K{r_extra}")
        tc = ws_res.cell(row=r_extra, column=1,
                         value="POSICIONES EN GARANTIA — MAYOR PLAZO (sin distribucion hoy)")
        tc.font = Font(bold=True, color="FFFFFF", size=11); tc.fill = _EXTRA_TITLE_FILL
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws_res.row_dimensions[r_extra].height = 22
        r_extra += 1
        for _ci, _h in {1: "Nro Comitente", 2: "GTIAS c/Aforo BYMA", 3: "GTIAS c/Aforo SAIL",
                        5: "Margen Cobertura", 6: "Estado", 8: "Account ID",
                        9: "BC - Total Margin Deficit", 11: "Validacion BC vs GTIAS BYMA"}.items():
            hc = ws_res.cell(row=r_extra, column=_ci, value=_h)
            hc.font = Font(bold=True, color="FFFFFF"); hc.fill = _EXTRA_HDR_FILL
            hc.alignment = Alignment(horizontal="center"); hc.border = thin_border()
        r_extra += 1
        for _ex in _rmc_extra:
            _saga_ex  = [s for s in saga_rows if s["ctte"] == _ex["ctte"]]
            _val_sail = _val_byma = 0.0
            for _sr in _saga_ex:
                _afs = especies.get(_sr["codigo"], {}).get("aforo_sail", 0.0)
                _afb = byma_dict.get(_sr["codigo"], (None, 0.0))[1]
                _val_sail += monto_garantizable(_sr["qty"], _sr["codigo"], _afs)
                _val_byma += monto_garantizable(_sr["qty"], _sr["codigo"], _afb)
            _margen = _val_byma - _val_sail
            _estado = "CUBRE" if _ex["deficit"] < _val_byma else "NO CUBRE"
            _val_k  = _ex["deficit"] - _val_byma
            for _ci, (_v, _fmt, _aln) in {
                1: (_ex["ctte"], None, "center"), 2: (_val_byma, FMT_PRICE, "right"),
                3: (_val_sail, FMT_PRICE, "right"), 5: (_margen, FMT_PRICE, "right"),
                6: (_estado, None, "center"), 8: (_ex["account_id"], None, "center"),
                9: (_ex["deficit"], FMT_PRICE, "right"), 11: (_val_k, FMT_PRICE, "right"),
            }.items():
                dc = ws_res.cell(row=r_extra, column=_ci, value=_v)
                dc.border = thin_border(); dc.alignment = Alignment(horizontal=_aln)
                if _fmt:
                    dc.number_format = _fmt
                if _ci == 6:
                    dc.fill = (GREEN_FILL if _estado == "CUBRE" else RED_FILL)
                    dc.font = (Font(bold=True, color="276221") if _estado == "CUBRE" else Font(bold=True, color="9C0006"))
                if _ci == 11:
                    if isinstance(_v, float) and _v > 0:
                        dc.fill = PatternFill("solid", fgColor="FCE4D6"); dc.font = Font(bold=True, color="9C0006")
                    else:
                        dc.fill = PatternFill("solid", fgColor="E2EFDA"); dc.font = Font(color="375623")
            r_extra += 1

    for col in ws_res.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value) if cell.value is not None else ""))
            except Exception:
                pass
        ws_res.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)
    ws_res.freeze_panes = "A2"

    # ── GALLO sheet ──────────────────────────────────────────────────────────────
    ws_gallo = wb_out.create_sheet(title="Gallo", index=1)
    for ci, h in enumerate(["CTTE", "TICKER", "Codigo CVSA", "VN", "Fecha Vto"], 1):
        apply_header(ws_gallo, 1, ci, h)

    _gallo_agg = {}
    for gr in gallo_rows:
        key = (gr["ctte"], gr["ticker"], gr["codigo"], gr["fecha_vto"])
        _gallo_agg[key] = _gallo_agg.get(key, 0) + gr["vn"]
    gallo_rows_consolidated = [
        {"ctte": k[0], "ticker": k[1], "codigo": k[2], "fecha_vto": k[3], "vn": v}
        for k, v in _gallo_agg.items()
    ]

    for i, gr in enumerate(gallo_rows_consolidated):
        r_idx     = i + 2
        fecha_str = gr["fecha_vto"].strftime("%d/%m/%Y") if gr["fecha_vto"] else ""
        alt_row   = (i % 2 == 1)
        for ci, v in enumerate([gr["ctte"], gr["ticker"], gr["codigo"], gr["vn"], fecha_str], 1):
            cell = ws_gallo.cell(row=r_idx, column=ci, value=v)
            cell.border = thin_border()
            if alt_row:
                cell.fill = ALT_FILL
            if ci == 4:
                cell.number_format = FMT_INT; cell.alignment = Alignment(horizontal="right")

    for col in ws_gallo.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value) if cell.value is not None else ""))
            except Exception:
                pass
        ws_gallo.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 25)
    ws_gallo.freeze_panes = "A2"

    # ── Save ─────────────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb_out.save(buf)
    buf.seek(0)

    resumen = {
        "fecha":              FECHA_PROCESO.strftime("%d/%m/%Y"),
        "n_comitentes":       len(comitentes),
        "summary_data":       summary_data,
        "validacion_issues":  _validacion_issues,
        "cpr_found":          cpr_found_global,
        "vto_cttes":          vto_cttes,
        "n_gallo_rows":       len(gallo_rows_consolidated),
        "tiene_rmc":          rmc_file is not None,
        "contbole_warning":   contbole_warning,
        "n_superavit":        sum(1 for s in summary_data if s["estado"] == "SUPERAVIT"),
        "n_deficit":          sum(1 for s in summary_data if s["estado"] == "DEFICIT"),
    }
    return buf, resumen
