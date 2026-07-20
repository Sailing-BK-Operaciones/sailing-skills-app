"""
Compras a Liquidar — versión web (Streamlit).

Identifica compras pendientes de liquidar para comitentes con saldo deudor
vencido > $99.900 (que van a recibir toma de caución).

Cruza: SALPESO.XLS (saldos), OPEVEN.XLS (ops a vencer), CONTBOLE.XLS (boletos),
       ESPECIES.XLS (ticker map).
"""

from io import BytesIO
from datetime import datetime, date

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


UMBRAL_DEUDOR = 99_900.0

# Conceptos a vencer y su moneda:
#   Compras: CPRA = ARS | CPU$ = MEP | CPUC = Cable  (importe en 'Importe Neto' o 'Dolar')
#   Ventas:  VTAS = ARS | VTU$ = MEP | VTUC = Cable  (importes vienen negativos en OPEVEN)
COMPRA_CONCEPTOS = {"CPRA": "ARS", "CPU$": "MEP", "CPUC": "Cable"}
VENTA_CONCEPTOS  = {"VTAS": "ARS", "VTU$": "MEP", "VTUC": "Cable"}
_MON_COMPRA      = {"ARS": "opeven_ars", "MEP": "opeven_mep", "Cable": "opeven_cable"}
_MON_VENTA       = {"ARS": "venta_ars",  "MEP": "venta_mep",  "Cable": "venta_cable"}

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
SUBTOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBTOTAL_FONT = Font(bold=True, size=10)
ORANGE_FILL   = PatternFill("solid", fgColor="FFE0B2")
VENTA_FILL    = PatternFill("solid", fgColor="E2EFDA")  # verde claro para ventas
# Cable → resaltado llamativo (violeta fuerte con texto blanco en negrita)
CABLE_FILL    = PatternFill("solid", fgColor="7030A0")
CABLE_FONT    = Font(bold=True, color="FFFFFF", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
FMT_MONEY = "#,##0.00"
FMT_NUM   = "#,##0"


def _read_xls(file_obj, sheet_name):
    """Lee un XLS desde un file-like; devuelve lista de dicts {header: value}."""
    file_obj.seek(0)
    wb = xlrd.open_workbook(file_contents=file_obj.read())
    sh = wb.sheet_by_name(sheet_name)
    headers = [str(sh.cell_value(0, c)).strip("'") for c in range(sh.ncols)]
    out = []
    for r in range(1, sh.nrows):
        out.append({headers[c]: sh.cell_value(r, c) for c in range(sh.ncols)})
    return out


def _clean_ctte(val):
    return str(val).strip().rstrip(".0").strip()


def _parse_date_str(s):
    s = str(s).strip()
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _apply_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER


def _style_data_row(ws, r, money_cols=(), num_cols=(), fill=None):
    for ci in range(1, ws.max_column + 1):
        cell = ws.cell(r, ci)
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="right" if ci in money_cols or ci in num_cols else "left",
            vertical="center",
        )
        if fill:
            cell.fill = fill
    for ci in money_cols:
        ws.cell(r, ci).number_format = FMT_MONEY
    for ci in num_cols:
        ws.cell(r, ci).number_format = FMT_NUM


def generar_reporte(salpeso_file, opeven_file, contbole_file, especies_file):
    """
    Parámetros
    ----------
    salpeso_file   : UploadedFile — SALPESO.XLS (saldos vencidos por comitente)
    opeven_file    : UploadedFile — OPEVEN.XLS (operaciones a vencer)
    contbole_file  : UploadedFile — CONTBOLE.XLS (boletos del día)
    especies_file  : UploadedFile — ESPECIES.XLS (maestro de especies)

    Devuelve (BytesIO con Excel, dict resumen).
    """
    # ── 1. ESPECIES.XLS → mapa código → ticker ───────────────────────────────
    ticker_map = {}   # codigo_5dig -> ticker (Norm. col 9)
    nombre_map = {}   # codigo_5dig -> nombre completo
    esp_rows = _read_xls(especies_file, "Datos_Fijos_Especies")
    for row in esp_rows:
        cod    = str(row.get("Codigo", "")).strip("'").strip().zfill(5)
        ticker = str(row.get("Norm.", "")).strip("'").strip()
        nombre = str(row.get("Nombre_de_la_Especie", "")).strip("'").strip()
        if cod:
            ticker_map[cod] = ticker
            nombre_map[cod] = nombre

    # ── 2. SALPESO → deudores > umbral ───────────────────────────────────────
    salpeso_rows = _read_xls(salpeso_file, "Listado_de_Saldos")
    deudores = {}   # ctte -> {"nombre": str, "saldo": float}
    for row in salpeso_rows:
        ctte = _clean_ctte(row.get("Numero", ""))
        if not ctte:
            continue
        try:
            saldo = float(row.get("Saldo Vencido", 0) or 0)
        except (ValueError, TypeError):
            saldo = 0.0
        if saldo > UMBRAL_DEUDOR:
            nombre = str(row.get("Nombre Comitente", "")).strip()
            deudores[ctte] = {"nombre": nombre, "saldo": saldo}

    # ── 3. OPEVEN → compras y ventas (ARS / MEP / Cable) de deudores ─────────
    opeven_rows = _read_xls(opeven_file, "Operaciones_Vencer")
    compras_opeven = []
    ventas_opeven  = []
    for row in opeven_rows:
        ctte   = _clean_ctte(row.get("Numero", ""))
        concep = str(row.get("Concepto", "")).strip()
        fecliq = str(row.get("Fec.Liq.", "")).strip()
        if not fecliq or ctte not in deudores:
            continue
        if concep in COMPRA_CONCEPTOS:
            moneda, tipo, target = COMPRA_CONCEPTOS[concep], "Compra", compras_opeven
        elif concep in VENTA_CONCEPTOS:
            moneda, tipo, target = VENTA_CONCEPTOS[concep], "Venta", ventas_opeven
        else:
            continue
        try:
            cantidad = float(row.get("Cantidad", 0) or 0)
            if moneda == "ARS":
                importe = float(row.get("Importe Neto", 0) or 0)
            else:  # MEP / Cable → el importe está en la columna 'Dolar'
                importe = float(row.get("Dolar", 0) or 0)
        except (ValueError, TypeError):
            importe = cantidad = 0.0
        # Ventas vienen negativas en OPEVEN → mostrar magnitud positiva
        importe  = abs(importe)
        cantidad = abs(cantidad)
        codigo = str(row.get("Codigo", "")).strip().zfill(5)
        target.append({
            "ctte":     ctte,
            "nombre":   deudores[ctte]["nombre"],
            "tipo":     tipo,
            "origen":   "OPEVEN",
            "moneda":   moneda,
            "codigo":   codigo,
            "ticker":   ticker_map.get(codigo, ""),
            "cantidad": cantidad,
            "importe":  importe,
            "fec_liq":  fecliq,
            "especie":  str(row.get("Nombre de la Especie", "")).strip(),
            "concepto": concep,
            "fec_ope":  str(row.get("Fec.Ope.", "")).strip(),
            "boleto":   str(row.get("Boleto", "")).strip(),
        })

    # ── 4. CONTBOLE → compras CI de deudores ─────────────────────────────────
    compras_ci = []
    cb_rows = _read_xls(contbole_file, "Control_de_Boletos")
    for row in cb_rows:
        operacion = str(row.get("peracion", "")).strip()
        if operacion != "CPRA":
            continue
        ctte = _clean_ctte(row.get("Comitente", ""))
        if ctte not in deudores:
            continue
        fec_ope = str(row.get("Fec_Ope", "")).strip()
        fec_liq = str(row.get("Fec_Liq", "")).strip()
        d_ope = _parse_date_str(fec_ope)
        d_liq = _parse_date_str(fec_liq)
        if d_ope is None or d_liq is None or d_ope != d_liq:
            continue
        try:
            importe  = float(row.get("Total_Neto", 0) or 0)
            cantidad = float(row.get("Valor_Nominal", 0) or 0)
        except (ValueError, TypeError):
            importe = cantidad = 0.0
        especie_cb = str(row.get("Especie", "")).strip()
        cod_cb = next((k for k, v in ticker_map.items() if v == especie_cb), "")
        compras_ci.append({
            "ctte":     ctte,
            "nombre":   deudores[ctte]["nombre"],
            "tipo":     "Compra",
            "origen":   "CONTBOLE CI",
            "moneda":   "ARS",
            "codigo":   cod_cb,
            "ticker":   especie_cb,
            "cantidad": cantidad,
            "importe":  importe,
            "fec_liq":  fec_liq,
            "especie":  nombre_map.get(cod_cb, especie_cb),
            "concepto": "CPRA CI",
            "fec_ope":  fec_ope,
            "boleto":   str(row.get("Boleto", "")).strip(),
        })

    # ── 5. Resumen por comitente (compras + ventas por moneda) ──────────────
    resumen = {}
    for c, d in deudores.items():
        resumen[c] = {"nombre": d["nombre"], "saldo_deudor": d["saldo"],
                      "opeven_ars": 0.0, "opeven_mep": 0.0, "opeven_cable": 0.0,
                      "total_ci": 0.0,
                      "venta_ars": 0.0,  "venta_mep": 0.0,  "venta_cable": 0.0}
    for op in compras_opeven:
        resumen[op["ctte"]][_MON_COMPRA[op["moneda"]]] += op["importe"]
    for op in compras_ci:
        resumen[op["ctte"]]["total_ci"] += op["importe"]
    for op in ventas_opeven:
        resumen[op["ctte"]][_MON_VENTA[op["moneda"]]] += op["importe"]

    all_ops = compras_opeven + compras_ci + ventas_opeven

    # ── 6. Excel ─────────────────────────────────────────────────────────────
    wb = Workbook()

    # Hoja Resumen — 10 cols (Compras ARS/MEP/Cable + CI + Ventas ARS/MEP/Cable)
    ws_res = wb.active
    ws_res.title = "Resumen"
    _apply_header(ws_res, ["Comitente", "Nombre", "Saldo Deudor",
                            "Compras ARS", "Compras MEP (USD)", "Compras Cable (USD)",
                            "Compras CI (ARS)",
                            "Ventas ARS", "Ventas MEP (USD)", "Ventas Cable (USD)"])
    for ctte in sorted(resumen.keys()):
        d = resumen[ctte]
        ws_res.append([ctte, d["nombre"], d["saldo_deudor"],
                       d["opeven_ars"], d["opeven_mep"], d["opeven_cable"],
                       d["total_ci"],
                       d["venta_ars"], d["venta_mep"], d["venta_cable"]])
        r = ws_res.max_row
        # Resaltar si tiene compras en dólares (MEP/Cable) o compras CI del día
        tiene_usd = d["opeven_mep"] > 0 or d["opeven_cable"] > 0
        fill = ORANGE_FILL if (d["total_ci"] > 0 or tiene_usd) else None
        _style_data_row(ws_res, r, money_cols=(3, 4, 5, 6, 7, 8, 9, 10), fill=fill)
        # Cable (compras col F=6 / ventas col J=10) → resaltado llamativo violeta
        tiene_cable = d["opeven_cable"] > 0 or d["venta_cable"] > 0
        for ci, val in ((6, d["opeven_cable"]), (10, d["venta_cable"])):
            if val > 0:
                cell = ws_res.cell(r, ci)
                cell.fill = CABLE_FILL
                cell.font = CABLE_FONT
        if tiene_cable:
            for ci in (1, 2):
                ws_res.cell(r, ci).font = Font(bold=True, color="7030A0", size=10)

    tot_saldo   = sum(d["saldo_deudor"] for d in resumen.values())
    tot_ars     = sum(d["opeven_ars"]   for d in resumen.values())
    tot_mep     = sum(d["opeven_mep"]   for d in resumen.values())
    tot_cable   = sum(d["opeven_cable"] for d in resumen.values())
    tot_ci      = sum(d["total_ci"]     for d in resumen.values())
    tot_v_ars   = sum(d["venta_ars"]    for d in resumen.values())
    tot_v_mep   = sum(d["venta_mep"]    for d in resumen.values())
    tot_v_cable = sum(d["venta_cable"]  for d in resumen.values())
    ws_res.append(["", "TOTAL", tot_saldo, tot_ars, tot_mep, tot_cable, tot_ci,
                   tot_v_ars, tot_v_mep, tot_v_cable])
    r = ws_res.max_row
    for ci in range(1, 11):
        cell = ws_res.cell(r, ci)
        cell.font   = SUBTOTAL_FONT
        cell.fill   = SUBTOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="right" if ci >= 3 else "left", vertical="center"
        )
        if ci >= 3:
            cell.number_format = FMT_MONEY
    # Totales de Cable también resaltados
    for ci, val in ((6, tot_cable), (10, tot_v_cable)):
        if val > 0:
            ws_res.cell(r, ci).fill = CABLE_FILL
            ws_res.cell(r, ci).font = CABLE_FONT

    ws_res.column_dimensions["A"].width = 12
    ws_res.column_dimensions["B"].width = 30
    for col in ["C", "D", "E", "F", "G", "H", "I", "J"]:
        ws_res.column_dimensions[col].width = 20
    ws_res.freeze_panes = "A2"
    if ws_res.max_row > 2:
        ws_res.auto_filter.ref = f"A1:J{ws_res.max_row - 1}"

    # Hoja Detalle — 15 cols (agrega Tipo en pos D)
    ws_det = wb.create_sheet("Detalle")
    _apply_header(ws_det, [
        "Comitente", "Nombre", "Saldo Deudor", "Tipo", "Origen",
        "Especie", "ticker", "Moneda", "Cantidad", "Importe",
        "Fec. Liq.", "Nombre Especie", "Concepto", "Fec. Ope.", "Boleto",
    ])
    for op in sorted(all_ops, key=lambda x: (x["ctte"], x["tipo"], x["moneda"], x["origen"], x["fec_liq"])):
        saldo = resumen[op["ctte"]]["saldo_deudor"]
        ws_det.append([
            op["ctte"],    op["nombre"],  saldo,          op["tipo"],     op["origen"],
            op["codigo"],  op["ticker"],  op["moneda"],   op["cantidad"], op["importe"],
            op["fec_liq"], op["especie"], op["concepto"], op["fec_ope"],  op["boleto"],
        ])
        r = ws_det.max_row
        # Prioridad: Cable (violeta) > Venta (verde) > CI/USD (naranja) > default
        es_cable = op["moneda"] == "Cable"
        if es_cable:
            fill = CABLE_FILL
        elif op["tipo"] == "Venta":
            fill = VENTA_FILL
        elif op["origen"] == "CONTBOLE CI" or op["moneda"] != "ARS":
            fill = ORANGE_FILL
        else:
            fill = None
        _style_data_row(ws_det, r, money_cols=(3, 10), num_cols=(9,), fill=fill)
        # Cable (compra o venta) → fila completa en negrita sobre fondo llamativo
        if es_cable:
            for ci in range(1, ws_det.max_column + 1):
                ws_det.cell(r, ci).font = CABLE_FONT

    col_widths = {
        "A": 12, "B": 30, "C": 18, "D": 8,  "E": 14, "F": 10, "G": 12, "H": 8,
        "I": 14, "J": 18, "K": 12, "L": 35, "M": 12, "N": 12, "O": 10,
    }
    for col, w in col_widths.items():
        ws_det.column_dimensions[col].width = w
    ws_det.freeze_panes = "A2"
    if ws_det.max_row >= 1:
        ws_det.auto_filter.ref = f"A1:O{max(ws_det.max_row, 1)}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    cttes_con_ci    = sorted({op["ctte"] for op in compras_ci})
    cttes_con_usd   = sorted({op["ctte"] for op in compras_opeven if op["moneda"] != "ARS"})
    cttes_con_venta = sorted({op["ctte"] for op in ventas_opeven})
    resumen_ui = {
        "fecha":            date.today().strftime("%d-%m-%Y"),
        "n_deudores":       len(deudores),
        "n_ops_opeven":     len(compras_opeven),
        "n_ops_ci":         len(compras_ci),
        "n_ops_venta":      len(ventas_opeven),
        "total_opeven_ars":   tot_ars,
        "total_opeven_mep":   tot_mep,
        "total_opeven_cable": tot_cable,
        "total_ci":         tot_ci,
        "total_venta_ars":   tot_v_ars,
        "total_venta_mep":   tot_v_mep,
        "total_venta_cable": tot_v_cable,
        "total_deudor":     tot_saldo,
        "cttes_con_ci":     cttes_con_ci,
        "cttes_con_usd":    cttes_con_usd,
        "cttes_con_venta":  cttes_con_venta,
        "detalle_deudores": sorted(
            [{"ctte": c, **d,
              "total_opeven": d["opeven_ars"] + d["opeven_mep"] + d["opeven_cable"],
              "total_venta":  d["venta_ars"]  + d["venta_mep"]  + d["venta_cable"],
              "total": d["opeven_ars"] + d["opeven_mep"] + d["opeven_cable"] + d["total_ci"]}
             for c, d in resumen.items()],
            key=lambda x: x["ctte"],
        ),
    }
    return output, resumen_ui
