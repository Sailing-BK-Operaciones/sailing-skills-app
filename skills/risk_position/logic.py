"""
Risk Position — versión web (Streamlit).
Compara saldos de garantía BYMA Clearing vs Saldos Gara a cubrir (Gallo).
Recibe 3 UploadedFiles; devuelve (BytesIO con el Excel modificado, dict con resumen).
"""

import re
from io import BytesIO

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

FILL_MISMATCH = PatternFill("solid", fgColor="FFEB9C")
FILL_ONLY_BC  = PatternFill("solid", fgColor="FFC7CE")
FMT_IMPORTE   = "#,##0.00"

AMOUNT_COL  = "Short Quantity"
ACCOUNT_COL = "Account"
DATA_START_ROW = 3


def _load_accounts_map(accounts_file):
    """Devuelve dict {ctte_str -> account_id}.
    El archivo es semicolon-delimited con formato: "AGENTE;COMITENTE;ID_CLEARING".
    Se filtra por AGENTE == "233" (Sailing).
    """
    content = accounts_file.read().decode("utf-8-sig")
    mapping = {}
    for line in content.splitlines()[1:]:   # saltar header
        line = line.strip().strip('"')
        if not line:
            continue
        parts = line.split(";")
        if len(parts) != 3:
            continue
        agente, ctte, acct_id = parts[0].strip(), parts[1].strip(), parts[2].strip()
        if agente == "233" and ctte.isdigit() and acct_id:
            mapping[ctte] = acct_id
    return mapping


def _extract_ctte(account_str):
    m = re.search(r"TM_SAILING ALYC_(\d+)", str(account_str))
    return m.group(1) if m else None


def _extract_bc_id(account_str):
    s = str(account_str)
    if "(" in s and s.endswith(")"):
        return s.rsplit("(", 1)[1].rstrip(")")
    return None


def generar_reporte(csv_risk_file, csv_accounts_file, saldos_file):
    """
    csv_risk_file   : table-riskPositions_*.csv
    csv_accounts_file: table-accounts_*.csv
    saldos_file     : Saldos Gara a cubrir.xlsx (se modifica y se devuelve)

    Devuelve (BytesIO, resumen_dict).
    """
    # 1. Tabla de cuentas
    accounts_map = _load_accounts_map(csv_accounts_file)

    # 2. CSV de risk position
    df_risk = pd.read_csv(csv_risk_file, dtype=str)
    cols = df_risk.columns.tolist()
    account_col = ACCOUNT_COL if ACCOUNT_COL in cols else None
    amount_col  = AMOUNT_COL  if AMOUNT_COL  in cols else None
    if account_col is None or amount_col is None:
        raise ValueError(
            f"No se encontraron las columnas esperadas en el CSV.\n"
            f"Columnas disponibles: {cols}\n"
            f"Se esperaban: '{ACCOUNT_COL}' y '{AMOUNT_COL}'"
        )

    # 3. Leer fecha de proceso y TC BYMA desde Saldos Gara (data_only para valores reales)
    saldos_bytes = saldos_file.read()
    wb_tmp = openpyxl.load_workbook(BytesIO(saldos_bytes), data_only=True)
    ws_tmp = wb_tmp.active
    fecha_proceso = None
    for r in range(DATA_START_ROW, ws_tmp.max_row + 1):
        val = ws_tmp.cell(r, 3).value
        if val is not None:
            fecha_proceso = val.strftime("%Y-%m-%d") if hasattr(val, "strftime") else str(val).strip()[:10]
            break
    tc_raw = ws_tmp.cell(1, 7).value
    wb_tmp.close()

    if fecha_proceso is None:
        raise ValueError("No se encontró fecha en columna C de Saldos Gara a cubrir.")
    if tc_raw is None:
        raise ValueError("No se encontró TC BYMA CLEARING en celda G1 de Saldos Gara a cubrir.")
    try:
        tc_byma = float(str(tc_raw).replace(",", "."))
    except (ValueError, TypeError):
        raise ValueError(f"TC BYMA CLEARING en G1 no es un número válido: {tc_raw!r}")

    # 4. Filtrar cauciones tomadoras del día
    mask_actual   = df_risk.get("Position Type", pd.Series(dtype=str)) == "ACTUAL"
    mask_fecha    = df_risk.get("Settlement Date", pd.Series(dtype=str)) == fecha_proceso
    lq = pd.to_numeric(df_risk.get("Long Quantity", 0), errors="coerce").fillna(0)
    mask_tomadora = lq <= 0
    node_col = "Market Filter Node"
    mask_caucion = (
        df_risk[node_col].str.startswith("AR.BYMA.CT.U.ARS", na=False) |
        df_risk[node_col].str.startswith("AR.BYMA.CT.U.USD", na=False)
    )
    df_filtrado = df_risk[mask_actual & mask_caucion & mask_fecha & mask_tomadora]

    risk_by_ctte    = {}
    risk_account_id = {}

    for _, row in df_filtrado.iterrows():
        raw_acct = str(row[account_col])
        ctte = _extract_ctte(raw_acct)
        if ctte is None:
            continue
        if ctte not in risk_account_id:
            acct_id = accounts_map.get(ctte) or _extract_bc_id(raw_acct)
            risk_account_id[ctte] = acct_id
        try:
            amount = abs(float(str(row[amount_col]).replace(",", ".")))
        except (ValueError, TypeError):
            amount = 0.0
        if str(row.get(node_col, "")).startswith("AR.BYMA.CT.U.USD"):
            amount = amount * tc_byma
        risk_by_ctte[ctte] = risk_by_ctte.get(ctte, 0.0) + amount

    # 5. Abrir y modificar Saldos Gara a cubrir
    wb = openpyxl.load_workbook(BytesIO(saldos_bytes))
    ws = wb.active

    # Determinar última fila real de Gallo
    last_gallo_row = DATA_START_ROW - 1
    for r in range(DATA_START_ROW, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).isdigit() and int(str(v)) > 0:
            last_gallo_row = r
        elif v is None and r > DATA_START_ROW:
            break

    # Limpiar sección solo-en-BC de corrida anterior
    for r in range(last_gallo_row + 1, ws.max_row + 2):
        for c in range(1, 10):
            cell = ws.cell(r, c)
            if hasattr(cell, "value"):
                cell.value = None

    cttes_en_saldos = set()
    last_data_row   = DATA_START_ROW - 1

    for row_idx in range(DATA_START_ROW, last_gallo_row + 1):
        ctte_val = ws.cell(row_idx, 1).value
        if ctte_val is None:
            break
        last_data_row = row_idx
        ctte_str = str(int(ctte_val)) if isinstance(ctte_val, (int, float)) else str(ctte_val)
        cttes_en_saldos.add(ctte_str)

        account_id = risk_account_id.get(ctte_str) or accounts_map.get(ctte_str, "")
        if account_id is None:
            account_id = ""
        ws.cell(row_idx, 5).value = int(account_id) if str(account_id).isdigit() else account_id
        ws.cell(row_idx, 5).alignment = Alignment(horizontal="center")

        importe_bc = risk_by_ctte.get(ctte_str)
        cell_f = ws.cell(row_idx, 6)
        if importe_bc is not None:
            cell_f.value = importe_bc
            cell_f.number_format = FMT_IMPORTE
        else:
            cell_f.value = "SIN DATO BC"
            cell_f.fill = FILL_MISMATCH

        importe_gallo = ws.cell(row_idx, 2).value or 0
        cell_h = ws.cell(row_idx, 8)
        if importe_bc is not None:
            diferencia = importe_bc - importe_gallo
            cell_h.value = diferencia
            cell_h.number_format = FMT_IMPORTE
            if abs(diferencia) > 0.01:
                cell_h.fill = FILL_MISMATCH
        else:
            cell_h.value = "N/A"

    # Sección solo-en-BC
    cttes_en_bc   = set(risk_by_ctte.keys())
    solo_en_bc    = cttes_en_bc - cttes_en_saldos
    solo_en_gallo = cttes_en_saldos - cttes_en_bc

    if solo_en_bc:
        row_idx = last_data_row + 2
        ws.cell(row_idx, 1).value = "SOLO EN BC (no están en Saldos Gara a cubrir)"
        ws.cell(row_idx, 1).font = Font(bold=True, color="9C0006")
        row_idx += 1
        ws.cell(row_idx, 1).value = "CTTE"
        ws.cell(row_idx, 5).value = "ACCOUNT ID BC"
        ws.cell(row_idx, 6).value = "IMPORTE BC"
        for c in [1, 5, 6]:
            ws.cell(row_idx, c).font = Font(bold=True)
        row_idx += 1
        for ctte_str in sorted(solo_en_bc, key=lambda x: int(x) if x.isdigit() else 0):
            ws.cell(row_idx, 1).value = int(ctte_str) if ctte_str.isdigit() else ctte_str
            ws.cell(row_idx, 5).value = accounts_map.get(ctte_str, "")
            cell_f = ws.cell(row_idx, 6)
            cell_f.value = risk_by_ctte.get(ctte_str, 0)
            cell_f.number_format = FMT_IMPORTE
            for c in range(1, 9):
                ws.cell(row_idx, c).fill = FILL_ONLY_BC
            row_idx += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    resumen = {
        "fecha_proceso": fecha_proceso,
        "tc_byma": tc_byma,
        "n_gallo": len(cttes_en_saldos),
        "n_bc": len(cttes_en_bc),
        "n_match": len(cttes_en_saldos & cttes_en_bc),
        "solo_en_gallo": sorted(solo_en_gallo),
        "solo_en_bc": sorted(solo_en_bc),
        "n_caucion_rows": len(df_filtrado),
    }
    return output, resumen
