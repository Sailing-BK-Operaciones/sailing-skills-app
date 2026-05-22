import re
import io
from collections import defaultdict

import openpyxl
import xlrd

HOUSE_COMITENTES = {1000, 1001, 1002, 1003}

HEADER = (
    "InstructingParty;SettlementParty;SecuritiesAccount;Instrument;"
    "InstrumentIdentifierType;CSDOfCounterparty;SettlementCounterparty;"
    "SecuritiesAccountOfCounterparty;InstructionReference;"
    "Instrument(MovementOfSecurities);Quantity;QuantityType;TransactionType;"
    "SettlementMethod;TradeDate;IntendedSettlementDate;PaymentType"
)


def _load_cvsa_qty_type(especies_file):
    especies_file.seek(0)
    wb = xlrd.open_workbook(file_contents=especies_file.read())
    ws = wb.sheet_by_name('Datos_Fijos_Especies')
    result = {}
    for r in range(1, ws.nrows):
        cvsa = str(ws.cell_value(r, 0)).strip()
        tipo = str(ws.cell_value(r, 4)).strip()
        if cvsa:
            result[cvsa] = 'FACE_AMOUNT' if tipo == 'Porc.' else 'UNIT'
    return result


def _extract_date(filename):
    # DD-MM-YYYY
    m = re.search(r'(\d{2})-(\d{2})-(\d{4})', filename)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        return f"{y}{mo}{d}", f"{d}-{mo}-{y[2:]}"
    # DD-MM-YY
    m = re.search(r'(\d{2})-(\d{2})-(\d{2})(?!\d)', filename)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        return f"20{y}{mo}{d}", f"{d}-{mo}-{y}"
    raise ValueError(f"No se pudo extraer fecha de: {filename}")


def _strip_zeros(cvsa):
    try:
        return str(int(str(cvsa).strip()))
    except (ValueError, TypeError):
        return str(cvsa)


def _parse_c_sheet(ws, cvsa_qty_type):
    all_rows = list(ws.iter_rows(values_only=True))
    especies_start = split_start = None
    for i, row in enumerate(all_rows):
        if row[0] == 'TABLA DE ESPECIES':
            especies_start = i
        elif row[0] == 'TABLA DE SPLIT -- COBERTURA DEL REQUERIMIENTO':
            split_start = i

    ticker_to_cvsa = {}
    if especies_start is not None:
        for row in all_rows[especies_start + 2:]:
            if row[0] is None or row[0] == 'TOTALES':
                break
            cvsa = str(row[0]).strip()
            if row[1]:
                ticker_to_cvsa[row[1]] = cvsa

    DELIVER = {'ENVIAR', 'ENVIAR CI'}
    RECEIVE = {'DISPONIBLE', 'DISPONIBLES', 'BAJAR DE GARA', 'DEVOLUCION'}
    SKIP    = {'STOCK', 'EN GARA'}

    enviar, disponible, unknowns = [], [], []
    if split_start is not None:
        for row in all_rows[split_start + 2:]:
            if row[0] is None and row[1] is None:
                break
            if row[0] is None:
                continue
            accion = str(row[9]).strip() if row[9] is not None else ''
            if accion in SKIP:
                continue
            if accion not in DELIVER and accion not in RECEIVE:
                unknowns.append(accion)
                continue
            ticker = row[1]
            cvsa = str(row[2]).strip() if row[2] is not None else ticker_to_cvsa.get(ticker, '')
            vn = row[3]
            entry = {
                'ticker': ticker,
                'cvsa': cvsa,
                'instrument': _strip_zeros(cvsa),
                'vn': int(vn) if vn else 0,
                'qty_type': cvsa_qty_type.get(cvsa, 'UNIT'),
            }
            (enviar if accion in DELIVER else disponible).append(entry)

    return enviar, disponible, unknowns


def _parse_vto_sheet(ws, cvsa_qty_type):
    all_rows = list(ws.iter_rows(values_only=True))
    data_start = None
    for i, row in enumerate(all_rows):
        if row[0] == 'Ticker':
            data_start = i + 1
            break
    if data_start is None:
        return []
    result = []
    for row in all_rows[data_start:]:
        if row[0] is None or str(row[0]).strip() == 'TOTALES':
            break
        ticker = row[0]
        cvsa = str(row[1]).strip() if row[1] is not None else ''
        vn = int(row[2]) if row[2] is not None else 0
        if ticker and cvsa and vn > 0:
            result.append({
                'ticker': ticker,
                'cvsa': cvsa,
                'instrument': _strip_zeros(cvsa),
                'vn': vn,
                'qty_type': cvsa_qty_type.get(cvsa, 'UNIT'),
            })
    return result


def _counterparty(comitente):
    if comitente in HOUSE_COMITENTES:
        return "80233/222222222", "HOUSE"
    return "80233/555555555", "CLIENT"


def _build_si2(rows_by_comitente, movement, date_yyyymmdd, date_ref, prefix, start):
    lines = [HEADER]
    counter = start
    for comitente, entries in rows_by_comitente:
        cp, _ = _counterparty(comitente)
        for e in entries:
            ref = f"{prefix}{date_ref}{counter:03d}"
            lines.append(
                f"233;233;233/{comitente};{e['instrument']};LOCAL_CODE;CVSA;233;"
                f"{cp};{ref};{movement};{e['vn']};"
                f"{e['qty_type']};TRAD;RTGS;{date_yyyymmdd};{date_yyyymmdd};NOTHING;NORMAL"
            )
            counter += 1
    n = counter - start
    return '\n'.join(lines).encode('utf-8'), n


def _build_resumen_totals(rows_by_comitente):
    client = defaultdict(lambda: {'ticker': '', 'vn': 0})
    house  = defaultdict(lambda: {'ticker': '', 'vn': 0})
    for comitente, entries in rows_by_comitente:
        _, nodo = _counterparty(comitente)
        target = client if nodo == 'CLIENT' else house
        for e in entries:
            target[e['cvsa']]['ticker'] = e['ticker']
            target[e['cvsa']]['vn'] += e['vn']
    return client, house


def _build_resumen_txt(client, house):
    lines = []
    if client:
        lines += ["NODO CLIENT 80233/555555555", "Codigo CVSA;Ticker;VN"]
        for cvsa in sorted(client):
            lines.append(f"{cvsa};{client[cvsa]['ticker']};{client[cvsa]['vn']}")
    if house:
        if lines:
            lines.append("")
        lines += ["NODO HOUSE 80233/222222222", "Codigo CVSA;Ticker;VN"]
        for cvsa in sorted(house):
            lines.append(f"{cvsa};{house[cvsa]['ticker']};{house[cvsa]['vn']}")
    return '\n'.join(lines).encode('utf-8')


def generar_archivos(distribucion_file, especies_file, counter_state=None):
    """
    Returns:
        outputs  — dict con claves: req_envio, resumen_deposit, ret_devolucion, resumen_withdraw
                   cada valor es (bytes, filename) o None; refs opcionales en *_refs
        resumen  — dict con métricas de la ejecución
        new_counter — counter_state actualizado (persistir en session_state)
    """
    if counter_state is None:
        counter_state = {}

    cvsa_qty_type = _load_cvsa_qty_type(especies_file)

    distribucion_file.seek(0)
    wb = openpyxl.load_workbook(io.BytesIO(distribucion_file.read()), data_only=True)

    filename = getattr(distribucion_file, 'name', 'Distribucion-Gara-Byma-01-01-2025.xlsx')
    date_yyyymmdd, date_dd_mm_aa = _extract_date(filename)
    date_ref = date_dd_mm_aa.replace('-', '')  # DDMMAA sin guiones

    enviar_by_ctte = []
    disponible_dict = {}
    unknown_all = []
    sheet_log = []
    vto_sheets = []

    for sname in wb.sheetnames:
        if not sname.startswith("C-"):
            continue
        try:
            ctte = int(sname[2:])
        except ValueError:
            continue
        enviar, disponible, unknowns = _parse_c_sheet(wb[sname], cvsa_qty_type)
        sheet_log.append(f"{sname}: {len(enviar)} ENVIAR, {len(disponible)} DISPONIBLE")
        unknown_all.extend(unknowns)
        if enviar:
            enviar_by_ctte.append((ctte, enviar))
        if disponible:
            disponible_dict.setdefault(ctte, []).extend(disponible)

    for sname in wb.sheetnames:
        if not sname.upper().startswith("VTO-"):
            continue
        try:
            ctte = int(sname.split("-", 1)[1])
        except (ValueError, IndexError):
            continue
        disponible = _parse_vto_sheet(wb[sname], cvsa_qty_type)
        if disponible:
            vto_sheets.append(sname)
            disponible_dict.setdefault(ctte, []).extend(disponible)

    disponible_by_ctte = [
        (ctte, disponible_dict[ctte]) for ctte in sorted(disponible_dict)
    ]

    outputs = {}
    new_counter = dict(counter_state)

    if enviar_by_ctte:
        tge_key = f"TGE{date_ref}"
        start = new_counter.get(tge_key, 0) + 1
        content, n = _build_si2(enviar_by_ctte, "DELIVER", date_yyyymmdd, date_ref, "TGE", start)
        new_counter[tge_key] = start + n - 1
        outputs['req_envio'] = (content, f"Req-envio de gtias {date_dd_mm_aa}.SI2")
        outputs['req_envio_refs'] = (start, start + n - 1, n)

        client_t, house_t = _build_resumen_totals(enviar_by_ctte)
        outputs['resumen_deposit'] = (_build_resumen_txt(client_t, house_t), f"Resumen DEPOSIT {date_dd_mm_aa}.txt")
        outputs['resumen_deposit_stats'] = (len(client_t), len(house_t))
    else:
        outputs['req_envio'] = None
        outputs['resumen_deposit'] = None

    if disponible_by_ctte:
        tgd_key = f"TGD{date_ref}"
        start = new_counter.get(tgd_key, 0) + 1
        content, n = _build_si2(disponible_by_ctte, "RECEIVE", date_yyyymmdd, date_ref, "TGD", start)
        new_counter[tgd_key] = start + n - 1
        outputs['ret_devolucion'] = (content, f"Ret-devolucion gtia {date_dd_mm_aa}.SI2")
        outputs['ret_devolucion_refs'] = (start, start + n - 1, n)

        client_t, house_t = _build_resumen_totals(disponible_by_ctte)
        outputs['resumen_withdraw'] = (_build_resumen_txt(client_t, house_t), f"Resumen WITHDRAW {date_dd_mm_aa}.txt")
        outputs['resumen_withdraw_stats'] = (len(client_t), len(house_t))
    else:
        outputs['ret_devolucion'] = None
        outputs['resumen_withdraw'] = None

    resumen = {
        'fecha': date_dd_mm_aa,
        'n_comitentes': len(set(c for c, _ in enviar_by_ctte) | set(disponible_dict)),
        'n_enviar': sum(len(e) for _, e in enviar_by_ctte),
        'n_disponible': sum(len(e) for _, e in disponible_by_ctte),
        'vto_sheets': vto_sheets,
        'unknown_actions': sorted(set(unknown_all)),
        'sheet_log': sheet_log,
    }

    return outputs, resumen, new_counter
