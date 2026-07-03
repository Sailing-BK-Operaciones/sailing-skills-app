"""
Reporte Operativo — Boletos
Procesa CONTBOLE (XLS de Gallo) y actualiza el reporte Excel acumulativo mensual.
"""

import io
from collections import defaultdict, Counter
from datetime import datetime

import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constantes ────────────────────────────────────────────────────────────────
MESES_ORDER = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]

# Cuentas de cartera propia (excluidas de la sección Clientes, van a Cartera)
CUENTAS_EXCLUIDAS = {"1000", "1003", "1002", "1060", "2583", "1854"}

HB_CANALS = {'WEB', 'APP', 'MgW'}

TC_HISTORICOS = {
    'Enero':    {'MEP': 1464.75, 'CCL': 1510.53},
    'Febrero':  {'MEP': 1436.64, 'CCL': 1476.11},
    'Marzo':    {'MEP': 1488.00, 'CCL': 1499.00},
    'Abril':    {'MEP': 1448.93, 'CCL': 1503.03},
}

# Colores (sin prefijo FF para openpyxl directo)
C_DARK   = "1A3A4A"
C_PRIME  = "0E7FAD"
C_LIGHT  = "E8F4F8"
C_GREEN  = "27AE60"
C_ORANGE = "F39C12"
C_VIOLET = "6E2F8A"
C_GREY   = "5D6D7E"
C_BLUE2  = "1A5276"
C_WHITE  = "FFFFFF"
C_BLACK  = "000000"
C_GREY_L = "7F8C8D"
C_CYAN   = "2BB5E0"


# ─── Style helpers ─────────────────────────────────────────────────────────────
def _fill(c):
    if not c:
        return PatternFill(fill_type=None)
    return PatternFill(fill_type='solid', fgColor=c.lstrip('F') if len(c) == 8 else c)

def _font(bold=False, italic=False, size=10, color=C_BLACK, name='Calibri'):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set(ws, r, c, value=None, bold=False, italic=False, size=10,
         color=C_BLACK, bg=None, halign='left', numfmt=None, wrap=False):
    cell = ws.cell(r, c)
    cell.value = value
    cell.font  = _font(bold=bold, italic=italic, size=size, color=color)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = _align(halign, 'center', wrap)
    if numfmt:
        cell.number_format = numfmt
    return cell

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# ─── Carga TABCOMPB ────────────────────────────────────────────────────────────
def load_tabcompb(file_obj):
    """Lee TABCOMPB.XLS y devuelve dict {abrev: {nombre, mercado, segmento, status}}."""
    file_obj.seek(0)
    wb = xlrd.open_workbook(file_contents=file_obj.read())
    ws = wb.sheet_by_name('Tabla_Comprobantes')
    d = {}
    for r in range(1, ws.nrows):
        abrev = str(ws.cell_value(r, 2)).strip()
        if abrev:
            d[abrev] = {
                'nombre':   str(ws.cell_value(r, 1)).strip(),
                'mercado':  str(ws.cell_value(r, 3)).strip(),
                'segmento': str(ws.cell_value(r, 4)).strip(),
                'status':   str(ws.cell_value(r, 5)).strip(),
            }
    return d


# ─── Carga CONTBOLE ────────────────────────────────────────────────────────────
def load_contbole(file_obj):
    """Lee CONTBOLE.XLS (hoja Control_de_Boletos) y devuelve lista de dicts."""
    file_obj.seek(0)
    wb = xlrd.open_workbook(file_contents=file_obj.read())
    ws = wb.sheet_by_name('Control_de_Boletos')
    rows = []
    for r in range(1, ws.nrows):
        comitente = str(ws.cell_value(r, 4)).strip()
        if not comitente:
            continue
        try:   imp_bruto   = float(ws.cell_value(r, 7))
        except: imp_bruto  = 0.0
        try:   val_nominal = float(ws.cell_value(r, 8))
        except: val_nominal = 0.0
        try:   arancel     = float(ws.cell_value(r, 9))
        except: arancel    = 0.0
        rows.append({
            'boleto':     str(ws.cell_value(r, 0)).strip(),
            'operacion':  str(ws.cell_value(r, 1)).strip(),
            'fec_ope':    str(ws.cell_value(r, 2)).strip(),
            'comitente':  comitente,
            'nombre':     str(ws.cell_value(r, 5)).strip(),
            'especie':    str(ws.cell_value(r, 6)).strip(),
            'imp_bruto':  imp_bruto,
            'val_nominal': val_nominal,
            'arancel':    arancel,
            'moneda_raw': str(ws.cell_value(r, 18)).strip(),
            'canal':      str(ws.cell_value(r, 24)).strip(),
        })
    return rows


# ─── Carga CONTBOLE con TODAS las filas (incluye subtotales) — para validación ─
def load_contbole_raw(file_obj):
    """Lee CONTBOLE.XLS SIN filtrar comitente vacío. Devuelve lista de dicts
    minimalista para la fase de validación previa.
    """
    file_obj.seek(0)
    wb = xlrd.open_workbook(file_contents=file_obj.read())
    ws = wb.sheet_by_name('Control_de_Boletos')
    rows = []
    for r in range(1, ws.nrows):
        try:   imp_bruto = float(ws.cell_value(r, 7))
        except: imp_bruto = 0.0
        rows.append({
            'boleto':    str(ws.cell_value(r, 0)).strip(),
            'operacion': str(ws.cell_value(r, 1)).strip(),
            'fec_ope':   str(ws.cell_value(r, 2)).strip(),
            'comitente': str(ws.cell_value(r, 4)).strip(),
            'nombre':    str(ws.cell_value(r, 5)).strip(),
            'imp_bruto': imp_bruto,
            'canal':     str(ws.cell_value(r, 24)).strip(),
        })
    return rows


# ─── Fase de validación previa (obligatoria) ──────────────────────────────────
_MERCADO_UNIVERSE   = {"BYMA", "MAV", "A3 - ROFEX", "EXTERIOR"}
_SUBTOTAL_KEYWORDS  = ("TOTAL", "TOT.", "VENTA", "COMPRA", "PARIDAD")


def validar_contbole(rows_raw, tabcompb, mes_esperado=None, anio_esperado=None):
    """Ejecuta la fase de validación previa sobre las filas crudas del CONTBOLE.
    Retorna lista de alertas: {tipo, severidad, valor, filas, ejemplo, accion}.
    """
    alertas = []

    # 1. Códigos de operación sin equivalencia en TABCOMPB (excluye ANUL)
    ops_no_map = defaultdict(lambda: {'filas': 0, 'ejemplo': ''})
    for r in rows_raw:
        op = r['operacion']
        if not op or 'ANUL' in op:
            continue
        if op not in tabcompb:
            ops_no_map[op]['filas'] += 1
            if not ops_no_map[op]['ejemplo']:
                ops_no_map[op]['ejemplo'] = r['boleto']
    for op, info in ops_no_map.items():
        alertas.append({
            'tipo':      'Código sin equivalencia en TABCOMPB',
            'severidad': '🔴 Crítico',
            'valor':     op,
            'filas':     info['filas'],
            'ejemplo':   info['ejemplo'],
            'accion':    'Excluido del reporte',
        })

    # 2. Valores de MERCADO inesperados
    mercados_desc = defaultdict(lambda: {'filas': 0, 'ejemplo': ''})
    for r in rows_raw:
        op = r['operacion']
        if not op or 'ANUL' in op or op not in tabcompb:
            continue
        merc = tabcompb[op].get('mercado', '').strip()
        if merc and merc not in _MERCADO_UNIVERSE:
            mercados_desc[merc]['filas'] += 1
            if not mercados_desc[merc]['ejemplo']:
                mercados_desc[merc]['ejemplo'] = r['boleto']
    for merc, info in mercados_desc.items():
        alertas.append({
            'tipo':      'Mercado no reconocido',
            'severidad': '🟡 Informativo',
            'valor':     merc,
            'filas':     info['filas'],
            'ejemplo':   info['ejemplo'],
            'accion':    'Procesado sin cambio',
        })

    # 3. Segmentos sin TC determinable (raro — se pesifica como ARS)
    segs_ambiguos = defaultdict(lambda: {'filas': 0, 'ejemplo': ''})
    for r in rows_raw:
        op = r['operacion']
        if not op or 'ANUL' in op or op not in tabcompb:
            continue
        seg = tabcompb[op].get('segmento', '').strip().upper()
        if 'CABLE' in seg and 'MEP' in seg:
            segs_ambiguos[seg]['filas'] += 1
            if not segs_ambiguos[seg]['ejemplo']:
                segs_ambiguos[seg]['ejemplo'] = r['boleto']
    for seg, info in segs_ambiguos.items():
        alertas.append({
            'tipo':      'Segmento ambiguo (contiene CABLE y MEP)',
            'severidad': '🟡 Informativo',
            'valor':     seg,
            'filas':     info['filas'],
            'ejemplo':   info['ejemplo'],
            'accion':    'Pesificado como AR$',
        })

    # 4. Filas con Comitente vacío que no son subtotales
    for r in rows_raw:
        if r['comitente']:
            continue
        nombre_upper = r['nombre'].upper()
        if any(kw in nombre_upper for kw in _SUBTOTAL_KEYWORDS):
            continue  # es un subtotal esperado
        # Fila sin comitente y sin keyword de subtotal → sospechosa
        alertas.append({
            'tipo':      'Fila sin Comitente que no parece subtotal',
            'severidad': '🔵 Aviso',
            'valor':     r['nombre'] or '(sin nombre)',
            'filas':     1,
            'ejemplo':   r['boleto'] or f"op={r['operacion']}",
            'accion':    'Excluida del reporte (no tiene comitente)',
        })

    # 5. Canales no reconocidos (los que no están en HB_CANALS y no son vacíos)
    canales_raros = defaultdict(lambda: {'filas': 0, 'ejemplo': ''})
    for r in rows_raw:
        if not r['comitente']:
            continue
        c = r['canal']
        if c and c not in HB_CANALS:
            # Ver si es un valor "raro" nunca visto — usamos un set predefinido de canales manuales
            # comunes para NO alertar. Si el canal contiene solo letras y no matcha nada conocido,
            # se reporta como aviso.
            if c.upper() not in {'MANUAL', 'IT', 'PHONE', 'FAX', 'PRE', 'BB', 'BT', 'EM', 'BLB', ''}:
                canales_raros[c]['filas'] += 1
                if not canales_raros[c]['ejemplo']:
                    canales_raros[c]['ejemplo'] = r['boleto']
    for c, info in canales_raros.items():
        alertas.append({
            'tipo':      'Canal no reconocido (fallback IT Manual)',
            'severidad': '🟡 Informativo',
            'valor':     c,
            'filas':     info['filas'],
            'ejemplo':   info['ejemplo'],
            'accion':    'Mapeado como IT Manual',
        })

    # 6. Fechas fuera del mes/año declarado
    if mes_esperado and anio_esperado:
        mes_idx = MESES_ORDER.index(mes_esperado) + 1  # 1-based
        anio_yy = int(str(anio_esperado)[-2:])
        fuera_rango = 0
        ejemplo = ''
        rango_min = rango_max = None
        for r in rows_raw:
            if not r['comitente']:
                continue
            d_str = r['fec_ope']
            d_obj = None
            for fmt in ('%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
                try:
                    d_obj = datetime.strptime(d_str, fmt).date()
                    break
                except Exception:
                    pass
            if d_obj is None:
                continue
            # Actualizar rango
            if rango_min is None or d_obj < rango_min: rango_min = d_obj
            if rango_max is None or d_obj > rango_max: rango_max = d_obj
            # Chequear match mes/año
            if d_obj.month != mes_idx or d_obj.year not in (int(anio_esperado), 2000 + anio_yy):
                fuera_rango += 1
                if not ejemplo:
                    ejemplo = r['boleto']
        if fuera_rango > 0:
            rango_txt = ''
            if rango_min and rango_max:
                rango_txt = f" (rango detectado: {rango_min:%d/%m/%y} → {rango_max:%d/%m/%y})"
            alertas.append({
                'tipo':      'Fecha fuera del mes/año declarado',
                'severidad': '🔵 Aviso',
                'valor':     f"{mes_esperado} {anio_esperado}{rango_txt}",
                'filas':     fuera_rango,
                'ejemplo':   ejemplo,
                'accion':    'Procesadas de todos modos',
            })

    return alertas


# ─── Pesificación ──────────────────────────────────────────────────────────────
def _get_tc(segmento, tc_mep, tc_ccl):
    s = segmento.upper()
    if 'CABLE' in s: return tc_ccl, 'Dolar CABLE'
    if 'MEP'   in s: return tc_mep, 'Dolar MEP'
    return 1.0, 'AR$'


# ─── Procesar datos del mes ────────────────────────────────────────────────────
def procesar(rows, tabcompb, tc_mep, tc_ccl):
    """
    Clasifica las filas del CONTBOLE y calcula estadísticas del mes.
    Retorna (datos_dict, advertencias).
    """
    advertencias = []
    anulados    = [r for r in rows if 'ANUL' in r['operacion']]
    no_anulados = [r for r in rows if 'ANUL' not in r['operacion']]

    validos_c = []   # clientes
    validos_p = []   # cartera propia
    desestimados = []
    ops_desconocidas = set()

    for r in no_anulados:
        op = r['operacion']
        if op not in tabcompb:
            desestimados.append(r)
            ops_desconocidas.add(op)
            continue
        info = tabcompb[op]
        if info['status'] == 'DESESTIMAR DEL REPORTE':
            desestimados.append(r)
            continue
        if r['comitente'] in CUENTAS_EXCLUIDAS:
            validos_p.append((r, info))
        else:
            validos_c.append((r, info))

    if ops_desconocidas:
        advertencias.append(
            f"Operaciones no encontradas en TABCOMPB ({len(ops_desconocidas)}): "
            + ', '.join(sorted(ops_desconocidas)[:10])
            + (' …' if len(ops_desconocidas) > 10 else '')
            + " — filas desestimadas."
        )

    n  = len(validos_c)
    hb  = sum(1 for r, _ in validos_c if r['canal'] in HB_CANALS)
    manual = n - hb
    cuentas = len(set(r['comitente'] for r, _ in validos_c))

    monto_ars = monto_mep = monto_cable = 0.0
    hb_ars = hb_mep = 0.0
    man_ars = man_mep = man_cable = 0.0
    # Aranceles: solo clientes (cartera propia no genera arancel).
    # Se acumulan en la moneda nativa del boleto según segmento (idéntico a get_tc).
    arancel_ars = arancel_mep = arancel_cable = 0.0
    canal_hb = {'WEB': [0, 0.0, 0.0], 'APP': [0, 0.0, 0.0], 'MgW': [0, 0.0, 0.0]}
    cnt_c   = Counter(); monto_c = defaultdict(float)
    mep_c   = defaultdict(float); cable_c = defaultdict(float); nom_c = {}
    merc_seg = defaultdict(lambda: {'count': 0, 'ars': 0.0, 'div': 0.0, 'moneda': ''})
    daily_c = defaultdict(lambda: {'M': 0, 'HB': 0, 'MgW': 0, 'APP': 0})

    for r, info in validos_c:
        tc, ml = _get_tc(info['segmento'], tc_mep, tc_ccl)
        m   = r['imp_bruto'] * tc
        u   = r['imp_bruto']
        monto_ars += m
        if ml == 'Dolar MEP':   monto_mep   += u
        if ml == 'Dolar CABLE': monto_cable += u

        # Aranceles: acumular en moneda nativa según segmento
        _ar = r.get('arancel', 0.0) or 0.0
        if ml == 'Dolar MEP':      arancel_mep   += _ar
        elif ml == 'Dolar CABLE':  arancel_cable += _ar
        else:                      arancel_ars   += _ar

        c = r['canal']
        if c in HB_CANALS:
            hb_ars += m
            if ml == 'Dolar MEP': hb_mep += u
            canal_hb[c][0] += 1; canal_hb[c][1] += m
            if ml == 'Dolar MEP': canal_hb[c][2] += u
        else:
            man_ars += m
            if ml == 'Dolar MEP':   man_mep   += u
            if ml == 'Dolar CABLE': man_cable += u

        cc = r['comitente']
        cnt_c[cc] += 1; monto_c[cc] += m
        if ml == 'Dolar MEP':   mep_c[cc]   += u
        if ml == 'Dolar CABLE': cable_c[cc] += u
        nom_c[cc] = r['nombre']

        key = (info['mercado'], info['segmento'])
        merc_seg[key]['count'] += 1; merc_seg[key]['ars'] += m
        if ml != 'AR$':
            merc_seg[key]['div'] += u; merc_seg[key]['moneda'] = ml

        d_day = r['fec_ope']
        if c == 'WEB':  daily_c[d_day]['HB']  += 1
        elif c == 'APP': daily_c[d_day]['APP'] += 1
        elif c == 'MgW': daily_c[d_day]['MgW'] += 1
        else:            daily_c[d_day]['M']   += 1

    cnt_p  = Counter(); monto_p = defaultdict(float)
    mep_p  = defaultdict(float); cable_p = defaultdict(float); nom_p = {}
    daily_p = defaultdict(lambda: {'M': 0, 'HB': 0, 'MgW': 0, 'APP': 0})

    for r, info in validos_p:
        tc, ml = _get_tc(info['segmento'], tc_mep, tc_ccl)
        cc = r['comitente']; cnt_p[cc] += 1
        monto_p[cc] += r['imp_bruto'] * tc
        if ml == 'Dolar MEP':   mep_p[cc]   += r['imp_bruto']
        if ml == 'Dolar CABLE': cable_p[cc] += r['imp_bruto']
        nom_p[cc] = r['nombre']
        d_day = r['fec_ope']
        c = r['canal']
        if c == 'WEB':  daily_p[d_day]['HB']  += 1
        elif c == 'APP': daily_p[d_day]['APP'] += 1
        elif c == 'MgW': daily_p[d_day]['MgW'] += 1
        else:            daily_p[d_day]['M']   += 1

    datos = {
        'boletos_totales':    len(rows),
        'anulados_count':     len(anulados),
        'boletos_analizados': n,
        'cuentas':            cuentas,
        'hb': hb, 'manual': manual,
        'monto_ars': monto_ars, 'monto_mep': monto_mep, 'monto_cable': monto_cable,
        'hb_ars': hb_ars, 'hb_mep': hb_mep,
        'man_ars': man_ars, 'man_mep': man_mep, 'man_cable': man_cable,
        'arancel_ars': arancel_ars, 'arancel_mep': arancel_mep, 'arancel_cable': arancel_cable,
        'canal_hb':   canal_hb,
        'top_cant':   cnt_c.most_common(20),
        'top_monto':  sorted(monto_c.items(), key=lambda x: -x[1])[:20],
        'nom_c': nom_c, 'monto_c': monto_c, 'mep_c': mep_c, 'cable_c': cable_c,
        'top_cant_p':  sorted(cnt_p.items(), key=lambda x: -x[1]),
        'top_monto_p': sorted(monto_p.items(), key=lambda x: -x[1])[:20],
        'nom_p': nom_p, 'monto_p': monto_p, 'mep_p': mep_p, 'cable_p': cable_p,
        'daily_c': daily_c, 'daily_p': daily_p,
        'merc_seg': merc_seg,
        'validos_c': validos_c, 'validos_p': validos_p,
        'rows_all':  rows,
        'desestimados': desestimados,
    }
    return datos, advertencias


# ─── Update: Panel de Control ──────────────────────────────────────────────────
def _write_aranceles_header(ws):
    """Escribe (si no existe) el encabezado del bloque ARANCELES (cols P-S, filas 6-9)."""
    # Fila 6: banner
    banner_cell = ws.cell(6, 16)
    if banner_cell.value == "ARANCELES":
        return  # ya está
    banner_cell.value = "ARANCELES"
    banner_cell.font  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    banner_cell.fill  = PatternFill(fill_type='solid', fgColor=C_DARK)
    banner_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=6, start_column=16, end_row=6, end_column=19)

    # Fila 7: título
    title_cell = ws.cell(7, 16)
    title_cell.value = "ARANCELES GENERADOS"
    title_cell.font  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    title_cell.fill  = PatternFill(fill_type='solid', fgColor=C_DARK)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=7, start_column=16, end_row=7, end_column=19)

    # Fila 8: sub-headers
    sub1 = ws.cell(8, 16)
    sub1.value = "En pesos AR$"
    sub1.font  = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    sub1.fill  = PatternFill(fill_type='solid', fgColor=C_BLUE2)
    sub1.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=8, start_column=16, end_row=8, end_column=17)
    sub2 = ws.cell(8, 18)
    sub2.value = "Dato original (en divisa)"
    sub2.font  = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    sub2.fill  = PatternFill(fill_type='solid', fgColor=C_BLUE2)
    sub2.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=8, start_column=18, end_row=8, end_column=19)

    # Fila 9: headers de columna
    headers = ["Total Aranceles AR$", "Aranceles AR$", "Aranceles MEP (USD)", "Aranceles USD Cable"]
    for ci, h in enumerate(headers, 16):
        cell = ws.cell(9, ci)
        cell.value = h
        cell.font  = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
        cell.fill  = PatternFill(fill_type='solid', fgColor=C_BLUE2)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Anchos
    for ci, w in [(16, 16), (17, 14), (18, 13), (19, 14)]:
        ws.column_dimensions[get_column_letter(ci)].width = w


def _update_panel(ws, d, mes_name, tc_mep, tc_ccl, row_mes, row_total, fecha_gen):
    ws.cell(3, 1).value = f"Reporte anual acumulativo  |  Generado: {fecha_gen}"

    _write_aranceles_header(ws)

    n = d['boletos_analizados']
    hb = d['hb']; man = d['manual']

    vals = [
        mes_name, d['cuentas'], d['boletos_totales'], n, d['anulados_count'],
        hb, hb / n if n else 0, man, man / n if n else 0,
        d['monto_ars'], d['monto_mep'], d['monto_cable'], tc_mep, tc_ccl,
    ]
    fmts = [None,'#,##0','#,##0','#,##0','#,##0','#,##0','0.0%',
            '#,##0','0.0%','#,##0','#,##0.00','#,##0.00','#,##0.00','#,##0.00']
    colors = [C_BLACK]*10 + [C_GREY_L, C_GREY_L, C_BLUE2, C_VIOLET]
    italics = [False]*10 + [True, True, True, True]

    for ci, (v, fmt, col, ital) in enumerate(zip(vals, fmts, colors, italics), 1):
        _set(ws, row_mes, ci, v, bold=True, italic=ital, color=col, numfmt=fmt,
             halign='center' if ci > 1 else 'left')

    # ── Bloque ARANCELES (cols P-S, misma fila que el mes) ────────────────────
    ar_ars   = d.get('arancel_ars', 0.0)
    ar_mep   = d.get('arancel_mep', 0.0)
    ar_cable = d.get('arancel_cable', 0.0)
    total_ars = ar_ars + ar_mep * (tc_mep or 0) + ar_cable * (tc_ccl or 0)

    # Col P: Total AR$ (pesificado con TC de la propia fila)
    _set(ws, row_mes, 16, total_ars, bold=True, color=C_BLUE2, numfmt='#,##0.00',
         halign='right')
    # Col Q: AR$ nativo (bold, negro)
    _set(ws, row_mes, 17, ar_ars if ar_ars else None, bold=True, color=C_BLACK,
         numfmt='#,##0.00', halign='right')
    # Col R: MEP en USD (cursiva gris)
    _set(ws, row_mes, 18, ar_mep if ar_mep else None, italic=True, color=C_GREY_L,
         numfmt='#,##0.00', halign='right')
    # Col S: Cable en USD (cursiva gris)
    _set(ws, row_mes, 19, ar_cable if ar_cable else None, italic=True, color=C_GREY_L,
         numfmt='#,##0.00', halign='right')

    # Recalcular TOTAL ACUMULADO leyendo filas previas
    prev_rows = list(range(10, row_mes))
    tot = [0, 0, 0, 0, 0, 0, 0.0, 0.0, 0.0]  # cuentas, btot, bana, anul, hb, man, ars, mep, cable
    # Acumular tambien aranceles: P, Q, R, S
    ar_tot = [0.0, 0.0, 0.0, 0.0]
    for pr in prev_rows:
        tot[0] += ws.cell(pr, 2).value  or 0
        tot[1] += ws.cell(pr, 3).value  or 0
        tot[2] += ws.cell(pr, 4).value  or 0
        tot[3] += ws.cell(pr, 5).value  or 0
        tot[4] += ws.cell(pr, 6).value  or 0
        tot[5] += ws.cell(pr, 8).value  or 0
        tot[6] += ws.cell(pr, 10).value or 0
        tot[7] += ws.cell(pr, 11).value or 0
        tot[8] += ws.cell(pr, 12).value or 0
        for _oi, _col in enumerate((16, 17, 18, 19)):
            _v = ws.cell(pr, _col).value
            if isinstance(_v, (int, float)):
                ar_tot[_oi] += _v
    tot[0] += d['cuentas'];   tot[1] += d['boletos_totales']
    tot[2] += n;              tot[3] += d['anulados_count']
    tot[4] += hb;             tot[5] += man
    tot[6] += d['monto_ars']; tot[7] += d['monto_mep']; tot[8] += d['monto_cable']
    ar_tot[0] += total_ars
    ar_tot[1] += ar_ars
    ar_tot[2] += ar_mep
    ar_tot[3] += ar_cable

    tot_vals = ['TOTAL ACUMULADO', tot[0], tot[1], tot[2], tot[3],
                tot[4], None, tot[5], None, tot[6], tot[7], tot[8], None, None]
    tot_fmts = [None,'#,##0','#,##0','#,##0','#,##0','#,##0',None,
                '#,##0',None,'#,##0','#,##0.00','#,##0.00',None,None]
    for ci, (v, fmt) in enumerate(zip(tot_vals, tot_fmts), 1):
        cell = ws.cell(row_total, ci)
        cell.value = v
        cell.font  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill  = PatternFill(fill_type='solid', fgColor=C_DARK)
        cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left', vertical='center')
        if fmt: cell.number_format = fmt

    # Total acumulado del bloque aranceles (cols P-S)
    ar_fmts = ['#,##0.00', '#,##0.00', '#,##0.00', '#,##0.00']
    for _oi, _col in enumerate((16, 17, 18, 19)):
        cell = ws.cell(row_total, _col)
        cell.value = ar_tot[_oi] if ar_tot[_oi] else None
        cell.font  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        cell.fill  = PatternFill(fill_type='solid', fgColor=C_DARK)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = ar_fmts[_oi]


# ─── Add: Canal Digital vs Manual ─────────────────────────────────────────────
def _add_canal(ws, d, mes_label, fecha_gen):
    ws.cell(3, 1).value = f"Reporte anual acumulativo  |  Generado: {fecha_gen}"
    start = ws.max_row + 3
    n = d['boletos_analizados']; hb = d['hb']; man = d['manual']

    r = start
    c = ws.cell(r, 1)
    c.value = f"CANAL DIGITAL VS MANUAL — {mes_label}"
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill(fill_type='solid', fgColor=C_CYAN)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 22

    r += 1
    ws.cell(r, 6).value = "Operaciones cursadas en:"
    ws.cell(r, 6).font  = Font(name='Calibri', size=9, italic=True, color=C_BLUE2)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

    r += 1
    for ci, (h, bg) in enumerate(zip(
        ['Mes','Canal','Cant. Op.','% del Total','Monto AR$','Dólar MEP','Dólar CABLE'],
        [C_DARK, C_DARK, C_PRIME, C_PRIME, C_PRIME, C_BLUE2, C_BLUE2]
    ), 1):
        _set(ws, r, ci, h, bold=True, color=C_WHITE, bg=bg, halign='center')

    mes_base = mes_label.split()[0].capitalize()

    r += 1
    hb_pct = hb / n if n else 0
    for ci, (v, fmt, ital, col) in enumerate(zip(
        [mes_base, 'Opera HB', hb, hb_pct, d['hb_ars'], d['hb_mep'], None],
        [None, None, '#,##0', '0.0%', '#,##0', '#,##0.00', None],
        [False]*7, [C_BLACK]*5 + [C_GREY_L, C_GREY_L]
    ), 1):
        _set(ws, r, ci, v, italic=ital, color=col, numfmt=fmt,
             bg=C_LIGHT if r % 2 == 0 else None)

    r += 1
    man_pct = man / n if n else 0
    for ci, (v, fmt, col) in enumerate(zip(
        [mes_base, 'IT Manual', man, man_pct, d['man_ars'], d['man_mep'], d['man_cable']],
        [None, None, '#,##0', '0.0%', '#,##0', '#,##0.00', '#,##0.00'],
        [C_BLACK]*5 + [C_GREY_L, C_GREY_L]
    ), 1):
        _set(ws, r, ci, v, italic=(ci >= 6), color=col, numfmt=fmt)

    r += 2
    c = ws.cell(r, 1)
    c.value = f"DESGLOSE CANAL HB — {mes_label}"
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill(fill_type='solid', fgColor=C_GREEN)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 20

    r += 1
    ws.cell(r, 5).value = "Op. cursadas en:"
    ws.cell(r, 5).font  = Font(name='Calibri', size=9, italic=True, color=C_BLUE2)

    r += 1
    for ci, (h, bg) in enumerate(zip(
        ['Tipo Orden','Cantidad','% del HB','Monto AR$','Dólar MEP'],
        [C_DARK, C_GREEN, C_GREEN, C_PRIME, C_BLUE2]
    ), 1):
        _set(ws, r, ci, h, bold=True, color=C_WHITE, bg=bg, halign='center')

    for key, label in [('WEB','WEB'), ('APP','APP'), ('MgW','Manager WEB')]:
        r += 1
        v = d['canal_hb'].get(key, [0, 0.0, 0.0])
        cnt_v, mnt, mep = v[0], v[1], v[2]
        pct = cnt_v / hb if hb else 0
        for ci, (val, fmt, ital, col) in enumerate(zip(
            [label, cnt_v, pct, mnt, mep or None],
            [None, '#,##0', '0.0%', '#,##0', '#,##0.00'],
            [False]*4 + [True],
            [C_BLACK]*4 + [C_GREY_L]
        ), 1):
            _set(ws, r, ci, val, italic=ital, color=col, numfmt=fmt,
                 bg=C_LIGHT if r % 2 == 0 else None)


# ─── Add: Mercado y Segmento ───────────────────────────────────────────────────
def _add_mercado(ws, d, mes_label, fecha_gen):
    ws.cell(3, 1).value = f"Reporte anual acumulativo  |  Generado: {fecha_gen}"
    start = ws.max_row + 3
    n = d['boletos_analizados']

    r = start
    c = ws.cell(r, 1)
    c.value = f"VOLUMEN POR MERCADO Y SEGMENTO — {mes_label}"
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill(fill_type='solid', fgColor=C_PRIME)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 22

    r += 1
    ws.cell(r, 6).value = "Dato original (en divisa)"
    ws.cell(r, 6).font  = Font(name='Calibri', size=9, italic=True, color=C_GREY)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

    r += 1
    for ci, (h, bg) in enumerate(zip(
        ['Mercado','Segmento','Cant. Op.','Monto AR$','% s/Total','Monto Divisa','Moneda'],
        [C_DARK, C_DARK, C_PRIME, C_PRIME, C_PRIME, C_GREY, C_GREY]
    ), 1):
        _set(ws, r, ci, h, bold=True, color=C_WHITE, bg=bg, halign='center')

    gran_total_ars = sum(v['ars'] for v in d['merc_seg'].values())
    for merc in ['BYMA', 'MAV', 'A3 - ROFEX', 'EXTERIOR']:
        segs = [(seg, v) for (m, seg), v in d['merc_seg'].items() if m == merc]
        if not segs:
            continue
        segs_sorted = sorted(segs, key=lambda x: -x[1]['ars'])

        r += 1
        c = ws.cell(r, 1)
        c.value = f"▶  {merc}"
        c.font  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        c.fill  = PatternFill(fill_type='solid', fgColor=C_DARK)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

        for seg, v in segs_sorted:
            r += 1
            pct = v['ars'] / gran_total_ars if gran_total_ars else 0
            bg  = C_LIGHT if r % 2 == 0 else None
            _set(ws, r, 1, None, bg=bg)
            _set(ws, r, 2, seg, bg=bg)
            _set(ws, r, 3, v['count'],  numfmt='#,##0', bg=bg, halign='center')
            _set(ws, r, 4, v['ars'],    numfmt='#,##0', bg=bg, halign='right')
            _set(ws, r, 5, pct,         numfmt='0.0%',  bg=bg, halign='center')
            if v['div'] and v['moneda']:
                _set(ws, r, 6, v['div'], numfmt='#,##0.00', italic=True,
                     color=C_GREY_L, bg=bg, halign='right')
                _set(ws, r, 7, v['moneda'], italic=True, color=C_GREY_L, bg=bg)

        r += 1
        sub_cnt = sum(v['count'] for _, v in segs)
        sub_ars = sum(v['ars'] for _, v in segs)
        sub_pct = sub_ars / gran_total_ars if gran_total_ars else 0
        _set(ws, r, 1, f"Subtotal {merc}", bold=True, bg=C_LIGHT)
        _set(ws, r, 2, None, bg=C_LIGHT)
        _set(ws, r, 3, sub_cnt, bold=True, numfmt='#,##0', bg=C_LIGHT, halign='center')
        _set(ws, r, 4, sub_ars, bold=True, color=C_BLUE2, numfmt='#,##0', bg=C_LIGHT, halign='right')
        _set(ws, r, 5, sub_pct, bold=True, numfmt='0.0%', bg=C_LIGHT, halign='center')
        r += 1

    r += 1
    _set(ws, r, 1, 'GRAN TOTAL', bold=True, color=C_WHITE, bg=C_DARK)
    _set(ws, r, 2, None, bg=C_DARK)
    _set(ws, r, 3, n, bold=True, color=C_WHITE, numfmt='#,##0', bg=C_DARK, halign='center')
    _set(ws, r, 4, gran_total_ars, bold=True, color=C_WHITE, numfmt='#,##0', bg=C_DARK, halign='right')
    _set(ws, r, 5, 1.0, bold=True, color=C_WHITE, numfmt='0.0%', bg=C_DARK, halign='center')
    for ci in range(6, 8):
        _set(ws, r, ci, None, bg=C_DARK)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)


# ─── Add: Rankings TOP 20 ─────────────────────────────────────────────────────
def _add_rankings(ws, d, mes_label, fecha_gen):
    ws.cell(3, 1).value = f"Reporte anual acumulativo  |  Generado: {fecha_gen}"
    start = ws.max_row + 3
    r = start

    c = ws.cell(r, 1)
    c.value = f"RANKINGS TOP 20 — {mes_label}"
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill(fill_type='solid', fgColor=C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)

    c2 = ws.cell(r, 13)
    c2.value = f"CARTERA PROPIA — {mes_label}"
    c2.font  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c2.fill  = PatternFill(fill_type='solid', fgColor=C_VIOLET)
    c2.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=23)

    r += 1
    _set(ws, r, 1, 'TOP CANTIDAD DE OPERACIONES', bold=True, color=C_WHITE, bg=C_PRIME, halign='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    _set(ws, r, 6, 'TOP MONTO TOTAL', bold=True, color=C_WHITE, bg=C_BLUE2, halign='center')
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=11)
    _set(ws, r, 13, 'TOP CANTIDAD DE OPERACIONES', bold=True, color=C_WHITE, bg=C_PRIME, halign='center')
    ws.merge_cells(start_row=r, start_column=13, end_row=r, end_column=16)
    _set(ws, r, 18, 'TOP MONTO TOTAL', bold=True, color=C_WHITE, bg=C_BLUE2, halign='center')
    ws.merge_cells(start_row=r, start_column=18, end_row=r, end_column=23)

    r += 1
    _set(ws, r, 10, 'Operaciones cursadas en:', italic=True, color=C_BLUE2, halign='center')
    ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=11)
    _set(ws, r, 22, 'Operaciones cursadas en:', italic=True, color=C_BLUE2, halign='center')
    ws.merge_cells(start_row=r, start_column=22, end_row=r, end_column=23)

    r += 1
    for ci, h in enumerate(['Rank','Comitente','Nombre','Cant. Op.'], 1):
        _set(ws, r, ci, h, bold=True, color=C_WHITE, bg=C_DARK, halign='center')
    for ci, h in enumerate(['Rank','Comitente','Nombre','Monto AR$','Dólar MEP','Dólar CABLE'], 1):
        _set(ws, r, ci+5, h, bold=True, color=C_WHITE, bg=C_DARK, halign='center')
    for ci, h in enumerate(['Rank','Comitente','Nombre','Cant. Op.'], 1):
        _set(ws, r, ci+12, h, bold=True, color=C_WHITE, bg=C_VIOLET, halign='center')
    for ci, h in enumerate(['Rank','Comitente','Nombre','Monto AR$','Dólar MEP','Dólar CABLE'], 1):
        _set(ws, r, ci+17, h, bold=True, color=C_WHITE, bg=C_VIOLET, halign='center')

    nom_c = d['nom_c']; monto_c = d['monto_c']; mep_c = d['mep_c']; cable_c = d['cable_c']
    nom_p = d['nom_p']; monto_p = d['monto_p']; mep_p = d['mep_p']; cable_p = d['cable_p']
    top_cant = d['top_cant']; top_monto = d['top_monto']
    top_cant_p = d['top_cant_p']; top_monto_p = d['top_monto_p']

    for i in range(20):
        r += 1
        bg = "FFF3CD" if i == 0 else (C_LIGHT if i % 2 == 0 else None)
        b  = (i == 0)

        if i < len(top_cant):
            cte, cnt = top_cant[i]
            _set(ws, r, 1, i+1, bold=b, bg=bg, halign='center')
            _set(ws, r, 2, cte, bold=b, bg=bg, halign='center')
            _set(ws, r, 3, nom_c.get(cte,'')[:25], bold=b, bg=bg)
            _set(ws, r, 4, cnt, bold=b, numfmt='#,##0', bg=bg, halign='center')

        if i < len(top_monto):
            cte, mnt = top_monto[i]
            mep_v = mep_c.get(cte, 0) or None; cab_v = cable_c.get(cte, 0) or None
            _set(ws, r, 6, i+1, bold=b, bg=bg, halign='center')
            _set(ws, r, 7, cte, bold=b, bg=bg, halign='center')
            _set(ws, r, 8, nom_c.get(cte,'')[:25], bold=b, bg=bg)
            _set(ws, r, 9, mnt, bold=b, numfmt='#,##0', bg=bg, halign='right')
            _set(ws, r, 10, mep_v, italic=True, color=C_GREY_L, numfmt='#,##0.00', bg=bg, halign='right')
            _set(ws, r, 11, cab_v, italic=True, color=C_GREY_L, numfmt='#,##0.00', bg=bg, halign='right')

        if i < len(top_cant_p):
            cte, cnt = top_cant_p[i]
            _set(ws, r, 13, i+1, bold=b, bg=bg, halign='center')
            _set(ws, r, 14, cte, bold=b, bg=bg, halign='center')
            _set(ws, r, 15, nom_p.get(cte,'')[:25], bold=b, bg=bg)
            _set(ws, r, 16, cnt, bold=b, numfmt='#,##0', bg=bg, halign='center')

        if i < len(top_monto_p):
            cte, mnt = top_monto_p[i]
            mep_v = mep_p.get(cte, 0) or None; cab_v = cable_p.get(cte, 0) or None
            _set(ws, r, 18, i+1, bold=b, bg=bg, halign='center')
            _set(ws, r, 19, cte, bold=b, bg=bg, halign='center')
            _set(ws, r, 20, nom_p.get(cte,'')[:25], bold=b, bg=bg)
            _set(ws, r, 21, mnt, bold=b, numfmt='#,##0', bg=bg, halign='right')
            _set(ws, r, 22, mep_v, italic=True, color=C_GREY_L, numfmt='#,##0.00', bg=bg, halign='right')
            _set(ws, r, 23, cab_v, italic=True, color=C_GREY_L, numfmt='#,##0.00', bg=bg, halign='right')


# ─── Add: Operaciones Diarias ─────────────────────────────────────────────────
def _add_diarias(ws, d, mes_label, fecha_gen):
    ws.cell(3, 1).value = f"Reporte anual acumulativo  |  Generado: {fecha_gen}"
    start = ws.max_row + 3

    r = start
    c = ws.cell(r, 1)
    c.value = f"VOLUMEN DIARIO DE OPERACIONES — {mes_label}"
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill(fill_type='solid', fgColor=C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 22

    c2 = ws.cell(r, 8)
    c2.value = f"CARTERA PROPIA — {mes_label}"
    c2.font  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c2.fill  = PatternFill(fill_type='solid', fgColor=C_VIOLET)
    c2.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=13)

    r += 1
    hdrs = ['Fecha','Manual','HB','Manager WEB','APP HB','TOTAL']
    bgs  = [C_DARK, C_DARK, C_GREEN, C_VIOLET, C_ORANGE, C_DARK]
    for ci, (h, bg) in enumerate(zip(hdrs, bgs), 1):
        _set(ws, r, ci, h, bold=True, color=C_WHITE, bg=bg, halign='center')
    for ci, (h, bg) in enumerate(zip(hdrs, bgs), 8):
        _set(ws, r, ci, h, bold=True, color=C_WHITE, bg=bg, halign='center')

    daily_c = d['daily_c']; daily_p = d['daily_p']
    days    = sorted(daily_c.keys())

    keys = ['M', 'HB', 'MgW', 'APP', 'T']
    col_vals_c = {k: [] for k in keys}
    col_vals_p = {k: [] for k in keys}
    for day in days:
        dc = daily_c[day]; dp = daily_p.get(day, {'M':0,'HB':0,'MgW':0,'APP':0})
        tc = dc['M'] + dc['HB'] + dc['MgW'] + dc['APP']
        tp = dp['M'] + dp['HB'] + dp['MgW'] + dp['APP']
        for k in ['M','HB','MgW','APP']:
            col_vals_c[k].append(dc[k]); col_vals_p[k].append(dp[k])
        col_vals_c['T'].append(tc); col_vals_p['T'].append(tp)

    def _highlight(v, vals):
        if not vals: return None
        mx = max(vals); mn = min(vals)
        if mx == mn: return None
        return 'max' if v == mx else ('min' if v == mn else None)

    for idx, day in enumerate(days):
        r += 1
        dc = daily_c[day]; dp = daily_p.get(day, {'M':0,'HB':0,'MgW':0,'APP':0})
        tc = dc['M'] + dc['HB'] + dc['MgW'] + dc['APP']
        tp = dp['M'] + dp['HB'] + dp['MgW'] + dp['APP']
        bg_row = C_LIGHT if idx % 2 == 0 else None

        # Parse date
        d_obj = None
        for fmt in ('%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
            try: d_obj = datetime.strptime(day, fmt).date(); break
            except: pass

        for col_pair, vals_dict, col_vals_dict in [
            (1, dc, col_vals_c), (8, dp, col_vals_p)
        ]:
            cell_date = ws.cell(r, col_pair)
            cell_date.value = d_obj
            cell_date.number_format = 'DD/MM/YYYY'
            cell_date.alignment = Alignment(horizontal='center', vertical='center')
            if bg_row: cell_date.fill = PatternFill(fill_type='solid', fgColor=bg_row)

            vals_list = [vals_dict['M'], vals_dict['HB'], vals_dict['MgW'],
                         vals_dict['APP'],
                         vals_dict['M']+vals_dict['HB']+vals_dict['MgW']+vals_dict['APP']]
            for ci_off, (v, key) in enumerate(zip(vals_list, keys)):
                ci = col_pair + 1 + ci_off
                h = _highlight(v, col_vals_dict[key])
                if h == 'max':   fg = C_GREEN; txt = C_WHITE; b = True
                elif h == 'min': fg = 'C0392B'; txt = C_WHITE; b = True
                else:            fg = bg_row;  txt = C_BLACK; b = False
                _set(ws, r, ci, v, bold=b, color=txt, bg=fg, numfmt='#,##0', halign='center')

    r += 1
    tot_c = [sum(col_vals_c[k]) for k in ['M','HB','MgW','APP','T']]
    tot_p = [sum(col_vals_p[k]) for k in ['M','HB','MgW','APP','T']]
    _set(ws, r, 1, 'TOTAL', bold=True, color=C_WHITE, bg=C_DARK, halign='center')
    for ci_off, v in enumerate(tot_c):
        _set(ws, r, ci_off+2, v, bold=True, color=C_WHITE, bg=C_DARK, numfmt='#,##0', halign='center')
    _set(ws, r, 8, 'TOTAL', bold=True, color=C_WHITE, bg=C_DARK, halign='center')
    for ci_off, v in enumerate(tot_p):
        _set(ws, r, ci_off+9, v, bold=True, color=C_WHITE, bg=C_DARK, numfmt='#,##0', halign='center')


# ─── Add: Detalle del mes (nueva hoja) ────────────────────────────────────────
def _add_detalle(wb, d, mes_short, mes_name, mes_label, anio, tabcompb, fecha_gen):
    ws = wb.create_sheet(mes_short)
    ws.sheet_properties.tabColor = "999999"

    _set(ws, 1, 1, "SAILING INVERSIONES", bold=True, size=16, color=C_DARK)
    _set(ws, 2, 1, f"Detalle de Operaciones — {mes_name} {anio}", size=11, color=C_PRIME)
    _set(ws, 3, 1, f"Reporte anual acumulativo  |  Generado: {fecha_gen}", size=9, color=C_GREY)

    hdrs = ['Boleto','Operación','Fecha Op.','Comitente','Nombre',
            'Especie','Imp. Bruto','Val. Nominal','Moneda','Canal',
            'Nombre Operación','Mercado','Rubro','Anulado']
    for ci, h in enumerate(hdrs, 1):
        _set(ws, 7, ci, h, bold=True, color=C_WHITE, bg=C_DARK, halign='center')
    ws.freeze_panes = "A8"
    ws.auto_filter.ref = f"A7:{get_column_letter(len(hdrs))}7"

    r = 7
    for row in d['rows_all']:
        r += 1
        op = row['operacion']
        is_anulado = 'ANUL' in op
        is_cartera = row['comitente'] in CUENTAS_EXCLUIDAS
        tab = tabcompb.get(op, {}) if not is_anulado else {}

        if is_anulado:
            anulado_val = "SI"; bg_row = "FFE6E6"; txt_col = "CC0000"; ital = True
        elif is_cartera:
            anulado_val = "CARTERA PROPIA"; bg_row = "FFF0CC"; txt_col = C_BLACK; ital = False
        else:
            anulado_val = None; bg_row = None; txt_col = C_BLACK; ital = False

        vals = [row['boleto'], op, row['fec_ope'], row['comitente'], row['nombre'],
                row['especie'], row['imp_bruto'], row['val_nominal'], row['moneda_raw'],
                row['canal'], tab.get('nombre',''), tab.get('mercado',''),
                tab.get('segmento',''), anulado_val]
        fmts = [None]*6 + ['#,##0.00','#,##0.00'] + [None]*6

        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r, ci)
            cell.value = v
            cell.font  = Font(name='Calibri', size=9, italic=ital, color=txt_col)
            if bg_row:
                cell.fill = PatternFill(fill_type='solid', fgColor=bg_row)
            if fmt:
                cell.number_format = fmt

    widths = [10, 8, 12, 10, 28, 12, 14, 14, 14, 8, 22, 10, 20, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─── Add: hoja ⚠️ Alertas (opcional, se agrega si hay alertas de validación) ─
_ALERT_SHEET = "⚠️ Alertas"


def _add_alertas(wb, alertas, mes_label, fecha_gen):
    """Crea o actualiza la hoja de alertas al final del workbook.
    Si ya existe, agrega las nuevas alertas debajo con un separador de mes.
    """
    if not alertas:
        return
    if _ALERT_SHEET in wb.sheetnames:
        ws = wb[_ALERT_SHEET]
        start_row = ws.max_row + 2
    else:
        ws = wb.create_sheet(_ALERT_SHEET)
        ws.sheet_properties.tabColor = "F39C12"
        _set(ws, 1, 1, "SAILING INVERSIONES", bold=True, size=16, color=C_DARK)
        _set(ws, 2, 1, "Alertas de validacion", size=11, color=C_PRIME)
        _set(ws, 3, 1, f"Reporte anual acumulativo  |  Generado: {fecha_gen}",
             size=9, color=C_GREY)
        # Fondo amarillo muy suave para toda la hoja
        for r in range(1, 4):
            for c in range(1, 7):
                ws.cell(r, c).fill = PatternFill(fill_type='solid', fgColor='FFF8E1')
        start_row = 6

    # Banner del mes
    banner = ws.cell(start_row, 1)
    banner.value = f"── {mes_label} ──"
    banner.font  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    banner.fill  = PatternFill(fill_type='solid', fgColor=C_ORANGE)
    banner.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    start_row += 1

    # Headers
    headers = ['Tipo', 'Severidad', 'Valor problemático', 'Filas afectadas',
               'Ejemplo boleto', 'Acción tomada']
    for ci, h in enumerate(headers, 1):
        _set(ws, start_row, ci, h, bold=True, color=C_WHITE, bg=C_ORANGE,
             halign='center')

    # Datos
    r = start_row
    for a in alertas:
        r += 1
        _set(ws, r, 1, a['tipo'])
        _set(ws, r, 2, a['severidad'], halign='center')
        _set(ws, r, 3, str(a['valor']))
        _set(ws, r, 4, a['filas'], halign='center', numfmt='#,##0')
        _set(ws, r, 5, a['ejemplo'], halign='center')
        _set(ws, r, 6, a['accion'])

    # Anchos
    widths = [40, 14, 30, 14, 16, 30]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─── Función principal ─────────────────────────────────────────────────────────
def actualizar_reporte(existing_bytes, datos, tabcompb, mes_name, tc_mep, tc_ccl,
                       anio=2026, alertas=None):
    """
    Carga el reporte existente (bytes), agrega el mes nuevo y devuelve los bytes actualizados.
    """
    wb = openpyxl.load_workbook(io.BytesIO(existing_bytes))

    mes_idx   = MESES_ORDER.index(mes_name)
    mes_short = f"{mes_name[:3]} {str(anio)[-2:]}"   # "Abr 26"
    mes_label = f"{mes_name.upper()} {anio}"          # "ABRIL 2026"
    row_mes   = 10 + mes_idx                           # Enero=10, Feb=11, ...
    row_total = 22
    fecha_gen = datetime.now().strftime('%d/%m/%Y')

    sheet_names = wb.sheetnames

    if 'Panel de Control' in sheet_names:
        _update_panel(wb['Panel de Control'], datos, mes_name, tc_mep, tc_ccl,
                      row_mes, row_total, fecha_gen)

    if 'Canal Digital vs Manual' in sheet_names:
        _add_canal(wb['Canal Digital vs Manual'], datos, mes_label, fecha_gen)

    if 'Mercado y Segmento' in sheet_names:
        _add_mercado(wb['Mercado y Segmento'], datos, mes_label, fecha_gen)

    if 'Rankings TOP 20' in sheet_names:
        _add_rankings(wb['Rankings TOP 20'], datos, mes_label, fecha_gen)

    if 'Operaciones Diarias' in sheet_names:
        _add_diarias(wb['Operaciones Diarias'], datos, mes_label, fecha_gen)

    _add_detalle(wb, datos, mes_short, mes_name, mes_label, anio, tabcompb, fecha_gen)

    # Hoja Alertas al final (si hubo alertas de validación)
    if alertas:
        _add_alertas(wb, alertas, mes_label, fecha_gen)

    # Reordenar: hojas fijas primero, luego detalles por mes, Alertas al final
    fixed = ['Panel de Control', 'Canal Digital vs Manual', 'Mercado y Segmento',
             'Rankings TOP 20', 'Operaciones Diarias']
    detail_sheets = [s for s in wb.sheetnames if s not in fixed and s != _ALERT_SHEET]
    desired = fixed + sorted(
        detail_sheets,
        key=lambda s: next((i for i, m in enumerate(MESES_ORDER)
                            if m[:3].lower() in s.lower()), 99)
    )
    if _ALERT_SHEET in wb.sheetnames:
        desired.append(_ALERT_SHEET)
    for name in desired:
        if name in wb.sheetnames:
            current_idx = wb.sheetnames.index(name)
            target_idx  = desired.index(name)
            if current_idx != target_idx:
                wb.move_sheet(name, offset=target_idx - current_idx)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
