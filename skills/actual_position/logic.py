"""
Actual Position — versión web (Streamlit).
Recibe UploadedFiles; devuelve (BytesIO xlsx, BytesIO xls_opciones, dict resumen).
"""

import warnings
from io import BytesIO
from datetime import date

import pandas as pd
import openpyxl
import xlwt
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Paleta ────────────────────────────────────────────────────────────────────
COL_BYMA_BLUE  = "1E3A5F"
COL_HDR_GRAY   = "D9D9D9"
COL_TOTAL_GRAY = "BFC2C9"
COL_UNKNOWN    = "FFD966"
COL_RED_HDR    = "C00000"

NUM_FMT  = "#,##0.00"
DATE_FMT = "DD/MM/YYYY"

MONEDAS = [
    ("ARS",       "Pesos ARS"),
    ("USD_MEP",   "Dólar MEP"),
    ("USD_CABLE", "USD Cable"),
]
MONEDA_LABEL   = {"ARS": "PESOS (ARS)", "USD_MEP": "DÓLAR MEP (USD)", "USD_CABLE": "USD CABLE (EXT)"}
MONEDA_DISPLAY = {"ARS": "ARS",         "USD_MEP": "USD MEP",          "USD_CABLE": "USD Cable"}

NUMERIC_COLS = [
    "Short Quantity", "Short Reserved Quantity", "Short Available Quantity",
    "Short Initial Value", "Short Market Value",
    "Long Quantity",  "Long Reserved Quantity",  "Long Available Quantity",
    "Long Initial Value",  "Long Market Value",
    "Net Quantity", "Short Initial Quantity", "Long Initial Quantity",
    "Purchase On The Day", "Sale On The Day",
]


# ══════════════════════════════════════════════════════════════════════════════
#  CLASIFICACIÓN (sin cambios respecto al script local)
# ══════════════════════════════════════════════════════════════════════════════

def _parsear_filter_node(filter_node: str) -> dict:
    parts = str(filter_node).upper().split(".")
    return {
        "comp3": parts[3] if len(parts) > 3 else "",
        "comp4": parts[4] if len(parts) > 4 else "",
        "comp5": parts[5] if len(parts) > 5 else "",
    }


def get_moneda_fn(filter_node: str) -> str:
    fn = _parsear_filter_node(filter_node)
    ccy_map = {"ARS": "ARS", "USD": "USD_MEP", "EXT": "USD_CABLE"}
    if fn["comp3"] in ("C", "U"):
        return ccy_map.get(fn["comp4"], "ARS")
    elif fn["comp3"] == "SB":
        return ccy_map.get(fn["comp4"], "ARS")
    elif fn["comp3"] == "DER":
        return "ARS"
    return "ARS"


def clasificar(asset: str, trade_dt: date, settle_dt: date, filter_node: str):
    fn      = _parsear_filter_node(filter_node)
    c3, c5  = fn["comp3"], fn["comp5"]

    if c3 == "U":
        return ("FC" if trade_dt == settle_dt else "FV"), "G", False, ""
    if c3 == "DER":
        return "OP", "G", False, ""
    if str(filter_node).upper().startswith("AR.BYMA.SB."):
        concepto = "CI" if (c5 == "T0" or trade_dt == settle_dt) else "CN"
        return concepto, "NG", False, ""
    if c3 == "C":
        concepto = "CI" if (c5 == "T0" or trade_dt == settle_dt) else "CN"
        return concepto, "G", False, ""
    return "UNKNOWN", "G", True, f"Filter Node no reconocido: {filter_node} | Asset: {asset}"


def parse_vto_caucion(asset: str, process_date: date):
    try:
        ddmm  = asset.split("-")[1]
        day, month = int(ddmm[:2]), int(ddmm[2:])
        vto = date(process_date.year, month, day)
        if vto < process_date:
            vto = date(process_date.year + 1, month, day)
        return vto
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  CARGA DE ARCHIVOS (adaptados a UploadedFile / BytesIO)
# ══════════════════════════════════════════════════════════════════════════════

def cargar_cotizaciones_bc(prices_file) -> dict:
    """
    Lee el table-prices_*.csv y extrae los tipos de cambio de BYMA Clearing.
      ARS/USD → TC MEP = 1 / precio
      ARS/EXT → TC CCL = 1 / precio
    Retorna dict con 'mep', 'ccl', 'archivo' (o vacío si no está disponible).
    """
    if prices_file is None:
        return {}
    try:
        prices_file.seek(0)
        df = pd.read_csv(prices_file, quotechar='"', encoding="utf-8-sig")
        df.columns = [c.strip().strip('"') for c in df.columns]
        for col in df.select_dtypes(include="object").columns:
            df[col] = df[col].str.strip().str.strip('"')
        df["Price"] = pd.to_numeric(df["Price"], errors="coerce")

        def _tc(asset_code):
            row = df[df["Asset"] == asset_code]
            if row.empty:
                return None
            precio = row.iloc[0]["Price"]
            return 1 / precio if precio and precio != 0 else None

        return {
            "mep":     _tc("ARS/USD"),
            "ccl":     _tc("ARS/EXT"),
            "archivo": getattr(prices_file, "name", "table-prices_*.csv"),
        }
    except Exception:
        return {}


def cargar_mapa_cvsa(especies_file) -> dict:
    if especies_file is None:
        return {}
    try:
        df = pd.read_excel(especies_file, sheet_name="Datos_Fijos_Especies", dtype=str)
        df.columns = [c.strip().strip("'") for c in df.columns]
        for col in df.columns:
            df[col] = df[col].fillna("").str.strip().str.strip("'")
        mapa = {}
        for _, row in df.iterrows():
            codigo = row.get("Codigo", "").strip()
            if not codigo:
                continue
            for col_ticker in ("Norm.", "Parid", "Cable"):
                ticker = row.get(col_ticker, "").strip()
                if ticker and ticker.upper() not in ("NAN", "NO EXISTE", ""):
                    mapa[ticker.upper()] = codigo
        return mapa
    except Exception:
        return {}


def cargar_saldos_inicio(saldos_file) -> dict:
    if saldos_file is None:
        return {}
    raw = saldos_file.read()
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            df = pd.read_csv(BytesIO(raw), sep=";", encoding=enc, dtype=str)
            break
        except (UnicodeDecodeError, Exception):
            continue
    else:
        return {}

    df.columns = [c.strip() for c in df.columns]
    moneda_col = next((c for c in df.columns if "Moneda" in c), None)
    pago_col   = next((c for c in df.columns if "pago"   in c.lower()), None)
    saldo_col  = next((c for c in df.columns if "disponible" in c.lower()), None)
    if not moneda_col or not saldo_col:
        return {}

    saldos = {}
    for _, row in df.iterrows():
        moneda_raw = str(row.get(moneda_col, "")).strip()
        pago_raw   = str(row.get(pago_col,   "")).strip().upper() if pago_col else ""
        saldo_str  = str(row.get(saldo_col,  "0")).strip().replace(".", "").replace(",", ".")
        try:
            saldo = float(saldo_str)
        except ValueError:
            continue
        if moneda_raw == "ARS":
            saldos["ARS"] = saldo
        elif moneda_raw == "USD":
            if "EXT" in pago_raw or "CABLE" in pago_raw:
                saldos["USD_CABLE"] = saldo
            else:
                saldos["USD_MEP"] = saldo
    return saldos


def cargar_y_clasificar(csv_file, process_date: date):
    df = pd.read_csv(csv_file, quotechar='"', encoding="utf-8-sig")
    df.columns = [c.strip().strip('"') for c in df.columns]
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip().str.strip('"')

    df["Trade Date"]      = pd.to_datetime(df["Trade Date"],      errors="coerce")
    df["Settlement Date"] = pd.to_datetime(df["Settlement Date"], errors="coerce")

    if df.empty:
        return pd.DataFrame(), pd.DataFrame()

    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    def _clasificar_row(r):
        trade_dt  = r["Trade Date"].date()      if pd.notna(r["Trade Date"])      else process_date
        settle_dt = r["Settlement Date"].date() if pd.notna(r["Settlement Date"]) else None
        return clasificar(r["Asset"], trade_dt, settle_dt, r.get("Filter Node", ""))

    clasif = df.apply(_clasificar_row, axis=1)
    df["Concepto"]         = clasif.apply(lambda x: x[0])
    df["Segmento"]         = clasif.apply(lambda x: x[1])
    df["Para_Verificar"]   = clasif.apply(lambda x: x[2])
    df["Motivo_Verificar"] = clasif.apply(lambda x: x[3])

    settle_dates = df["Settlement Date"].dt.date
    mask_caucion = df["Concepto"].isin(["FC", "FV"])
    df.loc[mask_caucion & (settle_dates == process_date), "Concepto"] = "FV"
    df.loc[mask_caucion & (settle_dates >  process_date), "Concepto"] = "FC"

    df["Moneda"] = df["Filter Node"].apply(get_moneda_fn)

    settle_dates = df["Settlement Date"].dt.date
    mask_keep = (
        (df["Concepto"].isin(["CI", "CN"]) & (settle_dates == process_date)) |
        ((df["Concepto"] == "FV")          & (settle_dates == process_date)) |
        (df["Concepto"] == "OP") |
        (df["Para_Verificar"])
    )
    df = df[mask_keep].copy()

    if df.empty:
        return pd.DataFrame(), pd.DataFrame()

    def calc_flujos(row):
        if row["Concepto"] in ("FC", "FV"):
            nq = row["Net Quantity"]
            return max(nq, 0.0), max(-nq, 0.0)
        else:
            return abs(row["Short Market Value"]), abs(row["Long Market Value"])

    flujos = df.apply(calc_flujos, axis=1)
    df["A_Recibir"]  = flujos.apply(lambda x: x[0])
    df["A_Entregar"] = flujos.apply(lambda x: x[1])
    df["Neto"]       = df["A_Recibir"] - df["A_Entregar"]
    df["Comitente"]  = df["Account"].str.extract(r"_(\d+)\s*\(")

    def calc_vto(row):
        if row["Concepto"] in ("FC", "FV"):
            return parse_vto_caucion(row["Asset"], process_date)
        return None
    df["Fecha_Vto"] = df.apply(calc_vto, axis=1)

    return df[~df["Para_Verificar"]].copy(), df[df["Para_Verificar"]].copy()


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _cel_header(ws, row, col, value, bg=COL_HDR_GRAY, bold=True,
                align="center", font_color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(bold=bold, color=font_color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border    = _borde()
    return c

def _cel_dato(ws, row, col, value, num_fmt=None, bg=None, bold=False, align="right"):
    c = ws.cell(row=row, column=col, value=value)
    if num_fmt:   c.number_format = num_fmt
    if bg:        c.fill = PatternFill("solid", fgColor=bg)
    c.font      = Font(bold=bold)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    return c


# ══════════════════════════════════════════════════════════════════════════════
#  ESCRITURA DE HOJAS (idéntica al script local, sin cambios)
# ══════════════════════════════════════════════════════════════════════════════

def escribir_hoja_cotizaciones(wb, cotizaciones: dict, process_date: date):
    """Primera hoja: cotizaciones Dólar MEP y CCL según BYMA Clearing."""
    ws = wb.create_sheet("Cotizaciones BC", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = f"COTIZACIONES BYMA CLEARING — {process_date.strftime('%d/%m/%Y')}"
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=COL_BYMA_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for i, h in enumerate(["Tipo de Cambio", "Valor (ARS)", "Fuente"], 1):
        _cel_header(ws, 3, i, h)

    mep     = cotizaciones.get("mep")
    ccl     = cotizaciones.get("ccl")
    archivo = cotizaciones.get("archivo", "table-prices_*.csv")

    for offset, (label, valor) in enumerate([("Dólar MEP (ARS/USD)", mep), ("Dólar CCL (ARS/EXT)", ccl)]):
        r = 4 + offset
        _cel_dato(ws, r, 1, label, align="left")
        if valor is not None:
            _cel_dato(ws, r, 2, valor, num_fmt=NUM_FMT, bold=True)
        else:
            c2 = ws.cell(row=r, column=2, value="N/D")
            c2.font      = Font(italic=True, color="595959")
            c2.alignment = Alignment(horizontal="right", vertical="center")
            c2.border    = _borde()
        _cel_dato(ws, r, 3, "BYMA Clearing", align="center")

    nota_r = 7
    ws.merge_cells(f"A{nota_r}:C{nota_r}")
    cn = ws.cell(row=nota_r, column=1,
                 value=f"Fuente: {archivo}  |  Cálculo: 1 / cotización ARS publicada por BC")
    cn.font      = Font(italic=True, size=9, color="595959")
    cn.alignment = Alignment(horizontal="left", vertical="center")

    for i, w in enumerate([28, 18, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def escribir_hoja_moneda(wb, sheet_name, df_ok, moneda, process_date, saldo_inicio=None):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    dm = df_ok[df_ok["Moneda"] == moneda].copy() if not df_ok.empty else pd.DataFrame()

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"LIQUIDACIÓN {MONEDA_LABEL[moneda]} — {process_date.strftime('%d/%m/%Y')}"
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=COL_BYMA_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    r = 3
    ws.merge_cells(f"A{r}:G{r}")
    c = ws[f"A{r}"]
    c.value     = "TOTALES"
    c.font      = Font(bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=COL_BYMA_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20

    r += 1
    for i, h in enumerate(["Segmento","A Recibir","A Entregar","Neto a Liquidar",
                            "Ingresos (*)","Egresos (*)","Saldo"], 1):
        _cel_header(ws, r, i, h)

    r += 1
    if not dm.empty:
        for seg, grp in dm.groupby("Segmento", sort=True):
            ar, ae = grp["A_Recibir"].sum(), grp["A_Entregar"].sum()
            nt = ar - ae
            _cel_dato(ws, r, 1, seg,  align="center", bold=True)
            _cel_dato(ws, r, 2, ar,   num_fmt=NUM_FMT)
            _cel_dato(ws, r, 3, ae,   num_fmt=NUM_FMT)
            _cel_dato(ws, r, 4, nt,   num_fmt=NUM_FMT, bold=True)
            _cel_dato(ws, r, 5, 0.0,  num_fmt=NUM_FMT)
            _cel_dato(ws, r, 6, 0.0,  num_fmt=NUM_FMT)
            _cel_dato(ws, r, 7, nt,   num_fmt=NUM_FMT, bold=True)
            r += 1

    tot_ar = dm["A_Recibir"].sum()  if not dm.empty else 0.0
    tot_ae = dm["A_Entregar"].sum() if not dm.empty else 0.0
    tot_nt = tot_ar - tot_ae
    for i, v in enumerate(["T", tot_ar, tot_ae, tot_nt, 0.0, 0.0, tot_nt], 1):
        _cel_dato(ws, r, i, v, num_fmt=NUM_FMT if i>1 else None,
                  bg=COL_TOTAL_GRAY, bold=True, align="center" if i==1 else "right")
    r += 1

    if saldo_inicio is not None:
        saldo_proy = saldo_inicio + tot_nt
        bg_ini  = "EBF3FB"
        bg_proy = "D6E4BC" if saldo_proy >= 0 else "FCE4D6"
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(row=r, column=1, value="Saldo al inicio (Nasdaq)")
        c.font      = Font(bold=True); c.fill = PatternFill("solid", fgColor=bg_ini)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = _borde()
        _cel_dato(ws, r, 7, saldo_inicio, num_fmt=NUM_FMT, bg=bg_ini, bold=True)
        r += 1
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(row=r, column=1, value="Saldo proyectado del día")
        c.font      = Font(bold=True); c.fill = PatternFill("solid", fgColor=bg_proy)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = _borde()
        _cel_dato(ws, r, 7, saldo_proy, num_fmt=NUM_FMT, bg=bg_proy, bold=True)
        r += 1

    ws.cell(row=r, column=1,
            value="(*) Ingresos y Egresos provienen de Liquidación de Fondos — no incluidos en esta versión."
    ).font = Font(italic=True, size=9, color="595959")
    r += 2

    ws.merge_cells(f"A{r}:G{r}")
    c = ws[f"A{r}"]
    c.value     = "DETALLE POR CONCEPTO"
    c.font      = Font(bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=COL_BYMA_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    for i, h in enumerate(["Segmento","Contraparte","Concepto",
                            "A Recibir","A Entregar","Neto a Liquidar","Fecha Vto"], 1):
        _cel_header(ws, r, i, h)
    r += 1

    if not dm.empty:
        orden_concepto = {"CI":0,"CN":1,"FC":2,"FV":3,"OP":4}
        dm["_ord"] = dm["Concepto"].map(orden_concepto).fillna(99)
        for (seg, concepto), grp in dm.sort_values(["Segmento","_ord"]).groupby(
                ["Segmento","Concepto"], sort=False):
            ar, ae = grp["A_Recibir"].sum(), grp["A_Entregar"].sum()
            vto_val = ""
            if concepto == "FV":
                vto_val = process_date
            elif concepto == "FC":
                vtos = grp["Fecha_Vto"].dropna().unique()
                vto_val = vtos[0] if len(vtos)==1 else ("Varias" if len(vtos)>1 else "")
            _cel_dato(ws, r, 1, seg,      align="center")
            _cel_dato(ws, r, 2, "BYMA",   align="center")
            _cel_dato(ws, r, 3, concepto, align="center", bold=True)
            _cel_dato(ws, r, 4, ar,       num_fmt=NUM_FMT)
            _cel_dato(ws, r, 5, ae,       num_fmt=NUM_FMT)
            _cel_dato(ws, r, 6, ar-ae,    num_fmt=NUM_FMT, bold=True)
            c7 = _cel_dato(ws, r, 7, vto_val, align="center")
            if isinstance(vto_val, date): c7.number_format = DATE_FMT
            r += 1

    for i, w in enumerate([12,14,12,18,18,18,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def escribir_hoja_movimientos(wb, df_ok, process_date, mapa_cvsa):
    ws = wb.create_sheet("Mov. por Especie")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"MOVIMIENTOS POR ESPECIE — VN — {process_date.strftime('%d/%m/%Y')}"
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=COL_BYMA_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    cvsa_status = "vinculado desde ESPECIES.XLS" if mapa_cvsa else "no disponible (ESPECIES.XLS no subido)"
    ws.cell(row=2, column=1,
            value=f"Código CVSA {cvsa_status}. Usar AutoFilter (▼) en fila 3 para buscar."
    ).font = Font(italic=True, size=9, color="595959")

    if not df_ok.empty:
        df_mov = df_ok[df_ok["Concepto"].isin(["CI","CN","OP"])].copy()
    else:
        df_mov = pd.DataFrame()

    if not df_mov.empty:
        df_mov["Ticker"] = df_mov["Asset"].str.split("-").str[0]
        resumen = (
            df_mov.groupby(["Ticker","Moneda","Concepto"], sort=True)
            .agg(Compras=("Long Quantity","sum"), Ventas_raw=("Short Quantity","sum"))
            .reset_index()
        )
        resumen["Ventas"] = resumen["Ventas_raw"].abs()
        resumen["Neto"]   = resumen["Compras"] - resumen["Ventas"]
        resumen = resumen[["Ticker","Moneda","Concepto","Compras","Ventas","Neto"]].sort_values(
            ["Ticker","Moneda","Concepto"]).reset_index(drop=True)
    else:
        resumen = pd.DataFrame(columns=["Ticker","Moneda","Concepto","Compras","Ventas","Neto"])

    r = 3
    for i, h in enumerate(["Ticker","Código CVSA","Moneda","Concepto",
                            "Compras (VN)","Ventas (VN)","Neto (VN)"], 1):
        _cel_header(ws, r, i, h)

    r = 4
    fill_pos = PatternFill("solid", fgColor="E2EFDA")
    fill_neg = PatternFill("solid", fgColor="FCE4D6")
    VN_FMT   = "#,##0.##"
    for _, fila in resumen.iterrows():
        neto   = fila["Neto"]
        ticker = fila["Ticker"]
        cvsa   = mapa_cvsa.get(ticker.upper(), "")
        _cel_dato(ws, r, 1, ticker, align="left", bold=True)
        _cel_dato(ws, r, 2, cvsa,   align="center")
        _cel_dato(ws, r, 3, MONEDA_DISPLAY.get(fila["Moneda"], fila["Moneda"]), align="center")
        _cel_dato(ws, r, 4, fila["Concepto"], align="center")
        _cel_dato(ws, r, 5, fila["Compras"],  num_fmt=VN_FMT)
        _cel_dato(ws, r, 6, fila["Ventas"],   num_fmt=VN_FMT)
        c_n = _cel_dato(ws, r, 7, neto, num_fmt=VN_FMT, bold=True)
        if neto > 0:   c_n.fill = fill_pos
        elif neto < 0: c_n.fill = fill_neg
        r += 1

    ws.auto_filter.ref = f"A3:G{max(r-1,3)}"
    for i, w in enumerate([18,16,12,12,16,16,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def escribir_hoja_verificacion(wb, df_verif):
    ws = wb.create_sheet("Verificación")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "HOJA DE VERIFICACIÓN — Assets no clasificados (requieren revisión)"
    c.font      = Font(bold=True, size=12, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=COL_RED_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    if df_verif.empty:
        ws.cell(row=3, column=1, value="Sin registros pendientes de verificación."
                ).font = Font(italic=True, color="006100")
        return

    cols = [c for c in [
        "Asset","Account","Concepto","Moneda","Trade Date","Settlement Date",
        "Short Market Value","Long Market Value","Net Quantity",
        "Filter Node","Motivo_Verificar"
    ] if c in df_verif.columns]

    r = 2
    for i, h in enumerate(cols, 1):
        _cel_header(ws, r, i, h, bg=COL_RED_HDR, font_color="FFFFFF")

    r = 3
    fill_am = PatternFill("solid", fgColor=COL_UNKNOWN)
    for _, row in df_verif.iterrows():
        for i, col in enumerate(cols, 1):
            val = row.get(col, "")
            if pd.isna(val): val = ""
            c = ws.cell(row=r, column=i, value=str(val))
            c.fill = fill_am; c.border = _borde()
            c.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    for i in range(1, len(cols)+1):
        ws.column_dimensions[get_column_letter(i)].width = 24
    ws.freeze_panes = "A3"


def _generar_xls_opciones(df_ok) -> BytesIO:
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Resultado")

    font_bold = xlwt.easyxf("font: bold true")
    style_hdr = xlwt.easyxf(
        "font: bold true; alignment: horizontal centre, vertical centre; "
        "borders: left thin, right thin, top thin, bottom thin")
    style_num = xlwt.easyxf(
        "alignment: horizontal right; borders: left thin, right thin, top thin, bottom thin",
        num_format_str="#,##0.##")
    style_txt = xlwt.easyxf(
        "alignment: horizontal left; borders: left thin, right thin, top thin, bottom thin")
    style_ctr = xlwt.easyxf(
        "alignment: horizontal centre; borders: left thin, right thin, top thin, bottom thin")

    ws.write(0, 0, "Agente 233 - Opciones Sobre Títulos Valores", font_bold)
    headers = ["Agente","Serie","Comitente","Op.Titular","Op.Lanz.Cub",
               "Op.Lanz.Desc","Saldo Titular","Saldo L.C.","Saldo L.D."]
    for col, h in enumerate(headers):
        ws.write(1, col, h, style_hdr)

    if df_ok.empty or "Concepto" not in df_ok.columns:
        buf = BytesIO(); wb.save(buf); buf.seek(0); return buf

    df_op = df_ok[df_ok["Concepto"] == "OP"].copy()
    if df_op.empty:
        buf = BytesIO(); wb.save(buf); buf.seek(0); return buf

    df_op["Comitente"] = df_op["Account"].str.extract(r"_(\d+)\s*\(")
    df_op["Serie"]     = df_op["Asset"].str.strip()
    df_op["op_tit"]    = df_op["Long Quantity"].abs()
    df_op["op_ld"]     = df_op["Short Quantity"].abs()
    df_op["sal_tit"]   = df_op["Long Available Quantity"].abs()
    df_op["sal_ld"]    = df_op["Short Available Quantity"].abs()
    df_op = df_op.sort_values(["Serie","Comitente"]).reset_index(drop=True)

    row = 2
    for serie, grp in df_op.groupby("Serie", sort=True):
        for _, fila in grp.iterrows():
            ctte = fila["Comitente"]
            ws.write(row, 0, 233,          style_ctr)
            ws.write(row, 1, serie,        style_txt)
            ws.write(row, 2, int(ctte) if str(ctte).isdigit() else ctte, style_ctr)
            ws.write(row, 3, fila["op_tit"],  style_num)
            ws.write(row, 4, 0.0,             style_num)
            ws.write(row, 5, fila["op_ld"],   style_num)
            ws.write(row, 6, fila["sal_tit"], style_num)
            ws.write(row, 7, 0.0,             style_num)
            ws.write(row, 8, fila["sal_ld"],  style_num)
            row += 1
        for col in range(9):
            ws.write(row, col, 233 if col==0 else (serie if col==1 else ("Cuenta Operativa" if col==2 else 0.0)),
                     style_ctr if col in (0,2) else (style_txt if col==1 else style_num))
        row += 1

    col_widths = [8,16,18,12,14,14,14,12,12]
    for col, w in enumerate(col_widths):
        ws.col(col).width = w * 256

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def generar_reporte(csv_file, especies_file=None, saldos_inicio_file=None,
                    process_date: date = None, prices_file=None):
    """
    csv_file          : table-currentActualPositions_*.csv (UploadedFile)
    especies_file     : ESPECIES.XLS (UploadedFile, opcional)
    saldos_inicio_file: saldos al inicio Nasdaq.csv (UploadedFile, opcional)
    prices_file       : table-prices_*.csv (UploadedFile, opcional) — cotizaciones MEP/CCL
    process_date      : date (default: hoy)

    Devuelve (BytesIO xlsx, BytesIO xls_opciones, dict resumen).
    """
    if process_date is None:
        process_date = date.today()

    mapa_cvsa    = cargar_mapa_cvsa(especies_file)
    saldos_inicio = cargar_saldos_inicio(saldos_inicio_file)
    cotizaciones  = cargar_cotizaciones_bc(prices_file)
    df_ok, df_verif = cargar_y_clasificar(csv_file, process_date)

    total_rows = len(df_ok) + len(df_verif)
    if total_rows == 0:
        raise ValueError(
            f"No hay registros para Settlement Date = {process_date}. "
            "Verificá que el CSV corresponde a la fecha de proceso."
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Hoja 1: cotizaciones BC (siempre primera, aunque no haya datos)
    escribir_hoja_cotizaciones(wb, cotizaciones, process_date)

    for moneda, sheet_name in MONEDAS:
        escribir_hoja_moneda(wb, sheet_name, df_ok, moneda, process_date,
                             saldo_inicio=saldos_inicio.get(moneda))
    escribir_hoja_movimientos(wb, df_ok, process_date, mapa_cvsa)
    escribir_hoja_verificacion(wb, df_verif)

    xlsx_buf = BytesIO()
    wb.save(xlsx_buf)
    xlsx_buf.seek(0)

    xls_buf = _generar_xls_opciones(df_ok)

    resumen = {
        "process_date":    process_date.strftime("%d/%m/%Y"),
        "total_rows":      total_rows,
        "clasificados":    len(df_ok),
        "verificacion":    len(df_verif),
        "tiene_especies":  bool(mapa_cvsa),
        "tiene_saldos":    bool(saldos_inicio),
        "tiene_cotiz":     bool(cotizaciones.get("mep") or cotizaciones.get("ccl")),
        "cotiz_mep":       cotizaciones.get("mep"),
        "cotiz_ccl":       cotizaciones.get("ccl"),
        "n_op":            int((df_ok["Concepto"] == "OP").sum()) if not df_ok.empty else 0,
        "verif_assets":    df_verif["Asset"].unique().tolist() if not df_verif.empty else [],
    }
    return xlsx_buf, xls_buf, resumen
