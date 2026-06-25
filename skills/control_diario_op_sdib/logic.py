"""
Control diario OP-SDIB (MA) — versión web (Streamlit).
Concilia operaciones PPT BYMA + SENEBI vs CONTBOLE (Gallo) del día.

Outputs (10 hojas):
  Control 999, Pesos ARS, Dolar MEP, USD Cable, Resumen Ope-Titulos,
  Diferencias a Verificar, Operaciones, Cauciones, Trading Intraday, MAV.

Recibe UploadedFiles; devuelve (BytesIO con el Excel, dict con resumen).
"""

from io import BytesIO
from datetime import datetime, date
from collections import defaultdict, Counter
import xlrd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOL      = 0.02
NUM_FMT  = "#,##0.00"
C_GREEN  = "C6EFCE"
C_RED    = "FFC7CE"
C_YELLOW = "FFEB9C"
C_BLUE   = "DDEEFF"
C_HEADER = "4472C4"
C_ORANGE = "FFE0B2"
C_TITLE  = "1E3A5F"
C_TOTAL  = "BFC2C9"
C_HDR_GRAY = "D9D9D9"

DNC_EXCEPTIONS = {
    "DNC5D": ("DNC5O", "DOLAR MEP"),
    "DNC7D": ("DNC7O", "DOLAR MEP"),
    "DNCAD": ("DNCAO", "DOLAR MEP"),
    "DNCAC": ("DNCAO", "USD"),
}

GALLO_CAU  = {"CCCD", "CCTE", "TCCD", "TOCT", "CTCD", "CTTD", "CCCP", "CCFP"}
GALLO_TRAS = {"VTTR", "COTR"}
GALLO_SNB  = {"CRCN", "VRCN"}
GALLO_MAV  = {"VCHM", "CCHM"}
GALLO_OPC  = {"COPR", "VTPR"}   # Opciones (compra/venta de prima) — siempre ARS, siempre T+0

GALLO_SEN = {
    "CPRA": "C", "CPU$": "C", "CPUC": "C", "CSU$": "C",
    "CRCN": "C", "COPR": "C", "CTCD": "C", "CTTD": "C",
    "CCHM": "C",
    "VTAS": "V", "VTU$": "V", "VTUC": "V", "VSU$": "V",
    "VRCN": "V", "VCHM": "V", "VTPR": "V",
    "COTR": "C", "VTTR": "V",
    "CCCD": "V", "CCTE": "V",
    "TCCD": "C", "TOCT": "C",
    "CCCP": "V", "CCFP": "V",
}

GALLO_MON = {
    "CPRA": "Pesos",     "VTAS": "Pesos",
    "CRCN": "Pesos",     "VRCN": "Pesos",
    "VCHM": "Pesos",     "CCHM": "Pesos",
    "COPR": "Pesos",     "VTPR": "Pesos",
    "CPU$": "DOLAR MEP", "VTU$": "DOLAR MEP",
    "CSU$": "DOLAR MEP", "VSU$": "DOLAR MEP",
    "CTCD": "DOLAR MEP", "CTTD": "DOLAR MEP",
    "CPUC": "USD",       "VTUC": "USD",
    "VTTR": "Pesos",     "COTR": "Pesos",
    "CCCD": "Pesos",     "CCTE": "Pesos",
    "TCCD": "Pesos",     "TOCT": "Pesos",
    "CCCP": "DOLAR MEP", "CCFP": "DOLAR MEP",
}

GALLO_CAU_CAP = {"CCCD", "TCCD", "CTCD", "CCCP"}
CTTES_CARTERA = {"1000", "1001", "1002", "1003"}

MON_TO_CODE = {"Pesos": "ARS", "DOLAR MEP": "USD_MEP", "USD": "USD_CABLE"}
MON_LABEL   = {"ARS": "Pesos ARS", "USD_MEP": "Dolar MEP", "USD_CABLE": "USD Cable"}
MON_TITLE   = {"ARS": "PESOS (ARS)", "USD_MEP": "DOLAR MEP (USD)", "USD_CABLE": "USD CABLE (EXT)"}

CONCEPTO_LABEL = {
    "CI":    "CI - Contado Inmediato",
    "OPC":   "CI - Opciones",
    "TI_CI": "CI - Trading Intraday",
    "CAU":   "CAU - Cauciones (apertura/capital)",
    "CN":    "CN - Contado Normal",
    "TI":    "TI - Trading Intraday (plazo T1)",
    "MAV":   "MAV - Operaciones MAV",
}
CONCEPTO_ORDER = {"CI": 0, "OPC": 1, "TI_CI": 2, "CAU": 3, "CN": 4, "TI": 5, "MAV": 6}
SEGMENTO_ORDER = {"G": 0, "NG": 1}


# ── Helpers de parseo ──────────────────────────────────────────────────────────

def _load_byma_ticker_map(especies_file):
    """Lee ESPECIES.XLS y construye {byma_ticker: (gallo_ticker, moneda)}.
    Hoja Datos_Fijos_Especies: col 9=Norm (Gallo pesos), col 10=Parid (BYMA MEP),
    col 11=Cable (BYMA Cable). Cubre excepciones como GOGLD→GOOGL.
    """
    result = {}
    if especies_file is None:
        return result
    try:
        especies_file.seek(0)
        wb = xlrd.open_workbook(file_contents=especies_file.read())
        ws = wb.sheet_by_name("Datos_Fijos_Especies")
    except Exception:
        return result
    SKIP = {"", "NO EXISTE", "0"}
    for r in range(1, ws.nrows):
        norm  = str(ws.cell_value(r, 9)).strip().strip("'")
        parid = str(ws.cell_value(r, 10)).strip().strip("'")
        cable = str(ws.cell_value(r, 11)).strip().strip("'")
        if not norm or norm in SKIP:
            continue
        if parid and parid not in SKIP and parid != norm:
            result[parid] = (norm, "DOLAR MEP")
        if cable and cable not in SKIP and cable != norm:
            result[cable] = (norm, "USD")
    return result


def _map_byma_esp(esp_raw, gallo_esp, byma_map=None):
    esp = esp_raw.strip()
    if esp == "PESOS":
        return ("PESOS", "Pesos")
    if esp in DNC_EXCEPTIONS:
        return DNC_EXCEPTIONS[esp]
    # Lookup explicito ESPECIES.XLS (cubre GOGLD→GOOGL y similares)
    if byma_map and esp in byma_map:
        return byma_map[esp]
    if len(esp) > 1 and esp.endswith("C") and esp[:-1] in gallo_esp:
        return (esp[:-1], "USD")
    if len(esp) > 1 and esp.endswith("D") and esp[:-1] in gallo_esp:
        return (esp[:-1], "DOLAR MEP")
    if len(esp) > 1 and esp.endswith("D") and (esp[:-1] + "O") in gallo_esp:
        return (esp[:-1] + "O", "DOLAR MEP")
    if len(esp) > 1 and esp.endswith("C") and (esp[:-1] + "O") in gallo_esp:
        return (esp[:-1] + "O", "USD")
    return (esp, "Pesos")


def _merge_dicts(a, b):
    """Suma dos dicts {key: [vn, imp]} por clave común."""
    result = defaultdict(lambda: [0., 0.])
    for k, v in a.items():
        result[k][0] += v[0]
        result[k][1] += v[1]
    for k, v in b.items():
        result[k][0] += v[0]
        result[k][1] += v[1]
    return dict(result)


def _detect_process_date(dat_lines):
    """Lee la trade date desde la primera línea de OPERSECEXT (campo [15])."""
    for line in dat_lines:
        p = line.split('"')
        if len(p) >= 16:
            try:
                return datetime.strptime(p[15].strip(), "%Y%m%d").date()
            except Exception:
                continue
    return date.today()


def _parse_byma(lines, gallo_esp, byma_map=None):
    """Parsea OPERSECEXT_GARA.DAT (segmento garantizado PPT BYMA).
    Keys: mer = (ctte, esp, mon, sen, concepto)   concepto in {CI, CN, OPC}
          cau = (ctte, sen, mon)
    Retorna (mer, cau, opt_esp) donde opt_esp es el conjunto de tickers
    de opciones (tipo='O') — siempre ARS, siempre concepto OPC.
    """
    mer = defaultdict(lambda: [0., 0.])
    cau = defaultdict(lambda: [0., 0.])
    opt_esp = set()
    for line in lines:
        p = line.split('"')
        if len(p) < 28:
            continue
        sen  = p[2].strip()
        tipo = p[6].strip()
        ctte = p[17].strip()
        try:
            vn  = float(p[12])
            imp = float(p[19])
        except Exception:
            continue
        if sen not in ("V", "C"):
            continue
        if tipo == "U" or p[7].strip() == "PESOS":
            mon_cau = "DOLAR MEP" if p[7].strip() == "DOLAR" else "Pesos"
            cau[(ctte, sen, mon_cau)][0] += vn
            cau[(ctte, sen, mon_cau)][1] += imp
        elif tipo == "O":
            # Opciones argentinas: siempre ARS, concepto OPC (plazo viene vacio en DAT)
            esp, mon = _map_byma_esp(p[7], gallo_esp, byma_map)
            opt_esp.add(esp)
            mer[(ctte, esp, mon, sen, "OPC")][0] += vn
            mer[(ctte, esp, mon, sen, "OPC")][1] += imp
        else:
            esp, mon = _map_byma_esp(p[7], gallo_esp, byma_map)
            plazo_code = p[9].strip()
            concepto = "CI" if plazo_code == "0" else "CN"
            mer[(ctte, esp, mon, sen, concepto)][0] += vn
            mer[(ctte, esp, mon, sen, concepto)][1] += imp
    return dict(mer), dict(cau), opt_esp


def _parse_senebi(lines, gallo_esp, byma_map=None):
    """Parsea OPERBILEXT_GARA.DAT (SENEBI bilateral).
    Infiere contraparte solo si campo [29] tiene CTTE (evita entradas fantasma
    ctte=None que nunca matchean Gallo).
    Cuando ctte_cli esta vacio (contraparte externa sin cuenta Gallo), registra
    el par en ext_pairs para compensacion posterior (SENEBI GARA EXTERNO).
    Retorna (snb_dict, ext_pairs) donde ext_pairs es lista de
    (cparty, esp, mon, sen_cli, concepto, imp).
    """
    snb = defaultdict(lambda: [0., 0.])
    ext_pairs = []
    for line in lines:
        p = line.split('"')
        if len(p) < 30:
            continue
        sen_cli    = p[4].strip()
        ctte_cli   = p[7].strip()
        esp_raw    = p[9].strip()
        plazo_code = p[10].strip()
        cparty     = p[29].strip()
        try:
            vn  = float(p[13])
            imp = float(p[15])
        except Exception:
            continue
        if sen_cli not in ("V", "C"):
            continue
        esp, mon = _map_byma_esp(esp_raw, gallo_esp, byma_map)
        concepto = "CI" if plazo_code == "0" else "CN"
        snb[(ctte_cli, esp, mon, sen_cli, concepto)][0] += vn
        snb[(ctte_cli, esp, mon, sen_cli, concepto)][1] += imp
        if cparty:
            sen_cp = "V" if sen_cli == "C" else "C"
            snb[(cparty, esp, mon, sen_cp, concepto)][0] += vn
            snb[(cparty, esp, mon, sen_cp, concepto)][1] += imp
            if not ctte_cli:
                # Contraparte externa (sin cuenta Gallo) → guardar par para compensacion
                ext_pairs.append((cparty, esp, mon, sen_cli, concepto, imp))
    return dict(snb), ext_pairs


def _parse_gallo(file_bytes):
    """Parsea CONTBOLE.XLS desde bytes. Devuelve (esp_set, mer, cau, tra, snb, mav, arancel_alerts).
    Concepto CI/CN para PPT/SENEBI según Fec_Ope==Fec_Liq (CI) vs Fec_Liq>Fec_Ope (CN).
    """
    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_name("Control_de_Boletos")
    esp_set = set()
    mer = defaultdict(lambda: [0., 0.])
    cau = defaultdict(lambda: {"capital": 0., "total": 0.})
    tra = defaultdict(lambda: [0., 0.])
    snb = defaultdict(lambda: [0., 0.])
    mav = []
    arancel_alerts = []
    for i in range(1, ws.nrows):
        op  = str(ws.cell_value(i, 1)).strip()
        raw = ws.cell_value(i, 4)
        if not op or not raw:
            continue
        try:
            ctte = str(int(float(raw)))
        except Exception:
            continue
        fope = str(ws.cell_value(i, 2)).strip()
        fliq = str(ws.cell_value(i, 3)).strip()
        esp     = str(ws.cell_value(i, 6)).strip()
        imp     = float(ws.cell_value(i, 7)) if ws.cell_value(i, 7) else 0.
        vn      = float(ws.cell_value(i, 8)) if ws.cell_value(i, 8) else 0.
        arancel = float(ws.cell_value(i, 9)) if ws.cell_value(i, 9) else 0.
        esp_set.add(esp)
        sen = GALLO_SEN.get(op)
        mon = GALLO_MON.get(op)
        concepto_ppt = "CI" if fope and fliq and fope == fliq else "CN"
        if ctte in CTTES_CARTERA and abs(arancel) > 0.001:
            boleto = str(ws.cell_value(i, 0)).strip()
            arancel_alerts.append({
                "boleto": boleto, "ctte": ctte, "op": op,
                "esp": esp, "arancel": arancel,
            })
        if op in GALLO_CAU:
            mon_cau = GALLO_MON.get(op, "Pesos")
            if op in GALLO_CAU_CAP:
                cau[(ctte, sen, mon_cau)]["capital"] += imp
            else:
                cau[(ctte, sen, mon_cau)]["total"] += imp
        elif op in GALLO_TRAS:
            if sen and mon:
                tra[(ctte, esp, mon, sen, concepto_ppt)][0] += vn
                tra[(ctte, esp, mon, sen, concepto_ppt)][1] += imp
        elif op in GALLO_MAV:
            mav.append({"ctte": ctte, "op": op, "esp": esp,
                        "sen": sen or "V", "mon": mon or "Pesos",
                        "vn": vn, "imp": imp})
        elif op in GALLO_SNB:
            if sen and mon:
                snb[(ctte, esp, mon, sen, concepto_ppt)][0] += vn
                snb[(ctte, esp, mon, sen, concepto_ppt)][1] += imp
        elif op in GALLO_OPC:
            # Opciones: siempre ARS, siempre concepto OPC (independiente del plazo)
            if sen and mon:
                mer[(ctte, esp, mon, sen, "OPC")][0] += vn
                mer[(ctte, esp, mon, sen, "OPC")][1] += imp
        elif sen and mon:
            mer[(ctte, esp, mon, sen, concepto_ppt)][0] += vn
            mer[(ctte, esp, mon, sen, concepto_ppt)][1] += imp
    return esp_set, dict(mer), dict(cau), dict(tra), dict(snb), mav, arancel_alerts


# ── Comparadores ──────────────────────────────────────────────────────────────

def _get_st(in_b, in_g, difs):
    if in_b and in_g:
        return "OK" if all(abs(d) <= TOL for d in difs) else "DIFERENCIA"
    return "SOLO BYMA" if in_b else "SOLO GALLO"


def _compare_ops(b, g, ppt_keys_b, ppt_keys_g, snb_keys_b, snb_keys_g):
    """Clave: (ctte, esp, mon, sen, concepto). Columnas SEGMENTO y VERIFICAR."""
    rows = []
    for k in sorted(set(b) | set(g)):
        ctte, esp, mon, sen, concepto = k
        bv = b.get(k, [0., 0.])
        gv = g.get(k, [0., 0.])
        dv = bv[0] - gv[0]
        di = bv[1] - gv[1]
        in_ppt = k in ppt_keys_b or k in ppt_keys_g
        in_snb = k in snb_keys_b or k in snb_keys_g
        seg_mer  = ("PPT+SENEBI" if in_ppt and in_snb
                    else ("PPT" if in_ppt else "SENEBI"))
        segmento = "G" if in_ppt else "NG"
        estado = _get_st(k in b, k in g, [dv, di])
        verificar = ""
        if ctte == "1002":
            verificar = "CTA 1002"
        elif estado == "SOLO BYMA":
            verificar = "FALTA BOLETO"
        elif estado not in ("OK", "TRADING INTRADAY"):
            verificar = "VERIFICAR"
        rows.append({
            "ctte": ctte, "especie": esp, "moneda": mon, "sentido": sen,
            "concepto": concepto, "segmento_mer": seg_mer, "segmento": segmento,
            "vn_b": bv[0], "vn_g": gv[0], "dif_vn": dv,
            "imp_b": bv[1], "imp_g": gv[1], "dif_imp": di,
            "estado": estado, "verificar": verificar,
        })
    return rows


def _second_pass_ticker(rows):
    """Re-empareja SOLO BYMA vs SOLO GALLO donde el ticker difiere solo en el
    último carácter (convención de moneda O/D/C). Mismo ctte, sentido, concepto, VN, IMP."""
    solo_b = [(i, r) for i, r in enumerate(rows) if r["estado"] == "SOLO BYMA"]
    solo_g = [(i, r) for i, r in enumerate(rows) if r["estado"] == "SOLO GALLO"]
    absorbed_g = set()
    for ib, rb in solo_b:
        esp_b = rb["especie"]
        if len(esp_b) <= 1:
            continue
        for ig, rg in solo_g:
            if ig in absorbed_g:
                continue
            esp_g = rg["especie"]
            if len(esp_g) <= 1:
                continue
            if (rb["ctte"]     == rg["ctte"] and
                rb["sentido"]  == rg["sentido"] and
                rb["concepto"] == rg["concepto"] and
                esp_b[:-1]     == esp_g[:-1] and
                esp_b          != esp_g and
                abs(rb["vn_b"]  - rg["vn_g"])  <= TOL and
                abs(rb["imp_b"] - rg["imp_g"]) <= TOL):
                rows[ib]["vn_g"]      = rg["vn_g"]
                rows[ib]["imp_g"]     = rg["imp_g"]
                rows[ib]["dif_vn"]    = rb["vn_b"]  - rg["vn_g"]
                rows[ib]["dif_imp"]   = rb["imp_b"] - rg["imp_g"]
                rows[ib]["especie"]   = rg["especie"]
                rows[ib]["moneda"]    = rg["moneda"]
                rows[ib]["estado"]    = "OK"
                rows[ib]["verificar"] = ""
                absorbed_g.add(ig)
                break
    return [r for i, r in enumerate(rows) if i not in absorbed_g]


def _compare_cau(b, g):
    rows = []
    for k in sorted(set(b) | set(g)):
        ctte, sen, mon = k
        bv = b.get(k, [0., 0.])
        gv = g.get(k, {"capital": 0., "total": 0.})
        dc = bv[0] - gv["capital"]
        dt = bv[1] - gv["total"]
        rows.append({
            "ctte": ctte, "sentido": sen, "moneda": mon,
            "vn_b": bv[0], "cap_g": gv["capital"], "dif_cap": dc,
            "imp_b": bv[1], "tot_g": gv["total"],  "dif_tot": dt,
            "estado": _get_st(k in b, k in g, [dc, dt]),
        })
    return rows


def _compare_tra(g):
    rows = []
    for k in sorted(g):
        ctte, esp, mon, sen, concepto = k
        gv = g[k]
        rows.append({
            "ctte": ctte, "especie": esp, "moneda": mon, "sentido": sen,
            "concepto": concepto,
            "vn_b": 0., "vn_g": gv[0], "dif_vn": -gv[0],
            "imp_b": 0., "imp_g": gv[1], "dif_imp": -gv[1],
            "estado": "TRADING INTRADAY",
        })
    return rows


def _analyze_1002(ops_rows):
    """Para cada (esp, mon, sen, concepto) con CTA 1002 verifica si la dif
    está compensada por otros comitentes. Suma neto debería ser 0."""
    rows_1002 = {(r["especie"], r["moneda"], r["sentido"], r["concepto"]): r
                 for r in ops_rows if r["ctte"] == "1002"}
    results = []
    for k, r1002 in rows_1002.items():
        esp, mon, sen, concepto = k
        dif_1002 = r1002["dif_vn"]
        otros = [(r["ctte"], r["dif_vn"])
                 for r in ops_rows
                 if r["ctte"] != "1002"
                 and r["especie"]  == esp
                 and r["moneda"]   == mon
                 and r["sentido"]  == sen
                 and r["concepto"] == concepto
                 and abs(r["dif_vn"]) > TOL]
        dif_otros = sum(v for _, v in otros)
        neto = dif_1002 + dif_otros
        results.append({
            "especie":     esp,
            "moneda":      mon,
            "sentido":     sen,
            "concepto":    concepto,
            "estado_1002": r1002["estado"],
            "vn_b_1002":   r1002["vn_b"],
            "vn_g_1002":   r1002["vn_g"],
            "dif_1002":    dif_1002,
            "otros_cttes": ", ".join(c for c, _ in otros) or "-",
            "dif_otros":   dif_otros,
            "neto":        neto,
            "balance":     "OK" if abs(neto) <= TOL else "REVISAR",
        })
    return results


def _apply_1002_compensation(ops_rows, analisis_1002):
    for a in analisis_1002:
        if a["balance"] != "OK":
            continue
        esp, mon, sen, concepto = a["especie"], a["moneda"], a["sentido"], a["concepto"]
        otros_set = {c.strip() for c in a["otros_cttes"].split(",")
                     if c.strip() and c.strip() != "-"}
        for r in ops_rows:
            if (r["especie"]  != esp or r["moneda"]   != mon
                or r["sentido"]  != sen or r["concepto"] != concepto):
                continue
            if r["estado"] == "OK":
                continue
            if r["ctte"] == "1002" or r["ctte"] in otros_set:
                r["estado"]    = "OK"
                r["verificar"] = "CTA 1002"
    return ops_rows


def _apply_senebi_gara_externo_compensation(ops_rows, ext_pairs):
    """Compensa pares OPERBILEXT donde ctte_cli es vacío (contraparte externa sin
    cuenta Gallo). El lado 'sin CTTE' (SOLO BYMA) y el lado de la contraparte
    (exceso de BYMA en DIFERENCIA) son las dos caras del mismo trade SENEBI gara
    y netean a cero. Ambas filas se marcan OK/'SENEBI GARA EXTERNO'.
    """
    n_comp = 0
    for (cparty, esp, mon, sen_cli, concepto, exp_imp) in ext_pairs:
        sen_cp = "V" if sen_cli == "C" else "C"
        row_empty = next(
            (r for r in ops_rows
             if r["ctte"] == "" and r["especie"] == esp and r["moneda"] == mon
             and r["sentido"] == sen_cli and r["concepto"] == concepto
             and abs(r["imp_b"] - exp_imp) <= TOL),
            None,
        )
        row_cp = next(
            (r for r in ops_rows
             if r["ctte"] == cparty and r["especie"] == esp and r["moneda"] == mon
             and r["sentido"] == sen_cp and r["concepto"] == concepto
             and abs(r["dif_imp"] - exp_imp) <= TOL),
            None,
        )
        if row_empty:
            row_empty["estado"]    = "OK"
            row_empty["verificar"] = "SENEBI GARA EXTERNO"
            n_comp += 1
        if row_cp:
            row_cp["estado"]    = "OK"
            row_cp["verificar"] = "SENEBI GARA EXTERNO"
            n_comp += 1
    return ops_rows, n_comp


def _apply_senebi_nogara_compensation(ops_rows, snb_keys_g):
    """Detecta SENEBI bilateral no garantizado: boletos Gallo (CRCN/VRCN) sin
    contraparte BYMA que se compensan entre sí (compras = ventas por especie/
    moneda/concepto). Estos back-to-back no-gara se marcan OK/SENEBI NO GARA
    y se excluyen de la hoja Diferencias a Verificar.
    """
    candidatas = [r for r in ops_rows
                  if r["estado"] == "SOLO GALLO"
                  and (r["ctte"], r["especie"], r["moneda"],
                       r["sentido"], r["concepto"]) in snb_keys_g]
    grupos = defaultdict(lambda: {"C": 0., "V": 0., "rows": []})
    for r in candidatas:
        key = (r["especie"], r["moneda"], r["concepto"])
        grupos[key][r["sentido"]] += r["vn_g"]
        grupos[key]["rows"].append(r)
    compensadas = 0
    for key, g in grupos.items():
        if abs(g["C"] - g["V"]) <= TOL:
            for r in g["rows"]:
                r["estado"]    = "OK"
                r["verificar"] = "SENEBI NO GARA"
            compensadas += len(g["rows"])
    return ops_rows, compensadas


# ── Agregados por moneda ───────────────────────────────────────────────────────

def _ar_ae(sen, imp):
    """PPT/SENEBI: Venta -> A Recibir cash, Compra -> A Entregar cash."""
    if sen == "V": return imp, 0.0
    if sen == "C": return 0.0, imp
    return 0.0, 0.0


def _ar_ae_cau(sen, cap):
    """Caución apertura: Tomadora (C) recibe capital; Colocadora (V) entrega capital."""
    if sen == "C": return cap, 0.0
    if sen == "V": return 0.0, cap
    return 0.0, 0.0


def _build_currency_agg(mer_b, snb_b, cau_b, mer_g, snb_g, cau_g, tra_g, mav_g):
    """{moneda_code: {(concepto, segmento): {'ar_b','ae_b','ar_g','ae_g','imp_b','imp_g'}}}.
    Ops BYMA con contraparte TI en Gallo -> bucket TI también del lado BYMA.
    """
    agg = {m: defaultdict(lambda: {"ar_b": 0., "ae_b": 0., "ar_g": 0., "ae_g": 0.,
                                    "imp_b": 0., "imp_g": 0.})
           for m in ("ARS", "USD_MEP", "USD_CABLE")}

    # tra_g ahora tiene clave 5-tuple (ctte, esp, mon, sen, concepto)
    # Para el reclasificador BYMA->TI usamos solo el 4-tuple (sin concepto)
    tra_keys_4 = {(c, e, m, s) for (c, e, m, s, _) in tra_g}

    def _add(side, mon_internal, concepto, segmento, sen, imp):
        code = MON_TO_CODE.get(mon_internal)
        if not code or code not in agg:
            return
        ar, ae = _ar_ae(sen, imp)
        cell = agg[code][(concepto, segmento)]
        if side == "B":
            cell["ar_b"] += ar; cell["ae_b"] += ae; cell["imp_b"] += imp
        else:
            cell["ar_g"] += ar; cell["ae_g"] += ae; cell["imp_g"] += imp

    def _add_cau(side, mon_internal, sen, cap):
        code = MON_TO_CODE.get(mon_internal)
        if not code:
            return
        ar, ae = _ar_ae_cau(sen, cap)
        cell = agg[code][("CAU", "G")]
        if side == "B":
            cell["ar_b"] += ar; cell["ae_b"] += ae; cell["imp_b"] += cap
        else:
            cell["ar_g"] += ar; cell["ae_g"] += ae; cell["imp_g"] += cap

    for k, v in mer_b.items():
        ctte, esp, mon, sen, concepto = k
        if (ctte, esp, mon, sen) in tra_keys_4:
            # TI plazo CI liquida hoy / TI plazo T1 (CN) liquida mañana
            ti_bucket = "TI_CI" if concepto == "CI" else "TI"
            _add("B", mon, ti_bucket, "G", sen, v[1])
        else:
            _add("B", mon, concepto, "G", sen, v[1])
    for k, v in mer_g.items():
        ctte, esp, mon, sen, concepto = k
        _add("G", mon, concepto, "G", sen, v[1])

    for k, v in snb_b.items():
        ctte, esp, mon, sen, concepto = k
        _add("B", mon, concepto, "NG", sen, v[1])
    for k, v in snb_g.items():
        ctte, esp, mon, sen, concepto = k
        _add("G", mon, concepto, "NG", sen, v[1])

    for k, v in cau_b.items():
        ctte, sen, mon_cau = k
        _add_cau("B", mon_cau, sen, v[0])
    for k, v in cau_g.items():
        ctte, sen, mon_cau = k
        _add_cau("G", mon_cau, sen, v["capital"])

    for k, v in tra_g.items():
        ctte, esp, mon, sen, concepto = k
        code = MON_TO_CODE.get(mon)
        if not code:
            continue
        ar, ae = _ar_ae(sen, v[1])
        ti_bucket = "TI_CI" if concepto == "CI" else "TI"
        cell = agg[code][(ti_bucket, "G")]
        cell["ar_g"] += ar; cell["ae_g"] += ae; cell["imp_g"] += v[1]

    for r in mav_g:
        code = MON_TO_CODE.get(r["mon"])
        if not code:
            continue
        ar, ae = _ar_ae(r["sen"], r["imp"])
        cell = agg[code][("MAV", "NG")]
        cell["ar_g"] += ar; cell["ae_g"] += ae; cell["imp_g"] += r["imp"]

    return agg


# ── Actual Position: saldo proyectado por moneda ───────────────────────────────

def _cargar_saldos_actual_position_file(ap_file):
    """Lee un Actual Position xlsx (UploadedFile) y extrae 'Saldo proyectado del dia'
    (col G) y datos auxiliares (col J/K) por moneda.
    Retorna (dict_saldos, dict_extras, filename) o ({}, {}, '')."""
    if ap_file is None:
        return {}, {}, ""
    try:
        ap_file.seek(0)
        wb = openpyxl.load_workbook(BytesIO(ap_file.read()), data_only=True)
    except Exception:
        return {}, {}, ""
    sheet_map = [
        ("ARS",       ("Pesos ARS",)),
        ("USD_MEP",   ("Dólar MEP", "Dolar MEP", "Dolar MEP (USD)")),
        ("USD_CABLE", ("USD Cable",)),
    ]
    saldos = {}
    extras = {}
    for code, candidates in sheet_map:
        sname = next((c for c in candidates if c in wb.sheetnames), None)
        if not sname:
            continue
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if not label:
                continue
            label_str = str(label).lower()
            if "saldo proyectado" in label_str:
                val = ws.cell(r, 7).value
                if isinstance(val, (int, float)):
                    saldos[code] = val
                aux_val_j = ws.cell(r, 9).value  if ws.max_column >= 9  else None
                aux_lbl_j = ws.cell(r, 10).value if ws.max_column >= 10 else None
                if isinstance(aux_val_j, (int, float)) and isinstance(aux_lbl_j, str) \
                        and "999" in aux_lbl_j.lower():
                    extras.setdefault(code, {})["saldo_999"] = aux_val_j
            if "saldo al inicio" in label_str:
                aux_val_j = ws.cell(r, 9).value  if ws.max_column >= 9  else None
                aux_lbl_j = ws.cell(r, 10).value if ws.max_column >= 10 else None
                if isinstance(aux_val_j, (int, float)) and isinstance(aux_lbl_j, str) \
                        and "pendiente" in aux_lbl_j.lower():
                    extras.setdefault(code, {})["pendiente"] = aux_val_j
    return saldos, extras, getattr(ap_file, "name", "Actual Position.xlsx")


# ── Helpers Excel ──────────────────────────────────────────────────────────────

def _fl(c):  return PatternFill("solid", fgColor=c)
def _hf():   return Font(bold=True, color="FFFFFF")
def _bf():   return Font(bold=True)

def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _wh(ws, hdrs, row=1):
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row, c, h)
        cell.font = _hf()
        cell.fill = _fl(C_HEADER)
        cell.alignment = Alignment(horizontal="center")

def _nf(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.number_format = NUM_FMT
    return cell

def _rf(ws, r, n, color):
    for c in range(1, n + 1):
        ws.cell(r, c).fill = _fl(color)

def _af(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, 42)

def _st_color(estado):
    return {"OK": C_GREEN, "DIFERENCIA": C_RED,
            "SOLO BYMA": C_YELLOW, "SOLO GALLO": C_BLUE,
            "TRADING INTRADAY": C_ORANGE}.get(estado, "FFFFFF")


# ── Hoja Control 999 ───────────────────────────────────────────────────────────

def _emit_row_999(ws, r, label, neto_b, neto_g, nota="", is_subtotal=False):
    """Escribe una fila de movimiento en Control 999 y retorna la siguiente fila."""
    dif = neto_b - neto_g
    ws.cell(row=r, column=1, value=label)
    cell_b = _nf(ws, r, 2, neto_b)
    cell_g = _nf(ws, r, 3, neto_g)
    cell_d = _nf(ws, r, 4, dif)
    if is_subtotal:
        for cell in [ws.cell(row=r, column=1), cell_b, cell_g, cell_d]:
            cell.font = Font(bold=True, italic=True)
    if nota:
        ws.cell(row=r, column=5, value=nota).font = Font(italic=True, size=9, color="595959")
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = _border()
        if is_subtotal:
            ws.cell(row=r, column=col).fill = _fl(C_TOTAL)
    if abs(dif) > TOL:
        ws.cell(row=r, column=4).fill = _fl(C_RED)
        ws.cell(row=r, column=4).font = Font(bold=True, italic=True) if is_subtotal else _bf()
    return r + 1


def _write_control_999(wb, agg, ap_saldos, ap_extras, ap_filename, process_date):
    """Hoja Control 999 — saldo proyectado a fin del dia por moneda.
    Estructura por moneda:
      SALDO ACTUAL POSITION
      CI / CI Trading Intraday / CAU   (liquida hoy)
      SALDO FINAL DIARIO  = Saldo AP + CI + TI_CI + CAU
      CN / TI plazo T1                  (informacional — liquida mañana)
    """
    ws = wb.create_sheet("Control 999")
    ws.sheet_view.showGridLines = False
    fecha = process_date.strftime('%d/%m/%Y')

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"CONTROL 999 - SALDO PROYECTADO A FIN DEL DIA - {fecha}"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = _fl(C_TITLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    src = ap_filename or "(Actual Position no subido)"
    c2 = ws.cell(row=2, column=1,
                 value=f"Saldo inicio leido de: {src}  |  "
                       f"Movimientos del dia desde OPERSECEXT/OPERBILEXT y CONTBOLE")
    c2.font = Font(italic=True, size=9, color="595959")

    r = 4
    for code in ("ARS", "USD_MEP", "USD_CABLE"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(row=r, column=1, value=MON_TITLE[code])
        c.font = Font(bold=True, size=12, color="FFFFFF")
        c.fill = _fl(C_TITLE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

        for i, h in enumerate(["Concepto", "BYMA", "Gallo x boletos", "Diferencia", "Notas"], 1):
            cell = ws.cell(row=r, column=i, value=h)
            cell.fill = _fl(C_HEADER)
            cell.font = _hf()
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()
        r += 1

        saldo_inicio = ap_saldos.get(code)
        ag = agg[code]

        def neto_buckets(buckets):
            b = sum(ag.get(bk, {}).get("ar_b", 0.) - ag.get(bk, {}).get("ae_b", 0.)
                    for bk in buckets)
            g = sum(ag.get(bk, {}).get("ar_g", 0.) - ag.get(bk, {}).get("ae_g", 0.)
                    for bk in buckets)
            return b, g

        # SALDO ACTUAL POSITION
        ws.cell(row=r, column=1,
                value="SALDO ACTUAL POSITION (proyectado del dia)").font = _bf()
        if saldo_inicio is not None:
            _nf(ws, r, 2, saldo_inicio).font = _bf()
            _nf(ws, r, 3, saldo_inicio).font = _bf()
            _nf(ws, r, 4, 0.0).font = _bf()
        else:
            ws.cell(row=r, column=2, value="N/D").alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=3, value="N/D").alignment = Alignment(horizontal="right")
        extras_code = ap_extras.get(code, {})
        nota_inicio = []
        if "saldo_999" in extras_code:
            nota_inicio.append(f"999 dia previo: {extras_code['saldo_999']:,.2f}")
        if "pendiente" in extras_code:
            nota_inicio.append(f"pend. liquidar: {extras_code['pendiente']:,.2f}")
        if nota_inicio:
            ws.cell(row=r, column=5, value=" | ".join(nota_inicio)).font = \
                Font(italic=True, size=9, color="595959")
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = _border()
            ws.cell(row=r, column=col).fill = _fl(C_HDR_GRAY)
        r += 1

        # CI (G + NG merged) — liquida hoy
        ci_b, ci_g = neto_buckets([("CI", "G"), ("CI", "NG")])
        if abs(ci_b) + abs(ci_g) > TOL:
            r = _emit_row_999(ws, r, "  CI - Contado Inmediato",
                              ci_b, ci_g, "liquida hoy (T+0)")

        # CI Opciones — liquida hoy (siempre ARS, desglosado de CI normal)
        opc_b, opc_g = neto_buckets([("OPC", "G")])
        if abs(opc_b) + abs(opc_g) > TOL:
            r = _emit_row_999(ws, r, "  CI - Opciones",
                              opc_b, opc_g, "opciones de prima - liquida hoy")

        # CI TRADING INTRADAY — liquida hoy
        ti_ci_b, ti_ci_g = neto_buckets([("TI_CI", "G")])
        if abs(ti_ci_b) + abs(ti_ci_g) > TOL:
            r = _emit_row_999(ws, r, "  CI - Trading Intraday",
                              ti_ci_b, ti_ci_g, "intraday - liquida hoy")

        # CAU (apertura/capital) — liquida hoy
        cau_b, cau_g = neto_buckets([("CAU", "G")])
        if abs(cau_b) + abs(cau_g) > TOL:
            r = _emit_row_999(ws, r, "  CAU - Cauciones (apertura/capital)",
                              cau_b, cau_g, "apertura - capital hoy")

        # SALDO FINAL DIARIO = Saldo AP + CI + OPC + CI_TI + CAU
        total_hoy_b = ci_b + opc_b + ti_ci_b + cau_b
        total_hoy_g = ci_g + opc_g + ti_ci_g + cau_g
        if saldo_inicio is not None:
            saldo_final_b = saldo_inicio + total_hoy_b
            saldo_final_g = saldo_inicio + total_hoy_g
        else:
            saldo_final_b = total_hoy_b
            saldo_final_g = total_hoy_g
        r = _emit_row_999(ws, r, "SALDO FINAL DIARIO",
                          saldo_final_b, saldo_final_g, is_subtotal=True)

        # CN (G + NG merged) — liquida mañana (informacional)
        cn_b, cn_g = neto_buckets([("CN", "G"), ("CN", "NG")])
        if abs(cn_b) + abs(cn_g) > TOL:
            r = _emit_row_999(ws, r, "  CN - Contado Normal",
                              cn_b, cn_g, "liquida manana (T+1)")

        # TI plazo T1 — liquida mañana (informacional)
        ti_b, ti_g = neto_buckets([("TI", "G")])
        if abs(ti_b) + abs(ti_g) > TOL:
            r = _emit_row_999(ws, r, "  TI - Trading Intraday (plazo T1)",
                              ti_b, ti_g, "trading intraday T1")

        r += 2

    for i, w in enumerate([44, 22, 22, 20, 32], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    return ws


# ── Hojas por moneda ───────────────────────────────────────────────────────────

def _write_currency_sheet(wb, code, agg_code, process_date):
    title_label = MON_LABEL[code]
    ws = wb.create_sheet(title_label)
    ws.sheet_view.showGridLines = False
    fecha_str = process_date.strftime("%d/%m/%Y")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"VALORIZACION OPERACIONES {MON_TITLE[code]} - {fecha_str}"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = _fl(C_TITLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    c2 = ws.cell(row=2, column=1,
                 value="A Recibir = ventas (PPT/SENEBI) + cauciones tomadoras (apertura). "
                       "A Entregar = compras + cauciones colocadoras. "
                       "Cauciones valorizadas por capital (apertura).")
    c2.font = Font(italic=True, size=9, color="595959")

    headers = ["Concepto", "Segmento",
               "A Recibir BYMA", "A Entregar BYMA", "Neto BYMA",
               "A Recibir Gallo", "A Entregar Gallo", "Neto Gallo",
               "Dif Neto", "Estado"]
    r = 4
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = _fl(C_HEADER); cell.font = _hf()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    ws.row_dimensions[r].height = 30

    keys = sorted(agg_code.keys(),
                  key=lambda x: (CONCEPTO_ORDER.get(x[0], 99),
                                 SEGMENTO_ORDER.get(x[1], 99)))

    tot_ar_b = tot_ae_b = tot_ar_g = tot_ae_g = 0.0
    r = 5
    for (concepto, segmento) in keys:
        d = agg_code[(concepto, segmento)]
        ar_b, ae_b = d["ar_b"], d["ae_b"]
        ar_g, ae_g = d["ar_g"], d["ae_g"]
        neto_b = ar_b - ae_b
        neto_g = ar_g - ae_g
        dif = neto_b - neto_g
        if abs(ar_b) + abs(ae_b) + abs(ar_g) + abs(ae_g) <= TOL:
            continue
        movs_ok = (abs(ar_b - ar_g) <= TOL) and (abs(ae_b - ae_g) <= TOL)
        estado = "OK" if (abs(dif) <= TOL and movs_ok) else "DIFERENCIA"
        co = C_GREEN if estado == "OK" else C_RED

        ws.cell(row=r, column=1, value=CONCEPTO_LABEL.get(concepto, concepto)).alignment = \
            Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=2, value=segmento).alignment = \
            Alignment(horizontal="center", vertical="center")
        _nf(ws, r, 3, ar_b)
        _nf(ws, r, 4, ae_b)
        _nf(ws, r, 5, neto_b).font = _bf()
        _nf(ws, r, 6, ar_g)
        _nf(ws, r, 7, ae_g)
        _nf(ws, r, 8, neto_g).font = _bf()
        cell_dif = _nf(ws, r, 9, dif); cell_dif.font = _bf()
        cell_est = ws.cell(row=r, column=10, value=estado)
        cell_est.fill = _fl(co)
        cell_est.alignment = Alignment(horizontal="center")
        for col in range(1, 11):
            ws.cell(row=r, column=col).border = _border()
        tot_ar_b += ar_b; tot_ae_b += ae_b
        tot_ar_g += ar_g; tot_ae_g += ae_g
        r += 1

    neto_b_tot = tot_ar_b - tot_ae_b
    neto_g_tot = tot_ar_g - tot_ae_g
    dif_tot = neto_b_tot - neto_g_tot
    est_tot = "OK" if abs(dif_tot) <= TOL else "DIFERENCIA"
    co_tot = C_GREEN if est_tot == "OK" else C_RED

    ws.cell(row=r, column=1, value="TOTAL").font = _bf()
    ws.cell(row=r, column=2, value="").alignment = Alignment(horizontal="center")
    for col, val in [(3, tot_ar_b), (4, tot_ae_b), (5, neto_b_tot),
                     (6, tot_ar_g), (7, tot_ae_g), (8, neto_g_tot), (9, dif_tot)]:
        cell = _nf(ws, r, col, val); cell.font = _bf(); cell.fill = _fl(C_TOTAL)
    ws.cell(row=r, column=1).fill = _fl(C_TOTAL)
    ws.cell(row=r, column=2).fill = _fl(C_TOTAL)
    cell_est_tot = ws.cell(row=r, column=10, value=est_tot)
    cell_est_tot.fill = _fl(co_tot); cell_est_tot.font = _bf()
    cell_est_tot.alignment = Alignment(horizontal="center")
    for col in range(1, 11):
        ws.cell(row=r, column=col).border = _border()

    for i, w in enumerate([34, 12, 18, 18, 18, 18, 18, 18, 18, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    return ws


# ── Hoja Diferencias a Verificar ───────────────────────────────────────────────

def _build_verification_rows(ops_rows, cau_rows):
    out = []
    for r in ops_rows:
        # Excluir filas OK sin verificar, y las compensadas (SENEBI NO GARA / SENEBI GARA EXTERNO)
        if r["estado"] == "OK" and r["verificar"] in ("", "SENEBI NO GARA", "SENEBI GARA EXTERNO"):
            continue
        if r["estado"] == "TRADING INTRADAY":
            continue
        out.append({
            "ctte": r["ctte"], "especie": r["especie"], "moneda": r["moneda"],
            "sentido": r["sentido"], "concepto": r["concepto"], "segmento": r["segmento"],
            "imp_b": r["imp_b"], "imp_g": r["imp_g"], "dif_imp": r["dif_imp"],
            "estado": r["estado"], "verificar": r["verificar"], "origen": "OPERACION",
        })
    for r in cau_rows:
        if r["estado"] == "OK":
            continue
        out.append({
            "ctte": r["ctte"], "especie": "(CAUCION)", "moneda": r["moneda"],
            "sentido": r["sentido"], "concepto": "CAU", "segmento": "G",
            "imp_b": r["imp_b"], "imp_g": r["tot_g"], "dif_imp": r["dif_tot"],
            "estado": r["estado"],
            "verificar": "CTA 1002" if r["ctte"] == "1002" else "VERIFICAR",
            "origen": "CAUCION",
        })
    estado_order = {"DIFERENCIA": 0, "SOLO BYMA": 1, "SOLO GALLO": 2, "OK": 3}
    out.sort(key=lambda x: (estado_order.get(x["estado"], 9), x["moneda"], x["ctte"]))
    return out


def _write_verification_sheet(wb, ver_rows, process_date):
    ws = wb.create_sheet("Diferencias a Verificar")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = (f"DIFERENCIAS A VERIFICAR - {process_date.strftime('%d/%m/%Y')} - "
               f"{len(ver_rows)} fila(s)")
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = _fl(C_TITLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    c2 = ws.cell(row=2, column=1,
                 value="Operaciones valorizadas con discrepancia BYMA vs Gallo. "
                       "AutoFilter activo. Focos: FALTA BOLETO | CTA 1002.")
    c2.font = Font(italic=True, size=9, color="595959")

    headers = ["CTTE", "ESPECIE", "MONEDA", "SENTIDO", "CONCEPTO", "SEGMENTO",
               "ORIGEN", "IMPORTE BYMA", "IMPORTE GALLO", "DIFERENCIA",
               "ESTADO", "VERIFICAR"]
    r = 3
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = _fl(C_HEADER); cell.font = _hf()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
    ws.row_dimensions[r].height = 24

    widths = [10, 14, 14, 10, 12, 12, 12, 18, 18, 16, 14, 18]
    if not ver_rows:
        ws.cell(row=4, column=1,
                value="Sin diferencias por verificar. Conciliacion OK.").font = \
            Font(italic=True, color="006100")
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = "A3:L3"
        return ws

    r = 4
    for row in ver_rows:
        co = _st_color(row["estado"])
        ws.cell(row=r, column=1, value=row["ctte"])
        ws.cell(row=r, column=2, value=row["especie"])
        ws.cell(row=r, column=3, value=row["moneda"])
        ws.cell(row=r, column=4, value=row["sentido"])
        ws.cell(row=r, column=5, value=row["concepto"])
        ws.cell(row=r, column=6, value=row["segmento"])
        ws.cell(row=r, column=7, value=row["origen"])
        _nf(ws, r, 8, row["imp_b"])
        _nf(ws, r, 9, row["imp_g"])
        _nf(ws, r, 10, row["dif_imp"])
        ws.cell(row=r, column=11, value=row["estado"]).fill = _fl(co)
        ver_cell = ws.cell(row=r, column=12, value=row["verificar"])
        if row["verificar"] == "FALTA BOLETO":
            ver_cell.fill = _fl(C_RED)
            ver_cell.font = Font(bold=True, color="FFFFFF")
        elif row["verificar"] == "CTA 1002":
            ver_cell.fill = _fl(C_ORANGE); ver_cell.font = _bf()
        elif row["verificar"] == "VERIFICAR":
            ver_cell.fill = _fl(C_YELLOW)
        for col in range(1, 13):
            ws.cell(row=r, column=col).border = _border()
        r += 1

    ws.auto_filter.ref = f"A3:L{r-1}"
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    return ws


# ── Hojas detalle ──────────────────────────────────────────────────────────────

def _write_ops_sheet(ws, rows):
    _wh(ws, ["CTTE", "ESPECIE", "MONEDA", "SENTIDO", "CONCEPTO", "SEGMENTO",
             "VN BYMA", "VN GALLO", "DIF VN",
             "IMP BYMA", "IMP GALLO", "DIF IMP", "ESTADO", "VERIFICAR"])
    for i, r in enumerate(rows, 2):
        co = _st_color(r["estado"])
        ws.cell(i,  1, r["ctte"])
        ws.cell(i,  2, r["especie"])
        ws.cell(i,  3, r["moneda"])
        ws.cell(i,  4, r["sentido"])
        ws.cell(i,  5, r["concepto"])
        ws.cell(i,  6, r["segmento_mer"])
        _nf(ws, i,  7, r["vn_b"])
        _nf(ws, i,  8, r["vn_g"])
        _nf(ws, i,  9, r["dif_vn"])
        _nf(ws, i, 10, r["imp_b"])
        _nf(ws, i, 11, r["imp_g"])
        _nf(ws, i, 12, r["dif_imp"])
        ws.cell(i, 13, r["estado"])
        ws.cell(i, 14, r["verificar"])
        _rf(ws, i, 14, co)
    ws.auto_filter.ref = f"A1:N{len(rows)+1}"
    _af(ws)


def _write_cau_sheet(ws, rows):
    _wh(ws, ["CTTE", "SENTIDO", "MONEDA",
             "VN BYMA (capital)", "CAPITAL GALLO", "DIF CAPITAL",
             "IMP BYMA (total)", "TOTAL GALLO", "DIF TOTAL", "ESTADO"])
    for i, r in enumerate(rows, 2):
        co = _st_color(r["estado"])
        ws.cell(i, 1, r["ctte"])
        ws.cell(i, 2, r["sentido"])
        ws.cell(i, 3, r["moneda"])
        _nf(ws, i, 4, r["vn_b"])
        _nf(ws, i, 5, r["cap_g"])
        _nf(ws, i, 6, r["dif_cap"])
        _nf(ws, i, 7, r["imp_b"])
        _nf(ws, i, 8, r["tot_g"])
        _nf(ws, i, 9, r["dif_tot"])
        ws.cell(i, 10, r["estado"])
        _rf(ws, i, 10, co)
    ws.auto_filter.ref = f"A1:J{len(rows)+1}"
    _af(ws)


def _write_tra_sheet(ws, rows):
    _wh(ws, ["CTTE", "ESPECIE", "MONEDA", "SENTIDO", "CONCEPTO",
             "VN BYMA", "VN GALLO", "DIF VN",
             "IMP BYMA", "IMP GALLO", "DIF IMP", "ESTADO"])
    for i, r in enumerate(rows, 2):
        co = _st_color(r["estado"])
        ws.cell(i, 1, r["ctte"])
        ws.cell(i, 2, r["especie"])
        ws.cell(i, 3, r["moneda"])
        ws.cell(i, 4, r["sentido"])
        ws.cell(i, 5, r.get("concepto", ""))
        _nf(ws, i, 6, r["vn_b"])
        _nf(ws, i, 7, r["vn_g"])
        _nf(ws, i, 8, r["dif_vn"])
        _nf(ws, i, 9, r["imp_b"])
        _nf(ws, i, 10, r["imp_g"])
        _nf(ws, i, 11, r["dif_imp"])
        ws.cell(i, 12, r["estado"])
        _rf(ws, i, 12, co)
    ws.auto_filter.ref = f"A1:L{len(rows)+1}"
    _af(ws)


def _write_mav_sheet(ws, mav_rows):
    _wh(ws, ["CTTE", "ESPECIE", "SENTIDO", "MONEDA",
             "VN GALLO", "IMP GALLO", "OP GALLO"])
    for i, r in enumerate(mav_rows, 2):
        ws.cell(i, 1, r["ctte"])
        ws.cell(i, 2, r["esp"])
        ws.cell(i, 3, r["sen"])
        ws.cell(i, 4, r["mon"])
        _nf(ws, i, 5, r["vn"])
        _nf(ws, i, 6, r["imp"])
        ws.cell(i, 7, r["op"])
        _rf(ws, i, 7, C_BLUE)
    ws.auto_filter.ref = f"A1:G{len(mav_rows)+1}"
    _af(ws)


# ── Hoja Resumen Ope-Titulos ───────────────────────────────────────────────────

def _write_res(ws, ops_rows, cr, tr, mav_rows, analisis_1002, arancel_alerts, process_date,
               n_senebi_nogara=0, n_senebi_gara_ext=0):
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    for col in "CDEFGHIJK":
        ws.column_dimensions[col].width = 16

    row = 1
    ws.cell(row, 1, "RESUMEN CONCILIACION").font = _bf()
    row += 1
    ws.cell(row, 1, f"Fecha: {process_date.strftime('%d/%m/%Y')}")
    row += 2

    est_std = [("OK", C_GREEN), ("DIFERENCIA", C_RED),
               ("SOLO BYMA", C_YELLOW), ("SOLO GALLO", C_BLUE),
               ("TRADING INTRADAY", C_ORANGE)]

    def bloque(titulo, rows, estados):
        nonlocal row
        ws.cell(row, 1, titulo).fill = _fl(C_HEADER)
        ws.cell(row, 1).font = _hf()
        ws.cell(row, 2, "Cant").fill = _fl(C_HEADER)
        ws.cell(row, 2).font = _hf()
        row += 1
        cnt = Counter(x["estado"] for x in rows)
        shown = False
        for st, co in estados:
            n = cnt.get(st, 0)
            if n == 0:
                continue
            ws.cell(row, 1, st).fill = _fl(co)
            ws.cell(row, 2, n).fill = _fl(co)
            row += 1
            shown = True
        if not shown:
            ws.cell(row, 1, "(sin datos)")
            row += 1
        row += 1

    bloque("OPERACIONES (PPT BYMA + SENEBI)", ops_rows, est_std)
    bloque("CAUCIONES",                       cr,       est_std)
    bloque("TRADING INTRADAY",                tr,       est_std)

    n_falta  = sum(1 for r in ops_rows if r["verificar"] == "FALTA BOLETO")
    co_falta = C_RED if n_falta > 0 else C_GREEN
    label_falta = (f"FALTA BOLETO EN GALLO — {n_falta} op(s)" if n_falta > 0
                   else "SIN FALTANTES DE BOLETO EN GALLO")
    ws.cell(row, 1, label_falta).fill = _fl(co_falta)
    ws.cell(row, 1).font = Font(bold=True, color="FFFFFF" if n_falta > 0 else "000000")
    ws.cell(row, 2, n_falta).fill = _fl(co_falta)
    ws.cell(row, 2).font = Font(bold=True)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width,
                                          len(label_falta) + 2)
    row += 2

    ws.cell(row, 1, "MAV").fill = _fl(C_HEADER)
    ws.cell(row, 1).font = _hf()
    ws.cell(row, 2, "Cant").fill = _fl(C_HEADER)
    ws.cell(row, 2).font = _hf()
    row += 1
    ws.cell(row, 1, "Ops MAV Gallo (sin contr. BYMA)").fill = _fl(C_BLUE)
    ws.cell(row, 2, len(mav_rows)).fill = _fl(C_BLUE)
    row += 2

    # SENEBI NO GARA — informacional
    if n_senebi_nogara:
        label_sng = f"SENEBI NO GARA: {n_senebi_nogara} fila(s) back-to-back (neto=0, excluidas de Diferencias)"
        ws.cell(row, 1, label_sng).fill = _fl(C_GREEN)
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2, n_senebi_nogara).fill = _fl(C_GREEN)
        ws.cell(row, 2).font = Font(bold=True)
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width,
                                              len(label_sng) + 2)
        row += 1

    # SENEBI GARA EXTERNO — informacional
    if n_senebi_gara_ext:
        label_sge = (f"SENEBI GARA EXTERNO: {n_senebi_gara_ext} fila(s) "
                     f"(contraparte sin cuenta Gallo, neto=0, excluidas de Diferencias)")
        ws.cell(row, 1, label_sge).fill = _fl(C_GREEN)
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2, n_senebi_gara_ext).fill = _fl(C_GREEN)
        ws.cell(row, 2).font = Font(bold=True)
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width,
                                              len(label_sge) + 2)
        row += 1

    if n_senebi_nogara or n_senebi_gara_ext:
        row += 1

    n_ar   = len(arancel_alerts) if arancel_alerts else 0
    co_ar  = C_RED if n_ar > 0 else C_GREEN
    label_ar = (f"*** ARANCEL EN CUENTA CARTERA PROPIA — {n_ar} boleto(s) ***"
                if n_ar > 0 else "SIN ARANCEL EN CUENTAS CARTERA PROPIA (1000-1003)")
    ws.cell(row, 1, label_ar).fill = _fl(co_ar)
    ws.cell(row, 1).font = Font(bold=True, color="FFFFFF" if n_ar > 0 else "000000")
    ws.cell(row, 2, n_ar).fill = _fl(co_ar)
    ws.cell(row, 2).font = Font(bold=True)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width,
                                          len(label_ar) + 2)
    row += 1
    if arancel_alerts:
        for c, h in enumerate(["Boleto", "CTTE", "Operacion", "Especie", "Arancel"], 1):
            cell = ws.cell(row, c, h)
            cell.font = Font(bold=True); cell.fill = _fl("D9D9D9")
        row += 1
        for a in arancel_alerts:
            ws.cell(row, 1, a["boleto"])
            ws.cell(row, 2, a["ctte"])
            ws.cell(row, 3, a["op"])
            ws.cell(row, 4, a["esp"])
            cell_ar = ws.cell(row, 5, a["arancel"])
            cell_ar.number_format = NUM_FMT
            cell_ar.fill = _fl(C_RED); cell_ar.font = Font(bold=True, color="FFFFFF")
            row += 1
    row += 2

    if analisis_1002:
        ws.cell(row, 1, "ANALISIS CTA 1002").font = _hf()
        ws.cell(row, 1).fill = _fl(C_HEADER)
        row += 1
        hdrs_1002 = ["ESPECIE", "MONEDA", "SENTIDO", "CONCEPTO", "ESTADO 1002",
                     "VN BYMA 1002", "VN GALLO 1002", "DIF VN 1002",
                     "OTROS CTTES", "DIF VN OTROS", "NETO VN", "BALANCE"]
        for c, h in enumerate(hdrs_1002, 1):
            cell = ws.cell(row, c, h)
            cell.font = _bf(); cell.fill = _fl("D9D9D9")
        row += 1
        for a in sorted(analisis_1002, key=lambda x: (x["especie"], x["concepto"])):
            co = C_GREEN if a["balance"] == "OK" else C_RED
            ws.cell(row,  1, a["especie"])
            ws.cell(row,  2, a["moneda"])
            ws.cell(row,  3, a["sentido"])
            ws.cell(row,  4, a["concepto"])
            ws.cell(row,  5, a["estado_1002"])
            _nf(ws, row,  6, a["vn_b_1002"])
            _nf(ws, row,  7, a["vn_g_1002"])
            _nf(ws, row,  8, a["dif_1002"])
            ws.cell(row,  9, a["otros_cttes"])
            _nf(ws, row, 10, a["dif_otros"])
            _nf(ws, row, 11, a["neto"])
            ws.cell(row, 12, a["balance"]).fill = _fl(co)
            row += 1


# ── Punto de entrada ──────────────────────────────────────────────────────────

def generar_reporte(dat_file, xls_file, bil_file=None, ap_file=None, especies_file=None):
    """
    dat_file      : UploadedFile — OPERSECEXT_GARA.DAT
    xls_file      : UploadedFile — CONTBOLE.XLS
    bil_file      : UploadedFile or None — OPERBILEXT_GARA.DAT (opcional)
    ap_file       : UploadedFile or None — Actual Position DD-MM-AA.xlsx (opcional, Control 999)
    especies_file : UploadedFile or None — ESPECIES.XLS (opcional, byma_map GOGLD→GOOGL etc)

    Devuelve (BytesIO con Excel, dict resumen).
    """
    dat_lines = dat_file.read().decode("latin-1").splitlines()
    xls_bytes = xls_file.read()
    bil_lines = bil_file.read().decode("latin-1").splitlines() if bil_file else []

    # Mapa BYMA→Gallo (excepciones como GOGLD→GOOGL) — opcional
    byma_map = _load_byma_ticker_map(especies_file) if especies_file else {}

    process_date = _detect_process_date(dat_lines)

    esp_g, mer_g, cau_g, tra_g, snb_g, mav_g, arancel_alerts = _parse_gallo(xls_bytes)
    mer_b, cau_b, opt_esp = _parse_byma(dat_lines, esp_g, byma_map)
    if bil_lines:
        snb_b, ext_pairs = _parse_senebi(bil_lines, esp_g, byma_map)
    else:
        snb_b, ext_pairs = {}, []

    ppt_keys_b = set(mer_b.keys())
    ppt_keys_g = set(mer_g.keys())
    snb_keys_b = set(snb_b.keys())
    snb_keys_g = set(snb_g.keys())

    combined_b = _merge_dicts(mer_b, snb_b)
    combined_g = _merge_dicts(mer_g, snb_g)

    ops_rows = _compare_ops(combined_b, combined_g,
                            ppt_keys_b, ppt_keys_g, snb_keys_b, snb_keys_g)
    ops_rows = _second_pass_ticker(ops_rows)
    # tra_g ahora tiene clave 5-tuple; reducimos a 4-tuple para el match
    tra_keys_4_main = {(c, e, m, s) for (c, e, m, s, _) in tra_g}
    for r in ops_rows:
        if r["estado"] == "SOLO BYMA":
            k = (r["ctte"], r["especie"], r["moneda"], r["sentido"])
            if k in tra_keys_4_main:
                r["estado"]    = "TRADING INTRADAY"
                r["verificar"] = ""

    cr    = _compare_cau(cau_b, cau_g)
    tr    = _compare_tra(tra_g)
    a1002 = _analyze_1002(ops_rows)
    ops_rows = _apply_1002_compensation(ops_rows, a1002)
    ops_rows, n_senebi_gara_ext = _apply_senebi_gara_externo_compensation(ops_rows, ext_pairs)
    ops_rows, n_senebi_nogara   = _apply_senebi_nogara_compensation(ops_rows, snb_keys_g)

    agg = _build_currency_agg(mer_b, snb_b, cau_b, mer_g, snb_g, cau_g, tra_g, mav_g)
    ver_rows = _build_verification_rows(ops_rows, cr)
    ap_saldos, ap_extras, ap_filename = _cargar_saldos_actual_position_file(ap_file)

    # ── Excel ─────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws_res = wb.active
    ws_res.title = "Resumen Ope-Titulos"

    _write_control_999(wb, agg, ap_saldos, ap_extras, ap_filename, process_date)
    for code in ("ARS", "USD_MEP", "USD_CABLE"):
        _write_currency_sheet(wb, code, agg[code], process_date)
    _write_verification_sheet(wb, ver_rows, process_date)

    ws_ops = wb.create_sheet("Operaciones")
    ws_cau = wb.create_sheet("Cauciones")
    ws_tra = wb.create_sheet("Trading Intraday")
    ws_mav = wb.create_sheet("MAV")
    _write_res(ws_res, ops_rows, cr, tr, mav_g, a1002, arancel_alerts, process_date,
               n_senebi_nogara, n_senebi_gara_ext)
    _write_ops_sheet(ws_ops, ops_rows)
    _write_cau_sheet(ws_cau, cr)
    _write_tra_sheet(ws_tra, tr)
    _write_mav_sheet(ws_mav, mav_g)

    orden = [
        "Control 999", "Pesos ARS", "Dolar MEP", "USD Cable",
        "Resumen Ope-Titulos", "Diferencias a Verificar",
        "Operaciones", "Cauciones", "Trading Intraday", "MAV",
    ]
    wb._sheets = [wb[name] for name in orden if name in wb.sheetnames]

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # ── Resumen para la UI ────────────────────────────────────────────────────
    cnt_ops = Counter(r["estado"] for r in ops_rows)
    cnt_cau = Counter(r["estado"] for r in cr)
    # Diferencia neta por moneda (Neto BYMA vs Neto Gallo, suma de todos los conceptos)
    moneda_estado = {}
    for code in ("ARS", "USD_MEP", "USD_CABLE"):
        ar_b = sum(d["ar_b"] for d in agg[code].values())
        ae_b = sum(d["ae_b"] for d in agg[code].values())
        ar_g = sum(d["ar_g"] for d in agg[code].values())
        ae_g = sum(d["ae_g"] for d in agg[code].values())
        neto_b = ar_b - ae_b
        neto_g = ar_g - ae_g
        dif = neto_b - neto_g
        moneda_estado[code] = {
            "label":  MON_LABEL[code],
            "neto_b": neto_b,
            "neto_g": neto_g,
            "dif":    dif,
            "estado": "OK" if abs(dif) <= TOL else "DIFERENCIA",
        }
    # Saldo final proyectado por moneda (si hay AP) — usa SOLO lo que liquida hoy:
    # CI (G + NG) + TI plazo CI + CAU. CN y TI plazo T1 son informativos (T+1).
    saldos_finales = {}
    if ap_saldos:
        for code in ("ARS", "USD_MEP", "USD_CABLE"):
            si = ap_saldos.get(code)
            if si is None:
                continue
            ag = agg[code]
            def _neto_g_b(buckets):
                return sum(ag.get(bk, {}).get("ar_g", 0.) - ag.get(bk, {}).get("ae_g", 0.)
                           for bk in buckets)
            ci_g    = _neto_g_b([("CI", "G"), ("CI", "NG")])
            opc_g   = _neto_g_b([("OPC", "G")])
            ti_ci_g = _neto_g_b([("TI_CI", "G")])
            cau_g   = _neto_g_b([("CAU", "G")])
            total_hoy_g = ci_g + opc_g + ti_ci_g + cau_g
            saldos_finales[code] = {
                "label":       MON_LABEL[code],
                "saldo_ini":   si,
                "saldo_final": si + total_hoy_g,
            }

    resumen = {
        "fecha":               process_date.strftime("%d-%m-%Y"),
        "process_date":        process_date.strftime("%d/%m/%Y"),
        "n_ops_ok":            cnt_ops.get("OK", 0),
        "n_ops_dif":           cnt_ops.get("DIFERENCIA", 0),
        "n_solo_byma":         cnt_ops.get("SOLO BYMA", 0),
        "n_solo_gallo":        cnt_ops.get("SOLO GALLO", 0),
        "n_ti":                cnt_ops.get("TRADING INTRADAY", 0),
        "n_cau_ok":            cnt_cau.get("OK", 0),
        "n_cau_dif":           cnt_cau.get("DIFERENCIA", 0),
        "n_mav":               len(mav_g),
        "n_falta_boleto":      sum(1 for r in ops_rows if r["verificar"] == "FALTA BOLETO"),
        "n_arancel":           len(arancel_alerts),
        "arancel_alerts":      arancel_alerts,
        "falta_boleto_detail": [r for r in ops_rows if r["verificar"] == "FALTA BOLETO"],
        "con_senebi":          bool(bil_lines),
        "moneda_estado":       moneda_estado,
        "saldos_finales":      saldos_finales,
        "ap_filename":         ap_filename,
        "n_ver_rows":          len(ver_rows),
        "n_senebi_nogara":     n_senebi_nogara,
        "n_senebi_gara_ext":   n_senebi_gara_ext,
        "n_opt_esp":           len(opt_esp),
    }
    return output, resumen
