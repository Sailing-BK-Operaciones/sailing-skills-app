"""
Tesorería — Reporte de Gestión BackOffice
Procesa archivos MOVICTA (XLS de Gallo) y genera el reporte Excel mensual.
"""

import io
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    from PIL import Image as PILImage
    _HAS_PIL = True
except ImportError:
    _HAS_PIL = False

# ─── Constantes ────────────────────────────────────────────────────────────────
COLORS = {
    'azul_oscuro': '1A3A4A',
    'azul_prim':   '0E7FAD',
    'verde':       '27AE60',
    'rojo':        'E74C3C',
    'naranja':     'F39C12',
    'blanco':      'FFFFFF',
    'gris_fila':   'F2F2F2',
    'violeta':     '6E2F8A',
    'azul_tc':     '1A5276',
    'negro':       '000000',
}

MERCADOS = {
    999001: 'Mercado Argentino de Valores (MAV)',
    999002: 'A3 Mercados S.A. Garantía',
    999007: 'U34 Interactive Brokers LLC',
    999012: 'A3 Mercados S.A.',
    999051: 'Banco Comafi',
    999999: 'BYMA',
}

HB_REFS    = {'TRANSF.RECIBIDA', 'Solicitado x WEB', 'TRANSFERENCIA RECIBID', 'DOLAR MEP'}
ECHEQ_REFS = {'ECHEQ A CTTE'}
# Comparación case-insensitive: en Gallo aparecen variantes de casing
# (ej. 'Solicitado x Web' vs 'Solicitado x WEB') para la misma referencia.
HB_REFS_U    = {r.upper() for r in HB_REFS}
ECHEQ_REFS_U = {r.upper() for r in ECHEQ_REFS}

MESES_ORDER = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]

TC_HISTORICOS = {
    'Enero':    {'MEP': 1464.75, 'CCL': 1510.53},
    'Febrero':  {'MEP': 1436.64, 'CCL': 1476.11},
    'Marzo':    {'MEP': 1488.00, 'CCL': 1499.00},
    'Abril':    {'MEP': 1448.93, 'CCL': 1503.03},
}

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'assets', 'sailing_logo.jpg')


# ─── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, italic=False, color='000000', size=11, name='Calibri'):
    return Font(bold=bold, italic=italic, color=color, size=size, name=name)

def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    s = Side(style='thin', color='D0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def _merge_set(ws, cell_range, value='', bg=None, fnt=None, aln=None):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(':')[0]]
    if value != '':
        c.value = value
    if bg:
        c.fill = _fill(bg)
    if fnt:
        c.font = fnt
    if aln:
        c.alignment = aln
    return c

def _add_logo(ws, anchor='J1'):
    if not os.path.exists(LOGO_PATH):
        return
    try:
        if _HAS_PIL:
            img_pil = PILImage.open(LOGO_PATH)
            orig_w, orig_h = img_pil.size
            target_h = 90
            target_w = int(target_h * orig_w / orig_h)
            img_pil = img_pil.resize((target_w, target_h), PILImage.LANCZOS)
            buf = io.BytesIO()
            img_pil.save(buf, format='PNG')
            buf.seek(0)
            img = XLImage(buf)
        else:
            img = XLImage(LOGO_PATH)
            img.width, img.height = 138, 90
        img.anchor = anchor
        ws.add_image(img)
    except Exception:
        pass  # logo opcional

def _write_header(ws, sheet_title='', last_col='N'):
    ws.sheet_view.showGridLines = False
    for r, h in [(1,22),(2,16),(3,13),(4,8),(5,8),(6,6),(7,20)]:
        ws.row_dimensions[r].height = h

    ws['A1'].value     = 'SAILING INVERSIONES'
    ws['A1'].font      = _font(bold=True, color=COLORS['azul_oscuro'], size=14)
    ws['A1'].alignment = _align('left')

    ws['A2'].value     = 'Panel de Control - Tesorería BackOffice'
    ws['A2'].font      = _font(color=COLORS['azul_prim'], size=10)
    ws['A2'].alignment = _align('left')

    ws['A3'].value = (
        f"Consolidado AR$ + U$D MEP (TC MEP) + U$D CABLE (TC CCL) | "
        f"Generado: {datetime.now().strftime('%d/%m/%Y')}"
    )
    ws['A3'].font      = _font(size=8, color='777777')
    ws['A3'].alignment = _align('left')

    if sheet_title:
        try:
            _merge_set(ws, f'A7:{last_col}7', sheet_title.upper(),
                       bg=COLORS['azul_prim'],
                       fnt=_font(bold=True, color=COLORS['blanco'], size=10),
                       aln=_align('left'))
        except Exception:
            ws['A7'].value = sheet_title.upper()
            ws['A7'].fill  = _fill(COLORS['azul_prim'])
            ws['A7'].font  = _font(bold=True, color=COLORS['blanco'], size=10)

    _add_logo(ws)
    ws.freeze_panes = 'A8'


# ─── Carga y clasificación ─────────────────────────────────────────────────────
def _load_movicta(file_obj, usd_type=None):
    """Lee un archivo MOVICTA XLS (hoja Movimientos_Cuenta). Devuelve DataFrame limpio."""
    file_obj.seek(0)
    df = pd.read_excel(
        file_obj,
        sheet_name='Movimientos_Cuenta',
        header=None,
        engine='xlrd',
    )
    df.columns = ['Clte', 'Nombre', 'Especie', 'Cpbte', 'Nro', 'Fecha', 'Referencia', 'Importe']
    df = df.iloc[1:].copy()   # primera fila = encabezados de Gallo
    df = df[df['Referencia'].astype(str).str.strip() != 'Total']
    df['Clte']   = df['Clte'].ffill()
    df['Nombre'] = df['Nombre'].ffill()
    df = df.dropna(subset=['Cpbte', 'Importe'])
    df['Clte']       = pd.to_numeric(df['Clte'],    errors='coerce')
    df['Importe']    = pd.to_numeric(df['Importe'], errors='coerce')
    df['Importe_abs'] = df['Importe'].abs()
    df = df.dropna(subset=['Clte'])
    if usd_type:
        df['TipoUSD'] = usd_type
    return df


def _classify(cpbte, ref, ars_file=True):
    cpbte = str(cpbte).strip()
    ref   = str(ref).strip()
    ref_u = ref.upper()

    if ars_file:
        tipo = 'Ingreso' if cpbte == '066-COBR' else ('Egreso' if cpbte == '067-PAGO' else None)
    else:
        tipo = 'Ingreso' if cpbte == '050-CU$S' else ('Egreso' if cpbte == '051-PAU$' else None)

    if 'ANULADO' in ref_u or ref_u.startswith('ANULA'):
        canal = 'Anulacion'
    elif ref_u in ECHEQ_REFS_U:
        canal = 'eCheq'
    elif ref_u in HB_REFS_U:
        canal = 'Digital'
    else:
        canal = 'Manual'

    return tipo, canal


# ─── Procesar un mes ───────────────────────────────────────────────────────────
def procesar_mes(ars_file, mep_file, tc_mep, cable_file=None, tc_ccl=None):
    """
    Procesa los archivos MOVICTA de un mes.
    Retorna (DataFrame, advertencias).
    """
    advertencias = []
    rows = []

    def _add_rows(df, ars_file_flag, moneda, tipo_usd, tc):
        for _, row in df.iterrows():
            tipo, canal = _classify(row['Cpbte'], row['Referencia'], ars_file=ars_file_flag)
            if tipo is None:
                continue
            rows.append({
                'Clte':       int(row['Clte']),
                'Nombre':     row['Nombre'],
                'Cpbte':      row['Cpbte'],
                'Fecha':      row['Fecha'],
                'Referencia': row['Referencia'],
                'Importe_abs': row['Importe_abs'],
                'Tipo':       tipo,
                'Canal':      canal,
                'Moneda':     moneda,
                'TipoUSD':    tipo_usd,
                'TC':         tc,
                'Importe_ARS': row['Importe_abs'] * tc,
                'EsMercado':  int(row['Clte']) >= 999001,
            })

    _add_rows(_load_movicta(ars_file),                 True,  'ARS', None,    1.0)
    _add_rows(_load_movicta(mep_file, 'MEP'),          False, 'USD', 'MEP',  tc_mep)
    if cable_file is not None and tc_ccl:
        _add_rows(_load_movicta(cable_file, 'CABLE'),  False, 'USD', 'CABLE', tc_ccl)

    df_all = pd.DataFrame(rows)

    if not df_all.empty:
        mercado_cltes = df_all[df_all['EsMercado']]['Clte'].unique()
        for c in mercado_cltes:
            if int(c) not in MERCADOS:
                nombre = df_all[df_all['Clte'] == c]['Nombre'].iloc[0]
                advertencias.append(
                    f"Mercado no registrado: comitente {int(c)} — '{nombre}' "
                    "(procesado como Mercado con nombre de Gallo)."
                )

    return df_all, advertencias


# ─── Métricas de panel ─────────────────────────────────────────────────────────
def _calc_panel_row(df, entity='Cliente'):
    """
    entity: 'Cliente' | 'Mercado'
    %Manual incluye eCheq (comportamiento del Panel de Control).
    """
    mask = df['EsMercado'] == (entity == 'Mercado')
    d    = df[mask]
    result = {}
    for tipo in ['Ingreso', 'Egreso']:
        valid = d[(d['Canal'] != 'Anulacion') & (d['Tipo'] == tipo)]
        anul  = d[(d['Canal'] == 'Anulacion')  & (d['Tipo'] == tipo)]
        cant  = len(valid)
        monto = valid['Importe_ARS'].sum()
        dig   = valid[valid['Canal'] == 'Digital']
        man   = valid[valid['Canal'].isin(['Manual', 'eCheq'])]
        result[tipo] = {
            'cant':    cant,
            'monto':   monto,
            'pct_hb':  (len(dig) / cant * 100) if cant > 0 else 0,
            'pct_man': (len(man) / cant * 100) if cant > 0 else 0,
            'errores': len(anul),
        }
    neto = result['Ingreso']['monto'] - result['Egreso']['monto']
    return result, neto


# ─── Hoja: Panel de Control ────────────────────────────────────────────────────
def _write_panel(wb, months, month_data, tc_data):
    """
    months    : list of month names (ordered)
    month_data: {mes: DataFrame}
    tc_data   : {mes: {'MEP': float, 'CCL': float}}
    """
    ws = wb.create_sheet('Panel de Control')
    ws.sheet_properties.tabColor = COLORS['azul_oscuro']
    for col, w in {'A':14,'B':11,'C':18,'D':10,'E':12,
                   'F':11,'G':18,'H':10,'I':12,'J':14,
                   'K':10,'L':10,'M':14,'N':14}.items():
        ws.column_dimensions[col].width = w

    _write_header(ws, 'Panel de Control — Tesorería BackOffice', last_col='N')

    data_rows_by_mes = {}

    for section_idx, entity in enumerate(['Clientes', 'Mercados']):
        base = 9 + section_idx * (len(months) + 6)

        _merge_set(ws, f'A{base}:L{base}', f'▌  {entity.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=11),
                   aln=_align('left'))
        ws.row_dimensions[base].height = 20

        h1 = base + 1
        ws.row_dimensions[h1].height = 16
        _merge_set(ws, f'B{h1}:E{h1}', 'INGRESOS', bg=COLORS['verde'],
                   fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        _merge_set(ws, f'F{h1}:I{h1}', 'EGRESOS',  bg=COLORS['rojo'],
                   fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())

        h2 = base + 2
        ws.row_dimensions[h2].height = 30
        hdrs = ['Mes','Cant.','Monto AR$','% HB','% Manual',
                'Cant.','Monto AR$','% HB','% Manual','Neto AR$','Err.Ing','Err.Egr']
        for ci, h in enumerate(hdrs):
            c = ws[f'{get_column_letter(ci+1)}{h2}']
            c.value = h; c.fill = _fill(COLORS['azul_prim'])
            c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
            c.alignment = _align(wrap=True); c.border = _border()

        for ri, mes in enumerate(months):
            row = h2 + 1 + ri
            ws.row_dimensions[row].height = 18
            if entity == 'Clientes':
                data_rows_by_mes[mes] = row

            df_mes        = month_data[mes]
            stats, neto   = _calc_panel_row(df_mes, entity.rstrip('s'))
            bg = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']

            row_vals = [mes,
                        stats['Ingreso']['cant'],  stats['Ingreso']['monto'],
                        stats['Ingreso']['pct_hb'] / 100, stats['Ingreso']['pct_man'] / 100,
                        stats['Egreso']['cant'],   stats['Egreso']['monto'],
                        stats['Egreso']['pct_hb'] / 100,  stats['Egreso']['pct_man'] / 100,
                        neto,
                        stats['Ingreso']['errores'], stats['Egreso']['errores']]
            pct_cols   = {3, 4, 7, 8}
            money_cols = {2, 6, 9}

            for ci, val in enumerate(row_vals):
                c = ws[f'{get_column_letter(ci+1)}{row}']
                c.value = val; c.fill = _fill(bg)
                c.border = _border(); c.alignment = _align()
                if ci in pct_cols:
                    c.number_format = '0.0%'; c.font = _font(size=9)
                elif ci in money_cols:
                    neto_color = (COLORS['verde'] if val >= 0 else COLORS['rojo']) if ci == 9 else COLORS['negro']
                    c.number_format = '#,##0'
                    c.font = _font(bold=(ci == 9), size=9, color=neto_color)
                elif ci == 0:
                    c.font = _font(bold=True, size=9)
                else:
                    c.font = _font(size=9)

        # Fila TOTAL ACUMULADO
        total_row = h2 + 1 + len(months)
        ws.row_dimensions[total_row].height = 20
        totals = ['TOTAL ACUMULADO', 0, 0, None, None, 0, 0, None, None, 0, 0, 0]
        for mes in months:
            stats, neto = _calc_panel_row(month_data[mes], entity.rstrip('s'))
            totals[1] += stats['Ingreso']['cant'];  totals[2]  += stats['Ingreso']['monto']
            totals[5] += stats['Egreso']['cant'];   totals[6]  += stats['Egreso']['monto']
            totals[9] += neto
            totals[10] += stats['Ingreso']['errores']; totals[11] += stats['Egreso']['errores']

        for ci, val in enumerate(totals):
            c = ws[f'{get_column_letter(ci+1)}{total_row}']
            c.fill = _fill(COLORS['azul_oscuro'])
            c.font = _font(bold=True, color=COLORS['blanco'], size=9)
            c.border = _border(); c.alignment = _align()
            c.value = '—' if val is None else val
            if isinstance(val, (int, float)) and ci in {2, 6, 9}:
                c.number_format = '#,##0'

    # Tabla de TC (columnas M-N, alineada con filas de datos Clientes)
    first_data = data_rows_by_mes[months[0]]
    tc_title_row = first_data - 2
    _merge_set(ws, f'M{tc_title_row}:N{tc_title_row}', 'TIPOS DE CAMBIO DE REFERENCIA',
               bg=COLORS['azul_oscuro'],
               fnt=_font(bold=True, color=COLORS['blanco'], size=9), aln=_align())

    tc_hdr_row = first_data - 1
    for col, lbl in [('M', 'TC MEP'), ('N', 'TC CCL')]:
        c = ws[f'{col}{tc_hdr_row}']
        c.value = lbl; c.fill = _fill(COLORS['azul_prim'])
        c.font = _font(bold=True, color=COLORS['blanco'], size=9)
        c.alignment = _align(); c.border = _border()

    for mes in months:
        row    = data_rows_by_mes[mes]
        tc_mes = tc_data[mes]
        tc_m   = ws[f'M{row}']
        tc_c   = ws[f'N{row}']
        tc_m.value = tc_mes['MEP']
        tc_c.value = tc_mes.get('CCL') or 0
        tc_m.number_format = '"$"#,##0.00'
        tc_c.number_format = '"$"#,##0.00'
        tc_m.font = _font(color=COLORS['azul_tc'])
        tc_c.font = _font(color=COLORS['violeta'])
        for c in [tc_m, tc_c]:
            c.alignment = _align(); c.border = _border()


# ─── Hoja: Análisis de Mercados ────────────────────────────────────────────────
def _write_mercados(wb, months, month_data):
    ws = wb.create_sheet('Análisis de Mercados')
    ws.sheet_properties.tabColor = COLORS['azul_prim']
    for col, w in {'A':28,'B':10,'C':16,'D':8,'F':28,'G':10,'H':16,'I':8}.items():
        ws.column_dimensions[col].width = w
    _write_header(ws, 'Análisis de Mercados', last_col='I')

    current_row = 9
    for mes in months:
        df_m = month_data[mes][month_data[mes]['EsMercado']]

        _merge_set(ws, f'A{current_row}:I{current_row}', f'▌  {mes.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10), aln=_align('left'))
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for col_offset, tipo in [(0, 'Ingreso'), (5, 'Egreso')]:
            bg_c = COLORS['verde'] if tipo == 'Ingreso' else COLORS['rojo']
            t1   = get_column_letter(1 + col_offset)
            t2   = get_column_letter(4 + col_offset)
            _merge_set(ws, f'{t1}{current_row}:{t2}{current_row}',
                       f'{tipo}s por Mercado', bg=bg_c,
                       fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        for col_offset in [0, 5]:
            for ci, h in enumerate(['Mercado', 'Cant.', 'Monto AR$', '%']):
                c = ws[f'{get_column_letter(1+col_offset+ci)}{current_row}']
                c.value = h; c.fill = _fill(COLORS['azul_prim'])
                c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
                c.alignment = _align(); c.border = _border()
        ws.row_dimensions[current_row].height = 15
        current_row += 1

        data_start  = current_row
        max_rows    = 0
        for tipo, col_offset in [('Ingreso', 0), ('Egreso', 5)]:
            sub         = df_m[df_m['Tipo'] == tipo]
            total_monto = sub['Importe_ARS'].sum()
            grouped = (sub.groupby('Clte')
                       .agg(Nombre=('Nombre','first'), Cant=('Importe_ARS','count'), Monto=('Importe_ARS','sum'))
                       .reset_index())
            grouped['NombreMerc'] = grouped['Clte'].apply(
                lambda x: MERCADOS.get(int(x), f'{int(x)} (desconocido)'))
            for ri, (_, r) in enumerate(grouped.iterrows()):
                row = data_start + ri
                bg  = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
                pct = r['Monto'] / total_monto if total_monto > 0 else 0
                for ci, val in enumerate([r['NombreMerc'], r['Cant'], r['Monto'], pct]):
                    c = ws[f'{get_column_letter(1+col_offset+ci)}{row}']
                    c.value = val; c.fill = _fill(bg)
                    c.font  = _font(size=9); c.alignment = _align(); c.border = _border()
                    if ci == 2: c.number_format = '#,##0'
                    elif ci == 3: c.number_format = '0.0%'
            max_rows = max(max_rows, len(grouped))

        current_row = data_start + max(max_rows, 1) + 2


# ─── Hoja: Análisis de Clientes ────────────────────────────────────────────────
def _write_clientes(wb, months, month_data):
    ws = wb.create_sheet('Análisis de Clientes')
    ws.sheet_properties.tabColor = '2BB5E0'
    for col, w in {'A':14,'B':10,'C':16,'D':12,'E':9,
                   'G':14,'H':10,'I':16,'J':12,'K':9,'M':14,'N':16}.items():
        ws.column_dimensions[col].width = w
    _write_header(ws, 'Análisis de Clientes — Canal por Tipo', last_col='N')

    current_row = 9
    for mes in months:
        df_cli = month_data[mes][(month_data[mes]['EsMercado'] == False) &
                                  (month_data[mes]['Canal'] != 'Anulacion')]

        _merge_set(ws, f'A{current_row}:N{current_row}', f'▌  {mes.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10), aln=_align('left'))
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        _merge_set(ws, f'A{current_row}:E{current_row}', 'INGRESOS',
                   bg=COLORS['verde'], fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        _merge_set(ws, f'G{current_row}:K{current_row}', 'EGRESOS',
                   bg=COLORS['rojo'],  fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        _merge_set(ws, f'M{current_row}:N{current_row}', 'NETO',
                   bg=COLORS['azul_oscuro'], fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        for col_offset in [0, 6]:
            for ci, h in enumerate(['Canal', 'Cant.', 'Monto AR$', 'Ticket Prom.', '%']):
                c = ws[f'{get_column_letter(1+col_offset+ci)}{current_row}']
                c.value = h; c.fill = _fill(COLORS['azul_prim'])
                c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
                c.alignment = _align(); c.border = _border()
        for col, lbl in [('M', 'Neto Cant.'), ('N', 'Neto Monto AR$')]:
            c = ws[f'{col}{current_row}']
            c.value = lbl; c.fill = _fill(COLORS['azul_prim'])
            c.font = _font(bold=True, color=COLORS['blanco'], size=9)
            c.alignment = _align(wrap=True); c.border = _border()
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Determinar canales presentes (Digital siempre, eCheq solo si hay datos)
        canales = ['Digital', 'Manual']
        if (df_cli['Canal'] == 'eCheq').any():
            canales.append('eCheq')

        data_start     = current_row
        canal_stats    = {}
        total_ing_cant = total_ing_monto = total_egr_cant = total_egr_monto = 0

        for canal in canales:
            canal_stats[canal] = {}
            for tipo in ['Ingreso', 'Egreso']:
                sub = df_cli[(df_cli['Canal'] == canal) & (df_cli['Tipo'] == tipo)]
                canal_stats[canal][tipo] = {'cant': len(sub), 'monto': sub['Importe_ARS'].sum()}
            total_ing_cant  += canal_stats[canal]['Ingreso']['cant']
            total_ing_monto += canal_stats[canal]['Ingreso']['monto']
            total_egr_cant  += canal_stats[canal]['Egreso']['cant']
            total_egr_monto += canal_stats[canal]['Egreso']['monto']

        for ri, canal in enumerate(canales):
            row = data_start + ri
            bg  = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
            ing = canal_stats[canal]['Ingreso']
            egr = canal_stats[canal]['Egreso']
            neto_cant  = ing['cant']  - egr['cant']
            neto_monto = ing['monto'] - egr['monto']
            ticket_ing = ing['monto'] / ing['cant'] if ing['cant'] > 0 else 0
            ticket_egr = egr['monto'] / egr['cant'] if egr['cant'] > 0 else 0
            pct_ing    = ing['cant']  / total_ing_cant  if total_ing_cant  > 0 else 0
            pct_egr    = egr['cant']  / total_egr_cant  if total_egr_cant  > 0 else 0

            for ci, val in enumerate([canal, ing['cant'], ing['monto'], ticket_ing, pct_ing]):
                c = ws[f'{get_column_letter(1+ci)}{row}']
                c.value = val; c.fill = _fill(bg); c.font = _font(size=9)
                c.alignment = _align(); c.border = _border()
                if ci in {2, 3}: c.number_format = '#,##0'
                elif ci == 4:    c.number_format = '0.0%'

            for ci, val in enumerate([canal, egr['cant'], egr['monto'], ticket_egr, pct_egr]):
                c = ws[f'{get_column_letter(7+ci)}{row}']
                c.value = val; c.fill = _fill(bg); c.font = _font(size=9)
                c.alignment = _align(); c.border = _border()
                if ci in {2, 3}: c.number_format = '#,##0'
                elif ci == 4:    c.number_format = '0.0%'

            neto_color = COLORS['verde'] if neto_monto >= 0 else COLORS['rojo']
            for col, val, fmt in [('M', neto_cant, '#,##0'), ('N', neto_monto, '#,##0')]:
                c = ws[f'{col}{row}']
                c.value = val; c.fill = _fill(bg)
                c.font  = _font(bold=True, color=neto_color, size=9)
                c.alignment = _align(); c.border = _border()
                c.number_format = fmt

        current_row = data_start + len(canales) + 2


# ─── Hoja: Ranking Clientes ────────────────────────────────────────────────────
def _write_ranking(wb, months, month_data):
    ws = wb.create_sheet('Ranking Clientes')
    ws.sheet_properties.tabColor = COLORS['verde']
    for col, w in {'A':6,'B':10,'C':26,'D':14,'E':10,
                   'G':6,'H':10,'I':26,'J':14,'K':10}.items():
        ws.column_dimensions[col].width = w
    _write_header(ws, 'Ranking Clientes — Ingresos y Egresos', last_col='K')

    current_row = 9
    for mes in months:
        df_cli = month_data[mes][month_data[mes]['EsMercado'] == False]

        _merge_set(ws, f'A{current_row}:K{current_row}', f'▌  {mes.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10), aln=_align('left'))
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for c_start, c_end, label, bg_c in [
            ('A','E','TOP 20 INGRESOS', COLORS['verde']),
            ('G','K','TOP 20 EGRESOS',  COLORS['rojo']),
        ]:
            _merge_set(ws, f'{c_start}{current_row}:{c_end}{current_row}', label,
                       bg=bg_c, fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        for col_offset in [0, 6]:
            for ci, h in enumerate(['Rank', 'Comitente', 'Nombre', 'Monto AR$', 'Cant. Ops']):
                c = ws[f'{get_column_letter(1+col_offset+ci)}{current_row}']
                c.value = h
                c.fill  = _fill(COLORS['azul_oscuro'] if ci == 0 else COLORS['azul_prim'])
                c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
                c.alignment = _align(); c.border = _border()
        ws.row_dimensions[current_row].height = 15
        current_row += 1

        data_start = current_row
        for tipo, col_offset in [('Ingreso', 0), ('Egreso', 6)]:
            sub = df_cli[(df_cli['Tipo'] == tipo) & (df_cli['Canal'] != 'Anulacion')]
            grouped = (sub.groupby(['Clte', 'Nombre'])
                       .agg(Monto=('Importe_ARS','sum'), CantOps=('Importe_ARS','count'))
                       .reset_index()
                       .sort_values('Monto', ascending=False)
                       .head(20))
            for ri, (_, r) in enumerate(grouped.iterrows()):
                row = data_start + ri
                bg  = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
                for ci, val in enumerate([ri+1, int(r['Clte']), r['Nombre'], r['Monto'], int(r['CantOps'])]):
                    c = ws[f'{get_column_letter(1+col_offset+ci)}{row}']
                    c.value = val; c.fill = _fill(bg)
                    c.font  = _font(bold=(ci == 0), color=COLORS['azul_oscuro'] if ci == 0 else COLORS['negro'], size=9)
                    c.alignment = _align('left' if ci == 2 else 'center')
                    c.border = _border()
                    if ci == 3: c.number_format = '#,##0'

        current_row = data_start + 21


# ─── Hoja: Detalle ARS ─────────────────────────────────────────────────────────
def _write_detalle_ars(wb, months, month_data):
    ws = wb.create_sheet('Detalle ARS')
    ws.sheet_properties.tabColor = '999999'
    for col, w in {'A':10,'B':26,'C':12,'D':10,'E':22,'F':14,'G':10}.items():
        ws.column_dimensions[col].width = w
    _write_header(ws, 'Detalle de Operaciones — ARS', last_col='G')

    headers = ['Comitente', 'Nombre', 'Comprobante', 'Fecha', 'Referencia', 'Importe', 'Canal']
    current_row = 9

    for mes in months:
        df_ars = month_data[mes][month_data[mes]['Moneda'] == 'ARS']

        _merge_set(ws, f'A{current_row}:G{current_row}', f'▌  {mes.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10), aln=_align('left'))
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for ci, h in enumerate(headers):
            c = ws[f'{get_column_letter(ci+1)}{current_row}']
            c.value = h; c.fill = _fill(COLORS['azul_prim'])
            c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
            c.alignment = _align(); c.border = _border()
        ws.row_dimensions[current_row].height = 15
        current_row += 1

        canal_colors = {'Digital': COLORS['azul_prim'], 'Manual': COLORS['negro'],
                        'eCheq': 'D4780A', 'Anulacion': COLORS['rojo']}
        for ri, (_, row) in enumerate(df_ars.iterrows()):
            bg   = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
            vals = [int(row['Clte']), row['Nombre'], row['Cpbte'],
                    row['Fecha'],    row['Referencia'], row['Importe_abs'], row['Canal']]
            for ci, val in enumerate(vals):
                c = ws[f'{get_column_letter(ci+1)}{current_row}']
                c.value = val; c.fill = _fill(bg)
                c.font  = _font(size=8, color=canal_colors.get(row['Canal'], COLORS['negro']) if ci == 6 else COLORS['negro'])
                c.alignment = _align('left' if ci in {1, 4, 6} else 'center')
                c.border = _border()
                if ci == 5: c.number_format = '#,##0'
            current_row += 1
        current_row += 1


# ─── Hoja: Detalle USD ─────────────────────────────────────────────────────────
def _write_detalle_usd(wb, months, month_data):
    ws = wb.create_sheet('Detalle USD')
    ws.sheet_properties.tabColor = '8E44AD'
    for col, w in {'A':10,'B':26,'C':12,'D':10,'E':22,'F':12,'G':10,'H':8,'I':14}.items():
        ws.column_dimensions[col].width = w
    _write_header(ws, 'Detalle de Operaciones — USD (MEP + Cable)', last_col='I')

    headers = ['Comitente','Nombre','Comprobante','Fecha','Referencia',
               'Importe USD','Canal','Tipo USD','Importe ARS']
    current_row = 9

    for mes in months:
        df_usd = month_data[mes][month_data[mes]['Moneda'] == 'USD']

        _merge_set(ws, f'A{current_row}:I{current_row}', f'▌  {mes.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10), aln=_align('left'))
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for ci, h in enumerate(headers):
            c = ws[f'{get_column_letter(ci+1)}{current_row}']
            c.value = h; c.fill = _fill(COLORS['azul_prim'])
            c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
            c.alignment = _align(); c.border = _border()
        ws.row_dimensions[current_row].height = 15
        current_row += 1

        for ri, (_, row) in enumerate(df_usd.iterrows()):
            bg        = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
            usd_color = COLORS['azul_tc'] if row.get('TipoUSD') == 'MEP' else COLORS['violeta']
            vals = [int(row['Clte']), row['Nombre'], row['Cpbte'], row['Fecha'],
                    row['Referencia'], row['Importe_abs'], row['Canal'],
                    row.get('TipoUSD', ''), row['Importe_ARS']]
            for ci, val in enumerate(vals):
                c = ws[f'{get_column_letter(ci+1)}{current_row}']
                c.value = val; c.fill = _fill(bg)
                c.font  = _font(size=8, color=usd_color if ci == 7 else COLORS['negro'])
                c.alignment = _align('left' if ci in {1, 4, 6, 7} else 'center')
                c.border = _border()
                if ci in {5, 8}: c.number_format = '#,##0'
            current_row += 1
        current_row += 1


# ─── Función principal ─────────────────────────────────────────────────────────
def generar_reporte(months_dict):
    """
    months_dict: dict ordenado {mes: {"df": DataFrame, "tc_mep": float, "tc_ccl": float|None}}
                 Las claves deben estar en orden cronológico.

    Retorna (xlsx_bytes, resumen, advertencias_acumuladas).
    """
    months     = list(months_dict.keys())
    month_data = {m: months_dict[m]['df'] for m in months}
    tc_data    = {m: {'MEP': months_dict[m]['tc_mep'],
                      'CCL': months_dict[m].get('tc_ccl', 0)}
                  for m in months}

    # Verificación
    advertencias = []
    for mes in months:
        if month_data[mes].empty:
            advertencias.append(f"Mes {mes}: sin datos procesados.")

    wb = Workbook()
    wb.remove(wb.active)

    _write_panel(wb, months, month_data, tc_data)
    _write_mercados(wb, months, month_data)
    _write_clientes(wb, months, month_data)
    _write_ranking(wb, months, month_data)
    _write_detalle_ars(wb, months, month_data)
    _write_detalle_usd(wb, months, month_data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Agregar mes a reporte existente ───────────────────────────────────────────
def agregar_mes(existing_bytes, mes_name, df_mes, tc_mep, tc_ccl):
    """
    Agrega un mes nuevo a un reporte de Tesorería existente (patrón update).
    existing_bytes : bytes del XLSX base
    mes_name       : nombre del mes a agregar (ej. 'Mayo')
    df_mes         : DataFrame del mes (output de procesar_mes)
    tc_mep, tc_ccl : tipos de cambio del mes
    Retorna bytes del XLSX actualizado.
    """
    from openpyxl import load_workbook
    from copy import copy as _copy_obj

    wb         = load_workbook(io.BytesIO(existing_bytes))
    tc_ccl_val = tc_ccl or 0
    today      = datetime.now().strftime('%d/%m/%Y')

    # Actualizar "Generado:" en todas las hojas
    for sh in wb.sheetnames:
        c = wb[sh].cell(3, 1)
        if c.value and 'Generado:' in str(c.value):
            c.value = (
                f"Consolidado AR$ + U$D MEP (TC MEP) + U$D CABLE (TC CCL) | "
                f"Generado: {today}"
            )

    # ── helpers locales ───────────────────────────────────────────────────────
    def _copy_style(ws, src_row, dst_row, ncols):
        for col in range(1, ncols + 1):
            src = ws.cell(src_row, col)
            dst = ws.cell(dst_row, col)
            if src.has_style:
                dst.font          = _copy_obj(src.font)
                dst.fill          = _copy_obj(src.fill)
                dst.border        = _copy_obj(src.border)
                dst.alignment     = _copy_obj(src.alignment)
                dst.number_format = src.number_format

    def _find_total(ws, start=9):
        for r in range(start, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v and 'TOTAL ACUMULADO' in str(v).upper():
                return r
        return None

    # ── 1. Panel de Control ───────────────────────────────────────────────────
    ws_p = wb['Panel de Control'] if 'Panel de Control' in wb.sheetnames else None
    if ws_p is not None:
        stats_c, neto_c = _calc_panel_row(df_mes, 'Cliente')
        stats_m, neto_m = _calc_panel_row(df_mes, 'Mercado')

        pct_cols   = {4, 5, 8, 9}   # 1-indexed
        money_cols = {3, 7, 10}

        def _write_panel_row(ws, row, stats, neto, include_tc):
            vals = [
                mes_name,
                stats['Ingreso']['cant'],  stats['Ingreso']['monto'],
                stats['Ingreso']['pct_hb']  / 100, stats['Ingreso']['pct_man'] / 100,
                stats['Egreso']['cant'],   stats['Egreso']['monto'],
                stats['Egreso']['pct_hb']  / 100,  stats['Egreso']['pct_man'] / 100,
                neto,
                stats['Ingreso']['errores'], stats['Egreso']['errores'],
                tc_mep     if include_tc else None,
                tc_ccl_val if include_tc else None,
            ]
            for ci, val in enumerate(vals, start=1):
                c = ws.cell(row, ci)
                c.value = val
                if val is None:
                    continue
                if ci in pct_cols:    c.number_format = '0.0%'
                elif ci in money_cols: c.number_format = '#,##0'
                elif ci in {13, 14}:  c.number_format = '"$"#,##0.00'

        def _recalc_total(ws, data_start, data_end, total_row):
            for col in range(2, 13):
                vals = [ws.cell(r, col).value
                        for r in range(data_start, data_end + 1)
                        if isinstance(ws.cell(r, col).value, (int, float))]
                if vals:
                    if col in (4, 5, 8, 9):
                        ws.cell(total_row, col).value = sum(vals) / len(vals)
                    elif col not in (13, 14):
                        ws.cell(total_row, col).value = sum(vals)

        def _insert_rows_keep_merges(ws, at_row):
            """Inserta una fila desplazando los rangos mergeados >= at_row.
            openpyxl NO reubica los merges al insertar filas, por lo que hay
            que desmergear antes y volver a mergear +1 después. Sin esto, el
            banner mergeado 'TOTAL ACUMULADO' de Clientes (A:L) queda mal
            ubicado y tapa los datos del mes nuevo."""
            saved = []
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row >= at_row:
                    saved.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))
                    ws.unmerge_cells(str(rng))
            ws.insert_rows(at_row)
            for r1, c1, r2, c2 in saved:
                ws.merge_cells(start_row=r1 + 1, start_column=c1,
                               end_row=r2 + 1, end_column=c2)

        def _is_merged_banner(ws, row):
            """True si la fila de total es un banner mergeado multi-columna
            (caso Clientes: A:L sin totales numéricos)."""
            return any(rng.min_row == rng.max_row == row and rng.min_col != rng.max_col
                       for rng in ws.merged_cells.ranges)

        def _insert_section(ws, stats, neto, include_tc, search_from=9):
            tot = _find_total(ws, start=search_from)
            if not tot:
                return None
            data_start = next(
                (r for r in range(search_from, tot)
                 if isinstance(ws.cell(r, 2).value, (int, float))),
                None
            )
            banner = _is_merged_banner(ws, tot)   # detectar antes de insertar
            _insert_rows_keep_merges(ws, tot)
            _copy_style(ws, tot - 1, tot, 14)
            _write_panel_row(ws, tot, stats, neto, include_tc)
            new_tot = tot + 1
            # Solo recalcular totales si la fila de total NO es un banner mergeado
            # (Clientes usa banner sin números; solo Mercados tiene total numérico).
            if data_start and not banner:
                _recalc_total(ws, data_start, tot, new_tot)
            return new_tot

        new_tot_c = _insert_section(ws_p, stats_c, neto_c, include_tc=True)
        if new_tot_c:
            _insert_section(ws_p, stats_m, neto_m, include_tc=False,
                            search_from=new_tot_c + 1)

    # ── 2. Análisis de Mercados ───────────────────────────────────────────────
    ws_merc = wb['Análisis de Mercados'] if 'Análisis de Mercados' in wb.sheetnames else None
    if ws_merc is not None:
        r    = ws_merc.max_row + 2
        df_m = df_mes[df_mes['EsMercado']]
        _merge_set(ws_merc, f'A{r}:I{r}', f'▌  {mes_name.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10),
                   aln=_align('left'))
        ws_merc.row_dimensions[r].height = 20;  r += 1
        for col_offset, tipo in [(0, 'Ingreso'), (5, 'Egreso')]:
            bg_c = COLORS['verde'] if tipo == 'Ingreso' else COLORS['rojo']
            t1 = get_column_letter(1 + col_offset)
            t2 = get_column_letter(4 + col_offset)
            _merge_set(ws_merc, f'{t1}{r}:{t2}{r}', f'{tipo}s por Mercado', bg=bg_c,
                       fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        r += 1
        for col_offset in [0, 5]:
            for ci, h in enumerate(['Mercado', 'Cant.', 'Monto AR$', '%']):
                c = ws_merc[f'{get_column_letter(1+col_offset+ci)}{r}']
                c.value = h;  c.fill = _fill(COLORS['azul_prim'])
                c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
                c.alignment = _align();  c.border = _border()
        r += 1
        data_start = r;  max_rows = 0
        for tipo, col_offset in [('Ingreso', 0), ('Egreso', 5)]:
            sub         = df_m[df_m['Tipo'] == tipo]
            total_monto = sub['Importe_ARS'].sum()
            grouped = (sub.groupby('Clte')
                       .agg(Nombre=('Nombre','first'), Cant=('Importe_ARS','count'),
                            Monto=('Importe_ARS','sum'))
                       .reset_index())
            grouped['NombreMerc'] = grouped['Clte'].apply(
                lambda x: MERCADOS.get(int(x), f'{int(x)} (desconocido)'))
            for ri, (_, rr) in enumerate(grouped.iterrows()):
                row = data_start + ri
                bg  = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
                pct = rr['Monto'] / total_monto if total_monto > 0 else 0
                for ci, val in enumerate([rr['NombreMerc'], rr['Cant'], rr['Monto'], pct]):
                    c = ws_merc[f'{get_column_letter(1+col_offset+ci)}{row}']
                    c.value = val;  c.fill = _fill(bg)
                    c.font  = _font(size=9);  c.alignment = _align();  c.border = _border()
                    if ci == 2: c.number_format = '#,##0'
                    elif ci == 3: c.number_format = '0.0%'
            max_rows = max(max_rows, len(grouped))

    # ── 3. Análisis de Clientes ───────────────────────────────────────────────
    ws_cli = wb['Análisis de Clientes'] if 'Análisis de Clientes' in wb.sheetnames else None
    if ws_cli is not None:
        r      = ws_cli.max_row + 2
        df_cli = df_mes[(df_mes['EsMercado'] == False) & (df_mes['Canal'] != 'Anulacion')]
        _merge_set(ws_cli, f'A{r}:N{r}', f'▌  {mes_name.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10), aln=_align('left'))
        ws_cli.row_dimensions[r].height = 20;  r += 1
        _merge_set(ws_cli, f'A{r}:E{r}', 'INGRESOS',
                   bg=COLORS['verde'], fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        _merge_set(ws_cli, f'G{r}:K{r}', 'EGRESOS',
                   bg=COLORS['rojo'],  fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        _merge_set(ws_cli, f'M{r}:N{r}', 'NETO',
                   bg=COLORS['azul_oscuro'], fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        ws_cli.row_dimensions[r].height = 16;  r += 1
        for col_offset in [0, 6]:
            for ci, h in enumerate(['Canal', 'Cant.', 'Monto AR$', 'Ticket Prom.', '%']):
                c = ws_cli[f'{get_column_letter(1+col_offset+ci)}{r}']
                c.value = h;  c.fill = _fill(COLORS['azul_prim'])
                c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
                c.alignment = _align();  c.border = _border()
        for col, lbl in [('M', 'Neto Cant.'), ('N', 'Neto Monto AR$')]:
            c = ws_cli[f'{col}{r}']
            c.value = lbl;  c.fill = _fill(COLORS['azul_prim'])
            c.font = _font(bold=True, color=COLORS['blanco'], size=9)
            c.alignment = _align(wrap=True);  c.border = _border()
        ws_cli.row_dimensions[r].height = 28;  r += 1
        canales = ['Digital', 'Manual']
        if (df_cli['Canal'] == 'eCheq').any():
            canales.append('eCheq')
        data_start = r
        canal_stats = {}
        total_ing_cant = total_ing_monto = total_egr_cant = total_egr_monto = 0
        for canal in canales:
            canal_stats[canal] = {}
            for tipo in ['Ingreso', 'Egreso']:
                sub = df_cli[(df_cli['Canal'] == canal) & (df_cli['Tipo'] == tipo)]
                canal_stats[canal][tipo] = {'cant': len(sub), 'monto': sub['Importe_ARS'].sum()}
            total_ing_cant  += canal_stats[canal]['Ingreso']['cant']
            total_ing_monto += canal_stats[canal]['Ingreso']['monto']
            total_egr_cant  += canal_stats[canal]['Egreso']['cant']
            total_egr_monto += canal_stats[canal]['Egreso']['monto']
        for ri, canal in enumerate(canales):
            row = data_start + ri
            bg  = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
            ing = canal_stats[canal]['Ingreso']
            egr = canal_stats[canal]['Egreso']
            neto_cant  = ing['cant']  - egr['cant']
            neto_monto = ing['monto'] - egr['monto']
            ticket_ing = ing['monto'] / ing['cant'] if ing['cant'] > 0 else 0
            ticket_egr = egr['monto'] / egr['cant'] if egr['cant'] > 0 else 0
            pct_ing = ing['cant'] / total_ing_cant if total_ing_cant > 0 else 0
            pct_egr = egr['cant'] / total_egr_cant if total_egr_cant > 0 else 0
            for ci, val in enumerate([canal, ing['cant'], ing['monto'], ticket_ing, pct_ing]):
                c = ws_cli[f'{get_column_letter(1+ci)}{row}']
                c.value = val;  c.fill = _fill(bg);  c.font = _font(size=9)
                c.alignment = _align();  c.border = _border()
                if ci in {2, 3}: c.number_format = '#,##0'
                elif ci == 4:    c.number_format = '0.0%'
            for ci, val in enumerate([canal, egr['cant'], egr['monto'], ticket_egr, pct_egr]):
                c = ws_cli[f'{get_column_letter(7+ci)}{row}']
                c.value = val;  c.fill = _fill(bg);  c.font = _font(size=9)
                c.alignment = _align();  c.border = _border()
                if ci in {2, 3}: c.number_format = '#,##0'
                elif ci == 4:    c.number_format = '0.0%'
            neto_color = COLORS['verde'] if neto_monto >= 0 else COLORS['rojo']
            for col, val, fmt in [('M', neto_cant, '#,##0'), ('N', neto_monto, '#,##0')]:
                c = ws_cli[f'{col}{row}']
                c.value = val;  c.fill = _fill(bg)
                c.font  = _font(bold=True, color=neto_color, size=9)
                c.alignment = _align();  c.border = _border()
                c.number_format = fmt

    # ── 4. Ranking Clientes ───────────────────────────────────────────────────
    ws_rank = wb['Ranking Clientes'] if 'Ranking Clientes' in wb.sheetnames else None
    if ws_rank is not None:
        r       = ws_rank.max_row + 2
        df_cli_r = df_mes[df_mes['EsMercado'] == False]
        _merge_set(ws_rank, f'A{r}:K{r}', f'▌  {mes_name.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10), aln=_align('left'))
        ws_rank.row_dimensions[r].height = 20;  r += 1
        for c_start, c_end, label, bg_c in [
            ('A', 'E', 'TOP 20 INGRESOS', COLORS['verde']),
            ('G', 'K', 'TOP 20 EGRESOS',  COLORS['rojo']),
        ]:
            _merge_set(ws_rank, f'{c_start}{r}:{c_end}{r}', label,
                       bg=bg_c, fnt=_font(bold=True, color=COLORS['blanco']), aln=_align())
        ws_rank.row_dimensions[r].height = 16;  r += 1
        for col_offset in [0, 6]:
            for ci, h in enumerate(['Rank', 'Comitente', 'Nombre', 'Monto AR$', 'Cant. Ops']):
                c = ws_rank[f'{get_column_letter(1+col_offset+ci)}{r}']
                c.value = h
                c.fill  = _fill(COLORS['azul_oscuro'] if ci == 0 else COLORS['azul_prim'])
                c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
                c.alignment = _align();  c.border = _border()
        ws_rank.row_dimensions[r].height = 15;  r += 1
        data_start = r
        for tipo, col_offset in [('Ingreso', 0), ('Egreso', 6)]:
            sub = df_cli_r[(df_cli_r['Tipo'] == tipo) & (df_cli_r['Canal'] != 'Anulacion')]
            grouped = (sub.groupby(['Clte', 'Nombre'])
                       .agg(Monto=('Importe_ARS', 'sum'), CantOps=('Importe_ARS', 'count'))
                       .reset_index()
                       .sort_values('Monto', ascending=False)
                       .head(20))
            for ri, (_, rr) in enumerate(grouped.iterrows()):
                row = data_start + ri
                bg  = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
                for ci, val in enumerate([ri+1, int(rr['Clte']), rr['Nombre'],
                                          rr['Monto'], int(rr['CantOps'])]):
                    c = ws_rank[f'{get_column_letter(1+col_offset+ci)}{row}']
                    c.value = val;  c.fill = _fill(bg)
                    c.font  = _font(bold=(ci == 0),
                                    color=COLORS['azul_oscuro'] if ci == 0 else COLORS['negro'],
                                    size=9)
                    c.alignment = _align('left' if ci == 2 else 'center')
                    c.border = _border()
                    if ci == 3: c.number_format = '#,##0'

    # ── 5. Detalle ARS ────────────────────────────────────────────────────────
    ws_dars = wb['Detalle ARS'] if 'Detalle ARS' in wb.sheetnames else None
    if ws_dars is not None:
        r      = ws_dars.max_row + 2
        df_ars = df_mes[df_mes['Moneda'] == 'ARS']
        _merge_set(ws_dars, f'A{r}:G{r}', f'▌  {mes_name.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10), aln=_align('left'))
        ws_dars.row_dimensions[r].height = 20;  r += 1
        headers = ['Comitente', 'Nombre', 'Comprobante', 'Fecha', 'Referencia', 'Importe', 'Canal']
        for ci, h in enumerate(headers):
            c = ws_dars[f'{get_column_letter(ci+1)}{r}']
            c.value = h;  c.fill = _fill(COLORS['azul_prim'])
            c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
            c.alignment = _align();  c.border = _border()
        ws_dars.row_dimensions[r].height = 15;  r += 1
        canal_colors = {'Digital': COLORS['azul_prim'], 'Manual': COLORS['negro'],
                        'eCheq': 'D4780A', 'Anulacion': COLORS['rojo']}
        for ri, (_, rd) in enumerate(df_ars.iterrows()):
            bg   = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
            vals = [int(rd['Clte']), rd['Nombre'], rd['Cpbte'],
                    rd['Fecha'],     rd['Referencia'], rd['Importe_abs'], rd['Canal']]
            for ci, val in enumerate(vals):
                c = ws_dars[f'{get_column_letter(ci+1)}{r}']
                c.value = val;  c.fill = _fill(bg)
                c.font  = _font(size=8,
                                color=canal_colors.get(rd['Canal'], COLORS['negro']) if ci == 6
                                else COLORS['negro'])
                c.alignment = _align('left' if ci in {1, 4, 6} else 'center')
                c.border = _border()
                if ci == 5: c.number_format = '#,##0'
            r += 1

    # ── 6. Detalle USD ────────────────────────────────────────────────────────
    ws_dusd = wb['Detalle USD'] if 'Detalle USD' in wb.sheetnames else None
    if ws_dusd is not None:
        r      = ws_dusd.max_row + 2
        df_usd = df_mes[df_mes['Moneda'] == 'USD']
        _merge_set(ws_dusd, f'A{r}:I{r}', f'▌  {mes_name.upper()}',
                   bg=COLORS['azul_oscuro'],
                   fnt=_font(bold=True, color=COLORS['blanco'], size=10), aln=_align('left'))
        ws_dusd.row_dimensions[r].height = 20;  r += 1
        headers_usd = ['Comitente', 'Nombre', 'Comprobante', 'Fecha', 'Referencia',
                       'Importe USD', 'Canal', 'Tipo USD', 'Importe ARS']
        for ci, h in enumerate(headers_usd):
            c = ws_dusd[f'{get_column_letter(ci+1)}{r}']
            c.value = h;  c.fill = _fill(COLORS['azul_prim'])
            c.font  = _font(bold=True, color=COLORS['blanco'], size=9)
            c.alignment = _align();  c.border = _border()
        ws_dusd.row_dimensions[r].height = 15;  r += 1
        for ri, (_, rd) in enumerate(df_usd.iterrows()):
            bg        = COLORS['gris_fila'] if ri % 2 == 0 else COLORS['blanco']
            usd_color = COLORS['azul_tc'] if rd.get('TipoUSD') == 'MEP' else COLORS['violeta']
            vals = [int(rd['Clte']), rd['Nombre'], rd['Cpbte'], rd['Fecha'],
                    rd['Referencia'], rd['Importe_abs'], rd['Canal'],
                    rd.get('TipoUSD', ''), rd['Importe_ARS']]
            for ci, val in enumerate(vals):
                c = ws_dusd[f'{get_column_letter(ci+1)}{r}']
                c.value = val;  c.fill = _fill(bg)
                c.font  = _font(size=8, color=usd_color if ci == 7 else COLORS['negro'])
                c.alignment = _align('left' if ci in {1, 4, 6, 7} else 'center')
                c.border = _border()
                if ci in {5, 8}: c.number_format = '#,##0'
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
