"""
Collateral Position — versión web (Streamlit).
Recibe UploadedFiles; devuelve (BytesIO xlsx, dict resumen).
"""

import re
import csv
from io import BytesIO
from collections import OrderedDict
from datetime import date

import xlrd
import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Estilos ───────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
TITLE_FILL    = PatternFill("solid", fgColor="2E75B6")
TITLE_FONT    = Font(color="FFFFFF", bold=True, size=11)
SUBTOT_FILL   = PatternFill("solid", fgColor="D6E4F0")
SUBTOT_FONT   = Font(bold=True, size=10, color="1F3864")
TOTAL_FILL    = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FONT    = Font(bold=True, size=10)
ODD_FILL      = PatternFill("solid", fgColor="FFFFFF")
EVEN_FILL     = PatternFill("solid", fgColor="F2F2F2")
YELLOW_FILL   = PatternFill("solid", fgColor="FFEB9C")
USD_FILL      = PatternFill("solid", fgColor="E2EFDA")
ESPECIE_HDR_FILL = PatternFill("solid", fgColor="9DC3E6")
ESPECIE_HDR_FONT = Font(bold=True, size=10, color="1F3864")
DIFF_OK_FILL     = PatternFill("solid", fgColor="C6EFCE")
DIFF_WARN_FILL   = PatternFill("solid", fgColor="FFEB9C")

FMT_INT   = "#,##0"
FMT_PRICE = "#,##0.00"
FMT_MONEY = "#,##0.00"

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

NODE_SHORT = {
    "TM_SAILING ALYC_Conc 233 Client (556)":                    "Client (556)",
    "TM_SAILING ALYC_Conc 233 House (557)":                     "House (557)",
    "CM_9999233 SAILING SA_9999233 SAILING SA Def Fund (1121)": "Def Fund (1121)",
}
NODE_ORDER  = ["Client (556)", "House (557)", "Def Fund (1121)"]
HOUSE_CTTES = set(range(1000, 1020))
DEFFUND_CTTES = {888888888}

COLS       = ["CVSA", "Ticker", "Nombre", "VN / Cant.", "Precio Cierre",
              "Tipo Precio", "Monto a Precio Cierre"]
COL_WIDTHS = [8, 10, 36, 14, 14, 11, 24]
COL_MONEY  = 7

CONC_COLS   = ["Nodo", "CVSA", "Ticker", "Nombre", "VN BC", "VN Gallo", "Diferencia"]
CONC_WIDTHS = [18, 8, 10, 32, 14, 14, 14]
DET_COLS    = ["CVSA", "Ticker", "Nombre", "VN BC", "Nro. Ctte", "Account ID BC",
               "Nombre Ctte", "VN Gallo", "Diferencia"]
DET_WIDTHS  = [8, 10, 32, 14, 10, 14, 26, 14, 14]
DET_TITLES  = {
    "Client (556)":    "Concil. Client",
    "House (557)":     "Concil. House",
    "Def Fund (1121)": "Concil. Def Fund",
}


# ── Helpers ───────────────────────────────────────────────────────────────────
def get_node_short(node_key):
    if node_key in NODE_SHORT:
        return NODE_SHORT[node_key]
    lk = node_key.lower()
    if "client"   in lk: return "Client (556)"
    if "house"    in lk: return "House (557)"
    if "def fund" in lk: return "Def Fund (1121)"
    return node_key

def ctte_to_node_short(ctte_raw):
    try:
        n = int(float(str(ctte_raw).strip()))
    except (ValueError, TypeError):
        return "Client (556)"
    if n in DEFFUND_CTTES: return "Def Fund (1121)"
    if n in HOUSE_CTTES:   return "House (557)"
    return "Client (556)"

def _diff_fill(val):
    return DIFF_OK_FILL if val == 0 else DIFF_WARN_FILL

def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill      = fill
    cell.font      = font
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_data(cell, row_fill, fmt=None, align="right"):
    cell.fill      = row_fill
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt


# ══════════════════════════════════════════════════════════════════════════════
#  CARGA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════════════════

def _leer_collateral_csv(csv_file):
    collateral   = OrderedDict()
    cash_by_node = {}
    content = csv_file.read().decode("utf-8-sig")
    reader  = csv.reader(content.splitlines())
    next(reader, None)
    for row in reader:
        if len(row) < 3:
            continue
        account, asset, balance_str = row[0].strip(), row[1].strip(), row[2].strip()
        if account not in collateral:
            collateral[account]   = []
            cash_by_node[account] = 0.0
        if asset == "USD":
            try:
                cash_by_node[account] += float(balance_str)
            except ValueError:
                pass
            continue
        m = re.match(r"^(\d+)\s*-\s*(.+)$", asset)
        if not m:
            continue
        try:
            balance = float(balance_str)
        except ValueError:
            balance = 0.0
        collateral[account].append((m.group(1).strip(), m.group(2).strip(), balance))
    return collateral, cash_by_node


def _leer_especies(especies_file):
    especies = {}
    if especies_file is None:
        return especies
    wb = xlrd.open_workbook(file_contents=especies_file.read())
    sh = wb.sheet_by_name("Datos_Fijos_Especies")
    for i in range(1, sh.nrows):
        row  = sh.row_values(i)
        cod  = str(row[0]).strip().strip("'")
        nom  = str(row[1]).strip().strip("'")
        tprc = str(row[4]).strip().strip("'")
        especies[cod] = {"nombre": nom, "tipo_precio": tprc}
    return especies


def _leer_sagaclte(sagaclte_file, especies):
    gallo_by_node = {n: {} for n in NODE_ORDER}
    if sagaclte_file is None:
        return gallo_by_node
    wb = xlrd.open_workbook(file_contents=sagaclte_file.read())
    sh = wb.sheet_by_name("Saldos_de_Garantias")
    for i in range(1, sh.nrows):
        row      = sh.row_values(i)
        ctte_raw = row[0]
        nom_ctte = str(row[1]).strip()
        cvsa_raw = str(row[2]).strip().strip("'")
        nom_esp  = str(row[3]).strip()
        cant_raw = row[6]
        try:
            ctte_num = int(float(str(ctte_raw).strip()))
        except (ValueError, TypeError):
            continue
        try:
            cantidad = float(cant_raw)
        except (ValueError, TypeError):
            continue
        if cantidad <= 0:
            continue
        cvsa5 = cvsa_raw.zfill(5) if cvsa_raw.replace(".", "").isdigit() else cvsa_raw
        node  = ctte_to_node_short(ctte_num)
        if cvsa5 not in gallo_by_node[node]:
            gallo_by_node[node][cvsa5] = {"rows": [], "nombre_especie": nom_esp}
        gallo_by_node[node][cvsa5]["rows"].append((ctte_num, nom_ctte, cantidad))
    return gallo_by_node


def _leer_accounts(accounts_file):
    ctte_to_account_id = {}
    if accounts_file is None:
        return ctte_to_account_id
    content = accounts_file.read().decode("utf-8-sig")
    for line in content.splitlines():
        line  = line.strip().strip('"')
        parts = line.split(";")
        if len(parts) == 3 and parts[0] != "AGENTE":
            _, ctte, id_clearing = parts
            if ctte.strip().isdigit() and id_clearing.strip():
                ctte_to_account_id[ctte.strip()] = id_clearing.strip()
    # Fallback: CSV tabular con Account Name / Account ID
    if not ctte_to_account_id:
        accounts_file.seek(0)
        try:
            import pandas as pd
            df = pd.read_csv(accounts_file, dtype=str)
            if "Account Name" in df.columns and "Account ID" in df.columns:
                for _, row in df.iterrows():
                    name   = str(row.get("Account Name", "")).strip()
                    acc_id = str(row.get("Account ID",   "")).strip()
                    if name.isdigit() and acc_id:
                        ctte_to_account_id[name] = acc_id
        except Exception:
            pass
    return ctte_to_account_id


def _leer_pdf_aforos(pdf_file):
    byma_dict = {}
    if pdf_file is None:
        return byma_dict
    pdf_bytes = pdf_file.read()
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        full_text = "\n".join(p.extract_text() for p in pdf.pages if p.extract_text())
    pat = re.compile(r"([A-Za-z0-9][A-Za-z0-9._-]*)\s+(\d{2,6})\s+(\d+)%\s+\d+%")
    for m in pat.finditer(full_text):
        byma_dict[m.group(2).zfill(5)] = int(m.group(3)) / 100.0
    return byma_dict


def _leer_precios(pc_file):
    precios = {}
    if pc_file is None:
        return precios
    wb = xlrd.open_workbook(file_contents=pc_file.read())
    sh = wb.sheet_by_name("Precios_de_Cierre")
    for i in range(1, sh.nrows):
        row = sh.row_values(i)
        esp = str(row[0]).strip().strip("'")
        cod = esp.split()[0] if esp else ""
        precios[cod] = float(row[1]) if row[1] else 0.0
    return precios


# ══════════════════════════════════════════════════════════════════════════════
#  ENRIQUECIMIENTO
# ══════════════════════════════════════════════════════════════════════════════

def enrich_security(cvsa_raw, ticker, balance, especies, precios, tc_usd):
    cod5        = cvsa_raw.zfill(5)
    esp_info    = especies.get(cod5, {})
    nombre      = esp_info.get("nombre", ticker)
    tipo_precio = esp_info.get("tipo_precio", "Normal")
    precio      = precios.get(cod5, 0.0)
    if precio > 0:
        monto = balance * (precio / 100.0) if tipo_precio.lower().startswith("porc") else balance * precio
    else:
        monto = None
    return {"cvsa": cvsa_raw, "ticker": ticker, "nombre": nombre,
            "balance": balance, "tipo_precio": tipo_precio,
            "precio": precio, "monto": monto, "is_usd": False}

def enrich_usd(balance_usd, tc):
    monto = balance_usd * tc
    return {"cvsa": "USD", "ticker": "USD",
            "nombre": f"Dolares (Cash)  [TC: {tc:,.0f}]",
            "balance": balance_usd, "tipo_precio": "TC",
            "precio": tc, "monto": monto, "is_usd": True}


# ══════════════════════════════════════════════════════════════════════════════
#  ESCRITURA DE HOJAS
# ══════════════════════════════════════════════════════════════════════════════

def write_node_sheet(ws, node_key, items):
    short   = get_node_short(node_key)
    n_cols  = len(COLS)
    col_ltr = get_column_letter(COL_MONEY)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(row=1, column=1, value=short)
    t.fill = TITLE_FILL; t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ci, col_name in enumerate(COLS, 1):
        style_header(ws.cell(row=2, column=ci, value=col_name))
    ws.row_dimensions[2].height = 28
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}2"

    for ci in range(1, n_cols + 1):
        c = ws.cell(row=3, column=ci)
        c.fill = SUBTOT_FILL; c.font = SUBTOT_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=1, value="SUBTOTAL").alignment = Alignment(horizontal="center", vertical="center")
    first_data, last_data = 4, 3 + len(items)
    ws.cell(row=3, column=COL_MONEY,
            value=f"=SUBTOTAL(9,{col_ltr}{first_data}:{col_ltr}{last_data})"
            ).number_format = FMT_MONEY
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    total_monto = 0.0
    for ri, row in enumerate(items, start=first_data):
        if row["is_usd"]:
            fill = USD_FILL
        elif row["precio"] == 0 or row["monto"] is None:
            fill = YELLOW_FILL
        else:
            fill = ODD_FILL if (ri % 2 == 0) else EVEN_FILL

        balance_fmt = FMT_PRICE if row["is_usd"] else FMT_INT
        vals   = [row["cvsa"], row["ticker"], row["nombre"],
                  row["balance"], row["precio"] if row["precio"] > 0 else "",
                  row["tipo_precio"], row["monto"]]
        fmts   = [None, None, None, balance_fmt, FMT_PRICE, None, FMT_MONEY]
        aligns = ["center","center","left","right","right","center","right"]
        for ci, (val, fmt, align) in enumerate(zip(vals, fmts, aligns), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            style_data(c, fill, fmt, align)
            if row["is_usd"]:
                c.font = Font(bold=True, size=10)
        if row["monto"] is not None:
            total_monto += row["monto"]

    total_row = last_data + 1
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=total_row, column=ci)
        c.fill = TOTAL_FILL; c.font = TOTAL_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=total_row, column=1).value     = "TOTAL"
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=COL_MONEY).value         = total_monto
    ws.cell(row=total_row, column=COL_MONEY).number_format = FMT_MONEY
    return total_monto


def write_resumen(wb, summary_data, fecha_proceso, tc_usd):
    ws_sum = wb.create_sheet(title="Resumen")
    ws_sum.merge_cells("A1:C1")
    t = ws_sum.cell(row=1, column=1,
                    value=f"Collateral Position  {fecha_proceso.strftime('%d/%m/%Y')}   |   TC USD: {tc_usd:,.0f}")
    t.fill = TITLE_FILL; t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 22

    sum_headers = ["Nodo", "Nro. Activos", "Monto a Precio Cierre"]
    sum_widths  = [22, 14, 26]
    for ci, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        style_header(ws_sum.cell(row=2, column=ci, value=h))
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.row_dimensions[2].height = 22

    grand_monto = 0.0
    for ri, (short, n_items, tm) in enumerate(summary_data, start=3):
        fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        ws_sum.cell(row=ri, column=1, value=short).fill = fill
        ws_sum.cell(row=ri, column=2, value=n_items).fill = fill
        c3 = ws_sum.cell(row=ri, column=3, value=tm)
        c3.fill = fill; c3.number_format = FMT_MONEY
        for ci in range(1, 4):
            ws_sum.cell(row=ri, column=ci).border    = BORDER
            ws_sum.cell(row=ri, column=ci).alignment = Alignment(horizontal="right", vertical="center")
        ws_sum.cell(row=ri, column=1).alignment = Alignment(horizontal="left",   vertical="center")
        ws_sum.cell(row=ri, column=2).alignment = Alignment(horizontal="center", vertical="center")
        grand_monto += tm

    total_row = 3 + len(summary_data)
    for ci, val in enumerate(["TOTAL GENERAL", sum(s[1] for s in summary_data), grand_monto], 1):
        c = ws_sum.cell(row=total_row, column=ci, value=val)
        c.fill = TOTAL_FILL; c.font = TOTAL_FONT; c.border = BORDER
        c.alignment = Alignment(horizontal="right", vertical="center")
    ws_sum.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.cell(row=total_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.cell(row=total_row, column=3).number_format = FMT_MONEY
    return grand_monto


def write_conciliacion(wb, bc_by_node, gallo_by_node, especies, fecha_proceso):
    ws = wb.create_sheet(title="Conciliación")
    nc = len(CONC_COLS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    t = ws.cell(row=1, column=1,
                value=f"Conciliación BC vs Gallo — {fecha_proceso.strftime('%d/%m/%Y')}")
    t.fill = TITLE_FILL; t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(zip(CONC_COLS, CONC_WIDTHS), 1):
        style_header(ws.cell(row=2, column=ci, value=h))
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22
    ws.auto_filter.ref = f"A2:{get_column_letter(nc)}2"
    ws.freeze_panes = "A3"

    ri = 3
    for node in NODE_ORDER:
        bc_n  = bc_by_node.get(node, {})
        gal_n = gallo_by_node.get(node, {})
        for cvsa5 in sorted(set(bc_n) | set(gal_n)):
            bc_info  = bc_n.get(cvsa5, {})
            gal_data = gal_n.get(cvsa5, {"rows": [], "nombre_especie": ""})
            vn_bc    = bc_info.get("balance", 0.0)
            vn_gal   = sum(r[2] for r in gal_data["rows"])
            diff     = vn_bc - vn_gal
            ticker   = bc_info.get("ticker", "")
            nombre   = (bc_info.get("nombre") or gal_data.get("nombre_especie")
                        or especies.get(cvsa5, {}).get("nombre", cvsa5))
            fill = ODD_FILL if ri % 2 == 0 else EVEN_FILL
            for ci, (val, aln, fmt) in enumerate(zip(
                [node, cvsa5, ticker, nombre, vn_bc, vn_gal, diff],
                ["left","center","center","left","right","right","right"],
                [None,  None,   None,   None,  FMT_INT, FMT_INT, FMT_INT]
            ), 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill; c.border = BORDER
                c.alignment = Alignment(horizontal=aln, vertical="center")
                if fmt: c.number_format = fmt
            ws.cell(row=ri, column=7).fill = _diff_fill(diff)
            ri += 1


def write_concil_detalle(wb, node, bc_by_node, gallo_by_node, especies, ctte_to_account_id, fecha_proceso):
    title  = DET_TITLES[node]
    ws     = wb.create_sheet(title=title)
    nd     = len(DET_COLS)
    bc_n   = bc_by_node.get(node, {})
    gal_n  = gallo_by_node.get(node, {})

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nd)
    t = ws.cell(row=1, column=1,
                value=f"Conciliación {node} — {fecha_proceso.strftime('%d/%m/%Y')}")
    t.fill = TITLE_FILL; t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(zip(DET_COLS, DET_WIDTHS), 1):
        style_header(ws.cell(row=2, column=ci, value=h))
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22
    ws.auto_filter.ref = f"A2:{get_column_letter(nd)}2"
    ws.freeze_panes = "A3"

    ri = 3
    for cvsa5 in sorted(set(bc_n) | set(gal_n)):
        bc_info  = bc_n.get(cvsa5, {})
        gal_data = gal_n.get(cvsa5, {"rows": [], "nombre_especie": ""})
        gal_rows = gal_data["rows"]
        vn_bc    = bc_info.get("balance", 0.0)
        vn_gal   = sum(r[2] for r in gal_rows)
        diff     = vn_bc - vn_gal
        ticker   = bc_info.get("ticker", "")
        nombre   = (bc_info.get("nombre") or gal_data.get("nombre_especie")
                    or especies.get(cvsa5, {}).get("nombre", cvsa5))

        for ci, (val, fmt, aln) in enumerate(zip(
            [cvsa5, ticker, nombre, vn_bc, "", "", "", vn_gal, diff],
            [None, None, None, FMT_INT, None, None, None, FMT_INT, FMT_INT],
            ["center","center","left","right","center","center","left","right","right"]
        ), 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = ESPECIE_HDR_FILL; c.font = ESPECIE_HDR_FONT
            c.border = BORDER
            c.alignment = Alignment(horizontal=aln, vertical="center")
            if fmt: c.number_format = fmt
        ws.cell(row=ri, column=9).fill = _diff_fill(diff)
        ri += 1

        for idx, (ctte_num, nom_ctte, cantidad) in enumerate(gal_rows):
            cfill  = ODD_FILL if idx % 2 == 0 else EVEN_FILL
            acc_id = ctte_to_account_id.get(str(ctte_num), "")
            for ci, (val, fmt, aln) in enumerate(zip(
                [cvsa5, "", "", "", ctte_num, acc_id, nom_ctte, cantidad, ""],
                [None, None, None, None, None, None, None, FMT_INT, None],
                ["center","center","left","right","center","center","left","right","right"]
            ), 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = cfill; c.border = BORDER
                c.alignment = Alignment(horizontal=aln, vertical="center")
                if fmt: c.number_format = fmt
            ri += 1


# ══════════════════════════════════════════════════════════════════════════════
#  FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def generar_reporte(
    csv_file,
    pc_file,
    especies_file,
    pdf_aforos_file,
    sagaclte_file=None,
    accounts_file=None,
    tc_usd: float = 1400.0,
    fecha_proceso: date = None,
) -> tuple:
    """
    Devuelve (BytesIO xlsx, dict resumen).

    Requeridos: csv_file, pc_file, especies_file, pdf_aforos_file
    Opcionales: sagaclte_file, accounts_file  (activan hojas de conciliación)
    """
    if fecha_proceso is None:
        fecha_proceso = date.today()

    # Cargar inputs
    collateral, cash_by_node = _leer_collateral_csv(csv_file)
    especies       = _leer_especies(especies_file)
    precios        = _leer_precios(pc_file)
    byma_dict      = _leer_pdf_aforos(pdf_aforos_file)
    gallo_by_node  = _leer_sagaclte(sagaclte_file, especies)
    ctte_to_acc_id = _leer_accounts(accounts_file)

    # Enriquecer
    enriched = OrderedDict()
    for node, items in collateral.items():
        rows = [enrich_security(cvsa, ticker, bal, especies, precios, tc_usd)
                for (cvsa, ticker, bal) in items]
        usd = cash_by_node.get(node, 0.0)
        if usd > 0:
            rows.append(enrich_usd(usd, tc_usd))
        enriched[node] = rows

    # Construir bc_by_node para conciliación
    bc_by_node = {n: {} for n in NODE_ORDER}
    for nk, items in enriched.items():
        short = get_node_short(nk)
        if short not in bc_by_node:
            continue
        for it in items:
            if it["is_usd"]:
                continue
            c5 = it["cvsa"].zfill(5)
            bc_by_node[short][c5] = {
                "ticker":  it["ticker"],
                "nombre":  it["nombre"],
                "balance": it["balance"],
            }

    # Construir workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary_data = []
    for node_key, items in enriched.items():
        short = get_node_short(node_key)
        ws    = wb.create_sheet(title=short)
        tm    = write_node_sheet(ws, node_key, items)
        n_sec = sum(1 for i in items if not i["is_usd"])
        summary_data.append((short, n_sec, tm))

    grand_monto = write_resumen(wb, summary_data, fecha_proceso, tc_usd)
    wb.move_sheet("Resumen", offset=-len(wb.sheetnames) + 1)

    tiene_conciliacion = sagaclte_file is not None
    if tiene_conciliacion:
        write_conciliacion(wb, bc_by_node, gallo_by_node, especies, fecha_proceso)
        for node in NODE_ORDER:
            write_concil_detalle(wb, node, bc_by_node, gallo_by_node,
                                 especies, ctte_to_acc_id, fecha_proceso)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    n_sin_precio  = sum(1 for items in enriched.values()
                        for it in items if not it["is_usd"] and it["precio"] == 0)
    n_sin_aforo   = sum(1 for items in enriched.values()
                        for it in items if not it["is_usd"] and it["precio"] > 0
                        and byma_dict.get(it["cvsa"].zfill(5), 0) == 0)

    resumen = {
        "fecha":           fecha_proceso.strftime("%d/%m/%Y"),
        "tc_usd":          tc_usd,
        "grand_monto":     grand_monto,
        "n_nodes":         len(enriched),
        "n_sin_precio":    n_sin_precio,
        "n_sin_aforo":     n_sin_aforo,
        "tiene_conc":      tiene_conciliacion,
        "summary":         summary_data,   # [(short, n_sec, monto), ...]
        "n_aforos":        len(byma_dict),
        "n_precios":       len(precios),
    }
    return output, resumen
