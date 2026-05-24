"""
Arreglos Garantías NASDAQ
Genera archivos SI2 (DELIVER/RECEIVE) y resúmenes TXT para arreglos manuales
de garantías en NASDAQ BYMA, más archivo Gallo XLSX con las dos hojas.
Adaptado de run_arreglo_gara_nasdaq.py para funcionar con archivos en memoria.
"""

import io
from datetime import datetime
from collections import defaultdict

import openpyxl
import xlrd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── Constantes ───────────────────────────────────────────────────────────────
HOUSE_COMITENTES       = {1000, 1001, 1002, 1003}
DEFAULT_FUND_COMITENTE = 888888888

HEADER_SI2 = (
    "InstructingParty;SettlementParty;SecuritiesAccount;Instrument;"
    "InstrumentIdentifierType;CSDOfCounterparty;SettlementCounterparty;"
    "SecuritiesAccountOfCounterparty;InstructionReference;"
    "Instrument(MovementOfSecurities);Quantity;QuantityType;TransactionType;"
    "SettlementMethod;TradeDate;IntendedSettlementDate;PaymentType"
)


# ─── Helpers ──────────────────────────────────────────────────────────────────
def _normalize_cvsa(val):
    try:
        return str(int(float(str(val).strip()))).zfill(5)
    except (ValueError, TypeError):
        return str(val).strip().zfill(5)


def _strip_leading_zeros(cvsa):
    try:
        return str(int(str(cvsa).strip()))
    except (ValueError, TypeError):
        return str(cvsa)


def _is_numeric(val):
    try:
        int(float(str(val).strip()))
        return True
    except (ValueError, TypeError):
        return False


def _build_counterparty(comitente):
    """Devuelve (cuenta_contraparte, nombre_nodo)."""
    c = int(comitente)
    if c == DEFAULT_FUND_COMITENTE:
        return "880233/888888888", "DEFAULT_FUND"
    if c in HOUSE_COMITENTES:
        return "80233/222222222", "HOUSE"
    return "80233/555555555", "CLIENT"


def _build_resumen(rows):
    client_totals       = defaultdict(lambda: {"ticker": "", "vn": 0})
    house_totals        = defaultdict(lambda: {"ticker": "", "vn": 0})
    default_fund_totals = defaultdict(lambda: {"ticker": "", "vn": 0})
    for entry in rows:
        _, nodo = _build_counterparty(entry["ctte"])
        target = (default_fund_totals if nodo == "DEFAULT_FUND"
                  else house_totals   if nodo == "HOUSE"
                  else client_totals)
        target[entry["cvsa"]]["ticker"]  = entry["ticker"]
        target[entry["cvsa"]]["vn"]     += entry["vn"]
    return client_totals, house_totals, default_fund_totals


def _resumen_to_bytes(client_totals, house_totals, default_fund_totals):
    lines = []
    if client_totals:
        lines.append("NODO CLIENT 80233/555555555")
        lines.append("Codigo CVSA;Ticker;VN")
        for cvsa in sorted(client_totals.keys()):
            d = client_totals[cvsa]
            lines.append(f"{cvsa};{d['ticker']};{d['vn']}")
    if house_totals:
        if lines:
            lines.append("")
        lines.append("NODO HOUSE 80233/222222222")
        lines.append("Codigo CVSA;Ticker;VN")
        for cvsa in sorted(house_totals.keys()):
            d = house_totals[cvsa]
            lines.append(f"{cvsa};{d['ticker']};{d['vn']}")
    if default_fund_totals:
        if lines:
            lines.append("")
        lines.append("DEFAULT FUND 880233/888888888")
        lines.append("Codigo CVSA;Ticker;VN")
        for cvsa in sorted(default_fund_totals.keys()):
            d = default_fund_totals[cvsa]
            lines.append(f"{cvsa};{d['ticker']};{d['vn']}")
    return "\n".join(lines).encode("utf-8")


def _si2_to_bytes(rows, movement, date_yyyymmdd, date_ref, ref_prefix, start_counter):
    """Genera contenido SI2 y retorna (bytes, último_contador_usado)."""
    lines   = [HEADER_SI2]
    counter = start_counter
    for entry in rows:
        counterparty_account, _ = _build_counterparty(entry["ctte"])
        ref  = f"{ref_prefix}{date_ref}{counter:03d}"
        line = (
            f"233;233;233/{entry['ctte']};{entry['instrument']};LOCAL_CODE;CVSA;233;"
            f"{counterparty_account};{ref};{movement};{entry['vn']};"
            f"{entry['qty_type']};TRAD;RTGS;{date_yyyymmdd};{date_yyyymmdd};NOTHING;NORMAL"
        )
        lines.append(line)
        counter += 1
    return "\n".join(lines).encode("utf-8"), counter - 1


def _gallo_to_bytes(enviar, traer, date_label):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    title_fill  = PatternFill("solid", fgColor="1F4E79")
    title_font  = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    col_headers = ["CTTE", "TICKER", "Codigo CVSA", "VN", "Fecha de Vto"]
    col_widths  = [14, 14, 14, 14, 16]

    for sheet_name, rows in [("ENTREGA GTIAS", enviar), ("DEVOLUCIONES GTIAS", traer)]:
        ws = wb.create_sheet(title=sheet_name)

        # Fila 1: título fusionado A1:E1
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value     = sheet_name
        title_cell.font      = title_font
        title_cell.fill      = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        # Fila 2: encabezados
        for col_idx, header in enumerate(col_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Filas de datos
        for data_row in rows:
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=data_row["ctte"])
            ws.cell(row=r, column=2, value=data_row["ticker"])
            ws.cell(row=r, column=3, value=data_row["cvsa"])
            ws.cell(row=r, column=4, value=data_row["vn"])
            fecha = data_row["fecha_vto"]
            if fecha is not None:
                if hasattr(fecha, "date"):
                    fecha = fecha.date()
                fecha_cell = ws.cell(row=r, column=5, value=fecha)
                fecha_cell.number_format = "DD/MM/YYYY"
            else:
                ws.cell(row=r, column=5, value=None)

        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Función principal ────────────────────────────────────────────────────────
def generar_arreglo(inputs_file, especies_file, counter_state=None):
    """
    Parámetros
    ----------
    inputs_file   : file-like  — INPUTS ARREGLOS GARA.xlsx
    especies_file : file-like  — ESPECIES.XLS
    counter_state : dict | None — {"{PREFIX}{DDMMAA}": last_used}
                    Si None se usan contadores desde 1.

    Retorna
    -------
    outputs       : dict  — claves: req_envio, resumen_deposit,
                            ret_devolucion, resumen_withdraw, gallo
                    Cada valor: (bytes, filename) o None si no aplica
    new_counter   : dict  — estado actualizado de contadores
    resumen       : dict  — métricas para la UI
    advertencias  : list[str]
    """
    advertencias  = []
    counter_state = dict(counter_state) if counter_state else {}
    today         = datetime.today()
    date_yyyymmdd = today.strftime("%Y%m%d")
    date_dd_mm_aa = today.strftime("%d-%m-%y")
    date_ref      = date_dd_mm_aa.replace("-", "")   # DDMMAA sin guiones, ej: 090426

    # ── 1. ESPECIES.XLS ───────────────────────────────────────────────────────
    especies_file.seek(0)
    wb_esp = xlrd.open_workbook(file_contents=especies_file.read())
    hoja   = ("Datos_Fijos_Especies"
              if "Datos_Fijos_Especies" in wb_esp.sheet_names()
              else wb_esp.sheet_names()[0])
    ws_esp = wb_esp.sheet_by_name(hoja)

    cvsa_to_ticker  = {}
    ticker_to_cvsa  = {}
    cvsa_to_qty_type = {}

    for i in range(1, ws_esp.nrows):
        row    = ws_esp.row_values(i)
        cvsa   = str(row[0]).strip().strip("'")
        ticker = str(row[9]).strip().strip("'")   # col 9 = Norm.
        tipo   = str(row[4]).strip()              # col 4 = tipo precio
        if not cvsa:
            continue
        cvsa_to_qty_type[cvsa] = "FACE_AMOUNT" if tipo == "Porc." else "UNIT"
        if ticker and ticker not in ("", "None"):
            cvsa_to_ticker[cvsa]           = ticker
            ticker_to_cvsa[ticker.upper()] = cvsa

    # ── 2. INPUTS ARREGLOS GARA.xlsx ─────────────────────────────────────────
    inputs_file.seek(0)
    wb_inp = openpyxl.load_workbook(inputs_file, data_only=True)
    ws_inp = wb_inp["Hoja1"] if "Hoja1" in wb_inp.sheetnames else wb_inp.active

    enviar = []
    traer  = []

    for row in ws_inp.iter_rows(min_row=2, values_only=True):
        if len(row) < 4:
            continue
        ctte, movimiento, codigo, vn = row[0], row[1], row[2], row[3]
        fecha_vto = row[4] if len(row) > 4 else None

        if ctte is None or movimiento is None or codigo is None or vn is None:
            continue
        movimiento = str(movimiento).strip().upper()
        if movimiento not in ("ENVIAR", "TRAER"):
            continue

        codigo_str = str(codigo).strip()
        if _is_numeric(codigo_str):
            cvsa   = _normalize_cvsa(codigo_str)
            ticker = cvsa_to_ticker.get(cvsa, cvsa)
        else:
            ticker = codigo_str.upper()
            cvsa   = ticker_to_cvsa.get(ticker, "")
            if not cvsa:
                advertencias.append(
                    f"Ticker '{ticker}' no encontrado en ESPECIES.XLS — "
                    "se usó como código provisorio."
                )
                cvsa = ticker

        entry = {
            "ctte":      int(float(str(ctte))),
            "cvsa":      cvsa,
            "instrument": _strip_leading_zeros(cvsa),
            "ticker":    ticker,
            "vn":        int(float(str(vn))),
            "qty_type":  cvsa_to_qty_type.get(cvsa, "UNIT"),
            "fecha_vto": fecha_vto,
        }

        if movimiento == "ENVIAR":
            enviar.append(entry)
        else:
            traer.append(entry)

    if not enviar and not traer:
        raise ValueError(
            "INPUTS ARREGLOS GARA.xlsx no tiene filas válidas con "
            "MOVIMIENTO = ENVIAR o TRAER."
        )

    # ── 3. Generar outputs ────────────────────────────────────────────────────
    outputs     = {}
    new_counter = dict(counter_state)

    # [a] Req-envio (DELIVER) + [b] Resumen DEPOSIT
    if enviar:
        age_key   = f"AGE{date_ref}"
        start_age = new_counter.get(age_key, 0) + 1
        si2_bytes, last_age = _si2_to_bytes(
            enviar, "DELIVER", date_yyyymmdd, date_ref,
            "AGE", start_age
        )
        new_counter[age_key] = last_age

        fname_si2 = f"Req-envio de gtias {date_dd_mm_aa}.SI2"
        outputs["req_envio"] = (si2_bytes, fname_si2)

        ct, ht, dft = _build_resumen(enviar)
        resumen_bytes = _resumen_to_bytes(ct, ht, dft)
        fname_dep = f"Resumen DEPOSIT {date_dd_mm_aa}.txt"
        outputs["resumen_deposit"] = (resumen_bytes, fname_dep)
    else:
        outputs["req_envio"]       = None
        outputs["resumen_deposit"] = None

    # [c] Ret-devolucion (RECEIVE) + [d] Resumen WITHDRAW
    if traer:
        agd_key   = f"AGD{date_ref}"
        start_agd = new_counter.get(agd_key, 0) + 1
        si2_bytes, last_agd = _si2_to_bytes(
            traer, "RECEIVE", date_yyyymmdd, date_ref,
            "AGD", start_agd
        )
        new_counter[agd_key] = last_agd

        fname_si2 = f"Ret-devolucion gtia {date_dd_mm_aa}.SI2"
        outputs["ret_devolucion"] = (si2_bytes, fname_si2)

        ct, ht, dft = _build_resumen(traer)
        resumen_bytes = _resumen_to_bytes(ct, ht, dft)
        fname_wit = f"Resumen WITHDRAW {date_dd_mm_aa}.txt"
        outputs["resumen_withdraw"] = (resumen_bytes, fname_wit)
    else:
        outputs["ret_devolucion"]   = None
        outputs["resumen_withdraw"] = None

    # [e] Gallo XLSX
    gallo_bytes  = _gallo_to_bytes(enviar, traer, date_dd_mm_aa)
    fname_gallo  = f"Gallo {date_dd_mm_aa}.xlsx"
    outputs["gallo"] = (gallo_bytes, fname_gallo)

    # ── 4. Resumen para la UI ─────────────────────────────────────────────────
    # Rangos de refs generadas
    def _ref_range(prefix, key, start, last):
        if last < start:
            return ""
        return (f"{prefix}{date_ref}{start:03d}"
                + (f" – {prefix}{date_ref}{last:03d}" if last > start else ""))

    age_start = counter_state.get(f"AGE{date_ref}", 0) + 1 if enviar else None
    agd_start = counter_state.get(f"AGD{date_ref}", 0) + 1 if traer  else None

    resumen = {
        "fecha":        date_dd_mm_aa,
        "n_enviar":     len(enviar),
        "n_traer":      len(traer),
        "refs_age":     _ref_range("AGE", f"AGE{date_ref}",
                                   age_start,
                                   new_counter.get(f"AGE{date_ref}", 0)) if enviar else "",
        "refs_agd":     _ref_range("AGD", f"AGD{date_ref}",
                                   agd_start,
                                   new_counter.get(f"AGD{date_ref}", 0)) if traer  else "",
    }

    return outputs, new_counter, resumen, advertencias
