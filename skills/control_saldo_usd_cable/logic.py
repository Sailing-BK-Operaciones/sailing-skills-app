# -*- coding: utf-8 -*-
"""
Control Saldo USD Cable — versión web (Streamlit)
Adaptado de run_saldos_cable.py para inputs vía UploadedFile y outputs bytes.

Dos modos:
  - BASELINE (procesar_baseline): saldo de partida a partir de SALUSD + LISTADO GALLO.
  - DIARIO   (procesar_diario)  : incremental sobre baseline + CCs + control SALUSD.

Los archivos de estado (baseline_split.csv, estado_diario.csv,
asignaciones_byma_broker.csv, movimientos_backdated.csv) se pasan como
UploadedFile opcionales y se devuelven actualizados como bytes.
"""
from __future__ import annotations

import csv
import io
import os
import re
import datetime as dt
from io import BytesIO, StringIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Directorio de estado (bundled en el repo, editable en runtime) ────────────
# baseline_split.csv / estado_diario.csv / asignaciones_byma_broker.csv /
# movimientos_backdated.csv se leen/escriben desde acá por defecto.
STATE_DIR = Path(__file__).parent / "state"
STATE_DIR.mkdir(exist_ok=True)

BASELINE_PATH = STATE_DIR / "baseline_split.csv"
ESTADO_PATH   = STATE_DIR / "estado_diario.csv"
DEC_PATH      = STATE_DIR / "asignaciones_byma_broker.csv"
BD_PATH       = STATE_DIR / "movimientos_backdated.csv"


# ── Constantes ────────────────────────────────────────────────────────────────
CC_SHEET       = "Consulta_Cuenta_Corriente_Otra"
MONEY_FMT      = "#,##0.00"
LISTADO_SHEET  = "Saldos CTTES"

# Cuentas de SALUSD que NO son comitentes cliente:
#   - >= 999000 → cuentas de mercado (STONEX / IBKR / concentración 999)
#   - 1000      → cartera propia (house)
CUENTAS_NO_CLIENTE = {1000}

BASE_FIELDS    = ["Comitente", "BROKER", "BYMA", "Total", "FechaCorte"]
ESTADO_FIELDS  = ["Comitente", "BROKER", "BYMA", "Total", "FechaActualizacion"]
DEC_FIELDS     = ["Comitente", "Cpbt", "Numero", "Destino", "Fecha", "Importe", "Referencia"]
BD_FIELDS      = ["Comitente", "Cpbt", "Numero", "Fecha", "Importe", "Nota"]

# Estilos openpyxl
HDR_FILL  = PatternFill("solid", fgColor="1F4E78")
HDR_FONT  = Font(bold=True, color="FFFFFF")
SUB_FILL  = PatternFill("solid", fgColor="D9E1F2")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL  = PatternFill("solid", fgColor="F8CBAD")
_THIN     = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ── Helpers ───────────────────────────────────────────────────────────────────
def _read_bytes(f) -> bytes:
    f.seek(0)
    return f.read()


def _read_excel(f, **kw) -> pd.DataFrame:
    return pd.read_excel(BytesIO(_read_bytes(f)), **kw)


def _norm_num(v) -> str:
    """Normaliza el número de comprobante a str sin decimales."""
    try:
        return str(int(float(v)))
    except (ValueError, TypeError):
        return str(v).strip()


def _es_no_cliente(ctte: int) -> bool:
    return ctte >= 999000 or ctte in CUENTAS_NO_CLIENTE


def _parse_fecha(s):
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(str(s).strip(), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _comitente_from_name(fname: str) -> str:
    m = re.search(r"HD(\d+)", fname, re.IGNORECASE)
    return m.group(1) if m else fname


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _autofit(ws, ncols, minw=10, maxw=42):
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        longest = 0
        for cell in ws[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(minw, min(maxw, longest + 2))


# ── Loaders ───────────────────────────────────────────────────────────────────
def _load_tabcompb(tabcompb_file):
    """Dict Abrev → destino ('BROKER' / 'BYMA' / 'BYMA O BROKER')."""
    raw = _read_excel(tabcompb_file, sheet_name="Tabla_Comprobantes", header=None)
    header = raw.iloc[0].tolist()
    col_dest = None
    for i, h in enumerate(header):
        if isinstance(h, str) and "analisis" in h.lower():
            col_dest = i
    if col_dest is None:
        col_dest = len(header) - 1
    col_abrev = 2
    mp = {}
    for _, row in raw.iloc[1:].iterrows():
        abrev = str(row[col_abrev]).strip()
        dest = row[col_dest]
        if abrev and abrev.lower() != "nan":
            dest = str(dest).strip()
            if dest and dest.lower() != "nan":
                mp[abrev] = dest.upper()
    return mp


def _load_salusd(salusd_file):
    """(dict comitente → saldo vencido firmado, dict comitente → nombre)."""
    df = _read_excel(salusd_file, sheet_name="Listado_Saldos_Dolares", header=0)
    df.columns = [str(c).strip().strip("'") for c in df.columns]
    df["Numero"] = pd.to_numeric(df["Numero"], errors="coerce")
    df = df[df["Numero"].notna()].copy()
    df["Numero"] = df["Numero"].astype(int)
    df["Saldo Vencido"] = pd.to_numeric(df["Saldo Vencido"], errors="coerce").fillna(0.0)
    nombres = dict(zip(df["Numero"], df["Nombre"].astype(str)))
    return dict(zip(df["Numero"], df["Saldo Vencido"])), nombres


def _load_sub_acreedores(salusd_file):
    """Lee el subtotal 'SUB.ACREEDORES' del SALUSD (columna D).
    Devuelve el valor en convención positiva (|dato|), o None si no se encuentra.
    """
    if salusd_file is None:
        return None
    try:
        raw = _read_excel(salusd_file, sheet_name="Listado_Saldos_Dolares", header=None)
    except Exception:
        return None
    for i in range(len(raw)):
        for j in range(raw.shape[1]):
            if str(raw.iat[i, j]).strip().upper() == "SUB.ACREEDORES":
                val = pd.to_numeric(raw.iat[i, 3], errors="coerce")  # columna D
                return abs(float(val)) if pd.notna(val) else None
    return None


def _load_listado_gallo(listado_file):
    """Dict comitente → (broker, byma, valor) desde bloque 'LISTADO GALLO'."""
    raw = _read_excel(listado_file, sheet_name=LISTADO_SHEET, header=None)
    r0 = c0 = None
    for i in range(len(raw)):
        for j in range(raw.shape[1]):
            if str(raw.iat[i, j]).strip().upper() == "LISTADO GALLO":
                r0, c0 = i, j
    if r0 is None:
        raise ValueError("No se encontró el bloque 'LISTADO GALLO' en el archivo.")
    sub = raw.iloc[r0 + 2:, c0:c0 + 4].copy()
    sub.columns = ["CTTE", "B", "Y", "VALOR"]
    sub["CTTE"] = pd.to_numeric(sub["CTTE"], errors="coerce")
    sub = sub[sub["CTTE"].notna()].copy()
    sub["CTTE"] = sub["CTTE"].astype(int)
    for c in ("B", "Y", "VALOR"):
        sub[c] = pd.to_numeric(sub[c], errors="coerce").fillna(0.0)
    return {int(r.CTTE): (float(r.B), float(r.Y), float(r.VALOR)) for r in sub.itertuples()}


def _load_cc(cc_file):
    """(df_movimientos, total_cc_reportado)."""
    df = _read_excel(cc_file, sheet_name=CC_SHEET, header=0)
    df.columns = ["Liquida", "Cpbt", "Numero", "Importe", "Saldo",
                  "Especie", "Referencia"][: len(df.columns)]
    total_row = df[df["Cpbt"].astype(str).str.strip() == "Total"]
    total_cc = float(total_row["Importe"].iloc[0]) if len(total_row) else None
    mov = df[
        df["Liquida"].notna()
        & df["Cpbt"].notna()
        & (df["Cpbt"].astype(str).str.strip() != "Total")
    ].copy()
    mov["Cpbt"] = mov["Cpbt"].astype(str).str.strip()
    mov["Importe"] = pd.to_numeric(mov["Importe"], errors="coerce").fillna(0.0)
    return mov, total_cc


# ── CSV state (in / out) ──────────────────────────────────────────────────────
def _csv_read(source, fields):
    """Devuelve lista de dicts leyendo un CSV UploadedFile, Path o str (o [] si es None).
    Acepta también rutas del filesystem — devuelve [] si el archivo no existe.
    """
    if source is None:
        return []
    if isinstance(source, (str, os.PathLike)):
        p = Path(source)
        if not p.exists():
            return []
        with open(p, encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    source.seek(0)
    content = source.read().decode("utf-8-sig")
    return list(csv.DictReader(StringIO(content)))


def _csv_write_path(path, rows, fields):
    """Escribe rows a un archivo CSV en disco."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fields})


def _csv_write(rows, fields) -> bytes:
    """Serializa rows (lista de dicts) a bytes de CSV utf-8-sig."""
    buf = StringIO()
    w = csv.DictWriter(buf, fieldnames=fields)
    w.writeheader()
    for r in rows:
        w.writerow({k: r.get(k, "") for k in fields})
    return buf.getvalue().encode("utf-8-sig")


def _load_baseline(baseline_csv_file):
    """Dict comitente → {'BROKER','BYMA','Total','FechaCorte'}."""
    out = {}
    for row in _csv_read(baseline_csv_file, BASE_FIELDS):
        try:
            ct = int(float(row["Comitente"]))
        except (ValueError, TypeError, KeyError):
            continue
        out[ct] = {
            "BROKER":     float(row.get("BROKER") or 0),
            "BYMA":       float(row.get("BYMA") or 0),
            "Total":      float(row.get("Total") or 0),
            "FechaCorte": row.get("FechaCorte", ""),
        }
    return out


def _load_estado(estado_csv_file):
    out = {}
    for row in _csv_read(estado_csv_file, ESTADO_FIELDS):
        try:
            ct = int(float(row["Comitente"]))
        except (ValueError, TypeError, KeyError):
            continue
        out[ct] = {
            "BROKER": float(row.get("BROKER") or 0),
            "BYMA":   float(row.get("BYMA") or 0),
            "Total":  float(row.get("Total") or 0),
        }
    return out


def _load_decisiones(dec_csv_file):
    """(dict key→destino, dict key→record) donde key=(ctte,cpbt,numero)."""
    dec, records = {}, {}
    for row in _csv_read(dec_csv_file, DEC_FIELDS):
        destino = (row.get("Destino") or "").strip().upper()
        if destino not in ("BROKER", "BYMA"):
            continue
        key = (str(row.get("Comitente", "")).strip(),
               str(row.get("Cpbt", "")).strip(),
               _norm_num(row.get("Numero", "")))
        dec[key] = destino
        records[key] = row
    return dec, records


def _load_backdated(bd_csv_file):
    """(dict comitente → set((cpbt,numero)), dict key → record)."""
    incluidos, records = {}, {}
    for row in _csv_read(bd_csv_file, BD_FIELDS):
        try:
            ct = int(float(row["Comitente"]))
        except (ValueError, TypeError, KeyError):
            continue
        cpbt = str(row.get("Cpbt", "")).strip()
        num  = _norm_num(row.get("Numero", ""))
        incluidos.setdefault(ct, set()).add((cpbt, num))
        records[(str(ct), cpbt, num)] = row
    return incluidos, records


# ── Modo BASELINE ─────────────────────────────────────────────────────────────
def procesar_baseline(salusd_file, listado_file, fecha_corte=None):
    """
    Retorna: (xlsx_bytes, baseline_csv_bytes, resumen_dict)
    """
    if fecha_corte is None:
        fecha_corte = dt.date.today()
    salusd, nombres = _load_salusd(salusd_file)
    lst = _load_listado_gallo(listado_file)

    baseline_rows = []
    resumen       = []
    excluidas     = []
    falta_definir = []
    revisar = 0

    for ct in sorted(salusd):
        saldo_venc = salusd[ct]
        total = abs(saldo_venc)
        if _es_no_cliente(ct):
            excluidas.append((ct, nombres.get(ct, ""), saldo_venc))
            continue
        if ct not in lst:
            falta_definir.append((ct, nombres.get(ct, ""), saldo_venc))
            continue
        b, y, _valor = lst[ct]
        ctrl = "OK" if abs(total - (b + y)) < 0.01 else "REVISAR"
        if ctrl != "OK":
            revisar += 1
        resumen.append([ct, nombres.get(ct, ""), b, y, b + y, total, ctrl])
        baseline_rows.append({
            "Comitente":  ct,
            "BROKER":     round(b, 2),
            "BYMA":       round(y, 2),
            "Total":      round(b + y, 2),
            "FechaCorte": fecha_corte.strftime("%d/%m/%Y"),
        })

    xlsx_bytes = _escribir_baseline_excel(resumen, excluidas, falta_definir, fecha_corte)
    csv_bytes  = _csv_write(baseline_rows, BASE_FIELDS)

    ui = {
        "fecha_corte":  fecha_corte.strftime("%d-%m-%Y"),
        "n_baseline":   len(baseline_rows),
        "n_ok":         len(baseline_rows) - revisar,
        "n_revisar":    revisar,
        "n_excluidas":  len(excluidas),
        "n_falta":      len(falta_definir),
        "excluidas":    [{"ctte": ct, "nombre": nom, "saldo": s}
                         for ct, nom, s in excluidas],
        "falta_definir": [{"ctte": ct, "nombre": nom, "saldo": s}
                          for ct, nom, s in falta_definir],
    }
    return xlsx_bytes, csv_bytes, ui


def _escribir_baseline_excel(resumen, excluidas, falta_definir, fecha_corte) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Baseline"
    ws["A1"] = f"SALDO DE PARTIDA USD CABLE POR VENUE — corte {fecha_corte:%d/%m/%Y}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:G1")
    ws["A2"] = ("Total = SALUSD (Saldo Vencido). Split BROKER/BYMA = LISTADO GALLO manual. "
                "Control valida que coincidan.")
    ws["A2"].font = Font(italic=True, size=9, color="595959")
    ws.merge_cells("A2:G2")

    hdr = ["Comitente", "Nombre", "Disponible BROKER", "Disponible BYMA",
           "Total Split", "Saldo SALUSD (|venc.|)", "Control"]
    hrow = 4
    for c, h in enumerate(hdr, 1):
        ws.cell(row=hrow, column=c, value=h)
    _style_header(ws, hrow, len(hdr))
    r = hrow + 1
    for row in resumen:
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            if c in (3, 4, 5, 6) and isinstance(v, (int, float)):
                cell.number_format = MONEY_FMT
        if row[-1] != "OK":
            ws.cell(row=r, column=7).fill = BAD_FILL
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for c in (3, 4, 5, 6):
        col = get_column_letter(c)
        cell = ws.cell(row=r, column=c, value=f"=SUM({col}{hrow+1}:{col}{r-1})")
        cell.number_format = MONEY_FMT
        cell.font = Font(bold=True)
        cell.fill = SUB_FILL
    for c in (1, 2, 7):
        ws.cell(row=r, column=c).fill = SUB_FILL
    _autofit(ws, len(hdr))
    ws.freeze_panes = f"A{hrow+1}"
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(hdr))}{r-1}"

    # Falta definir split
    wf = wb.create_sheet("Falta Definir Split")
    wf["A1"] = ("Comitentes en SALUSD (cliente) sin split en el LISTADO GALLO — "
                "DEFINIR BROKER/BYMA")
    wf["A1"].font = Font(bold=True, color="C00000")
    for c, h in enumerate(["Comitente", "Nombre", "Saldo Vencido"], 1):
        wf.cell(row=2, column=c, value=h)
    _style_header(wf, 2, 3)
    i = 3
    for ct, nom, s in falta_definir:
        wf.cell(row=i, column=1, value=ct).fill = WARN_FILL
        wf.cell(row=i, column=2, value=nom)
        cell = wf.cell(row=i, column=3, value=s); cell.number_format = MONEY_FMT
        i += 1
    if not falta_definir:
        wf.cell(row=3, column=1, value="(ninguno — todos definidos)")
    _autofit(wf, 3)

    # Excluidas
    we = wb.create_sheet("Excluidas (mercado)")
    we["A1"] = "Cuentas de SALUSD excluidas del split (mercado 999 / cartera propia 1000)"
    we["A1"].font = Font(bold=True)
    for c, h in enumerate(["Cuenta", "Nombre", "Saldo Vencido"], 1):
        we.cell(row=2, column=c, value=h)
    _style_header(we, 2, 3)
    i = 3
    for ct, nom, s in excluidas:
        we.cell(row=i, column=1, value=ct)
        we.cell(row=i, column=2, value=nom)
        cell = we.cell(row=i, column=3, value=s); cell.number_format = MONEY_FMT
        i += 1
    _autofit(we, 3)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Modo DIARIO ───────────────────────────────────────────────────────────────
def get_state_info():
    """Info descriptiva del estado bundled en disco (para la UI)."""
    def _stat(p):
        if not p.exists():
            return None
        return {
            "path":  str(p),
            "size":  p.stat().st_size,
            "mtime": dt.datetime.fromtimestamp(p.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),
        }
    return {
        "baseline":     _stat(BASELINE_PATH),
        "estado":       _stat(ESTADO_PATH),
        "asignaciones": _stat(DEC_PATH),
        "backdated":    _stat(BD_PATH),
    }


def read_state_file(kind):
    """Lee un archivo de estado bundled y devuelve sus bytes (para descargas)."""
    p = {"baseline": BASELINE_PATH, "estado": ESTADO_PATH,
         "asignaciones": DEC_PATH, "backdated": BD_PATH}.get(kind)
    if p and p.exists():
        return p.read_bytes()
    return b""


def procesar_diario(
    cc_files,                # list of UploadedFile, con .name para extraer CTTE
    tabcompb_file,
    salusd_file            = None,
    baseline_csv_file      = None,  # override opcional del baseline en disco
    estado_csv_file        = None,  # override opcional
    asignaciones_csv_file  = None,  # override opcional
    backdated_csv_file     = None,  # override opcional
    decisiones_ui          = None,   # dict {(ctte_str, cpbt, num_str): "BROKER"/"BYMA"}
    fecha_proc             = None,
    persist_to_disk        = True,   # escribir estado actualizado a STATE_DIR
):
    """
    Retorna:
      xlsx_bytes,
      estado_csv_bytes,       # actualizado
      asignaciones_csv_bytes, # actualizado (memoria BYMA O BROKER)
      backdated_csv_bytes,    # actualizado
      ui (dict con panel, pendientes, faltantes, alertas)
    """
    if fecha_proc is None:
        fecha_proc = dt.date.today()

    mp        = _load_tabcompb(tabcompb_file)
    # Baseline: primero el override (upload), sino el bundled en disco
    baseline  = _load_baseline(baseline_csv_file if baseline_csv_file is not None else BASELINE_PATH)
    if not baseline:
        raise ValueError(
            "No se encontró baseline_split.csv. Debería estar bundleado en "
            f"{BASELINE_PATH}. Contactar al admin de la app."
        )

    # fecha de corte del baseline (movimientos ESTRICTAMENTE posteriores)
    cutoff = None
    for v in baseline.values():
        cutoff = _parse_fecha(v["FechaCorte"])
        if cutoff:
            break
    if cutoff is None:
        raise ValueError("No se pudo parsear la FechaCorte del baseline.")
    ts_cut = pd.Timestamp(cutoff)

    estado_prev              = _load_estado(estado_csv_file if estado_csv_file is not None else ESTADO_PATH)
    decisiones, dec_records  = _load_decisiones(asignaciones_csv_file if asignaciones_csv_file is not None else DEC_PATH)
    backdated, bd_records    = _load_backdated(backdated_csv_file if backdated_csv_file is not None else BD_PATH)

    # Aplicar decisiones vía UI (segunda pasada). Si viene decisiones_ui,
    # las mergea a la memoria.
    if decisiones_ui:
        for (ct_s, cpbt, num_s), dest in decisiones_ui.items():
            dest = str(dest or "").strip().upper()
            if dest not in ("BROKER", "BYMA"):
                continue
            key = (ct_s, cpbt, num_s)
            decisiones[key] = dest
            dec_records.setdefault(key, {
                "Comitente": ct_s, "Cpbt": cpbt, "Numero": num_s,
                "Destino": dest, "Fecha": "", "Importe": "", "Referencia": "",
            })["Destino"] = dest

    salusd, nombres = ({}, {})
    sub_acreedores  = None
    if salusd_file is not None:
        try:
            salusd, nombres = _load_salusd(salusd_file)
        except Exception:
            pass
        # SUB.ACREEDORES: control global. Re-abrir el file para evitar side-effects
        # del seek en el read anterior.
        try:
            sub_acreedores = _load_sub_acreedores(salusd_file)
        except Exception:
            sub_acreedores = None

    # Mapear CCs por comitente
    cc_by_ctte = {}
    for f in (cc_files or []):
        name = getattr(f, "name", "") or ""
        try:
            cc_by_ctte[int(_comitente_from_name(name))] = f
        except ValueError:
            pass

    bd_detectados = []
    detalle_rows  = []
    panel         = []
    faltantes     = {}
    pendientes    = []
    recordadas = nuevas = 0
    nuevo_estado = {}

    def clasificar_mov(ct, r):
        nonlocal recordadas, nuevas
        cpbt = r["Cpbt"]
        dest = mp.get(cpbt)
        if dest is None:
            faltantes.setdefault(cpbt, set()).add(ct)
            bucket = "A DEFINIR"
        elif dest == "BYMA O BROKER":
            key = (str(ct), cpbt, _norm_num(r["Numero"]))
            if key in decisiones:
                bucket = decisiones[key]
                recordadas += 1
                # sobrescribir record con datos frescos del CC
                fe = r["Liquida"]
                dec_records[key] = {
                    "Comitente": str(ct), "Cpbt": cpbt,
                    "Numero": _norm_num(r["Numero"]), "Destino": bucket,
                    "Fecha": fe.strftime("%d/%m/%Y") if hasattr(fe, "strftime") else str(fe),
                    "Importe": f"{float(r['Importe']):.2f}",
                    "Referencia": str(r["Referencia"]),
                }
            else:
                bucket = "A DEFINIR"
                pendientes.append({
                    "ctte":       ct,
                    "fecha":      r["Liquida"],
                    "cpbt":       cpbt,
                    "numero":     r["Numero"],
                    "importe":    float(r["Importe"]),
                    "referencia": str(r["Referencia"]),
                    "key":        (str(ct), cpbt, _norm_num(r["Numero"])),
                })
        else:
            bucket = dest
        detalle_rows.append([ct, r["Liquida"], cpbt, r["Numero"],
                             float(r["Importe"]), r["Especie"], r["Referencia"],
                             mp.get(cpbt, "FALTANTE"), bucket])
        return bucket

    universo = set(baseline) | {c for c in salusd if not _es_no_cliente(c)}

    for ct in sorted(universo):
        base = baseline.get(ct, {"BROKER": 0.0, "BYMA": 0.0, "Total": 0.0})
        prev = estado_prev.get(ct)
        ref_prev = prev if prev else base
        prev_total = ref_prev["Total"]
        sal_total = abs(salusd[ct]) if ct in salusd else None
        nombre = nombres.get(ct, "")
        has_cc = ct in cc_by_ctte

        if has_cc:
            mov_all, _ = _load_cc(cc_by_ctte[ct])
            bd_set = set(backdated.get(ct, set()))
            inc_rows = [r for _, r in mov_all.iterrows()
                        if (r["Liquida"] > ts_cut)
                        or ((r["Cpbt"], _norm_num(r["Numero"])) in bd_set)]
            sum_inc = sum(float(r["Importe"]) for r in inc_rows)
            total_prov = base["Total"] - sum_inc

            # Auto-detección back-dated
            if sal_total is not None and abs(sal_total - total_prov) >= 0.01:
                objetivo = -(sal_total - total_prov)
                ya = {(r["Cpbt"], _norm_num(r["Numero"])) for r in inc_rows}
                cand = [r for _, r in mov_all.iterrows()
                        if r["Liquida"] <= ts_cut
                        and (r["Cpbt"], _norm_num(r["Numero"])) not in bd_set
                        and (r["Cpbt"], _norm_num(r["Numero"])) not in ya
                        and abs(float(r["Importe"]) - objetivo) < 0.01]
                if len(cand) == 1:
                    r = cand[0]
                    knum = _norm_num(r["Numero"])
                    inc_rows.append(r)
                    fe = r["Liquida"]
                    bd_records[(str(ct), r["Cpbt"], knum)] = {
                        "Comitente": str(ct), "Cpbt": r["Cpbt"], "Numero": knum,
                        "Fecha": fe.strftime("%d/%m/%Y") if hasattr(fe, "strftime") else str(fe),
                        "Importe": f"{float(r['Importe']):.2f}",
                        "Nota": "auto: cierra dif. vs SALUSD (posible dividendo)",
                    }
                    bd_detectados.append({
                        "ctte": ct, "cpbt": r["Cpbt"], "numero": r["Numero"],
                        "importe": float(r["Importe"]), "fecha": fe,
                    })

            inc_rows.sort(key=lambda r: r["Liquida"])
            inc = {"BROKER": 0.0, "BYMA": 0.0, "A DEFINIR": 0.0}
            for r in inc_rows:
                bucket = clasificar_mov(ct, r)
                inc[bucket] += float(r["Importe"])
            broker    = base["BROKER"] - inc["BROKER"]
            byma      = base["BYMA"]   - inc["BYMA"]
            adef      = -inc["A DEFINIR"]
            procesado = True
        else:
            broker, byma, adef = ref_prev["BROKER"], ref_prev["BYMA"], 0.0
            procesado = False

        total = broker + byma + adef
        if sal_total is None:
            estado_txt = "SIN SALUSD"
        elif abs(total - sal_total) < 0.01:
            estado_txt = "OK"
        elif not procesado:
            estado_txt = "OPERO - FALTA CC"
        else:
            estado_txt = "DISCREPANCIA"
        panel.append([ct, nombre, broker, byma, adef, total, prev_total,
                      sal_total, "SI" if procesado else "", estado_txt])
        # Regla anti-arrastre: solo se sobrescribe el estado si queda OK
        if estado_txt == "OK":
            nuevo_estado[ct] = {"BROKER": broker, "BYMA": byma, "Total": total}
        else:
            nuevo_estado[ct] = {"BROKER": ref_prev["BROKER"],
                                "BYMA":   ref_prev["BYMA"],
                                "Total":  ref_prev["Total"]}

    # Serializar outputs
    xlsx_bytes    = _escribir_diario_excel(panel, detalle_rows, faltantes,
                                           fecha_proc, cutoff, sub_acreedores)
    estado_rows   = [
        {"Comitente": ct,
         "BROKER":    round(v["BROKER"], 2),
         "BYMA":      round(v["BYMA"], 2),
         "Total":     round(v["Total"], 2),
         "FechaActualizacion": fecha_proc.strftime("%d/%m/%Y")}
        for ct, v in sorted(nuevo_estado.items())
    ]
    dec_rows = [dec_records[k] for k in sorted(dec_records)]
    bd_rows  = [bd_records[k]  for k in sorted(bd_records)]

    estado_bytes = _csv_write(estado_rows, ESTADO_FIELDS)
    dec_bytes    = _csv_write(dec_rows,    DEC_FIELDS)
    bd_bytes     = _csv_write(bd_rows,     BD_FIELDS)

    # Persistir en disco (state/) para que la próxima corrida arranque desde acá.
    # En Streamlit Cloud persiste durante la vida del container; si el container
    # se reinicia, vuelve al estado bundleado en el repo.
    if persist_to_disk:
        try:
            _csv_write_path(ESTADO_PATH, estado_rows, ESTADO_FIELDS)
            _csv_write_path(DEC_PATH,    dec_rows,    DEC_FIELDS)
            _csv_write_path(BD_PATH,     bd_rows,     BD_FIELDS)
        except OSError:
            # Filesystem read-only o similar → seguir con los bytes en memoria
            pass

    grand_total = sum(row[5] for row in panel)
    sub_ok = (sub_acreedores is not None
              and abs(sub_acreedores - grand_total) < 0.01)
    ui = {
        "fecha_proc":    fecha_proc.strftime("%d-%m-%Y"),
        "cutoff":        cutoff.strftime("%d-%m-%Y"),
        "n_panel":       len(panel),
        "grand_total":       grand_total,
        "sub_acreedores":    sub_acreedores,
        "sub_acreedores_ok": sub_ok,
        "sub_acreedores_dif": (sub_acreedores - grand_total)
                              if sub_acreedores is not None else None,
        "n_procesados":  sum(1 for p in panel if p[8] == "SI"),
        "n_ok":          sum(1 for p in panel if p[9] == "OK"),
        "n_falta_cc":    sum(1 for p in panel if p[9] == "OPERO - FALTA CC"),
        "n_discrep":     sum(1 for p in panel if p[9] == "DISCREPANCIA"),
        "n_sin_salusd":  sum(1 for p in panel if p[9] == "SIN SALUSD"),
        "n_asig_recordadas": recordadas,
        "n_asig_nuevas":     nuevas,   # (siempre 0 en modo web salvo si pasás decisiones_ui)
        "backdated":         bd_detectados,
        "falta_cc": [
            {"ctte": p[0], "nombre": p[1], "guardado": p[6], "salusd": p[7],
             "diff": (p[7] or 0) - p[6]}
            for p in panel if p[9] == "OPERO - FALTA CC"
        ],
        "discrepancias": [
            {"ctte": p[0], "nombre": p[1], "output": p[5], "salusd": p[7],
             "diff": p[5] - (p[7] or 0)}
            for p in panel if p[9] == "DISCREPANCIA"
        ],
        "faltantes_tabcompb": [
            {"cpbt": cpbt, "comitentes": sorted(list(cttes))}
            for cpbt, cttes in sorted(faltantes.items())
        ],
        "pendientes": pendientes,
        "panel":      panel,
    }
    return xlsx_bytes, estado_bytes, dec_bytes, bd_bytes, ui


def _escribir_diario_excel(panel, detalle_rows, faltantes, fecha_proc, cutoff,
                            sub_acreedores=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = f"SALDO USD CABLE DISPONIBLE POR VENUE — {fecha_proc:%d/%m/%Y}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:J1")
    ws["A2"] = (f"Baseline al {cutoff:%d/%m/%Y} + movimientos posteriores del CC. "
                "Control: Total output vs Saldo SALUSD del día. "
                "Solo se guarda el saldo cuando queda OK.")
    ws["A2"].font = Font(italic=True, size=9, color="595959")
    ws.merge_cells("A2:J2")

    # Fila 3: control SUB.ACREEDORES (SALUSD) vs TOTAL calculado.
    # Verde si coincide; rojo si hay diferencia (= algún comitente quedó deudor).
    grand_total = sum(row[5] for row in panel)
    ws["A3"] = "SUB.ACREEDORES"
    ws["A3"].font = Font(bold=True)
    if sub_acreedores is not None:
        ok = abs(sub_acreedores - grand_total) < 0.01
        cell = ws["C3"]
        cell.value = sub_acreedores
        cell.number_format = MONEY_FMT
        cell.font = Font(bold=True, color="008000" if ok else "FF0000")

    hdr = ["Comitente", "Nombre", "Disponible BROKER", "Disponible BYMA",
           "A DEFINIR", "Total", "Saldo día anterior", "Saldo SALUSD",
           "CC procesado", "Control"]
    hrow = 4
    for c, h in enumerate(hdr, 1):
        ws.cell(row=hrow, column=c, value=h)
    _style_header(ws, hrow, len(hdr))
    r = hrow + 1
    for row in panel:
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            if c in (3, 4, 5, 6, 7, 8) and isinstance(v, (int, float)):
                cell.number_format = MONEY_FMT
        est = row[9]
        if est == "OPERO - FALTA CC":
            ws.cell(row=r, column=10).fill = BAD_FILL
        elif est in ("DISCREPANCIA", "SIN SALUSD"):
            ws.cell(row=r, column=10).fill = WARN_FILL
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for c in (3, 4, 5, 6, 7, 8):
        col = get_column_letter(c)
        cell = ws.cell(row=r, column=c, value=f"=SUM({col}{hrow+1}:{col}{r-1})")
        cell.number_format = MONEY_FMT
        cell.font = Font(bold=True)
        cell.fill = SUB_FILL
    for c in (1, 2, 9, 10):
        ws.cell(row=r, column=c).fill = SUB_FILL
    _autofit(ws, len(hdr))
    ws.freeze_panes = f"A{hrow+1}"
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(hdr))}{r-1}"

    # Detalle
    wd = wb.create_sheet("Detalle")
    dhdr = ["Comitente", "Liquida", "Cpbt", "Numero", "Importe",
            "Especie", "Referencia", "Destino TABCOMPB", "Bucket"]
    for c, h in enumerate(dhdr, 1):
        wd.cell(row=1, column=c, value=h)
    _style_header(wd, 1, len(dhdr))
    rr = 2
    for row in detalle_rows:
        for c, v in enumerate(row, 1):
            cell = wd.cell(row=rr, column=c, value=v)
            cell.border = BORDER
            if c == 2 and hasattr(v, "strftime"):
                cell.number_format = "dd/mm/yyyy"
            if c == 5 and isinstance(v, (int, float)):
                cell.number_format = MONEY_FMT
        if row[7] == "FALTANTE":
            wd.cell(row=rr, column=8).fill = BAD_FILL
        elif row[8] == "A DEFINIR":
            wd.cell(row=rr, column=9).fill = WARN_FILL
        rr += 1
    if not detalle_rows:
        wd.cell(row=2, column=1, value="(sin movimientos posteriores al corte)")
    _autofit(wd, len(dhdr))
    wd.freeze_panes = "A2"
    if detalle_rows:
        wd.auto_filter.ref = f"A1:{get_column_letter(len(dhdr))}{rr-1}"

    if faltantes:
        wf = wb.create_sheet("Faltantes TABCOMPB")
        wf.cell(row=1, column=1, value="Comprobante")
        wf.cell(row=1, column=2, value="Comitentes")
        _style_header(wf, 1, 2)
        i = 2
        for cpbt, cttes in sorted(faltantes.items()):
            wf.cell(row=i, column=1, value=cpbt).fill = WARN_FILL
            wf.cell(row=i, column=2, value=", ".join(str(c) for c in sorted(cttes)))
            i += 1
        _autofit(wf, 2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
