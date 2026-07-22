# -*- coding: utf-8 -*-
"""
Posicion Aforada por Comitente — logica (versión web).

Adapta el script standalone run_posicion_aforada.py para inputs vía UploadedFile
y outputs bytes.

Expone:
  - parse_all(tenaforada_file, especies_file, fecha_input=None)
      → (comitentes_all, ticker_map, fecha_input)
  - generar_reporte_bytes(comitentes_all, ticker_map, fecha_input)
      → xlsx_bytes  (Excel con filtro saldo>100k y no cartera propia)
  - filtrar_seleccionados(comitentes_all)
      → sel  (aplica el filtro estándar de saldo mínimo y cartera propia)
  - consultar_comitente(comitentes_all, ctte)
      → dict o None  (consulta puntual sin filtro)
"""
from __future__ import annotations

from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


# ── Constantes ────────────────────────────────────────────────────────────────
SHEET_IN       = "Tenencia_Aforada"
SALDO_MINIMO   = 100_000                        # solo comitentes con saldo deudor > 100k
CARTERA_PROPIA = {1000, 1001, 1002, 1003}       # cuentas de cartera propia a excluir
MONEY_FMT      = "#,##0.00"
QTY_FMT        = "#,##0"
PCT_FMT        = "0"


# ── Helpers ───────────────────────────────────────────────────────────────────
def _num(v):
    """Convierte a float de forma robusta (soporta formato argentino 1.234.567,89)."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return 0.0
    s = str(v).strip().strip("'")
    if s == "" or s.lower() == "nan":
        return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _clean(v):
    if v is None:
        return ""
    s = str(v).strip().strip("'")
    return "" if s.lower() == "nan" else s


def _read_bytes(f) -> bytes:
    f.seek(0)
    return f.read()


# ── Tickers desde ESPECIES.XLS ────────────────────────────────────────────────
def _load_tickers(especies_file):
    """Mapa CVSA (5 dígitos, zero-pad) → ticker Normal (col 9 'Norm.')."""
    tk = {}
    if especies_file is None:
        return tk
    try:
        df = pd.read_excel(BytesIO(_read_bytes(especies_file)),
                           sheet_name="Datos_Fijos_Especies", header=None)
    except Exception:
        return tk
    for _, row in df.iterrows():
        cod = _clean(row[0])
        if cod == "" or cod == "Codigo":
            continue
        cod5 = cod.zfill(5)
        ticker = _clean(row[9])
        if ticker:
            tk[cod5] = ticker
    return tk


# ── Parse TENAFORADA.XLS ──────────────────────────────────────────────────────
def _parse_tenaforada(tenaforada_file):
    """Devuelve la lista de todos los comitentes parseados."""
    df = pd.read_excel(BytesIO(_read_bytes(tenaforada_file)),
                       sheet_name=SHEET_IN, header=None)

    comitentes = []
    current = None

    def add_especie(c, cod, especie, tenencia, precio, aforo_pct, valor_aforado):
        if cod != "" and tenencia > 0:
            c["especies"].append(dict(
                cod=cod, especie=especie, tenencia=tenencia,
                precio=precio, aforo_pct=aforo_pct, valor_aforado=valor_aforado,
            ))

    for _, row in df.iterrows():
        c0_raw = _clean(row[0])
        if c0_raw == "Numero":
            continue
        if c0_raw.lower().startswith("total gene"):
            break

        cod           = _clean(row[2])
        especie       = _clean(row[3])
        tenencia      = _num(row[4])
        precio        = _num(row[5])
        aforo_pct     = _num(row[6])
        valor_aforado = _num(row[7])
        saldo         = _num(row[8])

        if c0_raw != "" and c0_raw.isdigit():
            ctte = int(c0_raw)
            if current is None or current["ctte"] != ctte:
                current = {
                    "ctte":          ctte,
                    "nombre":        _clean(row[1]),
                    "saldo":         0.0,
                    "aforado_total": 0.0,
                    "especies":      [],
                }
                comitentes.append(current)
                # sin tenencias → saldo/aforado en la propia fila del comitente
                if not (cod != "" and tenencia > 0):
                    current["saldo"] = saldo
                    current["aforado_total"] = valor_aforado
            add_especie(current, cod, especie, tenencia, precio, aforo_pct, valor_aforado)
        elif c0_raw == "" and current is not None:
            # fila subtotal: col 0 vacía trae Valor Aforado total (col 7) y Saldo (col 8)
            current["saldo"] = saldo
            current["aforado_total"] = valor_aforado

    # Calcular diferencia
    for c in comitentes:
        c["diferencia"] = c["aforado_total"] - c["saldo"]

    return comitentes


# ── API pública ───────────────────────────────────────────────────────────────
def parse_all(tenaforada_file, especies_file, fecha_input=None):
    """Parsea TENAFORADA + ESPECIES. Devuelve (comitentes_all, ticker_map, fecha_input)."""
    comitentes = _parse_tenaforada(tenaforada_file)
    tickers    = _load_tickers(especies_file)
    if fecha_input is None:
        fecha_input = datetime.now()
    return comitentes, tickers, fecha_input


def filtrar_seleccionados(comitentes_all):
    """Aplica el filtro estándar del reporte: saldo > 100k y no cartera propia.
    Ordena por diferencia ascendente (descubiertos primero)."""
    sel = [c for c in comitentes_all
           if c["saldo"] > SALDO_MINIMO and c["ctte"] not in CARTERA_PROPIA]
    sel.sort(key=lambda c: c["diferencia"])
    return sel


def consultar_comitente(comitentes_all, ctte):
    """Devuelve el dict de un comitente por número, o None."""
    try:
        ctte_int = int(ctte)
    except (ValueError, TypeError):
        return None
    for c in comitentes_all:
        if c["ctte"] == ctte_int:
            return c
    return None


# ── Excel ─────────────────────────────────────────────────────────────────────
def generar_reporte_bytes(comitentes_all, ticker_map, fecha_input):
    """Genera el Excel con una fila por comitente (filtro estándar).
    Detalle de especies agrupado y colapsado. Devuelve bytes.
    """
    sel = filtrar_seleccionados(comitentes_all)

    wb = Workbook()
    ws = wb.active
    ws.title = "Posicion Aforada"
    ws.sheet_properties.outlinePr.summaryBelow = False   # resumen ARRIBA del detalle

    # estilos
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr_fill   = PatternFill("solid", fgColor="2E75B6")
    ctte_font  = Font(name="Calibri", size=10, bold=True)
    det_font   = Font(name="Calibri", size=9, color="404040")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(name="Calibri", size=10, bold=True, color="006100")
    red_fill   = PatternFill("solid", fgColor="FFC7CE")
    red_font   = Font(name="Calibri", size=10, bold=True, color="9C0006")
    thin       = Side(style="thin", color="D9D9D9")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    right      = Alignment(horizontal="right")
    left       = Alignment(horizontal="left")

    headers  = ["Comitente", "Nombre / Especie", "Tenencia", "Precio",
                "Aforo %", "Saldo Deudor", "Posicion Garantizable", "Diferencia"]
    ncol     = len(headers)
    last_col = get_column_letter(ncol)

    # fila 1: título
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "POSICION AFORADA POR COMITENTE"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # fila 2: reporte + fecha/hora
    ws.merge_cells(f"A2:{last_col}2")
    label_font = InlineFont(rFont="Calibri", sz=11, b=True, color="1F4E78")
    date_font  = InlineFont(rFont="Calibri", sz=11, b=True, color="FF0000")
    ws["A2"] = CellRichText(
        TextBlock(label_font,
                  "Reporte de saldo DEUDOR y Tenencias aceptadas en garantía a fecha: "),
        TextBlock(date_font, fecha_input.strftime("%d/%m/%Y %H:%M")),
    )
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # fila 3
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = "Desplegar (+) para ver los activos que componen la garantía"
    ws["A3"].font = Font(name="Calibri", size=9, italic=True, color="595959")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 16

    # fila 4: headers
    HDR = 4
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=HDR, column=j, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[HDR].height = 20

    r = HDR + 1
    tot_saldo = tot_afor = tot_dif = 0.0

    for c in sel:
        # fila resumen
        ws.cell(row=r, column=1, value=c["ctte"]).font = ctte_font
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=2, value=c["nombre"]).font = ctte_font
        ws.cell(row=r, column=2).alignment = left
        cs = ws.cell(row=r, column=6, value=round(c["saldo"], 2))
        cg = ws.cell(row=r, column=7, value=round(c["aforado_total"], 2))
        cd = ws.cell(row=r, column=8, value=round(c["diferencia"], 2))
        for cc in (cs, cg):
            cc.number_format = MONEY_FMT
            cc.font = ctte_font
            cc.alignment = right
        cd.number_format = MONEY_FMT
        cd.alignment = right
        if c["diferencia"] >= 0:
            cd.fill = green_fill
            cd.font = green_font
        else:
            cd.fill = red_fill
            cd.font = red_font
        for j in range(1, ncol + 1):
            ws.cell(row=r, column=j).border = border
        for j in range(1, 6):
            if ws.cell(row=r, column=j).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor="EAF1F8")
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor="EAF1F8")
        ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor="EAF1F8")

        tot_saldo += c["saldo"]
        tot_afor  += c["aforado_total"]
        tot_dif   += c["diferencia"]
        r += 1

        # filas detalle (agrupadas, colapsadas)
        for e in c["especies"]:
            ticker = ticker_map.get(str(e["cod"]).zfill(5), "")
            tk = ws.cell(row=r, column=1, value=ticker)
            tk.font = Font(name="Calibri", size=9, bold=True, color="1F4E78")
            tk.alignment = center
            ws.cell(row=r, column=2, value=f"   {e['cod']}  {e['especie']}").font = det_font
            ws.cell(row=r, column=2).alignment = left
            ce = ws.cell(row=r, column=3, value=round(e["tenencia"], 2))
            ce.number_format = QTY_FMT
            cp = ws.cell(row=r, column=4, value=round(e["precio"], 4))
            cp.number_format = MONEY_FMT
            ca = ws.cell(row=r, column=5, value=round(e["aforo_pct"], 0))
            ca.number_format = PCT_FMT
            cv = ws.cell(row=r, column=7, value=round(e["valor_aforado"], 2))
            cv.number_format = MONEY_FMT
            for j in (3, 4, 5, 7):
                ws.cell(row=r, column=j).font = det_font
                ws.cell(row=r, column=j).alignment = right
            for j in range(1, ncol + 1):
                ws.cell(row=r, column=j).border = border
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = True
            r += 1

    # fila TOTAL SELECCION
    r += 1
    ws.cell(row=r, column=2, value="TOTAL SELECCION").font = Font(bold=True, size=10)
    ws.cell(row=r, column=2).alignment = right
    for col, val in ((6, tot_saldo), (7, tot_afor), (8, tot_dif)):
        cell = ws.cell(row=r, column=col, value=round(val, 2))
        cell.number_format = MONEY_FMT
        cell.font = Font(bold=True, size=10)
        cell.alignment = right
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="D9E1F2")
    for j in range(1, ncol + 1):
        ws.cell(row=r, column=j).border = border

    widths = {"A": 11, "B": 34, "C": 15, "D": 13, "E": 8,
              "F": 18, "G": 20, "H": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{HDR + 1}"
    ws.auto_filter.ref = f"A{HDR}:{last_col}{HDR}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
