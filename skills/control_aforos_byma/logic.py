"""
Control Aforos BYMA vs Gallo
Lee haircut desde col 26 (Aforo) de ESPECIES.XLS — informado por la API BYMA via Gallo.
Aforo BYMA = 100 - haircut  (ej. haircut 15% -> aforo 85%)
El PDF de circular BYMA ya no se utiliza.

Genera dos Excel:
  1. DIFERENCIAS AFOROS.xlsx       — inconsistencias de aforo / sin lista / lista sin BYMA
  2. Reporte Comercial ...xlsx      — todas las especies BYMA agrupadas por categoria + buscador
"""

import io
import re
import datetime

import xlrd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Tablas Lista <-> Aforo BYMA ─────────────────────────────────────────────
TABLAS = {
    "Renta Variable": {
        1: 85, 2: 80, 3: 75, 4: 70, 5: 60, 6: 50, 7: 40, 8: 30,
    },
    "Renta Fija Publicos": {
        10: 95, 11: 90, 12: 85, 13: 80, 14: 75, 15: 70, 16: 65, 17: 60,
    },
    "Renta Fija Privados": {
        22: 85, 23: 80, 24: 75, 25: 70, 26: 65, 27: 60,
    },
    "Letras y Bonos del tesoro": {
        85: 85, 90: 90, 95: 95,
    },
}


def _tabla_para_lista(lista):
    if 1 <= lista <= 8:       return "Renta Variable"
    if 10 <= lista <= 17:     return "Renta Fija Publicos"
    if 22 <= lista <= 27:     return "Renta Fija Privados"
    if lista in (85, 90, 95): return "Letras y Bonos del tesoro"
    return None


def _tabla_para_tipo(tipo_col2, tipo_activo):
    t2 = (tipo_col2 or "").strip().upper()
    ta = (tipo_activo or "").lower()
    if t2 == "LETR":
        return "Letras y Bonos del tesoro"
    if t2 == "O/N":
        return "Renta Fija Privados"
    if t2 == "PUBL":
        if any(x in ta for x in ["letra", "tesoro", "17-letra"]):
            return "Letras y Bonos del tesoro"
        return "Renta Fija Publicos"
    if t2 in ("PRIV", "FDO."):
        return "Renta Variable"
    if t2 == "FIDE":
        return "Renta Fija Publicos"
    if any(x in ta for x in ["obliga", "negoci", "05-obli"]):
        return "Renta Fija Privados"
    if any(x in ta for x in ["letra", "tesoro"]):
        return "Letras y Bonos del tesoro"
    if any(x in ta for x in ["publi", "bono", "provincial", "03-titulo", "titulos de deuda", "19-titulo"]):
        return "Renta Fija Publicos"
    return "Renta Variable"


def _aforo_para_lista(tabla, lista):
    return TABLAS.get(tabla, {}).get(lista)


def _lista_para_aforo(tabla, aforo_pct):
    for l, a in TABLAS.get(tabla, {}).items():
        if a == aforo_pct:
            return l
    return "REVISAR"


# ─── Helpers xlrd ────────────────────────────────────────────────────────────
def _xls_int(cell):
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return int(cell.value)
    if cell.ctype == xlrd.XL_CELL_TEXT:
        v = str(cell.value).strip().lstrip("'")
        try:
            return int(float(v))
        except ValueError:
            return None
    return None


def _xls_str(cell):
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    return str(cell.value).strip()


_DATE_RE = re.compile(r'\b(\d{2})[/\-](\d{2})[/\-](\d{2,4})\b')


def _fecha_vencimiento(wb_xls, ws_xls, row_idx):
    """Devuelve date de vencimiento o None."""
    cell_v = ws_xls.cell(row_idx, 15)
    if cell_v.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell_v.value, wb_xls.datemode).date()
        except Exception:
            pass
    elif cell_v.ctype == xlrd.XL_CELL_NUMBER and cell_v.value > 1000:
        try:
            return xlrd.xldate_as_datetime(cell_v.value, wb_xls.datemode).date()
        except Exception:
            pass
    elif cell_v.ctype == xlrd.XL_CELL_TEXT:
        s = str(cell_v.value).strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    nombre = _xls_str(ws_xls.cell(row_idx, 1))
    m = _DATE_RE.search(nombre)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return datetime.date(y, mo, d)
        except ValueError:
            pass
    return None


def generar_control(especies_file):
    """
    Parametros
    ----------
    especies_file : file-like  — ESPECIES.XLS

    Retorna
    -------
    xlsx_dif_bytes : bytes     Excel DIFERENCIAS AFOROS
    xlsx_rc_bytes  : bytes     Excel Reporte Comercial Garantias BYMA
    resumen        : dict      Conteos para la UI
    advertencias   : list[str] Avisos de procesamiento
    """
    HOY = datetime.date.today()
    advertencias = []

    # ── Leer ESPECIES.XLS ────────────────────────────────────────────────────
    especies_file.seek(0)
    wb_xls = xlrd.open_workbook(file_contents=especies_file.read())
    hoja = ("Datos_Fijos_Especies"
            if "Datos_Fijos_Especies" in wb_xls.sheet_names()
            else wb_xls.sheet_names()[0])
    ws_xls = wb_xls.sheet_by_name(hoja)

    # Mapa de columnas por encabezado (fallback a indices fijos)
    COL_H = {str(ws_xls.cell_value(0, c)).strip("'"): c for c in range(ws_xls.ncols)}

    diferencias    = []
    sin_lista      = []
    lista_sin_byma = []
    rc_rows        = []   # para el Reporte Comercial
    vencidos_filtrados = 0
    total_byma = 0
    total_ok   = 0

    for r in range(1, ws_xls.nrows):
        cell_a = ws_xls.cell(r, 0)
        if cell_a.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            continue

        cvsa = _xls_int(cell_a)
        if cvsa is None:
            continue

        cvsa_str    = str(cvsa).zfill(5)
        ticker_norm = _xls_str(ws_xls.cell(r, 9))    # col Norm.
        nombre      = _xls_str(ws_xls.cell(r, 1))     # Nombre_de_la_Especie
        tipo_col2   = _xls_str(ws_xls.cell(r, 2))     # Tipo
        tipo_activo = _xls_str(ws_xls.cell(r, 18))    # Tipo_de_Activo
        lista       = _xls_int(ws_xls.cell(r, 5)) or 0

        # Campos extra para Reporte Comercial
        ticker_mep  = _xls_str(ws_xls.cell(r, COL_H.get("Parid", 10)))
        ticker_cbl  = _xls_str(ws_xls.cell(r, COL_H.get("Cable", 11)))
        emisor      = _xls_str(ws_xls.cell(r, COL_H.get("Emisor", 17)))
        precio_tipo = _xls_str(ws_xls.cell(r, 4))  # col Precio (Normal/Porc.)
        vencim_raw  = _xls_str(ws_xls.cell(r, 15))  # col Vencim

        # Col 26 = haircut BYMA API
        try:
            raw26 = ws_xls.cell(r, 26)
            haircut = _xls_int(raw26) or 0
        except IndexError:
            haircut = 0

        ticker = ticker_norm or str(cvsa)

        if haircut > 0:
            total_byma += 1
            aforo_byma = 100 - haircut

            # ── Acumular para Reporte Comercial ──────────────────────────────
            rc_rows.append({
                "cvsa": cvsa_str, "tipo_activo": tipo_activo, "precio": precio_tipo,
                "ticker_ars": ticker_norm, "ticker_mep": ticker_mep,
                "ticker_cbl": ticker_cbl, "nombre": nombre, "emisor": emisor,
                "aforo": aforo_byma, "haircut": haircut, "vencim": vencim_raw,
                "tiene_ars": bool(ticker_norm), "tiene_mep": bool(ticker_mep),
                "tiene_cbl": bool(ticker_cbl),
            })

            # ── Control de aforos ────────────────────────────────────────────
            if lista == 0:
                tabla_inf = _tabla_para_tipo(tipo_col2, tipo_activo)
                lista_sug = _lista_para_aforo(tabla_inf, aforo_byma) if tabla_inf else "REVISAR"
                sin_lista.append({
                    "cvsa": cvsa, "ticker": ticker, "nombre": nombre,
                    "tipo_activo": tipo_activo, "tipo_col2": tipo_col2,
                    "haircut": haircut, "aforo_byma": aforo_byma,
                    "lista_sugerida": lista_sug,
                })
            else:
                tabla = _tabla_para_lista(lista)
                if tabla is None:
                    advertencias.append(
                        f"Lista {lista} fuera de rango — CVSA {cvsa} ({ticker}), "
                        f"haircut={haircut}%. Especie omitida del control."
                    )
                    continue

                aforo_sailing = _aforo_para_lista(tabla, lista)
                if aforo_sailing is None:
                    advertencias.append(
                        f"Lista {lista} no mapeada en tabla '{tabla}' — CVSA {cvsa}. Omitida."
                    )
                    continue

                if aforo_sailing == aforo_byma:
                    total_ok += 1
                else:
                    lista_sug = _lista_para_aforo(tabla, aforo_byma)
                    diferencias.append({
                        "cvsa": cvsa, "ticker": ticker, "nombre": nombre,
                        "tipo_activo": tipo_activo, "tabla": tabla,
                        "haircut": haircut, "aforo_byma": aforo_byma,
                        "lista_actual": lista, "aforo_sailing": aforo_sailing,
                        "diferencia_pp": aforo_byma - aforo_sailing,
                        "lista_sugerida": lista_sug,
                    })
                    if lista_sug == "REVISAR":
                        advertencias.append(
                            f"Lista sugerida REVISAR — CVSA {cvsa} ({ticker}), "
                            f"aforo BYMA={aforo_byma}%, tabla '{tabla}'."
                        )

        elif lista > 0:
            fvenc = _fecha_vencimiento(wb_xls, ws_xls, r)
            if fvenc is not None and fvenc < HOY:
                vencidos_filtrados += 1
                continue
            lista_sin_byma.append({
                "cvsa": cvsa, "ticker": ticker, "nombre": nombre,
                "tipo_activo": tipo_activo, "lista": lista,
            })

    # ════════════════════════════════════════════════════════════════════════════
    # Excel 1 — DIFERENCIAS AFOROS.xlsx
    # ════════════════════════════════════════════════════════════════════════════
    AZUL_FILL    = PatternFill("solid", fgColor="4472C4")
    ROJO_FILL    = PatternFill("solid", fgColor="C00000")
    NARANJA_FILL = PatternFill("solid", fgColor="ED7D31")
    VERDE_FILL   = PatternFill("solid", fgColor="00B050")
    ROJO_CELDA   = PatternFill("solid", fgColor="FF0000")
    BLANCO_FONT  = Font(name="Arial", bold=True, color="FFFFFF")
    NORMAL_FONT  = Font(name="Arial", size=11)
    BORDER_THIN  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")

    def _hdr(ws, headers, fill=None):
        f = fill or AZUL_FILL
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = f; cell.font = BLANCO_FONT
            cell.alignment = CENTER; cell.border = BORDER_THIN

    def _row(ws, fila, valores, fills=None, fonts=None):
        for col, v in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=col, value=v)
            cell.font      = (fonts or {}).get(col, NORMAL_FONT)
            cell.alignment = CENTER; cell.border = BORDER_THIN
            if fills and fills.get(col):
                cell.fill = fills[col]

    def _widths(ws, col_widths):
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    wb_dif = openpyxl.Workbook()

    ws_dif = wb_dif.active
    ws_dif.title = "Diferencias Aforo"
    _hdr(ws_dif, [
        "Ticker", "Nombre", "Tipo Activo", "Codigo CVSA",
        "Aforo BYMA %", "Lista Actual (Gallo)", "Aforo Lista Actual %",
        "Diferencia (pp)", "LISTA SUGERIDA",
    ])
    ws_dif.cell(1, 9).fill = VERDE_FILL
    for i, d in enumerate(diferencias, 2):
        diff_str = (f"+{d['diferencia_pp']}" if d["diferencia_pp"] > 0
                    else str(d["diferencia_pp"])) + "pp"
        _row(ws_dif, i, [
            d["ticker"], d["nombre"], d["tipo_activo"], d["cvsa"],
            d["aforo_byma"], d["lista_actual"], d["aforo_sailing"],
            diff_str, d["lista_sugerida"],
        ], fills={7: ROJO_CELDA, 9: VERDE_FILL},
           fonts={9: Font(name="Arial", color="FFFFFF", bold=True)})
    _widths(ws_dif, [12, 30, 25, 14, 15, 22, 22, 16, 16])

    ws_sl = wb_dif.create_sheet("Sin Lista Gallo")
    _hdr(ws_sl, [
        "Ticker", "Nombre", "Tipo Activo", "Codigo CVSA",
        "Haircut BYMA %", "Aforo BYMA %", "LISTA SUGERIDA",
    ], fill=ROJO_FILL)
    ws_sl.cell(1, 7).fill = VERDE_FILL
    for i, esp in enumerate(sin_lista, 2):
        _row(ws_sl, i, [
            esp["ticker"], esp["nombre"], esp["tipo_activo"], esp["cvsa"],
            esp["haircut"], esp["aforo_byma"], esp["lista_sugerida"],
        ], fills={7: VERDE_FILL},
           fonts={7: Font(name="Arial", color="FFFFFF", bold=True)})
    _widths(ws_sl, [12, 30, 25, 14, 16, 15, 16])

    ws_lb = wb_dif.create_sheet("Lista Sin BYMA")
    _hdr(ws_lb, ["Ticker", "Nombre", "Tipo Activo", "Codigo CVSA", "Lista Gallo"],
         fill=NARANJA_FILL)
    for i, esp in enumerate(lista_sin_byma, 2):
        _row(ws_lb, i, [
            esp["ticker"], esp["nombre"], esp["tipo_activo"], esp["cvsa"], esp["lista"],
        ])
    _widths(ws_lb, [12, 30, 25, 14, 14])

    buf_dif = io.BytesIO()
    wb_dif.save(buf_dif)
    xlsx_dif_bytes = buf_dif.getvalue()

    # ════════════════════════════════════════════════════════════════════════════
    # Excel 2 — Reporte Comercial Garantias BYMA DD-MM-AAAA.xlsx
    # ════════════════════════════════════════════════════════════════════════════

    # ── Colores reporte comercial ────────────────────────────────────────────
    RC_HEADER  = "1F4E79"
    RC_SUBHDR  = "2E75B6"
    RC_INPUT   = "FFFF00"
    RC_RESULT  = "DEEAF1"
    RC_WHITE   = "FFFFFF"

    def rc_fill(hex_c): return PatternFill("solid", fgColor=hex_c)
    def rc_border_thin():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def rc_aforo_color(a):
        if a >= 95: return "E2EFDA"
        if a >= 85: return "EBF3E8"
        if a >= 75: return "FFF2CC"
        if a >= 65: return "FCE4D6"
        return "F2DCDB"

    def rc_hdr_cell(ws, row, col, value, bg=RC_HEADER, fg="FFFFFF", bold=True):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = rc_fill(bg); c.font = Font(bold=bold, color=fg, name="Calibri", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = rc_border_thin(); return c

    def rc_data_cell(ws, row, col, value, bg=RC_WHITE, bold=False, center=False):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = rc_fill(bg); c.font = Font(bold=bold, name="Calibri", size=9)
        c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        c.border = rc_border_thin(); return c

    # ── Clasificacion en grupos comerciales ──────────────────────────────────
    _ACCIONES = {"Titulos Privados"}
    _CEDEAR   = {"Cert.Deposito Argentino (Cedea", "23-CEDEARS",
                 "Cert.de Depositos de Acciones", "CEDEAR Obligaciones"}
    _TPUB     = {"03-TITULOS PUBLICOS", "Titulos Publicos", "Titulos Publico",
                 "Titulos Provinciales", "Bono", "19-TITULOS DE DEUDA"}
    _ON       = {"05-OBLI. NEGOCIABLES", "Obligaciones Negociables"}
    _LETRAS   = {"17-LETRAS TESORO NAC", "13-LETRAS", "Letras del Banco Central"}
    _FCI      = {"18-FONDOS INVERSION", "Fondos Comunes de Inversion"}

    RC_GRUPO_LABEL = {
        "Titulos Publicos":         "Títulos Públicos",
        "Letras":                   "Letras del Tesoro",
        "Obligaciones Negociables": "Obligaciones Negociables",
        "Acciones":                 "Acciones",
        "CEDEARs":                  "CEDEARs",
        "FCI":                      "Fondos Comunes de Inversión",
        "Otros":                    "Otros",
    }
    RC_GRUPOS_ORDER = ["Titulos Publicos", "Letras", "Obligaciones Negociables",
                       "Acciones", "CEDEARs", "FCI", "Otros"]
    RC_TIENE_VENCIM = {"Titulos Publicos", "Letras", "Obligaciones Negociables", "Otros"}

    def rc_clasificar(ta):
        if ta in _ACCIONES: return "Acciones"
        if ta in _CEDEAR:   return "CEDEARs"
        if ta in _TPUB:     return "Titulos Publicos"
        if ta in _ON:       return "Obligaciones Negociables"
        if ta in _LETRAS:   return "Letras"
        if ta in _FCI:      return "FCI"
        return "Otros"

    for r in rc_rows:
        r["grupo"] = rc_clasificar(r["tipo_activo"])

    rc_grupos = {g: [] for g in RC_GRUPOS_ORDER}
    for r in rc_rows:
        rc_grupos[r["grupo"]].append(r)
    for g in rc_grupos:
        rc_grupos[g].sort(key=lambda x: (-x["aforo"], x["nombre"]))

    # ── Formulas INDEX/MATCH para buscador ───────────────────────────────────
    RC_LK_ROWS = len(rc_rows) + 1

    def rc_lk_formula(input_cell, result_col_letter):
        n = RC_LK_ROWS
        rng_ars  = f"_Lookup!$B$2:$B${n}"
        rng_mep  = f"_Lookup!$C$2:$C${n}"
        rng_cbl  = f"_Lookup!$D$2:$D${n}"
        rng_cvsa = f"_Lookup!$A$2:$A${n}"
        res      = f"_Lookup!${result_col_letter}$2:${result_col_letter}${n}"
        def im(srng, lval): return f"INDEX({res},MATCH({lval},{srng},0))"
        up = f"UPPER({input_cell})"
        pd = f'TEXT(VALUE({input_cell}),"00000")'
        return ("=IFERROR(" + im(rng_ars, up) +
                ",IFERROR(" + im(rng_mep, up) +
                ",IFERROR(" + im(rng_cbl, up) +
                ",IFERROR(" + im(rng_cvsa, up) +
                ",IFERROR(" + im(rng_cvsa, pd) +
                ',"No encontrado")))))')

    # ── Workbook Reporte Comercial ────────────────────────────────────────────
    rc_wb = openpyxl.Workbook()
    rc_wb.remove(rc_wb.active)

    # Hoja oculta _Lookup
    ws_lk = rc_wb.create_sheet("_Lookup")
    ws_lk.sheet_state = "hidden"
    ws_lk.append(["CVSA", "Ticker_ARS", "Ticker_MEP", "Ticker_Cable",
                  "Nombre", "Aforo", "Tipo_de_activo"])
    for r in rc_rows:
        ws_lk.append([r["cvsa"], r["ticker_ars"], r["ticker_mep"], r["ticker_cbl"],
                      r["nombre"], r["aforo"], RC_GRUPO_LABEL[r["grupo"]]])

    # ── Hoja Resumen ──────────────────────────────────────────────────────────
    ws_res = rc_wb.create_sheet("Resumen")
    ws_res.sheet_view.showGridLines = False

    ws_res.merge_cells("A1:G1")
    c = ws_res["A1"]
    c.value = f"Activos Aceptados en Garantía — BYMA Clearing   |   Al {HOY:%d/%m/%Y}"
    c.fill = rc_fill(RC_HEADER)
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[1].height = 28

    ws_res.merge_cells("A2:G2")
    c = ws_res["A2"]
    c.value = "Aforo = porcentaje del valor de mercado que BYMA reconoce como garantía  |  Aforo = 100% − Haircut BYMA"
    c.fill = rc_fill("EBF3E8")
    c.font = Font(italic=True, color="374151", name="Calibri", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[2].height = 16
    ws_res.row_dimensions[3].height = 8

    hdr_row = 4
    rc_hdr_cell(ws_res, hdr_row, 1, "Categoría",               bg=RC_SUBHDR)
    rc_hdr_cell(ws_res, hdr_row, 2, "Cantidad de especies",    bg=RC_HEADER)
    rc_hdr_cell(ws_res, hdr_row, 3, "Aforo mínimo",            bg=RC_HEADER)
    rc_hdr_cell(ws_res, hdr_row, 4, "Aforo máximo",            bg=RC_HEADER)
    rc_hdr_cell(ws_res, hdr_row, 5, "Disponible en ARS",       bg=RC_HEADER)
    rc_hdr_cell(ws_res, hdr_row, 6, "Disponible en USD MEP",   bg=RC_HEADER)
    rc_hdr_cell(ws_res, hdr_row, 7, "Disponible en USD Cable", bg=RC_HEADER)
    ws_res.row_dimensions[hdr_row].height = 32

    r_idx = hdr_row + 1
    for g in RC_GRUPOS_ORDER:
        items = rc_grupos[g]
        if not items:
            continue
        aforos = sorted(set(x["aforo"] for x in items))
        row_bg = "F7FAFD"
        rc_data_cell(ws_res, r_idx, 1, RC_GRUPO_LABEL[g], bg=row_bg, bold=True)
        rc_data_cell(ws_res, r_idx, 2, len(items),         bg=row_bg, center=True)
        rc_data_cell(ws_res, r_idx, 3, f"{min(aforos)}%",  bg=rc_aforo_color(min(aforos)), center=True, bold=True)
        rc_data_cell(ws_res, r_idx, 4, f"{max(aforos)}%",  bg=rc_aforo_color(max(aforos)), center=True, bold=True)
        rc_data_cell(ws_res, r_idx, 5, sum(1 for x in items if x["tiene_ars"]), bg=row_bg, center=True)
        rc_data_cell(ws_res, r_idx, 6, sum(1 for x in items if x["tiene_mep"]), bg=row_bg, center=True)
        rc_data_cell(ws_res, r_idx, 7, sum(1 for x in items if x["tiene_cbl"]), bg=row_bg, center=True)
        ws_res.row_dimensions[r_idx].height = 18
        r_idx += 1

    r_idx += 1
    TOTAL_ROW = r_idx
    rc_hdr_cell(ws_res, r_idx, 1, "TOTAL",                                      bg=RC_SUBHDR)
    rc_hdr_cell(ws_res, r_idx, 2, len(rc_rows),                                  bg=RC_SUBHDR)
    rc_hdr_cell(ws_res, r_idx, 3, "",                                             bg=RC_SUBHDR)
    rc_hdr_cell(ws_res, r_idx, 4, "",                                             bg=RC_SUBHDR)
    rc_hdr_cell(ws_res, r_idx, 5, sum(1 for x in rc_rows if x["tiene_ars"]),     bg=RC_SUBHDR)
    rc_hdr_cell(ws_res, r_idx, 6, sum(1 for x in rc_rows if x["tiene_mep"]),     bg=RC_SUBHDR)
    rc_hdr_cell(ws_res, r_idx, 7, sum(1 for x in rc_rows if x["tiene_cbl"]),     bg=RC_SUBHDR)

    # ── Buscador ──────────────────────────────────────────────────────────────
    br = TOTAL_ROW + 2
    ws_res.row_dimensions[TOTAL_ROW + 1].height = 10

    ws_res.merge_cells(f"A{br}:C{br}")
    c = ws_res.cell(row=br, column=1, value="🔍  Consulta rápida de especie")
    c.fill = rc_fill(RC_HEADER); c.border = rc_border_thin()
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[br].height = 26
    br += 1

    ws_res.merge_cells(f"A{br}:C{br}")
    c = ws_res.cell(row=br, column=1,
                    value="Ingresá el ticker (cualquier moneda) o el código CVSA:")
    c.font = Font(italic=True, color="374151", name="Calibri", size=9)
    c.fill = rc_fill("F0F6FB"); c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = rc_border_thin()
    ws_res.row_dimensions[br].height = 14
    br += 1

    input_row  = br
    INPUT_CELL = f"B{input_row}"

    c_lbl = ws_res.cell(row=br, column=1, value="TICKER / CVSA")
    c_lbl.fill = rc_fill(RC_SUBHDR)
    c_lbl.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c_lbl.alignment = Alignment(horizontal="center", vertical="center")
    c_lbl.border = rc_border_thin()

    ws_res.merge_cells(f"B{br}:C{br}")
    c_inp = ws_res.cell(row=br, column=2, value="")
    c_inp.fill = rc_fill(RC_INPUT)
    c_inp.font = Font(bold=True, name="Calibri", size=11, color="1F4E79")
    c_inp.alignment = Alignment(horizontal="center", vertical="center")
    sm2 = Side(style="medium", color="2E75B6")
    c_inp.border = Border(left=sm2, right=sm2, top=sm2, bottom=sm2)
    ws_res.row_dimensions[br].height = 24
    br += 1

    ws_res.row_dimensions[br].height = 4
    br += 1

    s_out = Side(style="thin", color="2E75B6")
    b_out = Border(left=s_out, right=s_out, top=s_out, bottom=s_out)

    for lbl, lk_col, pct_fmt in [("NOMBRE", "E", False), ("AFORO BYMA", "F", True), ("TIPO DE ACTIVO", "G", False)]:
        c_lbl = ws_res.cell(row=br, column=1, value=lbl)
        c_lbl.fill = rc_fill(RC_SUBHDR)
        c_lbl.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c_lbl.alignment = Alignment(horizontal="center", vertical="center")
        c_lbl.border = b_out

        ws_res.merge_cells(f"B{br}:C{br}")
        c_res = ws_res.cell(row=br, column=2, value=rc_lk_formula(INPUT_CELL, lk_col))
        c_res.fill = rc_fill(RC_RESULT)
        c_res.font = Font(bold=(lbl == "AFORO BYMA"), name="Calibri", size=10, color="1F4E79")
        c_res.alignment = Alignment(horizontal="center", vertical="center")
        c_res.border = b_out
        if pct_fmt:
            c_res.number_format = '0"%"'
        ws_res.row_dimensions[br].height = 22
        br += 1

    ws_res.merge_cells(f"A{br}:C{br}")
    c = ws_res.cell(row=br, column=1, value="Mayúsculas o minúsculas — el buscador no distingue")
    c.font = Font(italic=True, color="808080", name="Calibri", size=8)
    c.fill = rc_fill("F0F6FB"); c.alignment = Alignment(horizontal="center", vertical="center")
    st2 = Side(style="thin", color="BFBFBF")
    c.border = Border(left=st2, right=st2, top=st2, bottom=st2)
    ws_res.row_dimensions[br].height = 13

    ws_res.column_dimensions["A"].width = 30
    ws_res.column_dimensions["B"].width = 22
    ws_res.column_dimensions["C"].width = 16
    ws_res.column_dimensions["D"].width = 14
    ws_res.column_dimensions["E"].width = 18
    ws_res.column_dimensions["F"].width = 18
    ws_res.column_dimensions["G"].width = 20

    # ── Hojas por grupo ───────────────────────────────────────────────────────
    _gcl = get_column_letter

    for g in RC_GRUPOS_ORDER:
        items = rc_grupos[g]
        if not items:
            continue
        label = RC_GRUPO_LABEL[g]
        ws = rc_wb.create_sheet(label)
        ws.sheet_view.showGridLines = False
        ncols = 8

        ws.merge_cells(f"A1:{_gcl(ncols)}1")
        c = ws["A1"]
        c.value = label
        c.fill = rc_fill(RC_HEADER)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        ws.merge_cells(f"A2:{_gcl(ncols)}2")
        c = ws["A2"]
        c.value = (f"Fuente: BYMA Clearing — ESPECIES.XLS   |   "
                   f"Fecha: {HOY:%d/%m/%Y}   |   Total: {len(items)} especies")
        c.fill = rc_fill("EBF3E8")
        c.font = Font(italic=True, color="374151", name="Calibri", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 6

        hrow = 4
        ws.row_dimensions[hrow].height = 30
        rc_hdr_cell(ws, hrow, 1, "Código CVSA",  bg=RC_SUBHDR)
        rc_hdr_cell(ws, hrow, 2, "Ticker ARS",   bg=RC_SUBHDR)
        rc_hdr_cell(ws, hrow, 3, "Ticker MEP",   bg=RC_SUBHDR)
        rc_hdr_cell(ws, hrow, 4, "Ticker Cable", bg=RC_SUBHDR)
        rc_hdr_cell(ws, hrow, 5, "Nombre",       bg=RC_SUBHDR)
        rc_hdr_cell(ws, hrow, 6,
                    "Vencimiento" if g in RC_TIENE_VENCIM else "Emisor / Sector",
                    bg=RC_SUBHDR)
        rc_hdr_cell(ws, hrow, 7, "Haircut BYMA", bg=RC_SUBHDR)
        rc_hdr_cell(ws, hrow, 8, "Aforo BYMA",   bg=RC_SUBHDR)

        for i, item in enumerate(items):
            dr = hrow + 1 + i
            bg = rc_aforo_color(item["aforo"])
            rc_data_cell(ws, dr, 1, item["cvsa"],       bg=bg, center=True)
            rc_data_cell(ws, dr, 2, item["ticker_ars"], bg=bg, center=True)
            rc_data_cell(ws, dr, 3, item["ticker_mep"], bg=bg, center=True)
            rc_data_cell(ws, dr, 4, item["ticker_cbl"], bg=bg, center=True)
            rc_data_cell(ws, dr, 5, item["nombre"],     bg=bg)
            col6 = item["vencim"] if g in RC_TIENE_VENCIM else (item["emisor"] or item["tipo_activo"])
            rc_data_cell(ws, dr, 6, col6, bg=bg, center=(g in RC_TIENE_VENCIM))
            rc_data_cell(ws, dr, 7, f'{item["haircut"]:.0f}%', bg=bg, center=True)
            c_af = ws.cell(row=dr, column=8, value=f'{item["aforo"]}%')
            c_af.fill = rc_fill(bg)
            c_af.font = Font(bold=True, name="Calibri", size=9)
            c_af.alignment = Alignment(horizontal="center", vertical="center")
            c_af.border = rc_border_thin()
            ws.row_dimensions[dr].height = 15

        ws.column_dimensions["A"].width = 13
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 13
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 13
        ws.column_dimensions["H"].width = 13
        ws.freeze_panes = f"A{hrow + 1}"
        ws.auto_filter.ref = f"A{hrow}:{_gcl(ncols)}{hrow + len(items)}"

    buf_rc = io.BytesIO()
    rc_wb.save(buf_rc)
    xlsx_rc_bytes = buf_rc.getvalue()

    # ── Resumen para la UI ────────────────────────────────────────────────────
    resumen = {
        "total_byma":         total_byma,
        "total_ok":           total_ok,
        "diferencias":        len(diferencias),
        "sin_lista":          len(sin_lista),
        "lista_sin_byma":     len(lista_sin_byma),
        "vencidos_filtrados": vencidos_filtrados,
        "fecha":              HOY.strftime("%d-%m-%Y"),
    }

    return xlsx_dif_bytes, xlsx_rc_bytes, resumen, advertencias
