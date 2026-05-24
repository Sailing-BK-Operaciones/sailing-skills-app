"""
Conversion Dólares Renta
Lee Diario.xlsx y genera:
  - 3 archivos XLS para Gallo (7K MEP / 7K USD Cable / 10K)
  - 1 archivo ICT para transferencias masivas en NASDAQ BYMA
"""

import io
from datetime import datetime

import openpyxl
import xlwt

# ─── Constantes ────────────────────────────────────────────────────────────────
HEADER_GALLO = [
    "Nro Cte", "Nombre Cte", "CUIT Cte",
    "Cant. Tesoro", "Cant. Caja de Valores", "Fecha", "Mail",
]

HEADER_ICT = (
    "SourceCashAccount;ReceivingCashAccount;TransactionReference;"
    "PaymentSystem;Currency;Amount;SettlementDate;Description;"
    "CorporateActionReference;TransactionOnHoldCSD;TransactionOnHoldParticipant"
)


# ─── Helpers ───────────────────────────────────────────────────────────────────
def _valid_monto(v):
    """Devuelve float > 0 si v es válido, None si es vacío / cero."""
    if v is None:
        return None
    try:
        f = float(v)
        return f if f != 0.0 else None
    except (ValueError, TypeError):
        return None


def _gallo_xls_bytes(rows):
    """
    rows: list of (ctte_int, monto_float)
    Genera bytes de un .xls con 7 columnas (col E = monto).
    """
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Hoja1")
    for col_idx, header in enumerate(HEADER_GALLO):
        ws.write(0, col_idx, header)
    for row_idx, (ctte, monto) in enumerate(rows, start=1):
        ws.write(row_idx, 0, ctte)   # Nro Cte  (col A)
        ws.write(row_idx, 4, monto)  # Cant. Caja de Valores (col E)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_fecha(val):
    """Convierte el valor de la celda D1 a datetime."""
    if val is None:
        raise ValueError("Celda D1 del Diario está vacía — falta la fecha del día.")
    if hasattr(val, "year"):          # datetime / date
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"No se pudo interpretar la fecha '{val}' de la celda D1.")


# ─── Función principal ─────────────────────────────────────────────────────────
def generar_conversion(diario_file, counter_state=None):
    """
    Parámetros
    ----------
    diario_file   : file-like  — Diario.xlsx (subido por el usuario)
    counter_state : dict | None — {"AAMMDD": last_used}

    Retorna
    -------
    outputs      : dict  — gallo_7k, gallo_7k_cable, gallo_10k, ict
                   Cada valor: (bytes, filename) o None si no aplica
    new_counter  : dict  — estado actualizado
    resumen      : dict  — métricas para la UI
    advertencias : list[str]
    """
    advertencias  = []
    counter_state = dict(counter_state) if counter_state else {}

    # ── 1. Leer Diario.xlsx ───────────────────────────────────────────────────
    diario_file.seek(0)
    wb = openpyxl.load_workbook(diario_file, data_only=True)
    ws = wb.active

    # Fecha en celda D1 (row=1, column=4)
    fecha_dia     = _parse_fecha(ws.cell(row=1, column=4).value)
    fecha_yyyymmdd = fecha_dia.strftime("%Y%m%d")
    fecha_aammdd   = fecha_dia.strftime("%y%m%d")
    fecha_dd_mm_aa = fecha_dia.strftime("%d-%m-%y")

    rows_mep   = []   # (ctte_int, monto) — col B: DÓLAR USA a MEP
    rows_cable = []   # (ctte_int, monto) — col C: DÓLAR USA a USD CABLE
    rows_10k   = []   # (ctte_int, monto) — col D: DÓLAR DLR

    for row in ws.iter_rows(min_row=4, values_only=True):
        ctte_val = row[0] if len(row) > 0 else None
        if ctte_val is None:
            continue
        try:
            ctte_int = int(float(str(ctte_val)))
        except (ValueError, TypeError):
            continue

        m_mep   = _valid_monto(row[1] if len(row) > 1 else None)
        m_cable = _valid_monto(row[2] if len(row) > 2 else None)
        m_dlr   = _valid_monto(row[3] if len(row) > 3 else None)

        if m_mep   is not None: rows_mep.append((ctte_int,   m_mep))
        if m_cable is not None: rows_cable.append((ctte_int, m_cable))
        if m_dlr   is not None: rows_10k.append((ctte_int,   m_dlr))

    if not rows_mep and not rows_cable and not rows_10k:
        raise ValueError(
            "Diario.xlsx no tiene importes válidos (columnas B, C y D "
            "vacías o en cero desde fila 4)."
        )

    # ── 2. Archivos XLS para Gallo ────────────────────────────────────────────
    outputs = {}

    if rows_mep:
        fname = f"GALLO Conversion Especie a Moneda 7K {fecha_dd_mm_aa}.xls"
        outputs["gallo_7k"] = (_gallo_xls_bytes(rows_mep), fname)
    else:
        outputs["gallo_7k"] = None

    if rows_cable:
        fname = f"GALLO Conversion Especie a Moneda 7K USD CABLE {fecha_dd_mm_aa}.xls"
        outputs["gallo_7k_cable"] = (_gallo_xls_bytes(rows_cable), fname)
    else:
        outputs["gallo_7k_cable"] = None

    if rows_10k:
        fname = f"GALLO Conversion Especie a Moneda 10K {fecha_dd_mm_aa}.xls"
        outputs["gallo_10k"] = (_gallo_xls_bytes(rows_10k), fname)
    else:
        outputs["gallo_10k"] = None

    # ── 3. Archivo ICT ────────────────────────────────────────────────────────
    start_counter = counter_state.get(fecha_aammdd, 0) + 1
    counter       = start_counter
    ict_lines     = [HEADER_ICT]

    # Grupo 1: DÓLAR USA a MEP — USD-EXTERNO, cuenta destino 233/1
    for ctte, monto in rows_mep:
        ref = f"Conv{fecha_aammdd}{counter:04d}"
        ict_lines.append(
            f"233/{ctte};233/1;{ref};USD-EXTERNO;USD;"
            f"{monto:.2f};{fecha_yyyymmdd};Description;;0;0"
        )
        counter += 1

    # Grupo 2: DÓLAR USA a USD CABLE — USD-EXTERNO, cuenta destino 70233/10000
    for ctte, monto in rows_cable:
        ref = f"Conv{fecha_aammdd}{counter:04d}"
        ict_lines.append(
            f"233/{ctte};70233/10000;{ref};USD-EXTERNO;USD;"
            f"{monto:.2f};{fecha_yyyymmdd};Description;;0;0"
        )
        counter += 1

    # Grupo 3: DÓLAR DLR — USD-LOCAL, cuenta destino 233/1
    for ctte, monto in rows_10k:
        ref = f"Conv{fecha_aammdd}{counter:04d}"
        ict_lines.append(
            f"233/{ctte};233/1;{ref};USD-LOCAL;USD;"
            f"{monto:.2f};{fecha_yyyymmdd};Description;;0;0"
        )
        counter += 1

    last_counter = counter - 1
    new_counter  = dict(counter_state)
    new_counter[fecha_aammdd] = last_counter

    ict_bytes = "\r\n".join(ict_lines).encode("utf-8")
    fname_ict = f"Archivo masivo - transferencias de efectivo {fecha_dd_mm_aa}.ICT"
    outputs["ict"] = (ict_bytes, fname_ict)

    # ── 4. Verificación ───────────────────────────────────────────────────────
    n_diario = len(rows_mep) + len(rows_cable) + len(rows_10k)
    n_ict    = len(ict_lines) - 1   # sin el encabezado
    if n_diario != n_ict:
        advertencias.append(
            f"Verificación fallida: el Diario tiene {n_diario} importes "
            f"pero el ICT generó {n_ict} líneas."
        )

    # ── 5. Resumen para la UI ─────────────────────────────────────────────────
    ref_primera = f"Conv{fecha_aammdd}{start_counter:04d}"
    ref_ultima  = f"Conv{fecha_aammdd}{last_counter:04d}"

    resumen = {
        "fecha":               fecha_dd_mm_aa,
        "n_7k":                len(rows_mep),
        "n_cable":             len(rows_cable),
        "n_10k":               len(rows_10k),
        "n_ict":               n_ict,
        "verificacion_ok":     n_diario == n_ict,
        "ref_primera":         ref_primera,
        "ref_ultima":          ref_ultima,
        "contador_continuado": start_counter > 1,
        "inicio_contador":     start_counter,
    }

    return outputs, new_counter, resumen, advertencias
