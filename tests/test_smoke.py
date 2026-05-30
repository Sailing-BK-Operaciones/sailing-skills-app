"""
Smoke tests — verifican que todos los módulos de skills se importan sin error
y que los patrones de openpyxl más comunes no crashean.

Estos tests NO requieren archivos reales ni datos de mercado.
Detectan errores como: wb.get(), imports rotos, typos en nombres de función.
"""

import importlib
import sys
import pytest
from io import BytesIO
from openpyxl import Workbook


# ── 1. Import checks ──────────────────────────────────────────────────────────
# Cada módulo logic.py debe importarse sin AttributeError ni SyntaxError.

SKILLS = [
    "skills.tesoreria.logic",
    "skills.distribucion_gara.logic",
    "skills.risk_monitoring_client.logic",
    "skills.control_aforos_byma.logic",
    "skills.Collateral_position.logic",
    "skills.actual_position.logic",
    "skills.arreglos_garantias.logic",
    "skills.control_diario_op_sdib.logic",
    "skills.control_margenes_gara_byma.logic",
    "skills.conversion_dolares_renta.logic",
    "skills.genera_txt_gara_nasdaq.logic",
    "skills.reporte_operativo.logic",
    "skills.risk_position.logic",
]


@pytest.mark.parametrize("module_path", SKILLS)
def test_import(module_path):
    """Cada logic.py debe importarse sin errores."""
    mod = importlib.import_module(module_path)
    assert mod is not None


# ── 2. openpyxl Workbook API ──────────────────────────────────────────────────
# Verifica que los patrones de acceso a hojas sean correctos.
# wb.get() no existe → debe usarse wb['nombre'] o wb.sheetnames.

def test_workbook_sheet_access():
    wb = Workbook()
    ws = wb.active
    ws.title = "Panel de Control"
    wb.create_sheet("Detalle ARS")
    wb.create_sheet("Detalle USD")

    # Patrón correcto: acceso directo
    assert wb["Panel de Control"].title == "Panel de Control"

    # Patrón correcto: guard con sheetnames
    ws_det = wb["Detalle ARS"] if "Detalle ARS" in wb.sheetnames else None
    assert ws_det is not None

    # Patrón incorrecto que causó el bug de hoy: wb.get() no existe
    assert not hasattr(wb, "get"), (
        "Workbook tiene .get() — este test debe fallar si openpyxl lo agrega "
        "para avisar que el guard ya no es necesario."
    )


def test_workbook_missing_sheet_returns_none():
    """El guard 'if X in wb.sheetnames else None' debe devolver None para hojas inexistentes."""
    wb = Workbook()
    wb.active.title = "Hoja1"

    result = wb["Hoja Inexistente"] if "Hoja Inexistente" in wb.sheetnames else None
    assert result is None
